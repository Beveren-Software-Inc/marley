"""Import Oracle PATIENT_ADJUSTMENT_01 Excel into Patient Adjustment rows."""

from __future__ import annotations

from typing import Any

import frappe
from frappe import _
from frappe.utils import flt, getdate

from healthcare.api.discharge_checklist_import import _resolve_cost_center
from healthcare.api.patient_info_import import (
	_cell_text,
	_clean_oracle_num,
	_excel_file_path,
	_parse_date_value,
	_parse_datetime_value,
	_require_admin,
)
from healthcare.api.visit_diagnosis_sync import _resolve_inpatient_admission

DOCTYPE = "Patient Adjustment"

EXCEL_HEADER_MAP = {
	"TRANS_NUM": "trans_num",
	"TRANS_DATE": "trans_date",
	"TRANS_TYPE": "trans_type",
	"FROM_PATIENT_NUM": "from_patient_num",
	"FROM_ADMISSION_NUM": "from_admission_num",
	"FROM_DUE_AMT": "from_due_amt",
	"FROM_BAL_AMT": "from_bal_amt",
	"TO_ADMISSION_NUM": "to_admission_num",
	"TO_PATIENT_NUM": "to_patient_num",
	"TO_DUE_AMT": "to_due_amt",
	"TO_BAL_AMT": "to_bal_amt",
	"ADJUSTED_AMT": "adjusted_amt",
	"REMARKS_HEADER": "remarks_header",
	"VCH_STATUS": "vch_status",
	"BRANCH_NUM": "branch_num",
	"CR_ID": "cr_id",
	"CR_DATE": "cr_date",
	"UP_ID": "up_id",
	"UP_DATE": "up_date",
	"AP_ID": "ap_id",
	"AP_DATE": "ap_date",
	"ADJUSTED_REFF_NUM": "adjusted_reff_num",
	"ADJUSTED_REFF_AMT": "adjusted_reff_amt",
	"ADJUSTED_REFF_AMT_ALREADY_ADJU": "adjusted_reff_amt_already_adjusted",
	"ADJUSTED_REFF_AMT_BAL": "adjusted_reff_amt_bal",
	"DOWN_INV_BRANCH": "down_inv_branch",
}


def _normalize_header(cell: Any) -> str:
	if cell is None:
		return ""
	text = str(cell).strip().upper().replace(" ", "_")
	return EXCEL_HEADER_MAP.get(text, text.lower())


def _trans_no(value: Any) -> str:
	text = _cell_text(value).strip()
	return text


def _format_legacy_datetime(value: Any) -> str | None:
	dt = _parse_datetime_value(value)
	if dt:
		return dt.strftime("%Y-%m-%d %H:%M:%S")
	parsed_date = _parse_date_value(value)
	if parsed_date:
		return parsed_date.strftime("%Y-%m-%d")
	text = _cell_text(value)
	return text or None


def _parse_date_field(value: Any):
	dt = _parse_datetime_value(value)
	if dt:
		return getdate(dt)
	parsed_date = _parse_date_value(value)
	if parsed_date:
		return parsed_date
	return None


def _amount(value: Any) -> float | None:
	if value in (None, ""):
		return None
	return flt(value)


def _resolve_patient(patient_num: Any) -> str | None:
	patient = _clean_oracle_num(patient_num)
	if not patient:
		return None
	if frappe.db.exists("Patient", patient):
		return patient
	return None


def _resolve_admission(admission_num: Any) -> str | None:
	admission = _clean_oracle_num(admission_num)
	if not admission:
		return None
	if frappe.db.exists("Inpatient Admission", admission):
		return admission
	return _resolve_inpatient_admission(admission)


def _parse_sheet_rows(ws) -> list[dict]:
	rows_iter = ws.iter_rows(values_only=True)
	try:
		header_row = next(rows_iter)
	except StopIteration:
		return []

	headers = [_normalize_header(h) for h in header_row]
	parsed: list[dict] = []
	for raw in rows_iter:
		if not raw or all(cell is None or str(cell).strip() == "" for cell in raw):
			continue
		row: dict[str, Any] = {}
		for idx, key in enumerate(headers):
			if not key or idx >= len(raw):
				continue
			row[key] = raw[idx]

		trans_no = _trans_no(row.get("trans_num"))
		if not trans_no:
			continue
		row["trans_no"] = trans_no
		parsed.append(row)
	return parsed


def _parse_excel_rows(file_url: str) -> tuple[list[dict], dict[str, int]]:
	try:
		import openpyxl
	except ImportError:
		frappe.throw(
			_(
				"openpyxl is required to read Excel files. Install it in the bench environment: pip install openpyxl"
			)
		)

	path = _excel_file_path(file_url)
	wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
	by_key: dict[str, dict] = {}
	sheet_row_counts: dict[str, int] = {}
	try:
		for sheet_name in wb.sheetnames:
			sheet_rows = _parse_sheet_rows(wb[sheet_name])
			sheet_row_counts[sheet_name] = len(sheet_rows)
			for row in sheet_rows:
				by_key[row["trans_no"]] = row
	finally:
		wb.close()
	return list(by_key.values()), sheet_row_counts


def _build_fields(row: dict) -> dict[str, Any]:
	fields: dict[str, Any] = {
		"trans_no": row["trans_no"],
		"trans_type": _cell_text(row.get("trans_type")) or None,
		"remarks_header": _cell_text(row.get("remarks_header")) or None,
		"vch_status": _cell_text(row.get("vch_status")) or None,
		"adjusted_reff_num": _cell_text(row.get("adjusted_reff_num")) or None,
		"cr_id": _clean_oracle_num(row.get("cr_id")) or None,
		"up_id": _clean_oracle_num(row.get("up_id")) or None,
		"ap_id": _clean_oracle_num(row.get("ap_id")) or None,
	}

	trans_date = _parse_date_field(row.get("trans_date"))
	if trans_date:
		fields["trans_date"] = trans_date

	for amount_field in (
		"from_due_amt",
		"from_bal_amt",
		"to_due_amt",
		"to_bal_amt",
		"adjusted_amt",
		"adjusted_reff_amt",
		"adjusted_reff_amt_already_adjusted",
		"adjusted_reff_amt_bal",
	):
		amount = _amount(row.get(amount_field))
		if amount is not None:
			fields[amount_field] = amount

	for date_field in ("cr_date", "up_date", "ap_date"):
		legacy_date = _format_legacy_datetime(row.get(date_field))
		if legacy_date:
			fields[date_field] = legacy_date

	from_patient = _resolve_patient(row.get("from_patient_num"))
	if from_patient:
		fields["from_patient"] = from_patient
	to_patient = _resolve_patient(row.get("to_patient_num"))
	if to_patient:
		fields["to_patient"] = to_patient

	from_admission = _resolve_admission(row.get("from_admission_num"))
	if from_admission:
		fields["from_admission"] = from_admission
	to_admission = _resolve_admission(row.get("to_admission_num"))
	if to_admission:
		fields["to_admission"] = to_admission

	branch = _resolve_cost_center(row.get("branch_num"))
	if branch:
		fields["branch"] = branch
	down_inv_branch = _resolve_cost_center(row.get("down_inv_branch"))
	if down_inv_branch:
		fields["down_inv_branch"] = down_inv_branch

	return {key: value for key, value in fields.items() if value not in (None, "")}


def _preview_stats(rows: list[dict]) -> dict:
	existing = 0
	resolvable_from_patients = 0
	resolvable_to_patients = 0
	resolvable_from_admissions = 0
	resolvable_to_admissions = 0
	for row in rows:
		if frappe.db.exists(DOCTYPE, row["trans_no"]):
			existing += 1
		if _resolve_patient(row.get("from_patient_num")):
			resolvable_from_patients += 1
		if _resolve_patient(row.get("to_patient_num")):
			resolvable_to_patients += 1
		if _resolve_admission(row.get("from_admission_num")):
			resolvable_from_admissions += 1
		if _resolve_admission(row.get("to_admission_num")):
			resolvable_to_admissions += 1

	return {
		"existing_records": existing,
		"resolvable_from_patients": resolvable_from_patients,
		"resolvable_to_patients": resolvable_to_patients,
		"resolvable_from_admissions": resolvable_from_admissions,
		"resolvable_to_admissions": resolvable_to_admissions,
		"sample_trans_nos": [row.get("trans_no") for row in rows[:5]],
	}


def upsert_patient_adjustment(row: dict) -> dict:
	fields = _build_fields(row)
	if not fields.get("trans_no"):
		return {"status": "skip", "trans_no": row.get("trans_no")}

	trans_no = fields["trans_no"]
	if frappe.db.exists(DOCTYPE, trans_no):
		doc = frappe.get_doc(DOCTYPE, trans_no)
		for key, value in fields.items():
			doc.set(key, value)
		doc.flags.ignore_mandatory = True
		doc.flags.ignore_links = True
		doc.flags.ignore_validate = True
		doc.save(ignore_permissions=True)
		action = "updated"
	else:
		doc = frappe.get_doc({"doctype": DOCTYPE, **fields})
		doc.flags.ignore_mandatory = True
		doc.flags.ignore_links = True
		doc.flags.ignore_validate = True
		doc.insert(ignore_permissions=True)
		action = "created"

	return {"status": action, "trans_no": trans_no, "name": doc.name}


@frappe.whitelist()
def preview_patient_adjustment_import(file_url: str) -> dict:
	_require_admin()
	if not (file_url or "").strip():
		frappe.throw(_("Please upload the PATIENT_ADJUSTMENT_01 Excel file."))

	rows, sheet_row_counts = _parse_excel_rows(file_url)
	stats = _preview_stats(rows)
	raw_row_total = sum(sheet_row_counts.values())
	return {
		"excel_rows": len(rows),
		"raw_excel_rows": raw_row_total,
		"sheets": list(sheet_row_counts.keys()),
		"sheet_row_counts": sheet_row_counts,
		**stats,
	}


@frappe.whitelist()
def run_patient_adjustment_import(file_url: str) -> dict:
	"""Import all rows synchronously."""
	_require_admin()
	if not (file_url or "").strip():
		frappe.throw(_("Please upload the PATIENT_ADJUSTMENT_01 Excel file."))

	rows, _ = _parse_excel_rows(file_url)
	created = updated = skipped = 0
	errors: list[str] = []

	for row in rows:
		try:
			result = upsert_patient_adjustment(row)
			status = result.get("status")
			if status == "created":
				created += 1
			elif status == "updated":
				updated += 1
			else:
				skipped += 1
		except Exception:
			errors.append(f"{row.get('trans_no')}: {frappe.get_traceback()}")
			frappe.log_error(title=f"PATIENT_ADJUSTMENT_01 import failed: {row.get('trans_no')}")

	frappe.db.commit()
	return {
		"ok": True,
		"total": len(rows),
		"created": created,
		"updated": updated,
		"skipped": skipped,
		"errors": len(errors),
	}
