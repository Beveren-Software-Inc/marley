"""Import Oracle VISIT_POSITIVE_FINDING_01 Excel into Visit Positive Finding rows."""

from __future__ import annotations

import json
from typing import Any

import frappe
from frappe import _

from healthcare.api.discharge_checklist_import import _resolve_cost_center
from healthcare.api.patient_info_import import (
	_cell_text,
	_clean_oracle_num,
	_excel_file_path,
	_require_admin,
)
from healthcare.api.visit_diagnoses_op_import import _legacy_data_datetime

VISIT_POSITIVE_FINDING_IMPORT_BATCH_SIZE = 500
CACHE_TTL = 7200
CACHE_KEYS = {
	"file_url": "healthcare:data_migration:visit_positive_finding_import:file_url",
	"row_keys": "healthcare:data_migration:visit_positive_finding_import:row_keys",
	"rows": "healthcare:data_migration:visit_positive_finding_import:rows",
}

EXCEL_HEADER_MAP = {
	"VISIT_NUM": "visit_num",
	"SR_NUM": "sr_num",
	"CODE_NUM": "code_num",
	"CODE_MORE_DETAIL": "code_more_detail",
	"CR_ID": "cr_id",
	"CR_DATE": "cr_date",
	"UP_ID": "up_id",
	"UP_DATE": "up_date",
	"BRANCH_NUM": "branch_num",
}


def _normalize_header(cell: Any) -> str:
	if cell is None:
		return ""
	text = str(cell).strip().upper().replace(" ", "_")
	return EXCEL_HEADER_MAP.get(text, text.lower())


def _oracle_data_value(value: Any) -> str | None:
	if value in (None, ""):
		return None
	cleaned = _clean_oracle_num(value)
	if cleaned:
		return cleaned
	text = _cell_text(value)
	return text or None


def _text_value(value: Any) -> str | None:
	text = _cell_text(value)
	return text or None


def _row_key(row: dict) -> str:
	return "|".join(
		[
			row.get("visit_num") or "",
			row.get("sr_num") or "",
			row.get("code_num") or "",
		]
	)


def _existing_visit_positive_finding_name(row: dict) -> str | None:
	visit_num = row.get("visit_num")
	sr_num = row.get("sr_num")
	code_num = row.get("code_num")
	if not visit_num or not code_num:
		return None
	filters = {"visit_num": visit_num, "code_num": code_num}
	if sr_num not in (None, ""):
		filters["sr_num"] = sr_num
	return frappe.db.get_value("Visit Positive Finding", filters, "name")


def _parse_sheet_rows(ws) -> list[dict]:
	rows_iter = ws.iter_rows(values_only=True)
	try:
		header_row = next(rows_iter)
	except StopIteration:
		return []

	headers = [_normalize_header(h) for h in header_row]
	parsed: list[dict] = []
	for raw in rows_iter:
		if not raw or all(cell is None or str(cell).strip() == "" for cell in raw):
			continue
		row: dict[str, Any] = {}
		for idx, key in enumerate(headers):
			if not key or idx >= len(raw):
				continue
			row[key] = raw[idx]

		visit_num = _oracle_data_value(row.get("visit_num"))
		code_num = _oracle_data_value(row.get("code_num"))
		if not visit_num or not code_num:
			continue

		row["visit_num"] = visit_num
		row["sr_num"] = _oracle_data_value(row.get("sr_num"))
		row["code_num"] = code_num
		row["cr_id"] = _oracle_data_value(row.get("cr_id"))
		row["up_id"] = _oracle_data_value(row.get("up_id"))
		parsed.append(row)
	return parsed


def _parse_excel_rows(file_url: str) -> tuple[list[dict], dict[str, int]]:
	try:
		import openpyxl
	except ImportError:
		frappe.throw(
			_(
				"openpyxl is required to read Excel files. Install it in the bench environment: pip install openpyxl"
			)
		)

	path = _excel_file_path(file_url)
	wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
	all_rows: list[dict] = []
	sheet_row_counts: dict[str, int] = {}
	try:
		for sheet_name in wb.sheetnames:
			sheet_rows = _parse_sheet_rows(wb[sheet_name])
			sheet_row_counts[sheet_name] = len(sheet_rows)
			all_rows.extend(sheet_rows)
	finally:
		wb.close()
	return all_rows, sheet_row_counts


def _build_visit_positive_finding_fields(row: dict) -> dict[str, Any]:
	fields: dict[str, Any] = {
		"visit_num": row["visit_num"],
		"sr_num": row.get("sr_num"),
		"code_num": row["code_num"],
		"code_more_detail": _text_value(row.get("code_more_detail")),
		"cr_id": row.get("cr_id"),
		"up_id": row.get("up_id"),
	}

	cr_date = _legacy_data_datetime(row.get("cr_date"))
	if cr_date:
		fields["cr_date"] = cr_date
	up_date = _legacy_data_datetime(row.get("up_date"))
	if up_date:
		fields["up_date"] = up_date

	cc = _resolve_cost_center(row.get("branch_num"))
	if cc:
		fields["cost_centre"] = cc

	return {key: value for key, value in fields.items() if value not in (None, "")}


def _apply_legacy_import_flags(doc) -> None:
	doc.flags.ignore_validate = True
	doc.flags.ignore_mandatory = True
	doc.flags.ignore_links = True
	doc.flags.ignore_permissions = True
	doc.flags.legacy_import = True


def _persist_visit_positive_finding(doc, *, existing: bool) -> None:
	_apply_legacy_import_flags(doc)
	if existing:
		doc.save(ignore_permissions=True)
	else:
		doc.insert(ignore_permissions=True)


def upsert_visit_positive_finding_from_row(row: dict) -> dict:
	key = _row_key(row)
	if not row.get("visit_num") or not row.get("code_num"):
		return {"status": "skip_no_key", "row_key": key}

	fields = _build_visit_positive_finding_fields(row)
	if not fields:
		return {"status": "skip_error", "row_key": key}

	existing_name = _existing_visit_positive_finding_name(row)
	if existing_name:
		doc = frappe.get_doc("Visit Positive Finding", existing_name)
		for field, value in fields.items():
			doc.set(field, value)
		action = "updated"
		existing = True
	else:
		doc = frappe.get_doc({"doctype": "Visit Positive Finding", **fields})
		action = "created"
		existing = False

	_persist_visit_positive_finding(doc, existing=existing)
	return {
		"status": action,
		"row_key": key,
		"name": doc.name,
	}


def _preview_counts(rows: list[dict]) -> dict:
	existing = sum(1 for row in rows if _existing_visit_positive_finding_name(row))
	skip_no_cost_center = 0
	for row in rows:
		if row.get("branch_num") not in (None, "") and not _resolve_cost_center(row.get("branch_num")):
			skip_no_cost_center += 1

	return {
		"existing_findings": existing,
		"new_findings": len(rows) - existing,
		"skip_no_cost_center": skip_no_cost_center,
		"sample_visit_nums": sorted({row["visit_num"] for row in rows})[:5],
	}


def parse_and_cache_excel(file_url: str) -> dict:
	rows, sheet_row_counts = _parse_excel_rows(file_url)
	by_key = {_row_key(row): row for row in rows}
	row_keys = list(by_key.keys())

	frappe.cache().set_value(CACHE_KEYS["file_url"], file_url, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(CACHE_KEYS["row_keys"], row_keys, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(
		CACHE_KEYS["rows"],
		json.dumps(by_key, default=str),
		expires_in_sec=CACHE_TTL,
	)

	preview = _preview_counts(list(by_key.values()))
	raw_row_total = sum(sheet_row_counts.values())
	return {
		"excel_rows": len(by_key),
		"raw_excel_rows": raw_row_total,
		"sheets": list(sheet_row_counts.keys()),
		"sheet_row_counts": sheet_row_counts,
		**preview,
	}


def _load_cached_rows() -> dict[str, dict]:
	raw = frappe.cache().get_value(CACHE_KEYS["rows"])
	if not raw:
		return {}
	if isinstance(raw, dict):
		return raw
	return json.loads(raw)


@frappe.whitelist()
def preview_visit_positive_finding_import(file_url: str) -> dict:
	_require_admin()
	if not (file_url or "").strip():
		frappe.throw(_("Please upload the VISIT_POSITIVE_FINDING_01 Excel file."))
	return parse_and_cache_excel(file_url)


def run_visit_positive_finding_import_batch(*, offset: int = 0) -> dict:
	row_keys = frappe.cache().get_value(CACHE_KEYS["row_keys"]) or []
	rows_by_key = _load_cached_rows()
	if not row_keys or not rows_by_key:
		return {"processed": offset, "done": True, "batch_count": 0}

	batch_keys = row_keys[offset : offset + VISIT_POSITIVE_FINDING_IMPORT_BATCH_SIZE]
	if not batch_keys:
		return {"processed": offset, "done": True, "batch_count": 0}

	created = updated = 0
	errors: list[str] = []

	for key in batch_keys:
		row = rows_by_key.get(key) or {}
		try:
			result = upsert_visit_positive_finding_from_row(row)
			status = result.get("status")
			if status == "created":
				created += 1
			elif status == "updated":
				updated += 1
		except Exception:
			errors.append(f"{key}: {frappe.get_traceback()}")
			frappe.log_error(title=f"VISIT_POSITIVE_FINDING_01 import failed: {key}")

	frappe.db.commit()
	processed = offset + len(batch_keys)
	return {
		"processed": processed,
		"done": len(batch_keys) < VISIT_POSITIVE_FINDING_IMPORT_BATCH_SIZE,
		"batch_count": len(batch_keys),
		"created": created,
		"updated": updated,
		"errors": len(errors),
	}
