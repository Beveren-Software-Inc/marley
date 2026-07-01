"""Import Oracle IP_ADMISSION_02 Excel into Patient History (one per admission).

Reuses ``patient_history_import.process_admission_import`` so child ``history_detail``
lines are seeded from Default History Form and notes are applied by ``attrib_num``.
"""

from __future__ import annotations

import json
from datetime import date, datetime
from typing import Any

import frappe
from frappe import _
from frappe.utils import cint

from healthcare.api.discharge_checklist_import import _resolve_cost_center
from healthcare.api.patient_history_date_backfill import _parse_legacy_datetime, _skip_care_episode_guard
from healthcare.api.patient_history_import import (
	PATIENT_HISTORY_IMPORT_BATCH_SIZE,
	_default_template_name,
	_template_detail_rows,
	process_admission_import,
)
from healthcare.api.patient_info_import import (
	_cell_text,
	_clean_oracle_num,
	_excel_file_path,
	_require_admin,
)
from healthcare.api.visit_diagnosis_sync import _resolve_inpatient_admission

CACHE_TTL = 7200
CACHE_KEYS = {
	"file_url": "healthcare:data_migration:ip_admission_02_import:file_url",
	"admissions": "healthcare:data_migration:ip_admission_02_import:admissions",
	"grouped": "healthcare:data_migration:ip_admission_02_import:grouped",
}

EXCEL_HEADER_MAP = {
	"IP_ADMISSION_FORM_ID": "ip_admission_form_id",
	"ADMISSION_NUM": "admission_num",
	"PATIENT_NUM": "patient_num",
	"ATTRIB_NUM": "attrib_num",
	"ATT_NOTES": "att_notes",
	"FIELD1": "field1",
	"FIELD2": "field2",
	"FIELD3": "field3",
	"FIELD4": "field4",
	"FIELD5": "field5",
	"FIELD6": "field6",
	"BRANCH_NUM": "branch_num",
	"CR_ID": "cr_id",
	"CR_DATE": "cr_date",
	"UP_ID": "up_id",
	"UP_DATE": "up_date",
	"ATT_NOTES2": "att_notes2",
	"ORDER_OF_ATTRIB": "order_of_attrib",
	"ADMISSION_NUM_OLD": "admission_num_old",
}


def _normalize_header(cell: Any) -> str:
	if cell is None:
		return ""
	text = str(cell).strip().upper().replace(" ", "_")
	return EXCEL_HEADER_MAP.get(text, text.lower())


def _legacy_date_text(value: Any) -> str | None:
	if value in (None, ""):
		return None
	if isinstance(value, datetime):
		return value.strftime("%Y-%m-%d %H:%M:%S")
	if isinstance(value, date):
		return datetime.combine(value, datetime.min.time()).strftime("%Y-%m-%d %H:%M:%S")
	text = _cell_text(value)
	return text or None


def _template_attrib_names(template_name: str) -> dict[int, str]:
	return {
		cint(row.get("attrib_num")): row.get("attribute") or ""
		for row in _template_detail_rows(template_name)
		if cint(row.get("attrib_num"))
	}


def _build_import_row(row: dict, *, attrib_names: dict[int, str]) -> dict:
	attrib_num = cint(row.get("attrib_num"))
	admission_num = row.get("admission_num") or ""
	patient_num = row.get("patient_num") or ""
	cost_center = _resolve_cost_center(row.get("branch_num"))

	import_row = {
		"admission": admission_num,
		"patient": patient_num,
		"attrib_num": attrib_num,
		"description": _cell_text(row.get("att_notes")),
		"field_1": _cell_text(row.get("field1")),
		"attribute_note_2": _cell_text(row.get("att_notes2")),
		"attribute": attrib_names.get(attrib_num) or "",
		"old_admission_no": row.get("admission_num_old") or "",
		"ip_admission_form_id": row.get("ip_admission_form_id") or "",
		"cr_id": row.get("cr_id") or "",
		"cr_date": _legacy_date_text(row.get("cr_date")),
		"up_id": row.get("up_id") or "",
	}
	if cost_center:
		import_row["cost_center"] = cost_center
	return import_row


def _parse_sheet_rows(ws, *, attrib_names: dict[int, str]) -> list[dict]:
	rows_iter = ws.iter_rows(values_only=True)
	try:
		header_row = next(rows_iter)
	except StopIteration:
		return []

	headers = [_normalize_header(h) for h in header_row]
	parsed: list[dict] = []
	for raw in rows_iter:
		if not raw or all(cell is None or str(cell).strip() == "" for cell in raw):
			continue
		row: dict[str, Any] = {}
		for idx, key in enumerate(headers):
			if not key or idx >= len(raw):
				continue
			row[key] = raw[idx]

		admission_num = _clean_oracle_num(row.get("admission_num"))
		attrib_num = _clean_oracle_num(row.get("attrib_num"))
		if not admission_num or not attrib_num:
			continue

		row["admission_num"] = admission_num
		row["admission_num_old"] = _clean_oracle_num(row.get("admission_num_old"))
		row["patient_num"] = _clean_oracle_num(row.get("patient_num"))
		row["attrib_num"] = attrib_num
		row["ip_admission_form_id"] = _clean_oracle_num(row.get("ip_admission_form_id"))
		row["cr_id"] = _clean_oracle_num(row.get("cr_id"))
		row["up_id"] = _clean_oracle_num(row.get("up_id"))
		row["branch_num"] = _clean_oracle_num(row.get("branch_num"))
		parsed.append(_build_import_row(row, attrib_names=attrib_names))
	return parsed


def _parse_excel_rows(file_url: str) -> tuple[dict[str, list[dict]], dict[str, int], list[dict]]:
	try:
		import openpyxl
	except ImportError:
		frappe.throw(
			_(
				"openpyxl is required to read Excel files. Install it in the bench environment: pip install openpyxl"
			)
		)

	template_name = _default_template_name()
	attrib_names = _template_attrib_names(template_name)
	path = _excel_file_path(file_url)
	wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
	by_admission: dict[str, list[dict]] = {}
	all_rows: list[dict] = []
	sheet_row_counts: dict[str, int] = {}
	try:
		for sheet_name in wb.sheetnames:
			sheet_rows = _parse_sheet_rows(wb[sheet_name], attrib_names=attrib_names)
			sheet_row_counts[sheet_name] = len(sheet_rows)
			all_rows.extend(sheet_rows)
			for row in sheet_rows:
				admission = row.get("admission") or ""
				by_admission.setdefault(admission, []).append(row)
	finally:
		wb.close()
	return by_admission, sheet_row_counts, all_rows


def _resolve_admission_for_group(admission_num: str, import_rows: list[dict]) -> str | None:
	patient = (import_rows[0].get("patient") if import_rows else "") or None
	candidates: list[str] = []
	for value in (
		admission_num,
		(import_rows[0].get("old_admission_no") if import_rows else "") or "",
	):
		text = (value or "").strip()
		if text and text not in candidates:
			candidates.append(text)

	for candidate in candidates:
		resolved = _resolve_inpatient_admission(candidate, patient)
		if resolved:
			return resolved
	return None


def _apply_date_if_missing(ph_name: str, import_rows: list[dict]) -> bool:
	if frappe.db.get_value("Patient History", ph_name, "date"):
		return False

	earliest: str | None = None
	for row in import_rows:
		dt = _parse_legacy_datetime(row.get("cr_date"))
		if dt and (earliest is None or dt < earliest):
			earliest = dt
	if not earliest:
		return False

	ph = frappe.get_doc("Patient History", ph_name)
	_skip_care_episode_guard(ph)
	ph.date = earliest
	ph.save(ignore_permissions=True)
	return True


def _preview_counts(by_admission: dict[str, list[dict]], all_rows: list[dict]) -> dict:
	template_name = _default_template_name()
	template_nums = {cint(row.get("attrib_num")) for row in _template_detail_rows(template_name)}
	resolvable = 0
	unresolved = 0
	existing_histories = 0
	matching_attrib = 0

	for admission_num, rows in by_admission.items():
		resolved = _resolve_admission_for_group(admission_num, rows)
		if resolved:
			resolvable += 1
			if frappe.db.exists("Patient History", {"inpatient_admission": resolved}):
				existing_histories += 1
		else:
			unresolved += 1

	for row in all_rows:
		if cint(row.get("attrib_num")) in template_nums:
			matching_attrib += 1

	return {
		"admissions": len(by_admission),
		"resolvable_admissions": resolvable,
		"unresolved_admissions": unresolved,
		"existing_histories": existing_histories,
		"new_histories": resolvable - existing_histories,
		"matching_attrib_rows": matching_attrib,
		"unknown_attrib_rows": len(all_rows) - matching_attrib,
		"template": template_name,
		"sample_admissions": sorted(by_admission.keys())[:5],
	}


def parse_and_cache_excel(file_url: str) -> dict:
	by_admission, sheet_row_counts, all_rows = _parse_excel_rows(file_url)
	admission_keys = sorted(by_admission.keys())

	frappe.cache().set_value(CACHE_KEYS["file_url"], file_url, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(CACHE_KEYS["admissions"], admission_keys, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(
		CACHE_KEYS["grouped"],
		json.dumps(by_admission, default=str),
		expires_in_sec=CACHE_TTL,
	)

	preview = _preview_counts(by_admission, all_rows)
	raw_row_total = sum(sheet_row_counts.values())
	return {
		"excel_rows": len(all_rows),
		"raw_excel_rows": raw_row_total,
		"sheets": list(sheet_row_counts.keys()),
		"sheet_row_counts": sheet_row_counts,
		**preview,
	}


def _load_cached_groups() -> dict[str, list[dict]]:
	raw = frappe.cache().get_value(CACHE_KEYS["grouped"])
	if not raw:
		return {}
	if isinstance(raw, dict):
		return raw
	return json.loads(raw)


@frappe.whitelist()
def preview_ip_admission_02_import(file_url: str) -> dict:
	_require_admin()
	if not (file_url or "").strip():
		frappe.throw(_("Please upload the IP_ADMISSION_02 Excel file."))
	return parse_and_cache_excel(file_url)


def run_ip_admission_02_import_batch(*, offset: int = 0) -> dict:
	admission_keys = frappe.cache().get_value(CACHE_KEYS["admissions"]) or []
	by_admission = _load_cached_groups()
	if not admission_keys or not by_admission:
		return {"processed": offset, "done": True, "batch_count": 0}

	batch_keys = admission_keys[offset : offset + PATIENT_HISTORY_IMPORT_BATCH_SIZE]
	if not batch_keys:
		return {"processed": offset, "done": True, "batch_count": 0}

	template_name = _default_template_name()
	template_rows = _template_detail_rows(template_name)
	stats = {
		"created": 0,
		"updated": 0,
		"skipped_lines": 0,
		"unresolved_groups": 0,
		"dates_set": 0,
	}
	errors: list[str] = []

	frappe.flags.healthcare_patient_history_import = True
	try:
		for import_admission in batch_keys:
			import_rows = by_admission.get(import_admission) or []
			if not import_rows:
				continue

			resolved_admission = _resolve_admission_for_group(import_admission, import_rows)
			if not resolved_admission:
				stats["unresolved_groups"] += 1
				stats["skipped_lines"] += len(import_rows)
				continue

			try:
				result = process_admission_import(
					resolved_admission,
					import_rows,
					template_name,
					template_rows,
				)
				if result.get("unresolved_admission"):
					stats["unresolved_groups"] += 1
					stats["skipped_lines"] += len(import_rows)
					continue

				if result.get("created"):
					stats["created"] += 1
				else:
					stats["updated"] += 1
				stats["skipped_lines"] += result.get("skipped_lines") or 0

				ph_name = result.get("patient_history")
				if ph_name and _apply_date_if_missing(ph_name, import_rows):
					stats["dates_set"] += 1
			except Exception:
				errors.append(f"{import_admission}: {frappe.get_traceback()}")
				frappe.log_error(title=f"IP_ADMISSION_02 Patient History import failed: {import_admission}")
	finally:
		frappe.flags.healthcare_patient_history_import = False

	frappe.db.commit()
	processed = offset + len(batch_keys)
	return {
		"processed": processed,
		"done": len(batch_keys) < PATIENT_HISTORY_IMPORT_BATCH_SIZE,
		"batch_count": len(batch_keys),
		"total_admissions": len(admission_keys),
		"stats": stats,
		"errors_in_batch": len(errors),
	}
