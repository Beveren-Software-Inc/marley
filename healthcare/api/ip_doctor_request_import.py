"""Import Oracle IP_DOCTOR_REQUEST_01 Excel into Doctor Order rows."""

from __future__ import annotations

import json
from typing import Any

import frappe
from frappe import _
from frappe.utils import cint

from healthcare.api.discharge_checklist_import import _resolve_cost_center
from healthcare.api.patient_appointment_old_status_backfill import ensure_healthcare_practitioner
from healthcare.api.patient_info_import import (
	_cell_text,
	_clean_oracle_num,
	_excel_file_path,
	_require_admin,
)
from healthcare.api.patient_visit_import import ensure_patient_for_legacy_import
from healthcare.api.visit_diagnoses_op_import import (
	_legacy_data_datetime,
	_parse_posting_datetime,
	_truncate_data,
)
from healthcare.api.visit_diagnosis_sync import _resolve_inpatient_admission

IP_DOCTOR_REQUEST_IMPORT_BATCH_SIZE = 200
CACHE_TTL = 7200
CACHE_KEYS = {
	"file_url": "healthcare:data_migration:ip_doctor_request_import:file_url",
	"row_keys": "healthcare:data_migration:ip_doctor_request_import:row_keys",
	"rows": "healthcare:data_migration:ip_doctor_request_import:rows",
}

EXCEL_HEADER_MAP = {
	"TRANS_NUM": "trans_num",
	"TRANS_DATE": "trans_date",
	"DOC_ORDER_DESC": "doc_order_desc",
	"ADMISSION_NUM": "admission_num",
	"PATIENT_NUM": "patient_num",
	"DOC_NUM": "doc_num",
	"DOC_DATE": "doc_date",
	"NURSES_NUM": "nurses_num",
	"NURSES_DATE": "nurses_date",
	"NURSES_REMARKS": "nurses_remarks",
	"JOB_STATUS": "job_status",
	"BRANCH_NUM": "branch_num",
	"CR_ID": "cr_id",
	"CR_DATE": "cr_date",
	"UP_ID": "up_id",
	"UP_DATE": "up_date",
	"REQUEST_TYPE": "request_type",
}


def _normalize_header(cell: Any) -> str:
	if cell is None:
		return ""
	text = str(cell).strip().upper().replace(" ", "_")
	return EXCEL_HEADER_MAP.get(text, text.lower())


def _map_status(value: Any) -> str:
	text = _cell_text(value)
	if not text:
		return "Pending"
	lower = text.lower()
	if lower in ("finished", "complete", "completed"):
		return "Finished"
	if lower in ("canceled", "cancelled"):
		return "Canceled"
	if lower == "pending":
		return "Pending"
	return text


def _apply_patient_fetch_fields(fields: dict[str, Any], patient: str | None) -> None:
	if not patient or not frappe.db.exists("Patient", patient):
		return
	patient_name = frappe.db.get_value("Patient", patient, "patient_name")
	if patient_name:
		fields["patient_name"] = _truncate_data(patient_name)


def _apply_doctor_fetch_fields(fields: dict[str, Any], doctor: str | None) -> None:
	if not doctor or not frappe.db.exists("Healthcare Practitioner", doctor):
		return
	row = frappe.db.get_value(
		"Healthcare Practitioner",
		doctor,
		["practitioner_name", "department"],
		as_dict=True,
	)
	if not row:
		return
	if row.practitioner_name:
		fields["doctor_name"] = _truncate_data(row.practitioner_name)
	if row.department and frappe.db.exists("Medical Department", row.department):
		fields["department"] = row.department


def _resolve_doctor_code(doc_num: Any) -> str | None:
	"""Resolve DOC_NUM to Healthcare Practitioner without creating records (preview-safe)."""
	code = _clean_oracle_num(doc_num)
	if not code:
		return None
	if frappe.db.exists("Healthcare Practitioner", code):
		return code
	return frappe.db.get_value("Healthcare Practitioner", {"doctors_id": code}, "name")


def _order_description(value: Any) -> str:
	text = _cell_text(value)
	return text if text else "-"


def _parse_sheet_rows(ws) -> list[dict]:
	rows_iter = ws.iter_rows(values_only=True)
	try:
		header_row = next(rows_iter)
	except StopIteration:
		return []

	headers = [_normalize_header(h) for h in header_row]
	parsed: list[dict] = []
	for raw in rows_iter:
		if not raw or all(cell is None or str(cell).strip() == "" for cell in raw):
			continue
		row: dict[str, Any] = {}
		for idx, key in enumerate(headers):
			if not key or idx >= len(raw):
				continue
			row[key] = raw[idx]

		trans_num = _clean_oracle_num(row.get("trans_num"))
		if not trans_num:
			continue

		row["trans_num"] = trans_num
		row["admission_num"] = _clean_oracle_num(row.get("admission_num"))
		row["patient_num"] = _clean_oracle_num(row.get("patient_num"))
		row["doc_num"] = _clean_oracle_num(row.get("doc_num"))
		row["doc_order_desc"] = _cell_text(row.get("doc_order_desc"))
		row["job_status"] = _cell_text(row.get("job_status"))
		row["request_type"] = _cell_text(row.get("request_type"))
		row["cr_id"] = _clean_oracle_num(row.get("cr_id"))
		row["up_id"] = _clean_oracle_num(row.get("up_id"))
		row["trans_date"] = row.get("trans_date")
		row["doc_date"] = row.get("doc_date")
		row["cr_date"] = row.get("cr_date")
		row["up_date"] = row.get("up_date")
		row["branch_num"] = row.get("branch_num")
		parsed.append(row)
	return parsed


def _parse_excel_rows(file_url: str) -> tuple[list[dict], dict[str, int]]:
	try:
		import openpyxl
	except ImportError:
		frappe.throw(
			_("openpyxl is required to read Excel files. Install it in the bench environment: pip install openpyxl")
		)

	path = _excel_file_path(file_url)
	wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
	by_key: dict[str, dict] = {}
	sheet_row_counts: dict[str, int] = {}
	try:
		for sheet_name in wb.sheetnames:
			sheet_rows = _parse_sheet_rows(wb[sheet_name])
			sheet_row_counts[sheet_name] = len(sheet_rows)
			for row in sheet_rows:
				by_key[row["trans_num"]] = row
	finally:
		wb.close()
	return list(by_key.values()), sheet_row_counts


def _build_doctor_order_fields(row: dict) -> tuple[dict, dict]:
	stats = {"patient_created": 0, "practitioner_created": 0}
	patient = row.get("patient_num") or ""
	if patient:
		if not frappe.db.exists("Patient", patient):
			result = ensure_patient_for_legacy_import(patient)
			if result.get("status") == "created":
				stats["patient_created"] = 1
			elif result.get("status") != "existing":
				patient = ""
		if patient and not frappe.db.exists("Patient", patient):
			patient = ""

	admission = None
	if row.get("admission_num"):
		admission = _resolve_inpatient_admission(row["admission_num"], patient or None)

	doctor = ""
	doc_num = row.get("doc_num") or ""
	if doc_num:
		doctor, created = ensure_healthcare_practitioner(doc_num)
		if created:
			stats["practitioner_created"] = 1

	trans_date = _parse_posting_datetime(row.get("trans_date"))
	doc_date = _parse_posting_datetime(row.get("doc_date")) or trans_date
	cost_center = _resolve_cost_center(row.get("branch_num"))

	fields: dict[str, Any] = {
		"trans_no": row["trans_num"],
		"trans_date": trans_date,
		"doctor_entry_date": doc_date,
		"doctor_order": _order_description(row.get("doc_order_desc")),
		"inpatient_admission": admission,
		"patient": patient or None,
		"doctor": doctor or None,
		"status": _map_status(row.get("job_status")),
		"cost_center": cost_center,
		"request": row.get("request_type") or None,
		"cr_id": row.get("cr_id") or None,
		"up_id": row.get("up_id") or None,
	}
	cr_date = _legacy_data_datetime(row.get("cr_date"))
	if cr_date:
		fields["cr_date"] = cr_date
	up_date = _legacy_data_datetime(row.get("up_date"))
	if up_date:
		fields["up_date"] = up_date

	if patient:
		_apply_patient_fetch_fields(fields, patient)
	if doctor:
		_apply_doctor_fetch_fields(fields, doctor)

	return {key: value for key, value in fields.items() if value not in (None, "")}, stats


def upsert_doctor_order_from_row(row: dict) -> dict:
	trans_num = row.get("trans_num")
	if not trans_num:
		return {"status": "skip_no_trans_num"}

	fields, side_stats = _build_doctor_order_fields(row)
	used_placeholder = not _cell_text(row.get("doc_order_desc"))

	existing = frappe.db.exists("Doctor Order", trans_num)
	if existing:
		doc = frappe.get_doc("Doctor Order", trans_num)
		for key, value in fields.items():
			if key == "trans_no":
				continue
			doc.set(key, value)
		action = "updated"
	else:
		doc = frappe.new_doc("Doctor Order")
		doc.update(fields)
		action = "created"

	doc.flags.ignore_validate = True
	doc.flags.ignore_mandatory = True
	doc.flags.ignore_links = True
	doc.flags.legacy_import = True

	if existing:
		doc.save(ignore_permissions=True)
	else:
		doc.insert(ignore_permissions=True)

	return {
		"status": action,
		"trans_num": trans_num,
		"name": doc.name,
		"placeholder_description": used_placeholder,
		**side_stats,
	}


def _preview_counts(rows: list[dict]) -> dict:
	patient_nums = {row["patient_num"] for row in rows if row.get("patient_num")}
	missing_patients = [pn for pn in patient_nums if not frappe.db.exists("Patient", pn)]
	existing = sum(1 for row in rows if frappe.db.exists("Doctor Order", row["trans_num"]))
	resolved_admissions = 0
	unresolved_admissions = 0
	resolved_doctors = 0
	empty_description = 0
	empty_description_samples: list[str] = []
	status_counts: dict[str, int] = {}

	for row in rows:
		status = _map_status(row.get("job_status"))
		status_counts[status] = status_counts.get(status, 0) + 1
		if not _cell_text(row.get("doc_order_desc")):
			empty_description += 1
			if len(empty_description_samples) < 20:
				empty_description_samples.append(row["trans_num"])
		if row.get("admission_num"):
			if _resolve_inpatient_admission(row["admission_num"], row.get("patient_num")):
				resolved_admissions += 1
			else:
				unresolved_admissions += 1
		if row.get("doc_num") and _resolve_doctor_code(row["doc_num"]):
			resolved_doctors += 1

	return {
		"unique_patients": len(patient_nums),
		"patients_to_create": len(missing_patients),
		"sample_patients_to_create": missing_patients[:10],
		"existing_orders": existing,
		"new_orders": len(rows) - existing,
		"empty_description": empty_description,
		"importable_rows": len(rows) - existing,
		"sample_empty_description": empty_description_samples,
		"resolved_admissions": resolved_admissions,
		"unresolved_admissions": unresolved_admissions,
		"resolved_doctors": resolved_doctors,
		"with_admission_num": sum(1 for row in rows if row.get("admission_num")),
		"status_counts": status_counts,
	}


def parse_and_cache_excel(file_url: str) -> dict:
	rows, sheet_row_counts = _parse_excel_rows(file_url)
	row_keys = [row["trans_num"] for row in rows]
	by_key = {row["trans_num"]: row for row in rows}

	frappe.cache().set_value(CACHE_KEYS["file_url"], file_url, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(CACHE_KEYS["row_keys"], row_keys, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(
		CACHE_KEYS["rows"],
		json.dumps(by_key, default=str),
		expires_in_sec=CACHE_TTL,
	)

	preview = _preview_counts(rows)
	raw_row_total = sum(sheet_row_counts.values())
	return {
		"excel_rows": len(rows),
		"raw_excel_rows": raw_row_total,
		"sheets": list(sheet_row_counts.keys()),
		"sheet_row_counts": sheet_row_counts,
		**preview,
		"sample_trans_nums": row_keys[:5],
	}


def _load_cached_rows() -> dict[str, dict]:
	raw = frappe.cache().get_value(CACHE_KEYS["rows"])
	if not raw:
		return {}
	if isinstance(raw, dict):
		return raw
	return json.loads(raw)


@frappe.whitelist()
def preview_ip_doctor_request_import(file_url: str) -> dict:
	_require_admin()
	if not (file_url or "").strip():
		frappe.throw(_("Please upload the IP_DOCTOR_REQUEST_01 Excel file."))
	return parse_and_cache_excel(file_url)


def run_ip_doctor_request_import_batch(*, offset: int = 0) -> dict:
	row_keys = frappe.cache().get_value(CACHE_KEYS["row_keys"]) or []
	rows_by_key = _load_cached_rows()
	if not row_keys or not rows_by_key:
		return {"processed": offset, "done": True, "batch_count": 0}

	batch_keys = row_keys[offset : offset + IP_DOCTOR_REQUEST_IMPORT_BATCH_SIZE]
	if not batch_keys:
		return {"processed": offset, "done": True, "batch_count": 0}

	created = updated = placeholder_description = 0
	patients_created = practitioners_created = 0
	errors: list[str] = []

	for key in batch_keys:
		row = rows_by_key.get(key) or {}
		try:
			result = upsert_doctor_order_from_row(row)
			status = result.get("status")
			if status == "created":
				created += 1
			elif status == "updated":
				updated += 1
			if result.get("placeholder_description"):
				placeholder_description += 1
			patients_created += cint(result.get("patient_created"))
			practitioners_created += cint(result.get("practitioner_created"))
		except Exception:
			errors.append(f"{key}: {frappe.get_traceback()}")
			frappe.log_error(title=f"IP_DOCTOR_REQUEST_01 import failed: {key}")

	frappe.db.commit()
	processed = offset + len(batch_keys)
	return {
		"processed": processed,
		"done": len(batch_keys) < IP_DOCTOR_REQUEST_IMPORT_BATCH_SIZE,
		"batch_count": len(batch_keys),
		"created": created,
		"updated": updated,
		"placeholder_description": placeholder_description,
		"patients_created": patients_created,
		"practitioners_created": practitioners_created,
		"errors": len(errors),
	}
