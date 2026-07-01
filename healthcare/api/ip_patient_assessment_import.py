"""Import Oracle IP_PATIENT_ASSESSMENT Excel directly into Patient Assessment rows."""

from __future__ import annotations

import json
from typing import Any

import frappe
from frappe import _
from frappe.utils import cint, get_datetime, get_time, now_datetime

from healthcare.api.patient_info_import import (
	_cell_text,
	_clean_oracle_num,
	_excel_file_path,
	_require_admin,
)
from healthcare.api.visit_diagnosis_sync import _resolve_inpatient_admission

IP_PATIENT_ASSESSMENT_IMPORT_BATCH_SIZE = 100
CACHE_TTL = 7200
CACHE_KEYS = {
	"file_url": "healthcare:data_migration:ip_patient_assessment_import:file_url",
	"row_keys": "healthcare:data_migration:ip_patient_assessment_import:row_keys",
	"rows": "healthcare:data_migration:ip_patient_assessment_import:rows",
}

EXCEL_HEADER_MAP = {
	"TRANS_NUM": "trans_num",
	"ADMISSION_NUM": "admission_num",
	"ARRIVAL": "arrival",
	"COMFORTABLE": "comfortable",
	"INTRODUCTION": "introduction",
	"REST": "rest",
	"STAT": "stat",
	"DIET": "diet",
	"CAME_BY": "came_by",
	"OBSERVE": "observe",
	"PHYSICAL": "physical",
	"RISK_OF_FALL": "risk_of_fall",
	"MEDICAL": "medical",
	"FAMILY_HISTORY": "family_history",
	"DRUG_HISTORY": "drug_history",
	"HISTORY_DSCP": "history_dscp",
	"PSYCHIATRIC": "psychiatric",
	"ORIENTED": "oriented",
	"ABNORMALITY": "abnormality",
	"OTHERS": "others",
	"CR_ID": "cr_id",
	"CR_DATE": "cr_date",
	"UP_ID": "up_id",
	"UP_DATE": "up_date",
	"BRANCH_NUM": "branch_num",
	"GEN_APPER_DULL_ACTIVE": "gen_apper_dull_active",
	"PHYSICAL_DESC": "physical_desc",
	"RISK_OF_FALL_DESC": "risk_of_fall_desc",
	"MEDICAL_DESC": "medical_desc",
	"FAMILY_HISTORY_DESC": "family_history_desc",
	"PSYCHIATRIC_DESC": "psychiatric_desc",
	"ORIENTED_DESC": "oriented_desc",
	"ADMISSION_NUM_OLD": "admission_num_old",
}


def _normalize_header(cell: Any) -> str:
	if cell is None:
		return ""
	text = str(cell).strip().upper().replace(" ", "_")
	return EXCEL_HEADER_MAP.get(text, text.lower())


def _to_yes_no_flag(value: Any) -> int:
	"""Normalize imported true-ish values to 1 (Excel often has 1.0 floats)."""
	if value is None or value == "":
		return 0
	if isinstance(value, bool):
		return 1 if value else 0
	if isinstance(value, (int, float)):
		return 1 if int(value) != 0 else 0
	text = str(value).strip().upper()
	if text.endswith(".0"):
		text = text[:-2]
	return 1 if text in {"1", "Y", "YES", "TRUE", "T"} else 0


def _abbrev_to_row_key(abbrev: str) -> str:
	return (abbrev or "").strip().lower().replace(" ", "_")


def _row_value_for_abbrev(row: dict, abbrev: str):
	"""Map parameter abbrev (e.g. CAME BY) to Excel row keys (e.g. came_by)."""
	abbrev = (abbrev or "").strip()
	if not abbrev:
		return None
	for key in (abbrev, abbrev.lower(), _abbrev_to_row_key(abbrev)):
		if key in row:
			return row.get(key)
	return None


def _assessment_datetime_from_row(row: dict):
	raw = row.get("cr_date")
	if raw in (None, ""):
		return now_datetime()
	try:
		return get_datetime(raw)
	except Exception:
		return now_datetime()


def _assessment_time_from_row(row: dict) -> str | None:
	dt = _assessment_datetime_from_row(row)
	if not dt:
		return None
	return get_time(dt).strftime("%H:%M:%S")


def _safe_row_value(row: dict, key: str):
	if key in row:
		return row.get(key)
	lower = key.lower()
	if lower in row:
		return row.get(lower)
	normalized = _abbrev_to_row_key(key)
	if normalized in row:
		return row.get(normalized)
	return None


def _load_default_patient_assessment_template():
	template_name = (
		frappe.db.get_value("Patient Assessment Template", {"assessment_name": "Default Patient Evaluation"}, "name")
		or frappe.db.get_value("Patient Assessment Template", {"default": 1}, "name")
	)
	if not template_name:
		frappe.throw(_("Patient Assessment Template “Default Patient Evaluation” was not found."))
	return frappe.get_doc("Patient Assessment Template", template_name)


def _parameter_abbrev_index() -> dict[str, str]:
	rows = frappe.get_all(
		"Patient Assessment Parameter",
		fields=["name", "parameter_abbrev"],
		limit_page_length=10000,
	)
	return {
		(row.get("name") or ""): (_cell_text(row.get("parameter_abbrev")) or "").strip()
		for row in rows
	}


def _parse_sheet_rows(ws) -> list[dict]:
	rows_iter = ws.iter_rows(values_only=True)
	try:
		header_row = next(rows_iter)
	except StopIteration:
		return []

	headers = [_normalize_header(h) for h in header_row]
	parsed: list[dict] = []
	for raw in rows_iter:
		if not raw or all(cell is None or str(cell).strip() == "" for cell in raw):
			continue
		row: dict[str, Any] = {}
		for idx, key in enumerate(headers):
			if not key or idx >= len(raw):
				continue
			row[key] = raw[idx]

		trans_num = _clean_oracle_num(row.get("trans_num"))
		if not trans_num:
			continue

		row["trans_num"] = trans_num
		row["admission_num"] = _clean_oracle_num(row.get("admission_num"))
		row["admission_num_old"] = _clean_oracle_num(row.get("admission_num_old"))
		row["history_dscp"] = _cell_text(row.get("history_dscp"))
		row["others"] = _cell_text(row.get("others"))
		row["cr_id"] = _clean_oracle_num(row.get("cr_id"))
		row["up_id"] = _clean_oracle_num(row.get("up_id"))
		parsed.append(row)
	return parsed


def _admission_dedup_key(row: dict) -> str:
	admission_num = row.get("admission_num") or row.get("admission_num_old") or ""
	resolved = _resolve_inpatient_admission(admission_num) if admission_num else None
	if resolved:
		return f"adm::{resolved}"
	return f"raw::{admission_num or row.get('trans_num')}"


def _dedupe_rows_by_admission(rows: list[dict]) -> tuple[list[dict], dict[str, int]]:
	"""Keep one row per admission (highest TRANS_NUM wins)."""
	by_key: dict[str, dict] = {}
	duplicate_trans: dict[str, int] = {}
	for row in rows:
		key = _admission_dedup_key(row)
		existing = by_key.get(key)
		if existing:
			duplicate_trans[key] = duplicate_trans.get(key, 1) + 1
			if cint(row.get("trans_num")) >= cint(existing.get("trans_num")):
				by_key[key] = row
		else:
			by_key[key] = row
	return list(by_key.values()), {
		"duplicate_admission_rows": sum(duplicate_trans.values()),
		"duplicate_admission_groups": len(duplicate_trans),
	}


def _parse_excel_rows(file_url: str) -> tuple[list[dict], dict[str, int], dict[str, int]]:
	try:
		import openpyxl
	except ImportError:
		frappe.throw(
			_("openpyxl is required to read Excel files. Install it in the bench environment: pip install openpyxl")
		)

	path = _excel_file_path(file_url)
	wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
	all_rows: list[dict] = []
	sheet_row_counts: dict[str, int] = {}
	try:
		for sheet_name in wb.sheetnames:
			sheet_rows = _parse_sheet_rows(wb[sheet_name])
			sheet_row_counts[sheet_name] = len(sheet_rows)
			all_rows.extend(sheet_rows)
	finally:
		wb.close()

	deduped, dup_stats = _dedupe_rows_by_admission(all_rows)
	return deduped, sheet_row_counts, dup_stats


def _find_existing_patient_assessment(admission: str) -> str | None:
	rows = frappe.get_all(
		"Patient Assessment",
		filters={"admission": admission, "docstatus": ["!=", 2]},
		pluck="name",
		limit=1,
	)
	return rows[0] if rows else None


def _append_assessment_sheet(
	doc,
	row: dict,
	template,
	parameter_abbrev_by_name: dict[str, str],
	*,
	assessment_time: str | None = None,
) -> None:
	doc.set("assessment_sheet", [])
	for detail in template.get("parameters") or []:
		param_name = (_cell_text(detail.get("assessment_parameter")) or "").strip()
		if not param_name:
			continue
		abbrev = (parameter_abbrev_by_name.get(param_name) or "").strip()
		if not abbrev:
			continue

		flag_value = _row_value_for_abbrev(row, abbrev)
		yes_flag = _to_yes_no_flag(flag_value)

		desc_key = f"{_abbrev_to_row_key(abbrev)}_desc"
		comments = (_cell_text(_row_value_for_abbrev(row, desc_key)) or "").strip()
		if not comments and _abbrev_to_row_key(abbrev) == "history_dscp":
			comments = (_cell_text(row.get("history_dscp")) or "").strip()

		child_row = {
			"parameter": param_name,
			"yes": yes_flag,
			"comments": comments if yes_flag else "",
		}
		if assessment_time:
			child_row["time"] = assessment_time

		doc.append("assessment_sheet", child_row)


def _build_patient_assessment_doc(
	row: dict,
	*,
	template,
	parameter_abbrev_by_name: dict[str, str],
	existing_name: str | None = None,
):
	admission_num = row.get("admission_num") or row.get("admission_num_old") or ""
	admission = _resolve_inpatient_admission(admission_num)
	if not admission:
		return None, "skip_no_admission"

	inpatient = frappe.db.get_value(
		"Inpatient Admission",
		admission,
		["patient", "patient_name", "company", "primary_practitioner", "secondary_practitioner"],
		as_dict=True,
	) or {}
	patient = (_cell_text(inpatient.get("patient")) or "").strip()
	if not patient:
		return None, "skip_no_patient"

	if existing_name:
		doc = frappe.get_doc("Patient Assessment", existing_name)
	else:
		doc = frappe.new_doc("Patient Assessment")

	doc.patient = patient
	doc.patient_name = inpatient.get("patient_name")
	doc.assessment_template = template.name
	doc.reference_type = "Inpatient Admission"
	doc.encounter = admission
	doc.admission = admission
	doc.company = inpatient.get("company")
	doc.healthcare_practitioner = inpatient.get("primary_practitioner") or inpatient.get("secondary_practitioner")
	doc.assessment_datetime = _assessment_datetime_from_row(row)
	trans_num = row.get("trans_num") or ""
	if trans_num and frappe.db.exists("IP Patient Assessment", trans_num):
		doc.ip_patient_assessment = trans_num
	doc.assessment_description = (row.get("history_dscp") or "").strip() or (row.get("others") or "").strip()

	_append_assessment_sheet(
		doc,
		row,
		template,
		parameter_abbrev_by_name,
		assessment_time=_assessment_time_from_row(row),
	)
	return doc, None


def _persist_and_submit_patient_assessment(doc, *, existing: bool) -> bool:
	"""Insert or update, then submit draft assessments."""
	doc.flags.ignore_validate = True
	doc.flags.ignore_mandatory = True
	doc.flags.ignore_links = True
	doc.flags.legacy_import = True

	if existing:
		if doc.docstatus == 1:
			doc.cancel()
		doc.save(ignore_permissions=True)
	else:
		doc.insert(ignore_permissions=True)

	if doc.docstatus != 0:
		return False

	try:
		doc.flags.ignore_validate = True
		doc.flags.ignore_mandatory = True
		doc.flags.legacy_import = True
		doc.submit()
		return True
	except Exception:
		frappe.log_error(
			title=f"Patient Assessment submit failed: {doc.name}",
			message=frappe.get_traceback(),
		)
		return False


def upsert_patient_assessment_from_row(
	row: dict,
	*,
	template=None,
	parameter_abbrev_by_name: dict[str, str] | None = None,
) -> dict:
	if template is None:
		template = _load_default_patient_assessment_template()
	if parameter_abbrev_by_name is None:
		parameter_abbrev_by_name = _parameter_abbrev_index()

	admission_num = row.get("admission_num") or row.get("admission_num_old") or ""
	admission = _resolve_inpatient_admission(admission_num)
	if not admission:
		return {"status": "skip_no_admission", "trans_num": row.get("trans_num")}

	existing = _find_existing_patient_assessment(admission)
	doc, skip_reason = _build_patient_assessment_doc(
		row,
		template=template,
		parameter_abbrev_by_name=parameter_abbrev_by_name,
		existing_name=existing,
	)
	if skip_reason:
		return {"status": skip_reason, "trans_num": row.get("trans_num"), "admission": admission}
	if not doc:
		return {"status": "skip_error", "trans_num": row.get("trans_num")}

	action = "updated" if existing else "created"
	submitted = _persist_and_submit_patient_assessment(doc, existing=bool(existing))

	return {
		"status": action,
		"trans_num": row.get("trans_num"),
		"admission": admission,
		"name": doc.name,
		"submitted": submitted,
	}


def _preview_counts(rows: list[dict]) -> dict:
	template = _load_default_patient_assessment_template()
	existing = 0
	resolved_admissions = 0
	unresolved_admissions = 0
	for row in rows:
		admission_num = row.get("admission_num") or row.get("admission_num_old") or ""
		admission = _resolve_inpatient_admission(admission_num) if admission_num else None
		if admission:
			resolved_admissions += 1
			if _find_existing_patient_assessment(admission):
				existing += 1
		elif admission_num:
			unresolved_admissions += 1

	return {
		"assessment_template": template.name,
		"assessments": len(rows),
		"existing_assessments": existing,
		"new_assessments": len(rows) - existing,
		"resolved_admissions": resolved_admissions,
		"unresolved_admissions": unresolved_admissions,
	}


def parse_and_cache_excel(file_url: str) -> dict:
	rows, sheet_row_counts, dup_stats = _parse_excel_rows(file_url)
	row_keys = [_admission_dedup_key(row) for row in rows]
	by_key = {_admission_dedup_key(row): row for row in rows}

	frappe.cache().set_value(CACHE_KEYS["file_url"], file_url, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(CACHE_KEYS["row_keys"], row_keys, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(
		CACHE_KEYS["rows"],
		json.dumps(by_key, default=str),
		expires_in_sec=CACHE_TTL,
	)

	preview = _preview_counts(rows)
	raw_row_total = sum(sheet_row_counts.values())
	return {
		"excel_rows": len(rows),
		"raw_excel_rows": raw_row_total,
		"sheets": list(sheet_row_counts.keys()),
		"sheet_row_counts": sheet_row_counts,
		**dup_stats,
		**preview,
		"sample_trans_nums": [row["trans_num"] for row in rows[:5]],
	}


def _load_cached_rows() -> dict[str, dict]:
	raw = frappe.cache().get_value(CACHE_KEYS["rows"])
	if not raw:
		return {}
	if isinstance(raw, dict):
		return raw
	return json.loads(raw)


@frappe.whitelist()
def preview_ip_patient_assessment_import(file_url: str) -> dict:
	_require_admin()
	if not (file_url or "").strip():
		frappe.throw(_("Please upload the IP_PATIENT_ASSESSMENT Excel file."))
	return parse_and_cache_excel(file_url)


def run_ip_patient_assessment_import_batch(*, offset: int = 0) -> dict:
	row_keys = frappe.cache().get_value(CACHE_KEYS["row_keys"]) or []
	rows_by_key = _load_cached_rows()
	if not row_keys or not rows_by_key:
		return {"processed": offset, "done": True, "batch_count": 0}

	batch_keys = row_keys[offset : offset + IP_PATIENT_ASSESSMENT_IMPORT_BATCH_SIZE]
	if not batch_keys:
		return {"processed": offset, "done": True, "batch_count": 0}

	template = _load_default_patient_assessment_template()
	parameter_abbrev_by_name = _parameter_abbrev_index()

	created = updated = submitted = skip_no_admission = skip_no_patient = 0
	errors: list[str] = []

	for key in batch_keys:
		row = rows_by_key.get(key) or {}
		try:
			result = upsert_patient_assessment_from_row(
				row,
				template=template,
				parameter_abbrev_by_name=parameter_abbrev_by_name,
			)
			status = result.get("status")
			if status == "created":
				created += 1
			elif status == "updated":
				updated += 1
			elif status == "skip_no_admission":
				skip_no_admission += 1
			elif status == "skip_no_patient":
				skip_no_patient += 1
			if result.get("submitted"):
				submitted += 1
		except Exception:
			errors.append(f"{key}: {frappe.get_traceback()}")
			frappe.log_error(title=f"IP_PATIENT_ASSESSMENT import failed: {key}")

	frappe.db.commit()
	processed = offset + len(batch_keys)
	return {
		"processed": processed,
		"done": len(batch_keys) < IP_PATIENT_ASSESSMENT_IMPORT_BATCH_SIZE,
		"batch_count": len(batch_keys),
		"created": created,
		"updated": updated,
		"submitted": submitted,
		"skip_no_admission": skip_no_admission,
		"skip_no_patient": skip_no_patient,
		"errors": len(errors),
	}
