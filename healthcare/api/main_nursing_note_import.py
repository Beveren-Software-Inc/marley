"""Import Oracle nursing Excel into Main Nursing Note rows (one per TRANS_NUM)."""

from __future__ import annotations

import json
from datetime import date, datetime, time
from typing import Any

import frappe
from frappe import _
from frappe.utils import cint, get_datetime, getdate

from healthcare.api.discharge_checklist_import import _resolve_cost_center
from healthcare.api.patient_info_import import (
	_cell_text,
	_clean_oracle_num,
	_excel_file_path,
	_require_admin,
)
from healthcare.api.visit_diagnosis_sync import _resolve_inpatient_admission

MAIN_NURSING_NOTE_IMPORT_BATCH_SIZE = 200
CACHE_TTL = 7200
CACHE_KEYS = {
	"file_url": "healthcare:data_migration:main_nursing_note_import:file_url",
	"row_keys": "healthcare:data_migration:main_nursing_note_import:row_keys",
	"rows": "healthcare:data_migration:main_nursing_note_import:rows",
}

EXCEL_HEADER_MAP = {
	"TRANS_NUM": "trans_num",
	"ADMISSION_NUM": "admission_num",
	"NURSING_DESC": "nursing_desc",
	"NURSING_DATE": "nursing_date",
	"NURSING_TIME": "nursing_time",
	"USER_NAME": "user_name",
	"CR_ID": "cr_id",
	"CR_DATE": "cr_date",
	"UP_ID": "up_id",
	"UP_DATE": "up_date",
	"BRANCH_NUM": "branch_num",
	"SHIFT_CODE": "shift_code",
	"ADMISSION_NUM_OLD": "admission_num_old",
}

SHIFT_CODE_MAP = {
	"MOR": "Morning",
	"EVE": "Evening",
	"NGT": "Night",
	"MORNING": "Morning",
	"EVENING": "Evening",
	"NIGHT": "Night",
}


def _normalize_header(cell: Any) -> str:
	if cell is None:
		return ""
	text = str(cell).strip().upper().replace(" ", "_")
	return EXCEL_HEADER_MAP.get(text, text.lower())


def _parse_time_value(value: Any) -> time | None:
	if value in (None, ""):
		return None
	if isinstance(value, time):
		return value
	if isinstance(value, datetime):
		return value.time()
	text = _cell_text(value)
	if not text:
		return None
	for candidate in (text, text.split(" ")[-1]):
		for fmt in ("%H:%M:%S", "%H:%M"):
			try:
				return datetime.strptime(candidate, fmt).time()
			except ValueError:
				continue
	return None


def _parse_date_value(value: Any):
	if value in (None, ""):
		return None
	if isinstance(value, datetime):
		return value.date()
	if isinstance(value, date):
		return value
	try:
		return getdate(value)
	except Exception:
		return None


def _parse_datetime_value(value: Any) -> datetime | None:
	if value in (None, ""):
		return None
	if isinstance(value, datetime):
		return value
	if isinstance(value, date):
		return datetime.combine(value, time.min)
	try:
		return get_datetime(value)
	except Exception:
		return None


def _time_label(value: Any) -> str:
	parsed = _parse_time_value(value)
	if not parsed:
		return datetime.now().strftime("%H:%M")
	return parsed.strftime("%H:%M")


def _shift_from_time(value: Any) -> str:
	parsed = _parse_time_value(value)
	hour = parsed.hour if parsed else datetime.now().hour
	if 6 <= hour < 14:
		return "Morning"
	if 14 <= hour < 22:
		return "Evening"
	return "Night"


def _map_shift(shift_code: Any, nursing_time: Any) -> str:
	code = _cell_text(shift_code).upper()
	if code in SHIFT_CODE_MAP:
		return SHIFT_CODE_MAP[code]
	return _shift_from_time(nursing_time)


def _format_nursing_notes(desc: Any, nursing_time: Any) -> str | None:
	text = _cell_text(desc)
	if not text:
		return None
	return f"[{_time_label(nursing_time)}] {text}"


def _resolve_admission(row: dict) -> str | None:
	admission_num = row.get("admission_num") or ""
	old_num = row.get("admission_num_old") or ""
	for candidate in (admission_num, old_num):
		if not candidate:
			continue
		resolved = _resolve_inpatient_admission(candidate)
		if resolved:
			return resolved
	return None


def _resolve_user(user_name: str) -> str | None:
	name = (user_name or "").strip()
	if not name:
		return None
	if frappe.db.exists("User", name):
		return name
	for field in ("username", "full_name", "name"):
		match = frappe.db.get_value("User", {field: name}, "name")
		if match:
			return match
	return None


def _existing_main_nursing_note_name(trans_no: str) -> str | None:
	if frappe.db.exists("Main Nursing Note", trans_no):
		return trans_no
	return frappe.db.get_value("Main Nursing Note", {"trans_no": trans_no}, "name")


def _build_main_nursing_note_fields(row: dict, admission: str) -> dict[str, Any]:
	patient = frappe.db.get_value("Inpatient Admission", admission, "patient")
	patient_name = frappe.db.get_value("Patient", patient, "patient_name") if patient else None
	nursing_time = row.get("nursing_time")

	fields: dict[str, Any] = {
		"trans_no": row["trans_num"],
		"admission": admission,
		"shift": _map_shift(row.get("shift_code"), nursing_time),
	}

	if patient:
		fields["file_no"] = patient
	if patient_name:
		fields["patient_name"] = patient_name

	nursing_date = _parse_date_value(row.get("nursing_date"))
	if nursing_date:
		fields["date"] = nursing_date

	parsed_time = _parse_time_value(nursing_time)
	if parsed_time:
		fields["data"] = parsed_time

	nursing_notes = _format_nursing_notes(row.get("nursing_desc"), nursing_time)
	if nursing_notes:
		fields["nursing_notes"] = nursing_notes

	user_name = _cell_text(row.get("user_name"))
	if user_name:
		fields["user_name"] = user_name
		user = _resolve_user(user_name)
		if user:
			fields["user"] = user

	if row.get("admission_num_old"):
		fields["admission_old_no"] = row["admission_num_old"]

	cost_center = _resolve_cost_center(row.get("branch_num"))
	if cost_center:
		fields["cost_center"] = cost_center
	elif admission:
		cc = frappe.db.get_value("Inpatient Admission", admission, "cost_center")
		if cc:
			fields["cost_center"] = cc

	cr_id = row.get("cr_id")
	if cr_id not in (None, ""):
		fields["cr_id"] = cint(cr_id)
	up_id = row.get("up_id")
	if up_id not in (None, ""):
		fields["up_id"] = cint(up_id)

	cr_date = _parse_datetime_value(row.get("cr_date"))
	if cr_date:
		fields["cr_date"] = cr_date
	up_date = _parse_datetime_value(row.get("up_date"))
	if up_date:
		fields["up_date"] = up_date

	return fields


def _apply_legacy_import_flags(doc) -> None:
	doc.flags.ignore_validate = True
	doc.flags.ignore_mandatory = True
	doc.flags.ignore_links = True
	doc.flags.ignore_permissions = True
	doc.flags.legacy_import = True


def upsert_main_nursing_note_from_row(row: dict) -> dict:
	trans_no = row.get("trans_num") or ""
	if not trans_no:
		return {"status": "skip_no_trans_num"}

	admission = _resolve_admission(row)
	if not admission:
		return {"status": "skip_no_admission", "trans_no": trans_no}

	fields = _build_main_nursing_note_fields(row, admission)
	if not fields.get("nursing_notes"):
		return {"status": "skip_no_notes", "trans_no": trans_no}

	existing_name = _existing_main_nursing_note_name(trans_no)
	if existing_name:
		doc = frappe.get_doc("Main Nursing Note", existing_name)
		action = "updated"
	else:
		doc = frappe.get_doc({"doctype": "Main Nursing Note", "name": trans_no, **fields})
		action = "created"

	if existing_name:
		for key, value in fields.items():
			if key == "trans_no":
				continue
			doc.set(key, value)

	_apply_legacy_import_flags(doc)
	if existing_name:
		doc.save(ignore_permissions=True)
	else:
		doc.insert(ignore_permissions=True)

	return {
		"status": action,
		"trans_no": trans_no,
		"name": doc.name,
	}


def _parse_sheet_rows(ws) -> list[dict]:
	rows_iter = ws.iter_rows(values_only=True)
	try:
		header_row = next(rows_iter)
	except StopIteration:
		return []

	headers = [_normalize_header(h) for h in header_row]
	parsed: list[dict] = []
	for raw in rows_iter:
		if not raw or all(cell is None or str(cell).strip() == "" for cell in raw):
			continue
		row: dict[str, Any] = {}
		for idx, key in enumerate(headers):
			if not key or idx >= len(raw):
				continue
			row[key] = raw[idx]

		trans_num = _clean_oracle_num(row.get("trans_num"))
		if not trans_num:
			continue

		row["trans_num"] = trans_num
		row["admission_num"] = _clean_oracle_num(row.get("admission_num"))
		row["admission_num_old"] = _clean_oracle_num(row.get("admission_num_old"))
		row["branch_num"] = _clean_oracle_num(row.get("branch_num"))
		row["cr_id"] = _clean_oracle_num(row.get("cr_id"))
		row["up_id"] = _clean_oracle_num(row.get("up_id"))
		row["shift_code"] = _cell_text(row.get("shift_code")).strip()
		row["user_name"] = _cell_text(row.get("user_name")).strip()
		parsed.append(row)
	return parsed


def _parse_excel_rows(file_url: str) -> tuple[list[dict], dict[str, int]]:
	try:
		import openpyxl
	except ImportError:
		frappe.throw(
			_(
				"openpyxl is required to read Excel files. Install it in the bench environment: pip install openpyxl"
			)
		)

	path = _excel_file_path(file_url)
	wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
	all_rows: list[dict] = []
	sheet_row_counts: dict[str, int] = {}
	try:
		for sheet_name in wb.sheetnames:
			sheet_rows = _parse_sheet_rows(wb[sheet_name])
			sheet_row_counts[sheet_name] = len(sheet_rows)
			all_rows.extend(sheet_rows)
	finally:
		wb.close()
	return all_rows, sheet_row_counts


def _preview_counts(rows: list[dict]) -> dict:
	existing = 0
	resolved = 0
	unresolved = 0
	skip_no_notes = 0

	for row in rows:
		if not _cell_text(row.get("nursing_desc")):
			skip_no_notes += 1
		admission = _resolve_admission(row)
		if not admission:
			unresolved += 1
			continue
		resolved += 1
		if _existing_main_nursing_note_name(row.get("trans_num") or ""):
			existing += 1

	return {
		"existing_notes": existing,
		"new_notes": resolved - existing,
		"resolved_admissions": resolved,
		"unresolved_admissions": unresolved,
		"skip_no_notes": skip_no_notes,
		"sample_trans_nums": [row.get("trans_num") for row in rows[:5]],
	}


def parse_and_cache_excel(file_url: str) -> dict:
	rows, sheet_row_counts = _parse_excel_rows(file_url)
	by_key = {row["trans_num"]: row for row in rows}
	row_keys = list(by_key.keys())

	frappe.cache().set_value(CACHE_KEYS["file_url"], file_url, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(CACHE_KEYS["row_keys"], row_keys, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(
		CACHE_KEYS["rows"],
		json.dumps(by_key, default=str),
		expires_in_sec=CACHE_TTL,
	)

	preview = _preview_counts(list(by_key.values()))
	raw_row_total = sum(sheet_row_counts.values())
	return {
		"excel_rows": len(by_key),
		"raw_excel_rows": raw_row_total,
		"sheets": list(sheet_row_counts.keys()),
		"sheet_row_counts": sheet_row_counts,
		**preview,
	}


def _load_cached_rows() -> dict[str, dict]:
	raw = frappe.cache().get_value(CACHE_KEYS["rows"])
	if not raw:
		return {}
	if isinstance(raw, dict):
		return raw
	return json.loads(raw)


@frappe.whitelist()
def preview_main_nursing_note_import(file_url: str) -> dict:
	_require_admin()
	if not (file_url or "").strip():
		frappe.throw(_("Please upload the nursing Excel file."))
	return parse_and_cache_excel(file_url)


def run_main_nursing_note_import_batch(*, offset: int = 0) -> dict:
	row_keys = frappe.cache().get_value(CACHE_KEYS["row_keys"]) or []
	rows_by_key = _load_cached_rows()
	if not row_keys or not rows_by_key:
		return {"processed": offset, "done": True, "batch_count": 0}

	batch_keys = row_keys[offset : offset + MAIN_NURSING_NOTE_IMPORT_BATCH_SIZE]
	if not batch_keys:
		return {"processed": offset, "done": True, "batch_count": 0}

	created = updated = skip_no_admission = skip_no_notes = 0
	errors: list[str] = []

	for key in batch_keys:
		row = rows_by_key.get(key) or {}
		try:
			result = upsert_main_nursing_note_from_row(row)
			status = result.get("status")
			if status == "created":
				created += 1
			elif status == "updated":
				updated += 1
			elif status == "skip_no_admission":
				skip_no_admission += 1
			elif status == "skip_no_notes":
				skip_no_notes += 1
			else:
				errors.append(f"{key}: {status}")
		except Exception:
			errors.append(f"{key}: {frappe.get_traceback()}")
			frappe.log_error(title=f"Main Nursing Note import failed: {key}")

	frappe.db.commit()
	processed = offset + len(batch_keys)
	return {
		"processed": processed,
		"done": len(batch_keys) < MAIN_NURSING_NOTE_IMPORT_BATCH_SIZE,
		"batch_count": len(batch_keys),
		"created": created,
		"updated": updated,
		"skip_no_admission": skip_no_admission,
		"skip_no_notes": skip_no_notes,
		"errors": len(errors),
	}
