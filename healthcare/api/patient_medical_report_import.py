"""Import Oracle PATIENT_MEDICAL_REPORT_01 Excel into Patient Medical Report rows."""

from __future__ import annotations

import json
from typing import Any

import frappe
from frappe import _
from frappe.utils import getdate

from healthcare.api.discharge_checklist_import import _resolve_cost_center
from healthcare.api.ip_admission_discharge_import import _resolve_practitioner_by_doctors_id
from healthcare.api.patient_info_import import (
	_cell_text,
	_clean_oracle_num,
	_excel_file_path,
	_parse_date_value,
	_parse_datetime_value,
	_require_admin,
)
from healthcare.api.visit_diagnosis_sync import _resolve_inpatient_admission
from healthcare.api.patient_visit_practitioner import practitioner_name_from_link

DOCTYPE = "Patient Medical Report"
PATIENT_MEDICAL_REPORT_IMPORT_BATCH_SIZE = 500
CACHE_TTL = 7200
CACHE_KEYS = {
	"file_url": "healthcare:data_migration:patient_medical_report_import:file_url",
	"row_keys": "healthcare:data_migration:patient_medical_report_import:row_keys",
	"rows": "healthcare:data_migration:patient_medical_report_import:rows",
}

EXCEL_HEADER_MAP = {
	"TRANS_NUM": "trans_num",
	"TRANS_DATE": "trans_date",
	"IP_OP_SOURCE": "ip_op_source",
	"VISIT_NUM": "visit_num",
	"ADMISSION_NUM": "admission_num",
	"PATIENT_NUM": "patient_num",
	"REFF_NUM": "reference_no",
	"DOCTOR_NUM": "doctor_num",
	"REPORT_DATA_1_EN": "report_data_1_en",
	"REPORT_DATA_1_AR": "report_data_1_ar",
	"REPORT_DATA_2_EN": "report_data_2_en",
	"REPORT_DATA_2_AR": "report_data_2_ar",
	"BRANCH_NUM": "branch_num",
	"CR_ID": "cr_id",
	"CR_DATE": "cr_date",
	"UP_ID": "up_id",
	"UP_DATE": "up_date",
}


def _normalize_header(cell: Any) -> str:
	if cell is None:
		return ""
	text = str(cell).strip().upper().replace(" ", "_")
	return EXCEL_HEADER_MAP.get(text, text.lower())


def _report_html(value: Any) -> str | None:
	text = _cell_text(value)
	if not text:
		return None
	if "<" in text and ">" in text:
		return text
	return text.replace("\r\n", "\n").replace("\r", "\n").replace("\n", "<br>")


def _format_legacy_datetime(value: Any) -> str | None:
	dt = _parse_datetime_value(value)
	if dt:
		return dt.strftime("%Y-%m-%d %H:%M:%S")
	parsed_date = _parse_date_value(value)
	if parsed_date:
		return parsed_date.strftime("%Y-%m-%d")
	text = _cell_text(value)
	return text or None


def _parse_date_field(value: Any):
	dt = _parse_datetime_value(value)
	if dt:
		return getdate(dt)
	parsed_date = _parse_date_value(value)
	if parsed_date:
		return parsed_date
	return None


def _resolve_patient(patient_num: Any) -> str | None:
	patient = _clean_oracle_num(patient_num)
	if not patient:
		return None
	if frappe.db.exists("Patient", patient):
		return patient
	return None


def _resolve_admission(admission_num: Any) -> str | None:
	admission = _clean_oracle_num(admission_num)
	if not admission:
		return None
	if frappe.db.exists("Inpatient Admission", admission):
		return admission
	return _resolve_inpatient_admission(admission)


def _resolve_patient_visit(visit_num: Any) -> str | None:
	visit = _clean_oracle_num(visit_num)
	if not visit:
		return None
	if frappe.db.exists("Patient Visit", visit):
		return visit
	return None


def _parse_sheet_rows(ws) -> list[dict]:
	rows_iter = ws.iter_rows(values_only=True)
	try:
		header_row = next(rows_iter)
	except StopIteration:
		return []

	headers = [_normalize_header(h) for h in header_row]
	parsed: list[dict] = []
	for raw in rows_iter:
		if not raw or all(cell is None or str(cell).strip() == "" for cell in raw):
			continue
		row: dict[str, Any] = {}
		for idx, key in enumerate(headers):
			if not key or idx >= len(raw):
				continue
			row[key] = raw[idx]

		trans_no = _clean_oracle_num(row.get("trans_num"))
		if not trans_no:
			continue
		row["trans_no"] = trans_no
		parsed.append(row)
	return parsed


def _parse_excel_rows(file_url: str) -> tuple[list[dict], dict[str, int]]:
	try:
		import openpyxl
	except ImportError:
		frappe.throw(
			_(
				"openpyxl is required to read Excel files. Install it in the bench environment: pip install openpyxl"
			)
		)

	path = _excel_file_path(file_url)
	wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
	by_key: dict[str, dict] = {}
	sheet_row_counts: dict[str, int] = {}
	try:
		for sheet_name in wb.sheetnames:
			sheet_rows = _parse_sheet_rows(wb[sheet_name])
			sheet_row_counts[sheet_name] = len(sheet_rows)
			for row in sheet_rows:
				by_key[row["trans_no"]] = row
	finally:
		wb.close()
	return list(by_key.values()), sheet_row_counts


def _build_fields(row: dict) -> dict[str, Any]:
	fields: dict[str, Any] = {
		"trans_no": row["trans_no"],
		"ip_op_source": _cell_text(row.get("ip_op_source")) or None,
		"reference_no": _cell_text(row.get("reference_no")) or None,
		"cr_id": _clean_oracle_num(row.get("cr_id")) or None,
		"up_id": _clean_oracle_num(row.get("up_id")) or None,
	}

	trans_date = _parse_date_field(row.get("trans_date"))
	if trans_date:
		fields["trans_date"] = trans_date

	for report_field in (
		"report_data_1_en",
		"report_data_1_ar",
		"report_data_2_en",
		"report_data_2_ar",
	):
		html = _report_html(row.get(report_field))
		if html:
			fields[report_field] = html

	for date_field in ("cr_date", "up_date"):
		legacy_date = _format_legacy_datetime(row.get(date_field))
		if legacy_date:
			fields[date_field] = legacy_date

	patient = _resolve_patient(row.get("patient_num"))
	if patient:
		fields["patient"] = patient

	admission = _resolve_admission(row.get("admission_num"))
	if admission:
		fields["admission"] = admission

	patient_visit = _resolve_patient_visit(row.get("visit_num"))
	if patient_visit:
		fields["patient_visit"] = patient_visit

	practitioner = _resolve_practitioner_by_doctors_id(row.get("doctor_num"))
	if practitioner:
		fields["practitioner"] = practitioner
		practitioner_name = practitioner_name_from_link(practitioner)
		if practitioner_name:
			fields["practitioner_name"] = practitioner_name

	branch = _resolve_cost_center(row.get("branch_num"))
	if branch:
		fields["branch"] = branch

	return {key: value for key, value in fields.items() if value not in (None, "")}


def _preview_counts(rows: list[dict]) -> dict:
	existing = 0
	resolved_patients = 0
	resolved_admissions = 0
	resolved_visits = 0
	resolved_practitioners = 0
	with_report_text = 0
	for row in rows:
		if frappe.db.exists(DOCTYPE, row["trans_no"]):
			existing += 1
		if _resolve_patient(row.get("patient_num")):
			resolved_patients += 1
		if _resolve_admission(row.get("admission_num")):
			resolved_admissions += 1
		if _resolve_patient_visit(row.get("visit_num")):
			resolved_visits += 1
		if _resolve_practitioner_by_doctors_id(row.get("doctor_num")):
			resolved_practitioners += 1
		if _report_html(row.get("report_data_1_en")) or _report_html(row.get("report_data_1_ar")):
			with_report_text += 1

	return {
		"existing_records": existing,
		"resolved_patients": resolved_patients,
		"resolved_admissions": resolved_admissions,
		"resolved_visits": resolved_visits,
		"resolved_practitioners": resolved_practitioners,
		"with_report_text": with_report_text,
		"sample_trans_nos": [row.get("trans_no") for row in rows[:5]],
	}


def parse_and_cache_excel(file_url: str) -> dict:
	rows, sheet_row_counts = _parse_excel_rows(file_url)
	row_keys = [row["trans_no"] for row in rows]
	by_key = {row["trans_no"]: row for row in rows}

	frappe.cache().set_value(CACHE_KEYS["file_url"], file_url, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(CACHE_KEYS["row_keys"], row_keys, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(
		CACHE_KEYS["rows"],
		json.dumps(by_key, default=str),
		expires_in_sec=CACHE_TTL,
	)

	preview = _preview_counts(rows)
	raw_row_total = sum(sheet_row_counts.values())
	return {
		"excel_rows": len(rows),
		"raw_excel_rows": raw_row_total,
		"sheets": list(sheet_row_counts.keys()),
		"sheet_row_counts": sheet_row_counts,
		**preview,
	}


def _load_cached_rows() -> dict[str, dict]:
	raw = frappe.cache().get_value(CACHE_KEYS["rows"])
	if not raw:
		return {}
	if isinstance(raw, dict):
		return raw
	return json.loads(raw)


def upsert_patient_medical_report(row: dict) -> dict:
	fields = _build_fields(row)
	if not fields.get("trans_no"):
		return {"status": "skip", "trans_no": row.get("trans_no")}

	trans_no = fields["trans_no"]
	if frappe.db.exists(DOCTYPE, trans_no):
		doc = frappe.get_doc(DOCTYPE, trans_no)
		for key, value in fields.items():
			doc.set(key, value)
		doc.flags.ignore_mandatory = True
		doc.flags.ignore_links = True
		doc.flags.ignore_validate = True
		doc.save(ignore_permissions=True)
		action = "updated"
	else:
		doc = frappe.get_doc({"doctype": DOCTYPE, **fields})
		doc.flags.ignore_mandatory = True
		doc.flags.ignore_links = True
		doc.flags.ignore_validate = True
		doc.insert(ignore_permissions=True)
		action = "created"

	return {"status": action, "trans_no": trans_no, "name": doc.name}


@frappe.whitelist()
def preview_patient_medical_report_import(file_url: str) -> dict:
	_require_admin()
	if not (file_url or "").strip():
		frappe.throw(_("Please upload the PATIENT_MEDICAL_REPORT_01 Excel file."))
	return parse_and_cache_excel(file_url)


def run_patient_medical_report_import_batch(*, offset: int = 0) -> dict:
	row_keys = frappe.cache().get_value(CACHE_KEYS["row_keys"]) or []
	rows_by_key = _load_cached_rows()
	if not row_keys or not rows_by_key:
		return {"processed": offset, "done": True, "batch_count": 0}

	batch_keys = row_keys[offset : offset + PATIENT_MEDICAL_REPORT_IMPORT_BATCH_SIZE]
	if not batch_keys:
		return {"processed": offset, "done": True, "batch_count": 0}

	created = updated = 0
	errors: list[str] = []

	for key in batch_keys:
		row = rows_by_key.get(key) or {}
		try:
			result = upsert_patient_medical_report(row)
			status = result.get("status")
			if status == "created":
				created += 1
			elif status == "updated":
				updated += 1
		except Exception:
			errors.append(f"{key}: {frappe.get_traceback()}")
			frappe.log_error(title=f"PATIENT_MEDICAL_REPORT_01 import failed: {key}")

	frappe.db.commit()
	processed = offset + len(batch_keys)
	return {
		"processed": processed,
		"done": len(batch_keys) < PATIENT_MEDICAL_REPORT_IMPORT_BATCH_SIZE,
		"batch_count": len(batch_keys),
		"created": created,
		"updated": updated,
		"errors": len(errors),
	}
