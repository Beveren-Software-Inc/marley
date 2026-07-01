"""Import Oracle APPOINTMENTS_HOLD_01 Excel into Practitioner Unavailability rows."""

from __future__ import annotations

from typing import Any

import frappe
from frappe import _
from frappe.utils import getdate

from healthcare.api.patient_appointment_old_status_backfill import ensure_healthcare_practitioner
from healthcare.api.patient_info_import import (
	_cell_text,
	_clean_oracle_num,
	_excel_file_path,
	_parse_date_value,
	_parse_datetime_value,
	_require_admin,
)

EXCEL_HEADER_MAP = {
	"TRANS_NUM": "trans_num",
	"START_DATE": "start_date",
	"END_DATE": "end_date",
	"DOC_CODE": "doc_code",
	"IS_CANCEL": "is_cancel",
	"ANY_REMARKS": "any_remarks",
	"CR_ID": "cr_id",
	"CR_DATE": "cr_date",
	"UP_ID": "up_id",
	"UP_DATE": "up_date",
	"BRANCH_NUM": "branch_num",
}


def _normalize_header(cell: Any) -> str:
	if cell is None:
		return ""
	text = str(cell).strip().upper().replace(" ", "_")
	return EXCEL_HEADER_MAP.get(text, text.lower())


def _format_legacy_datetime(value: Any) -> str | None:
	dt = _parse_datetime_value(value)
	if dt:
		return dt.strftime("%Y-%m-%d %H:%M:%S")
	parsed_date = _parse_date_value(value)
	if parsed_date:
		return parsed_date.strftime("%Y-%m-%d")
	text = _cell_text(value)
	return text or None


def _parse_date_field(value: Any):
	dt = _parse_datetime_value(value)
	if dt:
		return getdate(dt)
	parsed_date = _parse_date_value(value)
	if parsed_date:
		return parsed_date
	return None


def _cancel_flag(value: Any) -> int:
	text = _cell_text(value).upper()
	return 1 if text in ("Y", "YES", "1", "TRUE", "T") else 0


def _resolve_cost_center(branch_label: Any) -> str | None:
	from healthcare.api.discharge_checklist_import import _resolve_cost_center as resolve_cc

	return resolve_cc(branch_label)


def _parse_sheet_rows(ws) -> list[dict]:
	rows_iter = ws.iter_rows(values_only=True)
	try:
		header_row = next(rows_iter)
	except StopIteration:
		return []

	headers = [_normalize_header(h) for h in header_row]
	parsed: list[dict] = []
	for raw in rows_iter:
		if not raw or all(cell is None or str(cell).strip() == "" for cell in raw):
			continue
		row: dict[str, Any] = {}
		for idx, key in enumerate(headers):
			if not key or idx >= len(raw):
				continue
			row[key] = raw[idx]
		tran_num = _clean_oracle_num(row.get("trans_num"))
		if not tran_num:
			continue
		row["tran_num"] = tran_num
		row["doc_code"] = _clean_oracle_num(row.get("doc_code"))
		row["cr_id"] = _clean_oracle_num(row.get("cr_id"))
		row["up_id"] = _clean_oracle_num(row.get("up_id"))
		parsed.append(row)
	return parsed


def _parse_excel_rows(file_url: str) -> tuple[list[dict], dict[str, int]]:
	try:
		import openpyxl
	except ImportError:
		frappe.throw(
			_("openpyxl is required to read Excel files. Install it in the bench environment: pip install openpyxl")
		)

	path = _excel_file_path(file_url)
	wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
	by_tran: dict[str, dict] = {}
	sheet_row_counts: dict[str, int] = {}
	try:
		for sheet_name in wb.sheetnames:
			sheet_rows = _parse_sheet_rows(wb[sheet_name])
			sheet_row_counts[sheet_name] = len(sheet_rows)
			for row in sheet_rows:
				by_tran[row["tran_num"]] = row
	finally:
		wb.close()
	return list(by_tran.values()), sheet_row_counts


def _resolve_practitioner(doc_code: str) -> tuple[str | None, str | None, bool]:
	doc_code = (doc_code or "").strip()
	if not doc_code:
		return None, None, False
	practitioner, created = ensure_healthcare_practitioner(doc_code)
	if not practitioner:
		return None, doc_code, False
	practitioner_name = frappe.db.get_value(
		"Healthcare Practitioner",
		practitioner,
		"practitioner_name",
	)
	return practitioner, practitioner_name or doc_code, created


def _build_fields(row: dict) -> tuple[dict[str, Any], bool]:
	start_date = _parse_date_field(row.get("start_date"))
	end_date = _parse_date_field(row.get("end_date"))
	doc_code = row.get("doc_code") or ""
	practitioner, practitioner_name, practitioner_created = _resolve_practitioner(doc_code)

	fields: dict[str, Any] = {
		"tran_num": row["tran_num"],
		"start_date": start_date,
		"end_date": end_date,
		"posting_date": start_date,
		"doctor_id": practitioner,
		"practitioner_name": practitioner_name or doc_code,
		"is_cancel": _cancel_flag(row.get("is_cancel")),
		"any_remarks": _cell_text(row.get("any_remarks")),
		"cr_id": row.get("cr_id") or "",
		"up_id": row.get("up_id") or "",
	}

	cr_date = _format_legacy_datetime(row.get("cr_date"))
	if cr_date:
		fields["cr_date"] = cr_date
	up_date = _format_legacy_datetime(row.get("up_date"))
	if up_date:
		fields["up_date"] = up_date

	branch = _resolve_cost_center(row.get("branch_num"))
	if branch:
		fields["branch"] = branch

	return {key: value for key, value in fields.items() if value not in (None, "")}, practitioner_created


def _preview_stats(rows: list[dict]) -> dict:
	existing = 0
	resolvable_practitioners = 0
	cancelled = 0
	doc_codes: set[str] = set()
	for row in rows:
		if row.get("doc_code"):
			doc_codes.add(row["doc_code"])
		if _cancel_flag(row.get("is_cancel")):
			cancelled += 1
		if frappe.db.exists("Practitioner Unavailability", row["tran_num"]):
			existing += 1
		practitioner, _, _ = _resolve_practitioner(row.get("doc_code") or "")
		if practitioner:
			resolvable_practitioners += 1

	return {
		"existing_records": existing,
		"resolvable_practitioners": resolvable_practitioners,
		"cancelled_rows": cancelled,
		"unique_doc_codes": len(doc_codes),
		"sample_tran_nums": sorted(r["tran_num"] for r in rows)[:5],
	}


def upsert_practitioner_unavailability(row: dict) -> dict:
	fields, practitioner_created = _build_fields(row)
	if not fields.get("tran_num"):
		return {"status": "skip", "tran_num": row.get("tran_num")}

	tran_num = fields["tran_num"]

	if frappe.db.exists("Practitioner Unavailability", tran_num):
		doc = frappe.get_doc("Practitioner Unavailability", tran_num)
		for key, value in fields.items():
			doc.set(key, value)
		doc.flags.ignore_mandatory = True
		doc.flags.ignore_links = True
		doc.save(ignore_permissions=True)
		action = "updated"
	else:
		doc = frappe.new_doc("Practitioner Unavailability")
		doc.update(fields)
		doc.flags.ignore_mandatory = True
		doc.flags.ignore_links = True
		doc.insert(ignore_permissions=True)
		action = "created"

	return {
		"status": action,
		"tran_num": tran_num,
		"name": doc.name,
		"practitioner_created": practitioner_created,
	}


@frappe.whitelist()
def preview_practitioner_unavailability_import(file_url: str) -> dict:
	_require_admin()
	if not (file_url or "").strip():
		frappe.throw(_("Please upload the APPOINTMENTS_HOLD_01 Excel file."))

	rows, sheet_row_counts = _parse_excel_rows(file_url)
	stats = _preview_stats(rows)
	raw_row_total = sum(sheet_row_counts.values())
	return {
		"excel_rows": len(rows),
		"raw_excel_rows": raw_row_total,
		"sheets": list(sheet_row_counts.keys()),
		"sheet_row_counts": sheet_row_counts,
		**stats,
	}


@frappe.whitelist()
def run_practitioner_unavailability_import(file_url: str) -> dict:
	"""Import all rows synchronously (small file — no background job)."""
	_require_admin()
	if not (file_url or "").strip():
		frappe.throw(_("Please upload the APPOINTMENTS_HOLD_01 Excel file."))

	rows, _ = _parse_excel_rows(file_url)
	created = updated = skipped = practitioners_created = 0
	errors: list[str] = []

	for row in rows:
		try:
			result = upsert_practitioner_unavailability(row)
			status = result.get("status")
			if status == "created":
				created += 1
			elif status == "updated":
				updated += 1
			else:
				skipped += 1
			if result.get("practitioner_created"):
				practitioners_created += 1
		except Exception:
			errors.append(f"{row.get('tran_num')}: {frappe.get_traceback()}")
			frappe.log_error(title=f"Practitioner Unavailability import failed: {row.get('tran_num')}")

	frappe.db.commit()
	return {
		"ok": True,
		"total": len(rows),
		"created": created,
		"updated": updated,
		"skipped": skipped,
		"practitioners_created": practitioners_created,
		"errors": len(errors),
	}
