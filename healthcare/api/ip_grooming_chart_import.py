"""Import Oracle IP_GROOMING_CHART Excel into IP Grooming Chart rows."""

from __future__ import annotations

import json
from datetime import date, datetime
from typing import Any

import frappe
from frappe import _

from healthcare.api.discharge_checklist_import import _resolve_cost_center
from healthcare.api.patient_info_import import (
	_cell_text,
	_clean_oracle_num,
	_excel_file_path,
	_require_admin,
)
from healthcare.api.visit_diagnoses_op_import import _legacy_data_datetime
from healthcare.api.visit_diagnosis_sync import _resolve_inpatient_admission

DOCTYPE = "IP Grooming Chart"
IP_GROOMING_CHART_IMPORT_BATCH_SIZE = 500
CACHE_TTL = 7200
CACHE_KEYS = {
	"file_url": "healthcare:data_migration:ip_grooming_chart_import:file_url",
	"row_keys": "healthcare:data_migration:ip_grooming_chart_import:row_keys",
	"rows": "healthcare:data_migration:ip_grooming_chart_import:rows",
}

EXCEL_HEADER_MAP = {
	"TRANS_NUM": "trans_num",
	"TRANS_DATE": "trans_date",
	"ADMISSION_NUM": "admission_num",
	"BRUSH_TEETH_M": "brush_teeth_morning",
	"CHANGE_CLOTHES_M": "change_clothes_morning",
	"BRUSH_TEETH_N": "brush_teeth_noon",
	"CHANGE_CLOTHES_N": "change_clothes_noon",
	"SHOWER": "shower",
	"BOWEL": "bowel",
	"BED_WETTING": "bed_wetting",
	"BREAKFAST": "breakfast",
	"SNACKS_1": "snack_1",
	"LUNCH": "lunch",
	"SNACKS_2": "snack_2",
	"DINNER": "dinner",
	"SNACKS_3": "snack_3",
	"WEIGHT": "weight",
	"LMP": "lmp",
	"CR_ID": "cr_id",
	"CR_DATE": "cr_date",
	"UP_ID": "up_id",
	"UP_DATE": "up_date",
	"BRANCH_NUM": "branch_num",
	"ADMISSION_NUM_OLD": "admission_num_old",
}


def _normalize_header(cell: Any) -> str:
	if cell is None:
		return ""
	text = str(cell).strip().upper().replace(" ", "_")
	return EXCEL_HEADER_MAP.get(text, text.lower())


def _oracle_data_value(value: Any) -> str | None:
	if value in (None, ""):
		return None
	if isinstance(value, float) and value == int(value):
		return str(int(value))
	if isinstance(value, int):
		return str(value)
	text = str(value).strip().replace(",", "")
	return text or None


def _parse_date_field(value: Any) -> date | None:
	if value in (None, ""):
		return None
	if isinstance(value, datetime):
		return value.date()
	if isinstance(value, date):
		return value
	return None


def _to_check(value: Any) -> int:
	if value in (None, "", 0, 0.0, "0", False):
		return 0
	text = str(value).strip().upper()
	if text in {"Y", "YES", "TRUE"}:
		return 1
	try:
		return 1 if float(text) != 0 else 0
	except Exception:
		return 0


def _resolve_admission(row: dict) -> str | None:
	admission_num = row.get("admission_num") or ""
	old_num = row.get("admission_num_old") or ""
	for candidate in (admission_num, old_num):
		if not candidate:
			continue
		if frappe.db.exists("Inpatient Admission", candidate):
			return candidate
		name = _resolve_inpatient_admission(candidate)
		if name:
			return name
	return None


def _parse_sheet_rows(ws) -> list[dict]:
	rows_iter = ws.iter_rows(values_only=True)
	try:
		header_row = next(rows_iter)
	except StopIteration:
		return []

	headers = [_normalize_header(h) for h in header_row]
	parsed: list[dict] = []
	for raw in rows_iter:
		if not raw or all(cell is None or str(cell).strip() == "" for cell in raw):
			continue
		row: dict[str, Any] = {}
		for idx, key in enumerate(headers):
			if not key or idx >= len(raw):
				continue
			row[key] = raw[idx]

		trans_num = _clean_oracle_num(row.get("trans_num"))
		if not trans_num:
			continue

		row["trans_num"] = trans_num
		row["admission_num"] = _clean_oracle_num(row.get("admission_num"))
		row["admission_num_old"] = _clean_oracle_num(row.get("admission_num_old"))
		row["cr_id"] = _clean_oracle_num(row.get("cr_id"))
		row["up_id"] = _clean_oracle_num(row.get("up_id"))
		parsed.append(row)
	return parsed


def _parse_excel_rows(file_url: str) -> tuple[list[dict], dict[str, int]]:
	try:
		import openpyxl
	except ImportError:
		frappe.throw(
			_("openpyxl is required to read Excel files. Install it in the bench environment: pip install openpyxl")
		)

	path = _excel_file_path(file_url)
	wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
	all_rows: list[dict] = []
	sheet_row_counts: dict[str, int] = {}
	try:
		for sheet_name in wb.sheetnames:
			sheet_rows = _parse_sheet_rows(wb[sheet_name])
			sheet_row_counts[sheet_name] = len(sheet_rows)
			all_rows.extend(sheet_rows)
	finally:
		wb.close()
	return all_rows, sheet_row_counts


def _build_fields(row: dict) -> tuple[dict[str, Any], dict[str, int]]:
	stats = {"skip_no_admission": 0, "skip_no_patient": 0}

	admission = _resolve_admission(row)
	if not admission:
		stats["skip_no_admission"] = 1
		return {}, stats

	patient = frappe.db.get_value("Inpatient Admission", admission, "patient")
	if not patient:
		stats["skip_no_patient"] = 1
		return {}, stats

	fields: dict[str, Any] = {
		"trans_num": row["trans_num"],
		"admission_no": admission,
		"file_no": patient,
		"admission_no_old": row.get("admission_num_old") or row.get("admission_num") or None,
		"brush_teeth_morning": _to_check(row.get("brush_teeth_morning")),
		"change_clothes_morning": _to_check(row.get("change_clothes_morning")),
		"brush_teeth_noon": _to_check(row.get("brush_teeth_noon")),
		"change_clothes_noon": _to_check(row.get("change_clothes_noon")),
		"shower": _to_check(row.get("shower")),
		"bowel": _to_check(row.get("bowel")),
		"bed_wetting": _to_check(row.get("bed_wetting")),
		"breakfast": _to_check(row.get("breakfast")),
		"snack_1": _to_check(row.get("snack_1")),
		"lunch": _to_check(row.get("lunch")),
		"snack_2": _to_check(row.get("snack_2")),
		"dinner": _to_check(row.get("dinner")),
		"snack_3": _to_check(row.get("snack_3")),
		"cr_id": row.get("cr_id"),
		"up_id": row.get("up_id"),
	}

	trans_date = _parse_date_field(row.get("trans_date"))
	if trans_date:
		fields["date"] = trans_date.isoformat()

	lmp = _parse_date_field(row.get("lmp"))
	if lmp:
		fields["lmp"] = lmp

	if row.get("weight") not in (None, ""):
		try:
			fields["weight"] = float(row.get("weight"))
		except Exception:
			pass

	cr_date = _legacy_data_datetime(row.get("cr_date"))
	if cr_date:
		fields["cr_date"] = cr_date
	up_date = _legacy_data_datetime(row.get("up_date"))
	if up_date:
		fields["up_date"] = up_date

	cost_center = _resolve_cost_center(row.get("branch_num"))
	if cost_center:
		fields["cost_center"] = cost_center

	return {key: value for key, value in fields.items() if value not in (None, "")}, stats


def _preview_counts(rows: list[dict]) -> dict:
	existing = 0
	resolved_admissions = 0
	skip_no_admission = 0
	for row in rows:
		if frappe.db.exists(DOCTYPE, {"trans_num": row["trans_num"]}):
			existing += 1
		if _resolve_admission(row):
			resolved_admissions += 1
		else:
			skip_no_admission += 1

	return {
		"existing_records": existing,
		"new_records": len(rows) - existing,
		"resolved_admissions": resolved_admissions,
		"skip_no_admission": skip_no_admission,
		"sample_trans_nums": [row.get("trans_num") for row in rows[:5]],
	}


def parse_and_cache_excel(file_url: str) -> dict:
	rows, sheet_row_counts = _parse_excel_rows(file_url)
	by_key = {row["trans_num"]: row for row in rows}
	row_keys = list(by_key.keys())

	frappe.cache().set_value(CACHE_KEYS["file_url"], file_url, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(CACHE_KEYS["row_keys"], row_keys, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(
		CACHE_KEYS["rows"],
		json.dumps(by_key, default=str),
		expires_in_sec=CACHE_TTL,
	)

	preview = _preview_counts(list(by_key.values()))
	raw_row_total = sum(sheet_row_counts.values())
	return {
		"excel_rows": len(by_key),
		"raw_excel_rows": raw_row_total,
		"sheets": list(sheet_row_counts.keys()),
		"sheet_row_counts": sheet_row_counts,
		**preview,
	}


def _load_cached_rows() -> dict[str, dict]:
	raw = frappe.cache().get_value(CACHE_KEYS["rows"])
	if not raw:
		return {}
	if isinstance(raw, dict):
		return raw
	return json.loads(raw)


def upsert_ip_grooming_chart(row: dict) -> dict:
	fields, stats = _build_fields(row)
	if not fields.get("trans_num"):
		return {"status": "skip", "trans_num": row.get("trans_num"), **stats}
	if stats.get("skip_no_admission") or stats.get("skip_no_patient"):
		return {"status": "skip", "trans_num": row.get("trans_num"), **stats}

	trans_num = fields["trans_num"]
	existing_name = frappe.db.get_value(DOCTYPE, {"trans_num": trans_num}, "name")
	if existing_name:
		doc = frappe.get_doc(DOCTYPE, existing_name)
		for key, value in fields.items():
			doc.set(key, value)
		doc.flags.ignore_mandatory = True
		doc.flags.ignore_links = True
		doc.flags.ignore_validate = True
		doc.save(ignore_permissions=True)
		action = "updated"
	else:
		doc = frappe.get_doc({"doctype": DOCTYPE, **fields})
		doc.flags.ignore_mandatory = True
		doc.flags.ignore_links = True
		doc.flags.ignore_validate = True
		doc.insert(ignore_permissions=True)
		action = "created"

	return {"status": action, "trans_num": trans_num, **stats}


@frappe.whitelist()
def preview_ip_grooming_chart_import(file_url: str) -> dict:
	_require_admin()
	if not (file_url or "").strip():
		frappe.throw(_("Please upload the IP_GROOMING_CHART Excel file."))
	return parse_and_cache_excel(file_url)


def run_ip_grooming_chart_import_batch(*, offset: int = 0) -> dict:
	row_keys = frappe.cache().get_value(CACHE_KEYS["row_keys"]) or []
	rows_by_key = _load_cached_rows()
	if not row_keys or not rows_by_key:
		return {"processed": offset, "done": True, "batch_count": 0}

	batch_keys = row_keys[offset : offset + IP_GROOMING_CHART_IMPORT_BATCH_SIZE]
	if not batch_keys:
		return {"processed": offset, "done": True, "batch_count": 0}

	created = updated = skipped = 0
	skip_no_admission = skip_no_patient = 0
	errors: list[str] = []

	for key in batch_keys:
		row = rows_by_key.get(key) or {}
		try:
			result = upsert_ip_grooming_chart(row)
			status = result.get("status")
			if status == "created":
				created += 1
			elif status == "updated":
				updated += 1
			else:
				skipped += 1
			skip_no_admission += cint(result.get("skip_no_admission", 0))
			skip_no_patient += cint(result.get("skip_no_patient", 0))
		except Exception:
			errors.append(f"{key}: {frappe.get_traceback()}")
			frappe.log_error(title=f"IP_GROOMING_CHART import failed: {key}")

	frappe.db.commit()
	processed = offset + len(batch_keys)
	return {
		"processed": processed,
		"done": len(batch_keys) < IP_GROOMING_CHART_IMPORT_BATCH_SIZE,
		"batch_count": len(batch_keys),
		"created": created,
		"updated": updated,
		"skipped": skipped,
		"skip_no_admission": skip_no_admission,
		"skip_no_patient": skip_no_patient,
		"errors": len(errors),
	}
