"""Import Oracle VISIT_00_05 Excel into Session Schedule (OP session clinical notes)."""

from __future__ import annotations

import json
from datetime import datetime
from typing import Any

import frappe
from frappe import _
from frappe.utils import getdate

from healthcare.api.discharge_checklist_import import _resolve_cost_center
from healthcare.api.patient_info_import import (
	_cell_text,
	_clean_oracle_num,
	_excel_file_path,
	_parse_date_value,
	_parse_datetime_value,
	_require_admin,
)
from healthcare.api.patient_visit_import import ensure_patient_for_legacy_import
from healthcare.api.service_request_visit_import import (
	_build_healthcare_service_template_index,
	_default_company,
	_order_date_time_from_cr_date,
	_practitioner_for_visit,
	_resolve_healthcare_service_template,
)
from healthcare.api.visit_diagnosis_sync import _resolve_patient_visit

DOCTYPE = "Session Schedule"
VISIT_00_05_IMPORT_BATCH_SIZE = 500
CACHE_TTL = 7200
CACHE_KEYS = {
	"file_url": "healthcare:data_migration:visit_00_05_import:file_url",
	"row_keys": "healthcare:data_migration:visit_00_05_import:row_keys",
	"rows": "healthcare:data_migration:visit_00_05_import:rows",
}

EXCEL_HEADER_MAP = {
	"PATIENT_NUM": "patient_num",
	"VISIT_NUM": "visit_num",
	"SERV_NUM": "serv_num",
	"SR_NUM": "sr_num",
	"DOC_REMARKS": "doc_remarks",
	"BRANCH_NUM": "branch_num",
	"CR_ID": "cr_id",
	"CR_DATE": "cr_date",
	"UP_ID": "up_id",
	"UP_DATE": "up_date",
	"FEEDBACK_REMARKS": "feedback_remarks",
	"FEEDBACK_CR_ID": "feedback_cr_id",
	"FEEDBACK_DATE_TIME": "feedback_date_time",
}


def _normalize_header(cell: Any) -> str:
	if cell is None:
		return ""
	text = str(cell).strip().upper().replace(" ", "_")
	return EXCEL_HEADER_MAP.get(text, text.lower())


def _normalize_visit_num(value: Any) -> str:
	if value in (None, ""):
		return ""
	if isinstance(value, datetime):
		try:
			import openpyxl.utils.datetime

			return str(int(openpyxl.utils.datetime.to_excel(value)))
		except Exception:
			return value.strftime("%Y-%m-%d")
	if isinstance(value, float) and value == int(value):
		return str(int(value))
	return _clean_oracle_num(value)


def _format_legacy_datetime(value: Any) -> str | None:
	dt = _parse_datetime_value(value)
	if dt:
		return dt.strftime("%Y-%m-%d %H:%M:%S")
	parsed_date = _parse_date_value(value)
	if parsed_date:
		return parsed_date.strftime("%Y-%m-%d")
	text = _cell_text(value)
	return text or None


def _row_key(visit_num: str, sr_num: str) -> str:
	return f"{visit_num}::{sr_num}"


def _resolve_patient(patient_num: Any) -> str | None:
	patient = _clean_oracle_num(patient_num)
	if not patient:
		return None
	if frappe.db.exists("Patient", patient):
		return patient
	result = ensure_patient_for_legacy_import(patient)
	if result.get("status") in ("existing", "created"):
		return patient
	return None


def _ensure_patient_visit(visit_num: str, patient: str | None, *, order_date: Any = None) -> str | None:
	visit_num = _normalize_visit_num(visit_num)
	if not visit_num:
		return None

	existing = _resolve_patient_visit(visit_num, patient)
	if existing:
		return existing

	if not patient:
		return None

	order_date_val, order_time_val = _order_date_time_from_cr_date(order_date)
	doc = frappe.new_doc("Patient Visit")
	doc.case_no = visit_num
	doc.patient = patient
	doc.encounter_date = order_date_val or getdate()
	if order_time_val:
		doc.encounter_time = order_time_val
	doc.status = "Completed"
	company = _default_company()
	if company:
		doc.company = company
	doc.flags.ignore_mandatory = True
	doc.flags.ignore_links = True
	doc.flags.from_legacy_import = True
	doc.insert(ignore_permissions=True)
	try:
		doc.flags.ignore_validate = True
		doc.flags.from_legacy_import = True
		doc.submit()
	except Exception:
		frappe.log_error(title=f"VISIT_00_05 legacy visit submit failed: {visit_num}")
	return doc.name


def _find_existing_session_schedule(patient_visit: str | None, sr_num: str) -> str | None:
	if not patient_visit or not sr_num:
		return None
	return frappe.db.get_value(
		DOCTYPE,
		{
			"visit_00_05": 1,
			"patient_visit": patient_visit,
			"sr_num": sr_num,
		},
		"name",
		order_by="modified desc",
	)


def _parse_sheet_rows(ws) -> list[dict]:
	rows_iter = ws.iter_rows(values_only=True)
	try:
		header_row = next(rows_iter)
	except StopIteration:
		return []

	headers = [_normalize_header(h) for h in header_row]
	parsed: list[dict] = []
	for raw in rows_iter:
		if not raw or all(cell is None or str(cell).strip() == "" for cell in raw):
			continue
		row: dict[str, Any] = {}
		for idx, key in enumerate(headers):
			if not key or idx >= len(raw):
				continue
			row[key] = raw[idx]

		visit_num = _normalize_visit_num(row.get("visit_num"))
		sr_num = _clean_oracle_num(row.get("sr_num"))
		if not visit_num or not sr_num:
			continue

		row["visit_num"] = visit_num
		row["sr_num"] = sr_num
		row["patient_num"] = _clean_oracle_num(row.get("patient_num")) or None
		row["serv_num"] = _cell_text(row.get("serv_num")) or None
		row["legacy_record_key"] = _row_key(visit_num, sr_num)
		parsed.append(row)
	return parsed


def _parse_excel_rows(file_url: str) -> tuple[list[dict], dict[str, int]]:
	try:
		import openpyxl
	except ImportError:
		frappe.throw(
			_(
				"openpyxl is required to read Excel files. Install it in the bench environment: pip install openpyxl"
			)
		)

	path = _excel_file_path(file_url)
	wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
	by_key: dict[str, dict] = {}
	sheet_row_counts: dict[str, int] = {}
	try:
		for sheet_name in wb.sheetnames:
			sheet_rows = _parse_sheet_rows(wb[sheet_name])
			sheet_row_counts[sheet_name] = len(sheet_rows)
			for row in sheet_rows:
				by_key[row["legacy_record_key"]] = row
	finally:
		wb.close()
	return list(by_key.values()), sheet_row_counts


def _build_fields(
	row: dict,
	*,
	template_index: dict[str, str] | None = None,
) -> dict[str, Any]:
	patient = _resolve_patient(row.get("patient_num"))
	if not patient:
		return {}

	patient_visit = _ensure_patient_visit(
		row.get("visit_num"),
		patient,
		order_date=row.get("cr_date"),
	)
	template_dn = _resolve_healthcare_service_template(
		row.get("serv_num"),
		template_index=template_index,
	)
	if not template_dn:
		return {}

	order_date, order_time = _order_date_time_from_cr_date(row.get("cr_date"))
	fields: dict[str, Any] = {
		"visit_00_05": 1,
		"sr_num": row.get("sr_num"),
		"patient_num": patient,
		"patient_visit": patient_visit,
		"session_type": template_dn,
		"transaction_status": "Completed",
	}

	if row.get("serv_num"):
		fields["doc_code"] = row["serv_num"]

	if order_date:
		fields["date"] = order_date
	if order_time:
		fields["from_time"] = order_time

	session_name = frappe.db.get_value("Healthcare Service Template", template_dn, "service_name")
	if session_name:
		fields["session_name"] = session_name
	elif row.get("serv_num"):
		fields["session_name"] = row["serv_num"]

	doc_remarks = _cell_text(row.get("doc_remarks"))
	if doc_remarks:
		fields["doc_remarks"] = doc_remarks

	feedback_remarks = _cell_text(row.get("feedback_remarks"))
	if feedback_remarks:
		fields["feedback_remarks"] = feedback_remarks

	cost_center = _resolve_cost_center(row.get("branch_num"))
	if cost_center:
		fields["cost_center"] = cost_center

	company = _default_company()
	if company:
		fields["company"] = company

	if patient_visit:
		practitioner = _practitioner_for_visit(patient_visit)
		if practitioner:
			fields["doctor"] = practitioner

	return {key: value for key, value in fields.items() if value not in (None, "")}


def _preview_counts(rows: list[dict], template_index: dict[str, str]) -> dict:
	existing = 0
	resolved_patients = 0
	resolved_visits = 0
	resolved_templates = 0
	with_doc_remarks = 0
	unmatched_templates: set[str] = set()

	for row in rows:
		patient = _resolve_patient(row.get("patient_num"))
		patient_visit = _resolve_patient_visit(row.get("visit_num"), patient)
		if _find_existing_session_schedule(patient_visit, row.get("sr_num")):
			existing += 1
		if patient:
			resolved_patients += 1
		if patient_visit:
			resolved_visits += 1
		if _resolve_healthcare_service_template(row.get("serv_num"), template_index=template_index):
			resolved_templates += 1
		elif row.get("serv_num"):
			unmatched_templates.add(row["serv_num"])
		if _cell_text(row.get("doc_remarks")):
			with_doc_remarks += 1

	return {
		"existing_records": existing,
		"resolved_patients": resolved_patients,
		"resolved_visits": resolved_visits,
		"resolved_templates": resolved_templates,
		"with_doc_remarks": with_doc_remarks,
		"unmatched_template_count": len(unmatched_templates),
		"sample_unmatched_templates": sorted(unmatched_templates)[:5],
		"sample_record_keys": [row.get("legacy_record_key") for row in rows[:5]],
	}


def parse_and_cache_excel(file_url: str) -> dict:
	rows, sheet_row_counts = _parse_excel_rows(file_url)
	template_index = _build_healthcare_service_template_index()
	row_keys = [row["legacy_record_key"] for row in rows]
	by_key = {row["legacy_record_key"]: row for row in rows}

	frappe.cache().set_value(CACHE_KEYS["file_url"], file_url, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(CACHE_KEYS["row_keys"], row_keys, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(
		CACHE_KEYS["rows"],
		json.dumps(by_key, default=str),
		expires_in_sec=CACHE_TTL,
	)

	preview = _preview_counts(rows, template_index)
	raw_row_total = sum(sheet_row_counts.values())
	return {
		"excel_rows": len(rows),
		"raw_excel_rows": raw_row_total,
		"sheets": list(sheet_row_counts.keys()),
		"sheet_row_counts": sheet_row_counts,
		**preview,
	}


def _load_cached_rows() -> dict[str, dict]:
	raw = frappe.cache().get_value(CACHE_KEYS["rows"])
	if not raw:
		return {}
	if isinstance(raw, dict):
		return raw
	return json.loads(raw)


def upsert_session_schedule_from_row(
	row: dict,
	*,
	template_index: dict[str, str] | None = None,
) -> dict:
	fields = _build_fields(row, template_index=template_index)
	if not fields:
		return {
			"status": "skip",
			"legacy_record_key": row.get("legacy_record_key"),
		}

	legacy_record_key = row.get("legacy_record_key")
	existing_name = _find_existing_session_schedule(
		fields.get("patient_visit"),
		fields.get("sr_num"),
	)
	if existing_name:
		doc = frappe.get_doc(DOCTYPE, existing_name)
		for key, value in fields.items():
			doc.set(key, value)
		doc.visit_00_05 = 1
		doc.flags.ignore_mandatory = True
		doc.flags.ignore_links = True
		doc.flags.ignore_validate = True
		doc.save(ignore_permissions=True)
		action = "updated"
	else:
		doc = frappe.get_doc({"doctype": DOCTYPE, **fields})
		doc.visit_00_05 = 1
		doc.flags.ignore_mandatory = True
		doc.flags.ignore_links = True
		doc.flags.ignore_validate = True
		doc.insert(ignore_permissions=True)
		action = "created"

	return {"status": action, "legacy_record_key": legacy_record_key, "name": doc.name}


@frappe.whitelist()
def preview_visit_00_05_import(file_url: str) -> dict:
	_require_admin()
	if not (file_url or "").strip():
		frappe.throw(_("Please upload the VISIT_00_05 Excel file."))
	return parse_and_cache_excel(file_url)


def run_visit_00_05_import_batch(*, offset: int = 0) -> dict:
	row_keys = frappe.cache().get_value(CACHE_KEYS["row_keys"]) or []
	rows_by_key = _load_cached_rows()
	if not row_keys or not rows_by_key:
		return {"processed": offset, "done": True, "batch_count": 0}

	batch_keys = row_keys[offset : offset + VISIT_00_05_IMPORT_BATCH_SIZE]
	if not batch_keys:
		return {"processed": offset, "done": True, "batch_count": 0}

	template_index = _build_healthcare_service_template_index()
	created = updated = skipped = 0
	errors: list[str] = []

	for key in batch_keys:
		row = rows_by_key.get(key) or {}
		try:
			result = upsert_session_schedule_from_row(row, template_index=template_index)
			status = result.get("status")
			if status == "created":
				created += 1
			elif status == "updated":
				updated += 1
			else:
				skipped += 1
		except Exception:
			errors.append(f"{key}: {frappe.get_traceback()}")
			frappe.log_error(title=f"VISIT_00_05 import failed: {key}")

	frappe.db.commit()
	processed = offset + len(batch_keys)
	return {
		"processed": processed,
		"done": len(batch_keys) < VISIT_00_05_IMPORT_BATCH_SIZE,
		"batch_count": len(batch_keys),
		"created": created,
		"updated": updated,
		"skipped": skipped,
		"errors": len(errors),
	}
