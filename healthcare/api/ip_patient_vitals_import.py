"""Import Oracle IP_PATIENT_VITALS Excel directly into Vital Signs rows."""

from __future__ import annotations

import json
from datetime import date, datetime
from typing import Any

import frappe
from frappe import _
from frappe.utils import cint, flt, getdate, get_datetime, get_time

from healthcare.api.discharge_checklist_import import _resolve_cost_center
from healthcare.api.patient_info_import import (
	_cell_text,
	_clean_oracle_num,
	_excel_file_path,
	_require_admin,
)
from healthcare.api.patient_visit_import import ensure_patient_for_legacy_import
from healthcare.api.visit_diagnoses_op_import import _legacy_data_datetime
from healthcare.api.visit_diagnosis_sync import _resolve_inpatient_admission

IP_PATIENT_VITALS_IMPORT_BATCH_SIZE = 500
CACHE_TTL = 7200
CACHE_KEYS = {
	"file_url": "healthcare:data_migration:ip_patient_vitals_import:file_url",
	"row_keys": "healthcare:data_migration:ip_patient_vitals_import:row_keys",
	"rows": "healthcare:data_migration:ip_patient_vitals_import:rows",
}

EXCEL_HEADER_MAP = {
	"TRANS_NUM": "trans_num",
	"PATIENT_FILE_NO": "patient_file_no",
	"VISIT_NUM": "visit_num",
	"ADMISSION_NUM": "admission_num",
	"ADMISSION_NUM_OLD": "admission_num_old",
	"RECORD_DATE": "record_date",
	"WEIGHT": "weight",
	"HEIGHT": "height",
	"BP_SYSTOLIC": "bp_systolic",
	"BP_DIASTOLIC": "bp_diastolic",
	"PULSE": "pulse",
	"TEMPERATURE": "temperature",
	"TEMP_LOCATION": "temp_location",
	"SPO2": "spo2",
	"REMARKS": "remarks",
	"RECORD_TIME": "record_time",
	"CR_ID": "cr_id",
	"CR_DATE": "cr_date",
	"UP_ID": "up_id",
	"UP_DATE": "up_date",
	"BRANCH_NUM": "branch_num",
	"RESP_RATE": "resp_rate",
}


def _normalize_header(cell: Any) -> str:
	if cell is None:
		return ""
	text = str(cell).strip().upper().replace(" ", "_")
	return EXCEL_HEADER_MAP.get(text, text.lower())


def _oracle_data_value(value: Any) -> str | None:
	if value in (None, ""):
		return None
	if isinstance(value, float) and value == int(value):
		return str(int(value))
	if isinstance(value, int):
		return str(value)
	text = str(value).strip().replace(",", "")
	return text or None


def _oracle_float(value: Any) -> float | None:
	if value in (None, ""):
		return None
	text = str(value).strip().replace(",", "")
	try:
		return flt(text)
	except (TypeError, ValueError):
		return None


def _parse_signs_date(value: Any):
	if value in (None, ""):
		return None
	if isinstance(value, datetime):
		return value.date()
	if isinstance(value, date):
		return value
	try:
		return getdate(value)
	except Exception:
		return None


def _oracle_record_time_to_hms(value: Any) -> str | None:
	if value in (None, ""):
		return None
	text = str(value).strip().replace(",", "")
	try:
		seconds = int(float(text))
	except (TypeError, ValueError):
		return _cell_text(value) or None
	seconds %= 86400
	hours, remainder = divmod(seconds, 3600)
	minutes, secs = divmod(remainder, 60)
	return f"{hours:02d}:{minutes:02d}:{secs:02d}"


def _record_time_raw(value: Any) -> str | None:
	cleaned = _oracle_data_value(value)
	return cleaned


def _resolve_patient_visit(visit_num: Any) -> str | None:
	code = _clean_oracle_num(visit_num)
	if not code:
		return None
	if frappe.db.exists("Patient Visit", code):
		return code
	return None


def _resolve_company(admission: str | None) -> str | None:
	if admission:
		company = frappe.db.get_value("Inpatient Admission", admission, "company")
		if company:
			return company
	return frappe.defaults.get_global_default("company")


def _set_bp_field(fields: dict[str, Any]) -> None:
	systolic = fields.get("bp_systolic")
	diastolic = fields.get("bp_diastolic")
	if systolic and diastolic:
		fields["bp"] = f"{systolic}/{diastolic} mmHg"


def _set_bmi_fields(fields: dict[str, Any]) -> None:
	height = fields.get("height")
	weight = fields.get("weight")
	if not height or not weight:
		return
	height_m = flt(height) / 100
	if not height_m:
		return
	bmi = round(flt(weight) / (height_m * height_m), 2)
	fields["bmi"] = bmi
	if bmi < 18.5:
		fields["nutrition_note"] = "Underweight"
	elif bmi < 25:
		fields["nutrition_note"] = "Normal"
	elif bmi < 30:
		fields["nutrition_note"] = "Overweight"
	else:
		fields["nutrition_note"] = "Obese"


def _parse_sheet_rows(ws) -> list[dict]:
	rows_iter = ws.iter_rows(values_only=True)
	try:
		header_row = next(rows_iter)
	except StopIteration:
		return []

	headers = [_normalize_header(h) for h in header_row]
	parsed: list[dict] = []
	for raw in rows_iter:
		if not raw or all(cell is None or str(cell).strip() == "" for cell in raw):
			continue
		row: dict[str, Any] = {}
		for idx, key in enumerate(headers):
			if not key or idx >= len(raw):
				continue
			row[key] = raw[idx]

		trans_num = _clean_oracle_num(row.get("trans_num"))
		if not trans_num:
			continue

		row["trans_num"] = trans_num
		row["patient_file_no"] = _clean_oracle_num(row.get("patient_file_no"))
		row["visit_num"] = _clean_oracle_num(row.get("visit_num"))
		row["admission_num"] = _clean_oracle_num(row.get("admission_num"))
		row["admission_num_old"] = _clean_oracle_num(row.get("admission_num_old"))
		row["remarks"] = _cell_text(row.get("remarks"))
		row["cr_id"] = _clean_oracle_num(row.get("cr_id"))
		row["up_id"] = _clean_oracle_num(row.get("up_id"))
		parsed.append(row)
	return parsed


def _parse_excel_rows(file_url: str) -> tuple[list[dict], dict[str, int]]:
	try:
		import openpyxl
	except ImportError:
		frappe.throw(
			_("openpyxl is required to read Excel files. Install it in the bench environment: pip install openpyxl")
		)

	path = _excel_file_path(file_url)
	wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
	all_rows: list[dict] = []
	sheet_row_counts: dict[str, int] = {}
	try:
		for sheet_name in wb.sheetnames:
			sheet_rows = _parse_sheet_rows(wb[sheet_name])
			sheet_row_counts[sheet_name] = len(sheet_rows)
			all_rows.extend(sheet_rows)
	finally:
		wb.close()
	return all_rows, sheet_row_counts


def _build_vital_signs_fields(row: dict) -> tuple[dict[str, Any], dict[str, int]]:
	stats = {"patient_created": 0}

	admission_num = row.get("admission_num") or row.get("admission_num_old") or ""
	admission = _resolve_inpatient_admission(admission_num, row.get("patient_file_no"))
	if not admission:
		return {}, {"skip_no_admission": 1}

	patient = row.get("patient_file_no") or None
	if patient and not frappe.db.exists("Patient", patient):
		patient_result = ensure_patient_for_legacy_import(patient)
		if patient_result.get("status") == "created":
			stats["patient_created"] = 1
		elif patient_result.get("status") != "existing":
			patient = None

	if not patient and admission:
		patient = frappe.db.get_value("Inpatient Admission", admission, "patient")
	if not patient:
		return {}, {"skip_no_patient": 1}

	patient_name = frappe.db.get_value("Patient", patient, "patient_name")
	visit = _resolve_patient_visit(row.get("visit_num"))
	signs_date = _parse_signs_date(row.get("record_date"))
	signs_time = _oracle_record_time_to_hms(row.get("record_time"))
	if not signs_date and row.get("cr_date"):
		try:
			signs_date = getdate(get_datetime(row.get("cr_date")))
		except Exception:
			pass
	if not signs_time and row.get("cr_date"):
		try:
			signs_time = get_time(get_datetime(row.get("cr_date"))).strftime("%H:%M:%S")
		except Exception:
			pass

	fields: dict[str, Any] = {
		"trans_no": row["trans_num"],
		"patient": patient,
		"patient_name": patient_name,
		"inpatient_record": admission,
		"old_admission_no": row.get("admission_num_old") or row.get("admission_num") or None,
		"encounter": visit,
		"company": _resolve_company(admission),
		"signs_date": signs_date,
		"signs_time": signs_time,
		"cost_center": _resolve_cost_center(row.get("branch_num")),
		"weight": _oracle_float(row.get("weight")),
		"height": _oracle_float(row.get("height")),
		"bp_systolic": _oracle_data_value(row.get("bp_systolic")),
		"bp_diastolic": _oracle_data_value(row.get("bp_diastolic")),
		"pulse": _oracle_data_value(row.get("pulse")),
		"temperature": _oracle_data_value(row.get("temperature")),
		"location_temperature": _oracle_data_value(row.get("temp_location")),
		"spo2": _oracle_data_value(row.get("spo2")),
		"respiratory_rate": _oracle_data_value(row.get("resp_rate")),
		"remarks": row.get("remarks") or None,
		"record_time": _record_time_raw(row.get("record_time")),
		"cr_id": row.get("cr_id") or None,
		"up_id": row.get("up_id") or None,
	}
	cr_date = _legacy_data_datetime(row.get("cr_date"))
	if cr_date:
		fields["cr_date"] = cr_date
	up_date = _legacy_data_datetime(row.get("up_date"))
	if up_date:
		fields["up_date"] = up_date

	_set_bp_field(fields)
	_set_bmi_fields(fields)

	return {key: value for key, value in fields.items() if value not in (None, "")}, stats


def _reopen_cancelled_vital_signs(doc) -> None:
	"""Legacy import may update rows that were submitted or cancelled; draft first."""
	if doc.docstatus != 2:
		return
	frappe.db.set_value("Vital Signs", doc.name, "docstatus", 0, update_modified=False)
	doc.docstatus = 0


def _persist_and_submit_vital_signs(doc, *, existing: bool) -> bool:
	doc.flags.ignore_validate = True
	doc.flags.ignore_mandatory = True
	doc.flags.ignore_links = True
	doc.flags.ignore_permissions = True
	doc.flags.legacy_import = True

	if existing:
		if doc.docstatus == 1:
			doc.cancel()
		_reopen_cancelled_vital_signs(doc)
		doc.save(ignore_permissions=True)
	else:
		doc.insert(ignore_permissions=True)

	if doc.docstatus != 0:
		return False

	try:
		doc.flags.ignore_validate = True
		doc.flags.ignore_mandatory = True
		doc.flags.legacy_import = True
		doc.submit()
		return True
	except Exception:
		frappe.log_error(
			title=f"Vital Signs submit failed: {doc.name}",
			message=frappe.get_traceback(),
		)
		return False


def upsert_vital_signs_from_row(row: dict) -> dict:
	trans_num = row.get("trans_num")
	if not trans_num:
		return {"status": "skip_no_trans_num"}

	fields, side_stats = _build_vital_signs_fields(row)
	if side_stats.get("skip_no_admission"):
		return {"status": "skip_no_admission", "trans_num": trans_num}
	if side_stats.get("skip_no_patient"):
		return {"status": "skip_no_patient", "trans_num": trans_num}
	if not fields:
		return {"status": "skip_error", "trans_num": trans_num}

	existing = frappe.db.exists("Vital Signs", trans_num)
	if existing:
		doc = frappe.get_doc("Vital Signs", trans_num)
		for key, value in fields.items():
			if key == "trans_no":
				continue
			doc.set(key, value)
		action = "updated"
	else:
		doc = frappe.new_doc("Vital Signs")
		doc.update(fields)
		action = "created"

	submitted = _persist_and_submit_vital_signs(doc, existing=bool(existing))
	return {
		"status": action,
		"trans_num": trans_num,
		"name": doc.name,
		"submitted": submitted,
		**side_stats,
	}


def _preview_counts(rows: list[dict]) -> dict:
	patient_nums = {row["patient_file_no"] for row in rows if row.get("patient_file_no")}
	missing_patients = [pn for pn in patient_nums if not frappe.db.exists("Patient", pn)]
	existing = sum(1 for row in rows if frappe.db.exists("Vital Signs", row["trans_num"]))
	resolved_admissions = 0
	unresolved_admissions = 0
	with_admission = 0

	for row in rows:
		admission_num = row.get("admission_num") or row.get("admission_num_old") or ""
		if not admission_num:
			continue
		with_admission += 1
		if _resolve_inpatient_admission(admission_num, row.get("patient_file_no")):
			resolved_admissions += 1
		else:
			unresolved_admissions += 1

	return {
		"unique_patients": len(patient_nums),
		"patients_to_create": len(missing_patients),
		"sample_patients_to_create": missing_patients[:10],
		"existing_vitals": existing,
		"new_vitals": len(rows) - existing,
		"resolved_admissions": resolved_admissions,
		"unresolved_admissions": unresolved_admissions,
		"with_admission_num": with_admission,
	}


def parse_and_cache_excel(file_url: str) -> dict:
	rows, sheet_row_counts = _parse_excel_rows(file_url)
	by_key = {row["trans_num"]: row for row in rows}
	row_keys = list(by_key.keys())

	frappe.cache().set_value(CACHE_KEYS["file_url"], file_url, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(CACHE_KEYS["row_keys"], row_keys, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(
		CACHE_KEYS["rows"],
		json.dumps(by_key, default=str),
		expires_in_sec=CACHE_TTL,
	)

	preview = _preview_counts(list(by_key.values()))
	raw_row_total = sum(sheet_row_counts.values())
	return {
		"excel_rows": len(by_key),
		"raw_excel_rows": raw_row_total,
		"sheets": list(sheet_row_counts.keys()),
		"sheet_row_counts": sheet_row_counts,
		**preview,
		"sample_trans_nums": row_keys[:5],
	}


def _load_cached_rows() -> dict[str, dict]:
	raw = frappe.cache().get_value(CACHE_KEYS["rows"])
	if not raw:
		return {}
	if isinstance(raw, dict):
		return raw
	return json.loads(raw)


@frappe.whitelist()
def preview_ip_patient_vitals_import(file_url: str) -> dict:
	_require_admin()
	if not (file_url or "").strip():
		frappe.throw(_("Please upload the IP_PATIENT_VITALS Excel file."))
	return parse_and_cache_excel(file_url)


def run_ip_patient_vitals_import_batch(*, offset: int = 0) -> dict:
	row_keys = frappe.cache().get_value(CACHE_KEYS["row_keys"]) or []
	rows_by_key = _load_cached_rows()
	if not row_keys or not rows_by_key:
		return {"processed": offset, "done": True, "batch_count": 0}

	batch_keys = row_keys[offset : offset + IP_PATIENT_VITALS_IMPORT_BATCH_SIZE]
	if not batch_keys:
		return {"processed": offset, "done": True, "batch_count": 0}

	created = updated = submitted = skip_no_admission = skip_no_patient = patients_created = 0
	errors: list[str] = []

	for key in batch_keys:
		row = rows_by_key.get(key) or {}
		try:
			result = upsert_vital_signs_from_row(row)
			status = result.get("status")
			if status == "created":
				created += 1
			elif status == "updated":
				updated += 1
			elif status == "skip_no_admission":
				skip_no_admission += 1
			elif status == "skip_no_patient":
				skip_no_patient += 1
			if result.get("submitted"):
				submitted += 1
			patients_created += cint(result.get("patient_created"))
		except Exception:
			errors.append(f"{key}: {frappe.get_traceback()}")
			frappe.log_error(title=f"IP_PATIENT_VITALS import failed: {key}")

	frappe.db.commit()
	processed = offset + len(batch_keys)
	return {
		"processed": processed,
		"done": len(batch_keys) < IP_PATIENT_VITALS_IMPORT_BATCH_SIZE,
		"batch_count": len(batch_keys),
		"created": created,
		"updated": updated,
		"submitted": submitted,
		"skip_no_admission": skip_no_admission,
		"skip_no_patient": skip_no_patient,
		"patients_created": patients_created,
		"errors": len(errors),
	}
