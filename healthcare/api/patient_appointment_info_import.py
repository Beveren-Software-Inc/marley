"""Import Oracle APPOINTMENTS_INFO_01 Excel into Patient Appointment rows."""

from __future__ import annotations

import json
from datetime import date, datetime
from typing import Any

import frappe
from frappe import _
from frappe.utils import cint, getdate

from healthcare.api.discharge_checklist_import import _resolve_cost_center
from healthcare.api.lab_test_legacy_import import _format_legacy_date_str
from healthcare.api.patient_appointment_old_status_backfill import (
	ensure_healthcare_practitioner,
	target_status_from_old_status,
)
from healthcare.api.patient_info_import import (
	_cell_text,
	_clean_oracle_num,
	_excel_file_path,
	_require_admin,
)
from healthcare.api.patient_visit_import import _default_gender

PATIENT_APPOINTMENT_INFO_IMPORT_BATCH_SIZE = 500
DATA_FIELD_MAX_LEN = 140
CACHE_TTL = 7200
CACHE_KEYS = {
	"file_url": "healthcare:data_migration:patient_appointment_info_import:file_url",
	"app_nums": "healthcare:data_migration:patient_appointment_info_import:app_nums",
	"rows": "healthcare:data_migration:patient_appointment_info_import:rows",
}

EXCEL_HEADER_MAP = {
	"APP_NUM": "app_num",
	"DOC_CODE": "doc_code",
	"APP_DATE": "app_date",
	"APP_TIME": "app_time",
	"PATIENT_NUM": "patient_num",
	"PATIENT_NAME": "patient_name",
	"PATIENT_CONTACT_NUM": "patient_contact_num",
	"APP_REMARKS": "app_remarks",
	"APP_STATUS": "app_status",
	"BRANCH_NUM": "branch_num",
	"CR_ID": "cr_id",
	"CR_DATE": "cr_date",
	"UP_ID": "up_id",
	"UP_DATE": "up_date",
	"STA_FLG": "sta_flg",
	"HH_TIME": "hh_time",
	"MI_TIME": "mi_time",
	"AMPM_TIME": "ampm_time",
}


def _normalize_header(cell: Any) -> str:
	if cell is None:
		return ""
	text = str(cell).strip().upper().replace(" ", "_")
	return EXCEL_HEADER_MAP.get(text, text.lower())


def _truncate_data(value: Any, max_len: int = DATA_FIELD_MAX_LEN) -> str:
	"""Truncate to Frappe Data field limit (varchar 140) so legacy rows do not fail import."""
	text = _cell_text(value)
	if len(text) <= max_len:
		return text
	return text[:max_len]


def _legacy_data_datetime(value: Any) -> str:
	"""Format Oracle date/datetime for Patient Appointment Data fields (cr_date, up_date)."""
	if value is None or value == "":
		return ""
	if isinstance(value, datetime):
		return value.strftime("%Y-%m-%d %H:%M:%S")
	if isinstance(value, date):
		return datetime.combine(value, datetime.min.time()).strftime("%Y-%m-%d %H:%M:%S")
	return _format_legacy_date_str(value)


def _apply_patient_fetch_fields(fields: dict[str, Any], patient: str | None) -> None:
	"""Mirror fetch_from on patient → patient_name, patient_sex (skipped when ignore_validate)."""
	if not patient or not frappe.db.exists("Patient", patient):
		return
	patient_row = frappe.db.get_value(
		"Patient",
		patient,
		["patient_name", "sex"],
		as_dict=True,
	)
	if not patient_row:
		return
	if patient_row.patient_name:
		fields["patient_name"] = _truncate_data(patient_row.patient_name)
	if patient_row.sex:
		fields["patient_sex"] = patient_row.sex


def _apply_practitioner_fetch_fields(fields: dict[str, Any], practitioner: str | None) -> None:
	if not practitioner or not frappe.db.exists("Healthcare Practitioner", practitioner):
		return
	practitioner_name = frappe.db.get_value("Healthcare Practitioner", practitioner, "practitioner_name")
	if practitioner_name:
		fields["practitioner_name"] = _truncate_data(practitioner_name)


def _default_company() -> str | None:
	company = frappe.defaults.get_global_default("company")
	if company and frappe.db.exists("Company", company):
		return company
	rows = frappe.get_all("Company", pluck="name", limit=1)
	return rows[0] if rows else None


def _parse_time_string(value: Any) -> str | None:
	text = _cell_text(value).replace(" ", "")
	if not text:
		return None
	for fmt in ("%I:%M%p", "%I:%M:%S%p", "%H:%M:%S", "%H:%M"):
		try:
			parsed = datetime.strptime(text, fmt)
			return parsed.strftime("%H:%M:%S")
		except ValueError:
			continue
	return None


def _resolve_appointment_time(row: dict) -> str | None:
	hh = row.get("hh_time")
	mi = row.get("mi_time")
	ampm = _cell_text(row.get("ampm_time")).upper()
	if hh not in (None, "") and mi not in (None, "") and ampm in ("AM", "PM"):
		hour = cint(hh)
		minute = cint(mi)
		if ampm == "PM" and hour < 12:
			hour += 12
		elif ampm == "AM" and hour == 12:
			hour = 0
		return f"{hour:02d}:{minute:02d}:00"
	return _parse_time_string(row.get("app_time"))


def _target_status(old_status: str | None, appointment_date) -> str | None:
	code = (old_status or "").strip().upper()
	if code == "C":
		return "Cancelled"
	return target_status_from_old_status(old_status, appointment_date)


def _sta_flg_value(value: Any) -> int:
	if value is None or value == "":
		return 0
	if isinstance(value, (int, float)):
		return 1 if int(value) != 0 else 0
	text = str(value).strip().upper()
	return 1 if text in ("1", "Y", "YES", "TRUE", "T") else 0


def ensure_patient_for_appointment_import(patient_num: str, patient_name: str | None = None) -> dict:
	"""Create a minimal Patient when legacy appointment import references an unknown file number."""
	patient_num = _clean_oracle_num(patient_num)
	if not patient_num:
		return {"status": "skip", "file_no": ""}
	if frappe.db.exists("Patient", patient_num):
		return {"status": "existing", "file_no": patient_num, "patient": patient_num}

	name = (patient_name or "").strip() or _("Patient {0}").format(patient_num)
	doc = frappe.new_doc("Patient")
	doc.file_no = patient_num
	doc.patient_name = name
	doc.first_name = name
	doc.sex = _default_gender()
	doc.flags.ignore_mandatory = True
	doc.flags.ignore_links = True
	doc.flags.from_legacy_import = True
	doc.insert(ignore_permissions=True)
	return {"status": "created", "file_no": patient_num, "patient": patient_num}


def _parse_sheet_rows(ws) -> list[dict]:
	rows_iter = ws.iter_rows(values_only=True)
	try:
		header_row = next(rows_iter)
	except StopIteration:
		return []

	headers = [_normalize_header(h) for h in header_row]
	parsed: list[dict] = []
	for raw in rows_iter:
		if not raw or all(cell is None or str(cell).strip() == "" for cell in raw):
			continue
		row: dict[str, Any] = {}
		for idx, key in enumerate(headers):
			if not key or idx >= len(raw):
				continue
			row[key] = raw[idx]

		app_num = _clean_oracle_num(row.get("app_num"))
		if not app_num:
			continue

		row["app_num"] = app_num
		row["doc_code"] = _clean_oracle_num(row.get("doc_code"))
		row["patient_num"] = _clean_oracle_num(row.get("patient_num"))
		row["patient_name"] = _cell_text(row.get("patient_name"))
		row["patient_contact_num"] = _cell_text(row.get("patient_contact_num"))
		row["app_remarks"] = _cell_text(row.get("app_remarks"))
		row["app_status"] = _cell_text(row.get("app_status")).upper()
		row["app_time"] = _cell_text(row.get("app_time"))
		row["cr_id"] = _clean_oracle_num(row.get("cr_id"))
		row["up_id"] = _clean_oracle_num(row.get("up_id"))
		row["cr_date"] = row.get("cr_date")
		row["up_date"] = row.get("up_date")
		row["hh_time"] = _cell_text(row.get("hh_time")) or row.get("hh_time")
		row["mi_time"] = _cell_text(row.get("mi_time")) or row.get("mi_time")
		row["ampm_time"] = _cell_text(row.get("ampm_time"))
		parsed.append(row)
	return parsed


def _parse_excel_rows(file_url: str) -> tuple[list[dict], dict[str, int]]:
	try:
		import openpyxl
	except ImportError:
		frappe.throw(
			_("openpyxl is required to read Excel files. Install it in the bench environment: pip install openpyxl")
		)

	path = _excel_file_path(file_url)
	wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
	by_app_num: dict[str, dict] = {}
	sheet_row_counts: dict[str, int] = {}
	try:
		for sheet_name in wb.sheetnames:
			sheet_rows = _parse_sheet_rows(wb[sheet_name])
			sheet_row_counts[sheet_name] = len(sheet_rows)
			for row in sheet_rows:
				by_app_num[row["app_num"]] = row
	finally:
		wb.close()
	return list(by_app_num.values()), sheet_row_counts


def _build_appointment_fields(row: dict) -> tuple[dict, dict]:
	"""Return ERP fields and side-effect stats (patient_created, practitioner_created)."""
	stats = {"patient_created": 0, "practitioner_created": 0}
	patient_num = row.get("patient_num") or ""
	patient_name = row.get("patient_name") or ""
	patient = None
	temporary_patient_name = None
	temporary_mobile_no = _truncate_data(row.get("patient_contact_num")) or None

	if patient_num:
		if not frappe.db.exists("Patient", patient_num):
			result = ensure_patient_for_appointment_import(patient_num, patient_name)
			if result.get("status") == "created":
				stats["patient_created"] = 1
			elif result.get("status") != "existing":
				patient_num = ""
		if patient_num and frappe.db.exists("Patient", patient_num):
			patient = patient_num
	elif patient_name:
		temporary_patient_name = _truncate_data(patient_name)

	appointment_date = None
	if row.get("app_date"):
		appointment_date = getdate(row.get("app_date"))

	doc_code = row.get("doc_code") or ""
	practitioner = ""
	if doc_code:
		practitioner, created = ensure_healthcare_practitioner(doc_code)
		if created:
			stats["practitioner_created"] = 1

	old_status = row.get("app_status") or ""
	status = _target_status(old_status, appointment_date) or "Scheduled"
	appointment_time = _resolve_appointment_time(row)
	company = _default_company()
	cost_center = _resolve_cost_center(row.get("branch_num"))

	fields: dict[str, Any] = {
		"trans_no": row["app_num"],
		"appointment_for": "Practitioner",
		"appointment_date": appointment_date,
		"appointment_time": appointment_time,
		"old_time": row.get("app_time") or None,
		"hh_time": _cell_text(row.get("hh_time")) or None,
		"mi_time": _cell_text(row.get("mi_time")) or None,
		"ampm_time": row.get("ampm_time") or None,
		"patient": patient,
		"temporary_patient_name": temporary_patient_name,
		"temporary_mobile_no": temporary_mobile_no,
		"remarks": row.get("app_remarks") or None,
		"old_status": old_status or None,
		"status": status,
		"sta_flg": _sta_flg_value(row.get("sta_flg")),
		"doc_code": doc_code or None,
		"practitioner": practitioner or None,
		"cost_center": cost_center,
		"cr_id": row.get("cr_id") or None,
		"up_id": row.get("up_id") or None,
	}
	cr_date = _legacy_data_datetime(row.get("cr_date"))
	if cr_date:
		fields["cr_date"] = cr_date
	up_date = _legacy_data_datetime(row.get("up_date"))
	if up_date:
		fields["up_date"] = up_date
	if company:
		fields["company"] = company

	if patient:
		_apply_patient_fetch_fields(fields, patient)

	if practitioner:
		_apply_practitioner_fetch_fields(fields, practitioner)
		dept = frappe.db.get_value("Healthcare Practitioner", practitioner, "department")
		if dept and frappe.db.exists("Medical Department", dept):
			fields["department"] = dept

	return {key: value for key, value in fields.items() if value not in (None, "")}, stats


def upsert_patient_appointment_from_row(row: dict) -> dict:
	app_num = row.get("app_num")
	if not app_num:
		return {"status": "skip_no_app_num"}

	fields, side_stats = _build_appointment_fields(row)
	if not fields.get("appointment_date"):
		return {"status": "skip_no_date", "app_num": app_num}

	existing = frappe.db.exists("Patient Appointment", app_num)
	if existing:
		doc = frappe.get_doc("Patient Appointment", app_num)
		for key, value in fields.items():
			if key == "trans_no":
				continue
			doc.set(key, value)
		action = "updated"
	else:
		doc = frappe.new_doc("Patient Appointment")
		doc.update(fields)
		action = "created"

	doc.flags.ignore_validate = True
	doc.flags.ignore_mandatory = True
	doc.flags.ignore_links = True
	doc.flags.legacy_import = True

	if existing:
		doc.save(ignore_permissions=True)
	else:
		doc.insert(ignore_permissions=True)

	return {
		"status": action,
		"app_num": app_num,
		"name": doc.name,
		**side_stats,
	}


def _preview_counts(rows: list[dict]) -> dict:
	patient_nums = {row["patient_num"] for row in rows if row.get("patient_num")}
	missing_patients = [pn for pn in patient_nums if not frappe.db.exists("Patient", pn)]
	walk_ins = sum(1 for row in rows if not row.get("patient_num"))
	existing = sum(1 for row in rows if frappe.db.exists("Patient Appointment", row["app_num"]))
	status_counts: dict[str, int] = {}
	for row in rows:
		code = (row.get("app_status") or "").strip().upper() or "(empty)"
		status_counts[code] = status_counts.get(code, 0) + 1

	return {
		"unique_patients": len(patient_nums),
		"patients_to_create": len(missing_patients),
		"sample_patients_to_create": missing_patients[:10],
		"walk_ins": walk_ins,
		"existing_appointments": existing,
		"new_appointments": len(rows) - existing,
		"status_counts": status_counts,
	}


def parse_and_cache_excel(file_url: str) -> dict:
	rows, sheet_row_counts = _parse_excel_rows(file_url)
	app_nums = [row["app_num"] for row in rows]
	by_app_num = {row["app_num"]: row for row in rows}

	frappe.cache().set_value(CACHE_KEYS["file_url"], file_url, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(CACHE_KEYS["app_nums"], app_nums, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(
		CACHE_KEYS["rows"],
		json.dumps(by_app_num, default=str),
		expires_in_sec=CACHE_TTL,
	)

	preview = _preview_counts(rows)
	raw_row_total = sum(sheet_row_counts.values())
	return {
		"excel_rows": len(rows),
		"raw_excel_rows": raw_row_total,
		"sheets": list(sheet_row_counts.keys()),
		"sheet_row_counts": sheet_row_counts,
		**preview,
		"sample_app_nums": app_nums[:5],
	}


def _load_cached_rows() -> dict[str, dict]:
	raw = frappe.cache().get_value(CACHE_KEYS["rows"])
	if not raw:
		return {}
	if isinstance(raw, dict):
		return raw
	return json.loads(raw)


@frappe.whitelist()
def preview_patient_appointment_info_import(file_url: str) -> dict:
	_require_admin()
	if not (file_url or "").strip():
		frappe.throw(_("Please upload the APPOINTMENTS_INFO_01 Excel file."))
	return parse_and_cache_excel(file_url)


def run_patient_appointment_info_import_batch(*, offset: int = 0) -> dict:
	app_nums = frappe.cache().get_value(CACHE_KEYS["app_nums"]) or []
	rows_by_app = _load_cached_rows()
	if not app_nums or not rows_by_app:
		return {"processed": offset, "done": True, "batch_count": 0}

	batch_nums = app_nums[offset : offset + PATIENT_APPOINTMENT_INFO_IMPORT_BATCH_SIZE]
	if not batch_nums:
		return {"processed": offset, "done": True, "batch_count": 0}

	created = updated = skip_no_date = 0
	patients_created = practitioners_created = 0
	errors: list[str] = []

	for app_num in batch_nums:
		row = rows_by_app.get(app_num) or {}
		try:
			result = upsert_patient_appointment_from_row(row)
			status = result.get("status")
			if status == "created":
				created += 1
			elif status == "updated":
				updated += 1
			elif status == "skip_no_date":
				skip_no_date += 1
			patients_created += cint(result.get("patient_created"))
			practitioners_created += cint(result.get("practitioner_created"))
		except Exception:
			errors.append(f"{app_num}: {frappe.get_traceback()}")
			frappe.log_error(title=f"Patient Appointment APPOINTMENTS_INFO_01 import failed: {app_num}")

	frappe.db.commit()
	processed = offset + len(batch_nums)
	return {
		"processed": processed,
		"done": len(batch_nums) < PATIENT_APPOINTMENT_INFO_IMPORT_BATCH_SIZE,
		"batch_count": len(batch_nums),
		"created": created,
		"updated": updated,
		"skip_no_date": skip_no_date,
		"patients_created": patients_created,
		"practitioners_created": practitioners_created,
		"errors": len(errors),
	}
