"""Import Oracle VISIT_00_01 Excel into Patient Visit records (all sheets)."""

from __future__ import annotations

import json
from datetime import date, datetime, time
from typing import Any

import frappe
from frappe import _
from frappe.utils import cint, flt, getdate

from healthcare.api.patient_info_import import (
	_cell_text,
	_clean_oracle_num,
	_excel_file_path,
	_parse_date_value,
	_parse_datetime_value,
	_require_admin,
	_yn_to_check,
)
from healthcare.api.patient_visit_practitioner import (
	consultant_name_from_admission,
	practitioner_name_from_link,
)
from healthcare.api.visit_diagnosis_sync import _resolve_inpatient_admission

PATIENT_VISIT_IMPORT_BATCH_SIZE = 500
CACHE_TTL = 7200
CACHE_KEYS = {
	"file_url": "healthcare:data_migration:patient_visit_import:file_url",
	"case_nos": "healthcare:data_migration:patient_visit_import:case_nos",
	"rows": "healthcare:data_migration:patient_visit_import:rows",
}

EXCEL_HEADER_MAP = {
	"VISIT_NUM": "case_no",
	"VISIT_DATE": "encounter_date",
	"VISIT_TYPE_NUM": "old_visit_type_",
	"VISIT_DOCTOR_NUM": "visit_doctor_num",
	"VISIT_PATIENT_NUM": "patient",
	"VISIT_DIAGNOSIS_DETAIL": "encounter_comment",
	"VISIT_DOCTOR_CHECKED_YN": "visit_doctor_checked",
	"VISIT_ALLERGIES_HISTORY": "visit_allergies_history",
	"VISIT_MEDICAL_HISTORY": "medical_history",
	"VISIT_PATIENT_SERIAL": "patient_serial",
	"SRV_TRANS_NUM": "srv_trans_no",
	"LAB_TRANS_NUM": "lab_trans_no",
	"IP_ADMISSION_NUM": "ip_admission_no",
	"RV_NUM": "rv_no",
	"BRANCH_NUM": "cost_center",
	"OP_VAT_NUM": "op_vat_no",
	"PATIENT_WEIGHT": "patient_weight",
	"VISIT_BP_1": "visit_bp_1",
	"VISIT_BP_2": "visit_bp2",
	"PRINT_COUNT": "print_count",
	"DAILY_TRANS_NUM": "daily_transfer_no",
}


def _normalize_header(cell: Any) -> str:
	if cell is None:
		return ""
	text = str(cell).strip().upper().replace(" ", "_")
	return EXCEL_HEADER_MAP.get(text, text.lower())


def _resolve_cost_center(branch_label: Any) -> str | None:
	from healthcare.api.discharge_checklist_import import _resolve_cost_center as resolve_cc

	return resolve_cc(branch_label)


def _resolve_practitioner_by_doctors_id(value: Any) -> str | None:
	code = _clean_oracle_num(value)
	if not code:
		return None
	if frappe.db.exists("Healthcare Practitioner", code):
		return code
	return frappe.db.get_value("Healthcare Practitioner", {"doctors_id": code}, "name")


def _default_company() -> str | None:
	company = frappe.defaults.get_global_default("company")
	if company and frappe.db.exists("Company", company):
		return company
	rows = frappe.get_all("Company", pluck="name", limit=1)
	return rows[0] if rows else None


def _parse_sheet_rows(ws, *, datemode: int = 0) -> list[dict]:
	rows_iter = ws.iter_rows(values_only=True)
	try:
		header_row = next(rows_iter)
	except StopIteration:
		return []

	headers = [_normalize_header(h) for h in header_row]
	parsed: list[dict] = []
	for raw in rows_iter:
		if not raw or all(cell is None or str(cell).strip() == "" for cell in raw):
			continue
		row: dict[str, Any] = {}
		for idx, key in enumerate(headers):
			if not key or idx >= len(raw):
				continue
			row[key] = raw[idx]
		case_no = _clean_oracle_num(row.get("case_no"))
		if not case_no:
			continue
		row["case_no"] = case_no
		row["patient"] = _clean_oracle_num(row.get("patient"))
		enc_date = _parse_date_value(row.get("encounter_date"), datemode=datemode)
		if enc_date:
			row["encounter_date"] = enc_date
		parsed.append(row)
	return parsed


def _parse_excel_rows(file_url: str) -> list[dict]:
	try:
		import openpyxl
	except ImportError:
		frappe.throw(
			_("openpyxl is required to read Excel files. Install it in the bench environment: pip install openpyxl")
		)

	path = _excel_file_path(file_url)
	wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
	seen: set[str] = set()
	parsed: list[dict] = []
	try:
		for sheet_name in wb.sheetnames:
			rows = _parse_sheet_rows(wb[sheet_name])
			for row in rows:
				case_no = row.get("case_no")
				if not case_no or case_no in seen:
					continue
				seen.add(case_no)
				parsed.append(row)
	finally:
		wb.close()
	return parsed


def _encounter_time_from_row(row: dict) -> str:
	raw = row.get("encounter_date")
	if isinstance(raw, datetime):
		return raw.strftime("%H:%M:%S")
	dt = _parse_datetime_value(raw)
	if dt:
		return dt.strftime("%H:%M:%S")
	return "00:00:00"


def _build_patient_visit_fields(row: dict) -> dict[str, Any]:
	patient = row.get("patient")
	encounter_date = row.get("encounter_date")
	if isinstance(encounter_date, datetime):
		encounter_date = getdate(encounter_date)
	elif not encounter_date:
		encounter_date = _parse_date_value(row.get("encounter_date"))

	fields: dict[str, Any] = {
		"case_no": row.get("case_no"),
		"patient": patient,
		"encounter_date": encounter_date,
		"encounter_time": _encounter_time_from_row(row),
		"status": "Completed",
		"old_visit_type_": _cell_text(row.get("old_visit_type_")),
		"encounter_comment": _cell_text(row.get("encounter_comment")),
		"visit_doctor_checked": _yn_to_check(row.get("visit_doctor_checked")),
		"visit_allergies_history": _cell_text(row.get("visit_allergies_history")),
		"medical_history": _cell_text(row.get("medical_history")),
		"patient_serial": _clean_oracle_num(row.get("patient_serial")),
		"srv_trans_no": _clean_oracle_num(row.get("srv_trans_no")),
		"lab_trans_no": _clean_oracle_num(row.get("lab_trans_no")),
		"ip_admission_no": _clean_oracle_num(row.get("ip_admission_no")),
		"rv_no": _clean_oracle_num(row.get("rv_no")),
		"op_vat_no": _clean_oracle_num(row.get("op_vat_no")),
		"patient_weight": _cell_text(row.get("patient_weight")),
		"visit_bp_1": _cell_text(row.get("visit_bp_1")),
		"visit_bp2": _cell_text(row.get("visit_bp2")),
		"print_count": _clean_oracle_num(row.get("print_count")),
		"daily_transfer_no": _clean_oracle_num(row.get("daily_transfer_no")),
	}

	practitioner = _resolve_practitioner_by_doctors_id(row.get("visit_doctor_num"))
	if practitioner:
		fields["practitioner"] = practitioner
		practitioner_name = practitioner_name_from_link(practitioner)
		if practitioner_name:
			fields["practitioner_name"] = practitioner_name

	cc = _resolve_cost_center(row.get("cost_center"))
	if cc:
		fields["cost_center"] = cc

	admission_num = fields.get("ip_admission_no")
	admission_name = None
	if admission_num:
		admission_name = _resolve_inpatient_admission(admission_num, patient)
		if admission_name:
			fields["inpatient_record"] = admission_name

	if not fields.get("practitioner_name") and admission_name:
		consultant_name = consultant_name_from_admission(admission_name)
		if consultant_name:
			fields["practitioner_name"] = consultant_name

	company = _default_company()
	if company:
		fields["company"] = company

	return {key: value for key, value in fields.items() if value not in (None, "")}


def _default_gender() -> str:
	for name in ("Male", "Female"):
		if frappe.db.exists("Gender", name):
			return name
	rows = frappe.get_all("Gender", pluck="name", limit=1)
	return rows[0] if rows else "Male"


def ensure_patient_for_legacy_import(file_no: str) -> dict:
	"""Create a minimal Patient when legacy import references an unknown file number."""
	file_no = _clean_oracle_num(file_no)
	if not file_no:
		return {"status": "skip", "file_no": ""}
	if frappe.db.exists("Patient", file_no):
		return {"status": "existing", "file_no": file_no, "patient": file_no}

	doc = frappe.new_doc("Patient")
	doc.file_no = file_no
	doc.patient_name = _("Patient {0}").format(file_no)
	doc.sex = _default_gender()
	doc.flags.ignore_mandatory = True
	doc.flags.ignore_links = True
	doc.flags.from_legacy_import = True
	doc.insert(ignore_permissions=True)
	return {"status": "created", "file_no": file_no, "patient": file_no}


def upsert_patient_visit_from_row(row: dict) -> dict:
	case_no = row.get("case_no")
	patient = row.get("patient")
	patient_created = False
	if not patient:
		return {"status": "skip_no_patient", "case_no": case_no}
	if not frappe.db.exists("Patient", patient):
		patient_result = ensure_patient_for_legacy_import(patient)
		if patient_result.get("status") == "created":
			patient_created = True
		elif patient_result.get("status") != "existing":
			return {"status": "skip_no_patient", "case_no": case_no}

	fields = _build_patient_visit_fields(row)
	if not fields.get("encounter_date"):
		return {"status": "skip_no_date", "case_no": case_no}

	if frappe.db.exists("Patient Visit", case_no):
		doc = frappe.get_doc("Patient Visit", case_no)
		for key, value in fields.items():
			if key == "case_no":
				continue
			doc.set(key, value)
		doc.flags.ignore_mandatory = True
		doc.flags.ignore_links = True
		doc.flags.from_legacy_import = True
		doc.save(ignore_permissions=True)
		action = "updated"
	else:
		doc = frappe.new_doc("Patient Visit")
		doc.update(fields)
		doc.flags.ignore_mandatory = True
		doc.flags.ignore_links = True
		doc.flags.from_legacy_import = True
		doc.insert(ignore_permissions=True)
		action = "created"

	submitted = False
	if doc.docstatus == 0:
		doc.flags.ignore_validate = True
		doc.flags.from_legacy_import = True
		try:
			doc.submit()
			submitted = True
		except Exception:
			frappe.log_error(title=f"Legacy patient visit submit failed: {case_no}")

	return {
		"status": action,
		"case_no": case_no,
		"submitted": submitted,
		"patient_created": patient_created,
	}


def _collect_patients_to_create(rows: list[dict], *, sample_limit: int = 20) -> dict:
	sample: list[dict] = []
	seen_cases: set[str] = set()
	unique_file_nos: set[str] = set()
	missing_count = 0

	for row in rows:
		patient = (row.get("patient") or "").strip()
		case_no = (row.get("case_no") or "").strip()
		if not patient:
			continue
		if frappe.db.exists("Patient", patient):
			continue
		missing_count += 1
		unique_file_nos.add(patient)
		if case_no not in seen_cases and len(seen_cases) < sample_limit:
			seen_cases.add(case_no)
			sample.append({"case_no": case_no, "patient": patient})

	return {
		"patients_to_create": missing_count,
		"unique_patients_to_create": len(unique_file_nos),
		"sample_patients_to_create": sample,
	}


def parse_and_cache_excel(file_url: str) -> dict:
	rows = _parse_excel_rows(file_url)
	by_case_no = {row["case_no"]: row for row in rows}
	case_nos = sorted(by_case_no.keys())

	frappe.cache().set_value(CACHE_KEYS["file_url"], file_url, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(CACHE_KEYS["case_nos"], case_nos, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(
		CACHE_KEYS["rows"],
		json.dumps(by_case_no, default=str),
		expires_in_sec=CACHE_TTL,
	)

	existing = sum(1 for case_no in case_nos if frappe.db.exists("Patient Visit", case_no))
	patients_detail = _collect_patients_to_create(rows)

	return {
		"excel_rows": len(rows),
		"visits": len(case_nos),
		"existing_visits": existing,
		"patients_to_create": patients_detail["patients_to_create"],
		"unique_patients_to_create": patients_detail["unique_patients_to_create"],
		"sample_patients_to_create": patients_detail["sample_patients_to_create"],
		"sample_case_nos": case_nos[:5],
	}


def _load_cached_rows() -> dict[str, dict]:
	raw = frappe.cache().get_value(CACHE_KEYS["rows"])
	if not raw:
		return {}
	if isinstance(raw, dict):
		return raw
	return json.loads(raw)


@frappe.whitelist()
def preview_patient_visit_import(file_url: str) -> dict:
	_require_admin()
	if not (file_url or "").strip():
		frappe.throw(_("Please upload the VISIT_00_01 Excel file."))
	return parse_and_cache_excel(file_url)


def run_patient_visit_import_batch(offset: int = 0) -> dict:
	case_nos = frappe.cache().get_value(CACHE_KEYS["case_nos"]) or []
	rows_by_case = _load_cached_rows()
	if not case_nos or not rows_by_case:
		return {"processed": offset, "done": True, "batch_count": 0}

	batch_keys = case_nos[offset : offset + PATIENT_VISIT_IMPORT_BATCH_SIZE]
	if not batch_keys:
		return {"processed": offset, "done": True, "batch_count": 0}

	created = updated = skip_no_patient = skip_no_date = submitted = patients_created = 0
	errors: list[str] = []

	for case_no in batch_keys:
		row = rows_by_case.get(case_no) or {}
		try:
			result = upsert_patient_visit_from_row(row)
			status = result.get("status")
			if status == "created":
				created += 1
			elif status == "updated":
				updated += 1
			elif status == "skip_no_patient":
				skip_no_patient += 1
			elif status == "skip_no_date":
				skip_no_date += 1
			if result.get("submitted"):
				submitted += 1
			if result.get("patient_created"):
				patients_created += 1
		except Exception:
			errors.append(f"{case_no}: {frappe.get_traceback()}")
			frappe.log_error(title=f"Patient visit import failed: {case_no}")

	frappe.db.commit()
	processed = offset + len(batch_keys)
	return {
		"processed": processed,
		"done": len(batch_keys) < PATIENT_VISIT_IMPORT_BATCH_SIZE,
		"batch_count": len(batch_keys),
		"created": created,
		"updated": updated,
		"skip_no_patient": skip_no_patient,
		"skip_no_date": skip_no_date,
		"submitted": submitted,
		"patients_created": patients_created,
		"errors": len(errors),
	}
