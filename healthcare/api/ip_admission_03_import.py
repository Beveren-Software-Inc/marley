"""Import Oracle IP_ADMISSION_03 Excel (all sheets) into IP Service + IP Service Detail rows."""

from __future__ import annotations

import json
from typing import Any

import frappe
from frappe import _
from frappe.utils import flt

from healthcare.api.discharge_checklist_import import _resolve_cost_center
from healthcare.api.ip_doctor_request_import import _resolve_doctor_code
from healthcare.api.ip_service_legacy_import import (
	_build_healthcare_service_template_index,
	_finalize_legacy_ip_service,
	_format_legacy_date_str,
	_line_date,
	_resolve_service_group,
	_resolve_service_template,
	_to_currency,
)
from healthcare.api.legacy_id_normalize import normalize_legacy_id
from healthcare.api.patient_info_import import (
	_cell_text,
	_clean_oracle_num,
	_excel_file_path,
	_require_admin,
)
from healthcare.api.visit_diagnosis_sync import _resolve_inpatient_admission

IP_ADMISSION_03_IMPORT_BATCH_SIZE = 100
CACHE_TTL = 7200
CACHE_KEYS = {
	"file_url": "healthcare:data_migration:ip_admission_03_import:file_url",
	"trans_nums": "healthcare:data_migration:ip_admission_03_import:trans_nums",
	"groups": "healthcare:data_migration:ip_admission_03_import:groups",
}

EXCEL_HEADER_MAP = {
	"TRANS_NUM": "trans_num",
	"TRANS_DATE": "trans_date",
	"INV_TYPE": "inv_type",
	"INV_NUM": "inv_num",
	"GL_CODE": "gl_code",
	"ADMISSION_NUM": "admission_num",
	"PATIENT_NUM": "patient_num",
	"DOC_NUM": "doc_num",
	"OLD_SERV_NUM": "old_serv_num",
	"SERV_NUM": "serv_num",
	"SERV_AMT": "serv_amt",
	"SERV_NOTE": "serv_note",
	"TIME_FROM": "time_from",
	"TIME_TO": "time_to",
	"FIELD1": "field1",
	"FIELD2": "field2",
	"FIELD3": "field3",
	"FIELD4": "field4",
	"FIELD5": "field5",
	"FIELD6": "field6",
	"FIELD7": "field7",
	"BRANCH_NUM": "branch_num",
	"CR_ID": "cr_id",
	"CR_DATE": "cr_date",
	"UP_ID": "up_id",
	"UP_DATE": "up_date",
	"OTH_TYPE": "oth_type",
	"TRANS_NUM_YEARLY": "trans_num_yearly",
}


def _normalize_header(cell: Any) -> str:
	if cell is None:
		return ""
	text = str(cell).strip().upper().replace(" ", "_")
	return EXCEL_HEADER_MAP.get(text, text.lower())


def _patient_matches_admission(patient_num: str | None, admission: str | None) -> bool:
	if not patient_num or not admission:
		return True
	adm_patient = frappe.db.get_value("Inpatient Admission", admission, "patient")
	if not adm_patient:
		return True
	return adm_patient == patient_num


def _resolve_patient_num(patient_num: str) -> str | None:
	if not patient_num:
		return None
	if frappe.db.exists("Patient", patient_num):
		return patient_num
	plain = normalize_legacy_id(patient_num)
	if plain and frappe.db.exists("Patient", plain):
		return plain
	name = frappe.db.get_value("Patient", {"file_no": patient_num}, "name")
	if name:
		return name
	if plain:
		return frappe.db.get_value("Patient", {"file_no": plain}, "name")
	return None


def _time_text(value: Any) -> str | None:
	text = _cell_text(value)
	return text or None


def _parse_sheet_rows(ws) -> list[dict]:
	rows_iter = ws.iter_rows(values_only=True)
	try:
		header_row = next(rows_iter)
	except StopIteration:
		return []

	headers = [_normalize_header(h) for h in header_row]
	parsed: list[dict] = []
	for raw in rows_iter:
		if not raw or all(cell is None or str(cell).strip() == "" for cell in raw):
			continue
		row: dict[str, Any] = {}
		for idx, key in enumerate(headers):
			if not key or idx >= len(raw):
				continue
			row[key] = raw[idx]

		trans_num = _clean_oracle_num(row.get("trans_num"))
		if not trans_num:
			continue

		row["trans_num"] = trans_num
		row["admission_num"] = _clean_oracle_num(row.get("admission_num"))
		row["patient_num"] = _clean_oracle_num(row.get("patient_num"))
		row["doc_num"] = _clean_oracle_num(row.get("doc_num"))
		row["old_serv_num"] = _clean_oracle_num(row.get("old_serv_num"))
		row["serv_num"] = _cell_text(row.get("serv_num")).replace(",", "").strip()
		row["branch_num"] = _clean_oracle_num(row.get("branch_num"))
		row["cr_id"] = _clean_oracle_num(row.get("cr_id"))
		row["up_id"] = _clean_oracle_num(row.get("up_id"))
		row["trans_num_yearly"] = _clean_oracle_num(row.get("trans_num_yearly"))
		row["oth_type"] = _cell_text(row.get("oth_type")).strip() or None
		parsed.append(row)
	return parsed


def _parse_excel_rows(file_url: str) -> tuple[dict[str, list[dict]], dict[str, int]]:
	try:
		import openpyxl
	except ImportError:
		frappe.throw(
			_(
				"openpyxl is required to read Excel files. Install it in the bench environment: pip install openpyxl"
			)
		)

	path = _excel_file_path(file_url)
	wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
	groups: dict[str, list[dict]] = {}
	sheet_row_counts: dict[str, int] = {}
	try:
		for sheet_name in wb.sheetnames:
			rows = _parse_sheet_rows(wb[sheet_name])
			sheet_row_counts[sheet_name] = len(rows)
			for row in rows:
				groups.setdefault(row["trans_num"], []).append(row)
	finally:
		wb.close()
	return groups, sheet_row_counts


def _build_service_detail_row(
	line: dict,
	*,
	template_index: dict[str, str],
	default_date: Any = None,
	line_no: int = 1,
) -> dict | None:
	serv_num = (line.get("serv_num") or "").strip()
	if not serv_num:
		return None

	service_type = _resolve_service_template(serv_num, template_index=template_index) or serv_num
	template_doc = None
	if frappe.db.exists("Healthcare Service Template", service_type):
		template_doc = frappe.get_cached_doc("Healthcare Service Template", service_type)

	item_code = (template_doc.item_code if template_doc else None) or serv_num
	service_name = (
		(template_doc.service_name or template_doc.display_name if template_doc else None) or serv_num
	)
	amount = _to_currency(line.get("serv_amt"))

	return {
		"date": _line_date(line.get("trans_date"), default_date),
		"service_group": _resolve_service_group(
			template_doc.service_group if template_doc else None,
			template_name=service_type if template_doc else None,
		),
		"service_type": service_type,
		"service_code": item_code,
		"service_name": service_name,
		"sr_no": str(line_no),
		"inv_type": _cell_text(line.get("inv_type")).strip() or None,
		"invoice_num": _cell_text(line.get("inv_num")).strip() or None,
		"gl_code": _cell_text(line.get("gl_code")).strip() or None,
		"old_service_num": line.get("old_serv_num") or None,
		"amount": amount,
		"total_amount": amount,
		"additional_amount": 0.0,
		"discount_amount": 0.0,
		"net_amount": amount,
		"note": _cell_text(line.get("serv_note")).strip() or None,
		"time_from": _time_text(line.get("time_from")),
		"time_to": _time_text(line.get("time_to")),
	}


def _sum_child_amounts(service_rows: list[dict], field: str) -> float:
	return flt(sum(_to_currency(row.get(field)) for row in service_rows), 2)


def upsert_ip_service_from_group(
	lines: list[dict],
	*,
	template_index: dict[str, str] | None = None,
) -> dict:
	if not lines:
		return {"status": "skip_no_data"}

	template_index = template_index or _build_healthcare_service_template_index()
	header = lines[0]
	trans_num = header.get("trans_num") or ""
	if not trans_num:
		return {"status": "skip_no_trans_num"}

	patient_num = header.get("patient_num") or ""
	admission_num = header.get("admission_num") or ""
	admission = _resolve_inpatient_admission(admission_num, patient_num or None)
	if not admission:
		return {"status": "skip_no_admission", "trans_num": trans_num}

	patient = _resolve_patient_num(patient_num) if patient_num else None
	if not patient:
		patient = frappe.db.get_value("Inpatient Admission", admission, "patient")
	if not patient:
		return {"status": "skip_no_patient", "trans_num": trans_num}

	if patient_num and not _patient_matches_admission(patient_num, admission):
		return {"status": "skip_patient_mismatch", "trans_num": trans_num}

	service_rows: list[dict] = []
	for idx, line in enumerate(lines, start=1):
		detail = _build_service_detail_row(
			line,
			template_index=template_index,
			default_date=header.get("trans_date") or header.get("cr_date"),
			line_no=idx,
		)
		if detail:
			service_rows.append(detail)

	if not service_rows:
		return {"status": "skip_no_lines", "trans_num": trans_num}

	existing = frappe.db.exists("IP Service", trans_num)
	if existing:
		doc = frappe.get_doc("IP Service", trans_num)
		action = "updated"
	else:
		doc = frappe.new_doc("IP Service")
		doc.trans_no = trans_num
		action = "created"

	doc.admission_no = admission
	doc.file_number = patient
	doc.type = "Internal Service"
	doc.category = "Medical Service"
	doc.ap_date = _format_legacy_date_str(header.get("trans_date"))
	doc.oth_type = header.get("oth_type") or ""
	doc.sr_num = header.get("trans_num_yearly") or ""
	doc.dr_gl_code = _cell_text(header.get("gl_code")).strip()
	doc.cr_id = header.get("cr_id") or ""
	doc.cr_date = _format_legacy_date_str(header.get("cr_date"))
	doc.up_id = header.get("up_id") or ""
	doc.up_date = _format_legacy_date_str(header.get("up_date"))
	doc.remarks = _cell_text(header.get("serv_note")).strip()
	doc.field_1 = _cell_text(header.get("field1")).strip()

	practitioner = _resolve_doctor_code(header.get("doc_num"))
	if practitioner:
		doc.practioner = practitioner
		doc.practitioner_name = frappe.db.get_value(
			"Healthcare Practitioner", practitioner, "practitioner_name"
		)

	cc = _resolve_cost_center(header.get("branch_num"))
	if cc:
		doc.cost_center = cc
	elif admission:
		cc = frappe.db.get_value("Inpatient Admission", admission, "cost_center")
		if cc:
			doc.cost_center = cc

	doc.total_amount = _sum_child_amounts(service_rows, "total_amount")
	doc.additional_amount = _sum_child_amounts(service_rows, "additional_amount")
	doc.discount_amount = _sum_child_amounts(service_rows, "discount_amount")
	doc.net_amount = _sum_child_amounts(service_rows, "net_amount")

	doc.set("services", [])
	for row in service_rows:
		doc.append("services", row)

	doc.flags.ignore_validate = True
	doc.flags.ignore_mandatory = True
	doc.flags.ignore_links = True
	doc.flags.from_legacy_import = True

	if existing:
		doc.save(ignore_permissions=True)
	else:
		doc.insert(ignore_permissions=True)

	submitted = _finalize_legacy_ip_service(doc)

	return {
		"status": "ok",
		"trans_num": trans_num,
		"action": action,
		"name": doc.name,
		"submitted": submitted,
		"service_lines": len(service_rows),
	}


def _preview_counts(groups: dict[str, list[dict]]) -> dict:
	existing = 0
	resolved_admissions = 0
	unresolved_admissions = 0
	skip_no_patient = 0
	skip_patient_mismatch = 0
	multi_line = 0

	for trans_num, lines in groups.items():
		if frappe.db.exists("IP Service", trans_num):
			existing += 1
		if len(lines) > 1:
			multi_line += 1

		header = lines[0]
		patient_num = header.get("patient_num") or ""
		admission_num = header.get("admission_num") or ""
		admission = _resolve_inpatient_admission(admission_num, patient_num or None)
		if not admission:
			unresolved_admissions += 1
			continue

		resolved_admissions += 1
		patient = _resolve_patient_num(patient_num) if patient_num else None
		if not patient:
			patient = frappe.db.get_value("Inpatient Admission", admission, "patient")
		if not patient:
			skip_no_patient += 1
		elif patient_num and not _patient_matches_admission(patient_num, admission):
			skip_patient_mismatch += 1

	return {
		"existing_services": existing,
		"new_services": len(groups) - existing,
		"resolved_admissions": resolved_admissions,
		"unresolved_admissions": unresolved_admissions,
		"skip_no_patient": skip_no_patient,
		"skip_patient_mismatch": skip_patient_mismatch,
		"multi_line_transactions": multi_line,
		"sample_trans_nums": sorted(groups.keys())[:5],
	}


def parse_and_cache_excel(file_url: str) -> dict:
	groups, sheet_row_counts = _parse_excel_rows(file_url)
	trans_nums = sorted(groups.keys())
	raw_row_total = sum(sheet_row_counts.values())

	frappe.cache().set_value(CACHE_KEYS["file_url"], file_url, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(CACHE_KEYS["trans_nums"], trans_nums, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(
		CACHE_KEYS["groups"],
		json.dumps(groups, default=str),
		expires_in_sec=CACHE_TTL,
	)

	preview = _preview_counts(groups)
	return {
		"excel_rows": len(trans_nums),
		"raw_excel_rows": raw_row_total,
		"sheets": list(sheet_row_counts.keys()),
		"sheet_row_counts": sheet_row_counts,
		**preview,
	}


def _load_cached_groups() -> dict[str, list[dict]]:
	raw = frappe.cache().get_value(CACHE_KEYS["groups"])
	if not raw:
		return {}
	if isinstance(raw, dict):
		return raw
	return json.loads(raw)


@frappe.whitelist()
def preview_ip_admission_03_import(file_url: str) -> dict:
	_require_admin()
	if not (file_url or "").strip():
		frappe.throw(_("Please upload the IP_ADMISSION_03 Excel file."))
	return parse_and_cache_excel(file_url)


def run_ip_admission_03_import_batch(*, offset: int = 0) -> dict:
	trans_nums = frappe.cache().get_value(CACHE_KEYS["trans_nums"]) or []
	groups = _load_cached_groups()
	if not trans_nums or not groups:
		return {"processed": offset, "done": True, "batch_count": 0}

	batch_keys = trans_nums[offset : offset + IP_ADMISSION_03_IMPORT_BATCH_SIZE]
	if not batch_keys:
		return {"processed": offset, "done": True, "batch_count": 0}

	template_index = _build_healthcare_service_template_index()
	created = updated = submitted = 0
	skip_no_admission = skip_no_patient = skip_patient_mismatch = skip_no_lines = 0
	errors: list[str] = []

	for trans_num in batch_keys:
		lines = groups.get(trans_num) or []
		savepoint = f"ip_adm03_{offset}_{trans_num}".replace("/", "_").replace(" ", "_")[:60]
		try:
			frappe.db.savepoint(savepoint)
			result = upsert_ip_service_from_group(lines, template_index=template_index)
			status = result.get("status")
			if status == "ok":
				if result.get("action") == "created":
					created += 1
				elif result.get("action") == "updated":
					updated += 1
				if result.get("submitted"):
					submitted += 1
			elif status == "skip_no_admission":
				skip_no_admission += 1
			elif status == "skip_no_patient":
				skip_no_patient += 1
			elif status == "skip_patient_mismatch":
				skip_patient_mismatch += 1
			elif status == "skip_no_lines":
				skip_no_lines += 1
			else:
				errors.append(f"{trans_num}: {status}")
		except Exception:
			frappe.db.rollback(save_point=savepoint)
			errors.append(f"{trans_num}: {frappe.get_traceback()}")
			frappe.log_error(title=f"IP_ADMISSION_03 import failed: {trans_num}")

	frappe.db.commit()
	processed = offset + len(batch_keys)
	return {
		"processed": processed,
		"done": len(batch_keys) < IP_ADMISSION_03_IMPORT_BATCH_SIZE,
		"batch_count": len(batch_keys),
		"created": created,
		"updated": updated,
		"submitted": submitted,
		"skip_no_admission": skip_no_admission,
		"skip_no_patient": skip_no_patient,
		"skip_patient_mismatch": skip_patient_mismatch,
		"skip_no_lines": skip_no_lines,
		"errors": len(errors),
	}
