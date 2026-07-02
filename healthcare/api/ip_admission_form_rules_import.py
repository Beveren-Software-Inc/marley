"""Import Oracle IP_ADMISSION_FORM_RULES Excel into Admission Form Rules rows."""

from __future__ import annotations

import json
from typing import Any

import frappe
from frappe import _

from healthcare.api.patient_info_import import (
	_cell_text,
	_clean_oracle_num,
	_excel_file_path,
	_require_admin,
)
from healthcare.api.visit_diagnoses_op_import import _legacy_data_datetime

IP_ADMISSION_FORM_RULES_IMPORT_BATCH_SIZE = 500
CACHE_TTL = 7200
CACHE_KEYS = {
	"file_url": "healthcare:data_migration:ip_admission_form_rules_import:file_url",
	"row_keys": "healthcare:data_migration:ip_admission_form_rules_import:row_keys",
	"rows": "healthcare:data_migration:ip_admission_form_rules_import:rows",
}

EXCEL_HEADER_MAP = {
	"TRANS_TYPE": "trans_type",
	"TRANS_NUM": "trans_num",
	"ORDER_NUM": "order_num",
	"ENG_HEADER": "eng_header",
	"ENG_DETAIL": "eng_detail",
	"ARB_HEADER": "arb_header",
	"ARB_DETAIL": "arb_detail",
	"CR_ID": "cr_id",
	"CR_DATE": "cr_date",
	"UP_ID": "up_id",
	"UP_DATE": "up_date",
}


def _normalize_header(cell: Any) -> str:
	if cell is None:
		return ""
	text = str(cell).strip().upper().replace(" ", "_")
	return EXCEL_HEADER_MAP.get(text, text.lower())


def _oracle_data_value(value: Any) -> str | None:
	if value in (None, ""):
		return None
	cleaned = _clean_oracle_num(value)
	if cleaned:
		return cleaned
	text = _cell_text(value)
	return text or None


def _text_value(value: Any) -> str | None:
	text = _cell_text(value)
	return text or None


def _existing_admission_form_rules_name(trans_no: str) -> str | None:
	if frappe.db.exists("Admission Form Rules", trans_no):
		return trans_no
	return frappe.db.get_value("Admission Form Rules", {"trans_no": trans_no}, "name")


def _parse_sheet_rows(ws) -> list[dict]:
	rows_iter = ws.iter_rows(values_only=True)
	try:
		header_row = next(rows_iter)
	except StopIteration:
		return []

	headers = [_normalize_header(h) for h in header_row]
	parsed: list[dict] = []
	for raw in rows_iter:
		if not raw or all(cell is None or str(cell).strip() == "" for cell in raw):
			continue
		row: dict[str, Any] = {}
		for idx, key in enumerate(headers):
			if not key or idx >= len(raw):
				continue
			row[key] = raw[idx]

		trans_no = _oracle_data_value(row.get("trans_num"))
		if not trans_no:
			continue

		row["trans_no"] = trans_no
		row["trans_type"] = _cell_text(row.get("trans_type")) or None
		row["order_num"] = _oracle_data_value(row.get("order_num"))
		row["cr_id"] = _oracle_data_value(row.get("cr_id"))
		row["up_id"] = _oracle_data_value(row.get("up_id"))
		parsed.append(row)
	return parsed


def _parse_excel_rows(file_url: str) -> tuple[list[dict], dict[str, int]]:
	try:
		import openpyxl
	except ImportError:
		frappe.throw(
			_(
				"openpyxl is required to read Excel files. Install it in the bench environment: pip install openpyxl"
			)
		)

	path = _excel_file_path(file_url)
	wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
	all_rows: list[dict] = []
	sheet_row_counts: dict[str, int] = {}
	try:
		for sheet_name in wb.sheetnames:
			sheet_rows = _parse_sheet_rows(wb[sheet_name])
			sheet_row_counts[sheet_name] = len(sheet_rows)
			all_rows.extend(sheet_rows)
	finally:
		wb.close()
	return all_rows, sheet_row_counts


def _build_admission_form_rules_fields(row: dict) -> dict[str, Any]:
	fields: dict[str, Any] = {
		"trans_no": row["trans_no"],
		"trans_type": row.get("trans_type"),
		"order_num": row.get("order_num"),
		"eng_header": _text_value(row.get("eng_header")),
		"eng_detail": _text_value(row.get("eng_detail")),
		"arb_header": _text_value(row.get("arb_header")),
		"arb_detail": _text_value(row.get("arb_detail")),
		"cr_id": row.get("cr_id"),
		"up_id": row.get("up_id"),
	}

	cr_date = _legacy_data_datetime(row.get("cr_date"))
	if cr_date:
		fields["cr_date"] = cr_date
	up_date = _legacy_data_datetime(row.get("up_date"))
	if up_date:
		fields["up_date"] = up_date

	return {key: value for key, value in fields.items() if value not in (None, "")}


def _apply_legacy_import_flags(doc) -> None:
	doc.flags.ignore_validate = True
	doc.flags.ignore_mandatory = True
	doc.flags.ignore_links = True
	doc.flags.ignore_permissions = True
	doc.flags.legacy_import = True


def _persist_admission_form_rules(doc, *, existing: bool) -> None:
	_apply_legacy_import_flags(doc)
	if existing:
		doc.save(ignore_permissions=True)
	else:
		doc.insert(ignore_permissions=True)


def upsert_admission_form_rules_from_row(row: dict) -> dict:
	trans_no = row.get("trans_no")
	if not trans_no:
		return {"status": "skip_no_trans_no"}

	fields = _build_admission_form_rules_fields(row)
	if not fields:
		return {"status": "skip_error", "trans_no": trans_no}

	existing_name = _existing_admission_form_rules_name(trans_no)
	if existing_name:
		doc = frappe.get_doc("Admission Form Rules", existing_name)
		for key, value in fields.items():
			if key == "trans_no":
				continue
			doc.set(key, value)
		action = "updated"
		existing = True
	else:
		doc = frappe.get_doc({"doctype": "Admission Form Rules", "name": trans_no, **fields})
		action = "created"
		existing = False

	_persist_admission_form_rules(doc, existing=existing)
	return {
		"status": action,
		"trans_no": trans_no,
		"name": doc.name,
	}


def _preview_counts(rows: list[dict]) -> dict:
	existing = sum(1 for row in rows if _existing_admission_form_rules_name(row["trans_no"]))
	return {
		"existing_rules": existing,
		"new_rules": len(rows) - existing,
		"sample_trans_nos": [row["trans_no"] for row in rows[:5]],
	}


def parse_and_cache_excel(file_url: str) -> dict:
	rows, sheet_row_counts = _parse_excel_rows(file_url)
	by_key = {row["trans_no"]: row for row in rows}
	row_keys = list(by_key.keys())

	frappe.cache().set_value(CACHE_KEYS["file_url"], file_url, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(CACHE_KEYS["row_keys"], row_keys, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(
		CACHE_KEYS["rows"],
		json.dumps(by_key, default=str),
		expires_in_sec=CACHE_TTL,
	)

	preview = _preview_counts(list(by_key.values()))
	raw_row_total = sum(sheet_row_counts.values())
	return {
		"excel_rows": len(by_key),
		"raw_excel_rows": raw_row_total,
		"sheets": list(sheet_row_counts.keys()),
		"sheet_row_counts": sheet_row_counts,
		**preview,
	}


def _load_cached_rows() -> dict[str, dict]:
	raw = frappe.cache().get_value(CACHE_KEYS["rows"])
	if not raw:
		return {}
	if isinstance(raw, dict):
		return raw
	return json.loads(raw)


@frappe.whitelist()
def preview_ip_admission_form_rules_import(file_url: str) -> dict:
	_require_admin()
	if not (file_url or "").strip():
		frappe.throw(_("Please upload the IP_ADMISSION_FORM_RULES Excel file."))
	return parse_and_cache_excel(file_url)


def run_ip_admission_form_rules_import_batch(*, offset: int = 0) -> dict:
	row_keys = frappe.cache().get_value(CACHE_KEYS["row_keys"]) or []
	rows_by_key = _load_cached_rows()
	if not row_keys or not rows_by_key:
		return {"processed": offset, "done": True, "batch_count": 0}

	batch_keys = row_keys[offset : offset + IP_ADMISSION_FORM_RULES_IMPORT_BATCH_SIZE]
	if not batch_keys:
		return {"processed": offset, "done": True, "batch_count": 0}

	created = updated = 0
	errors: list[str] = []

	for key in batch_keys:
		row = rows_by_key.get(key) or {}
		try:
			result = upsert_admission_form_rules_from_row(row)
			status = result.get("status")
			if status == "created":
				created += 1
			elif status == "updated":
				updated += 1
		except Exception:
			errors.append(f"{key}: {frappe.get_traceback()}")
			frappe.log_error(title=f"IP_ADMISSION_FORM_RULES import failed: {key}")

	frappe.db.commit()
	processed = offset + len(batch_keys)
	return {
		"processed": processed,
		"done": len(batch_keys) < IP_ADMISSION_FORM_RULES_IMPORT_BATCH_SIZE,
		"batch_count": len(batch_keys),
		"created": created,
		"updated": updated,
		"errors": len(errors),
	}
