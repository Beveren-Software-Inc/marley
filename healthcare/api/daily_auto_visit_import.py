"""Import Oracle DAILY_PATIENTS_02 Excel into Patient Visit (Daily Auto Visit)."""

from __future__ import annotations

import json

import frappe
from frappe import _
from frappe.utils import getdate

from healthcare.api.patient_info_import import (
	_cell_text,
	_clean_oracle_num,
	_excel_file_path,
	_parse_date_value,
	_parse_datetime_value,
	_require_admin,
)
from healthcare.api.patient_visit_import import ensure_patient_for_legacy_import
from healthcare.api.utils.api_utility import get_next_transaction_number

DAILY_AUTO_VISIT_IMPORT_BATCH_SIZE = 500
DAILY_AUTO_VISIT_TYPE = "Daily Auto Visit"
CACHE_TTL = 7200
CACHE_KEYS = {
	"file_url": "healthcare:data_migration:daily_auto_visit_import:file_url",
	"keys": "healthcare:data_migration:daily_auto_visit_import:keys",
	"rows": "healthcare:data_migration:daily_auto_visit_import:rows",
}

EXCEL_HEADER_MAP = {
	"INV_NUM": "inv_num",
	"PATIENT_NUM": "patient_num",
	"TRANS_NUM": "trans_num",
	"YEAR_MONTH": "year_month",
	"CR_ID": "cr_id",
	"CR_DATE": "cr_date",
	"UP_ID": "up_id",
	"UP_DATE": "up_date",
}


def _normalize_header(cell) -> str:
	if cell is None:
		return ""
	text = str(cell).strip().upper().replace(" ", "_")
	return EXCEL_HEADER_MAP.get(text, text.lower())


def _row_key(row: dict) -> str | None:
	inv_num = _clean_oracle_num(row.get("inv_num"))
	trans_num = _clean_oracle_num(row.get("trans_num"))
	if inv_num and trans_num:
		return f"{inv_num}|{trans_num}"
	return inv_num or trans_num


def _visit_date_from_row(row: dict, *, datemode: int = 0):
	cr_date = _parse_date_value(row.get("cr_date"), datemode=datemode)
	if cr_date:
		return getdate(cr_date)
	year_month = _cell_text(row.get("year_month"))
	if year_month:
		digits = "".join(ch for ch in year_month if ch.isdigit())
		if len(digits) >= 6:
			try:
				return getdate(f"{digits[:4]}-{digits[4:6]}-01")
			except Exception:
				pass
	return None


def _serialize_row_for_cache(row: dict) -> dict:
	out = dict(row)
	encounter_date = out.get("encounter_date")
	if encounter_date:
		out["encounter_date"] = str(getdate(encounter_date))
	return out


def _parse_sheet_rows(ws, *, datemode: int = 0) -> list[dict]:
	rows_iter = ws.iter_rows(values_only=True)
	try:
		header_row = next(rows_iter)
	except StopIteration:
		return []

	headers = [_normalize_header(h) for h in header_row]
	parsed: list[dict] = []
	for raw in rows_iter:
		if not raw or all(cell is None or str(cell).strip() == "" for cell in raw):
			continue
		row: dict = {}
		for idx, key in enumerate(headers):
			if not key or idx >= len(raw):
				continue
			row[key] = raw[idx]
		key = _row_key(row)
		if not key:
			continue
		row["_key"] = key
		row["inv_num"] = _clean_oracle_num(row.get("inv_num"))
		row["trans_num"] = _clean_oracle_num(row.get("trans_num"))
		row["patient_num"] = _clean_oracle_num(row.get("patient_num"))
		row["year_month"] = _cell_text(row.get("year_month"))
		visit_date = _visit_date_from_row(row, datemode=datemode)
		if visit_date:
			row["encounter_date"] = visit_date
		for dt_key in ("cr_date", "up_date"):
			dt = _parse_datetime_value(row.get(dt_key), datemode=datemode)
			if dt:
				row[dt_key] = dt.isoformat(sep=" ", timespec="seconds")
			elif row.get(dt_key) not in (None, ""):
				row[dt_key] = _cell_text(row.get(dt_key))
		parsed.append(row)
	return parsed


def _parse_excel_rows(file_url: str) -> list[dict]:
	try:
		import openpyxl
	except ImportError:
		frappe.throw(
			_("openpyxl is required to read Excel files. Install it in the bench environment: pip install openpyxl")
		)

	path = _excel_file_path(file_url)
	wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
	seen: set[str] = set()
	parsed: list[dict] = []
	try:
		for sheet_name in wb.sheetnames:
			rows = _parse_sheet_rows(wb[sheet_name], datemode=wb.epoch)
			for row in rows:
				key = row.get("_key")
				if not key or key in seen:
					continue
				seen.add(key)
				parsed.append(row)
	finally:
		wb.close()
	return parsed


def _load_cached_rows() -> dict[str, dict]:
	raw = frappe.cache().get_value(CACHE_KEYS["rows"])
	if not raw:
		return {}
	if isinstance(raw, dict):
		return raw
	return json.loads(raw)


def _cache_is_warm(file_url: str) -> bool:
	cache = frappe.cache()
	return (
		cache.get_value(CACHE_KEYS["file_url"]) == file_url
		and bool(cache.get_value(CACHE_KEYS["keys"]))
		and bool(cache.get_value(CACHE_KEYS["rows"]))
	)


def _count_missing_patients(patient_nums: set[str]) -> int:
	if not patient_nums:
		return 0
	names = list(patient_nums)
	found = set(frappe.get_all("Patient", filters={"name": ["in", names]}, pluck="name"))
	remaining = patient_nums - found
	if remaining:
		found.update(
			frappe.get_all(
				"Patient",
				filters={"file_no": ["in", list(remaining)]},
				pluck="file_no",
			)
		)
	return len(patient_nums - found)


def _count_existing_visits(rows: list[dict]) -> int:
	case_nos: set[str] = set()
	trans_nums: set[str] = set()
	inv_nums: set[str] = set()
	for row in rows:
		inv_num = row.get("inv_num")
		trans_num = row.get("trans_num")
		if inv_num:
			case_nos.add(inv_num)
			inv_nums.add(inv_num)
		elif trans_num:
			case_nos.add(trans_num)
		if trans_num:
			trans_nums.add(trans_num)

	existing_names: set[str] = set()
	if case_nos:
		existing_names.update(
			frappe.get_all("Patient Visit", filters={"name": ["in", list(case_nos)]}, pluck="name")
		)
	existing_trans: set[str] = set()
	if trans_nums:
		existing_trans.update(
			frappe.get_all(
				"Patient Visit",
				filters={"old_trans_num": ["in", list(trans_nums)]},
				pluck="old_trans_num",
			)
		)
	existing_inv: set[str] = set()
	if inv_nums:
		existing_inv.update(
			frappe.get_all(
				"Patient Visit",
				filters={"inv_num": ["in", list(inv_nums)]},
				pluck="inv_num",
			)
		)

	existing = 0
	for row in rows:
		inv_num = row.get("inv_num")
		trans_num = row.get("trans_num")
		if inv_num and (inv_num in existing_names or inv_num in existing_inv):
			existing += 1
		elif trans_num and (trans_num in existing_names or trans_num in existing_trans):
			existing += 1
	return existing


def _build_preview_summary(rows: list[dict]) -> dict:
	patient_nums = {
		_clean_oracle_num(row.get("patient_num"))
		for row in rows
		if _clean_oracle_num(row.get("patient_num"))
	}
	return {
		"excel_rows": len(rows),
		"visits": len(rows),
		"existing_visits": _count_existing_visits(rows),
		"patients_to_create": _count_missing_patients(patient_nums),
	}


def _default_company() -> str | None:
	company = frappe.defaults.get_global_default("company")
	if company and frappe.db.exists("Company", company):
		return company
	rows = frappe.get_all("Company", pluck="name", limit=1)
	return rows[0] if rows else None


def _resolve_case_no(row: dict) -> str | None:
	"""Patient Visit is named by case_no — use INV_NUM, else TRANS_NUM, else generate."""
	inv_num = _clean_oracle_num(row.get("inv_num"))
	trans_num = _clean_oracle_num(row.get("trans_num"))
	if inv_num:
		return inv_num
	if trans_num:
		return trans_num
	return get_next_transaction_number("Patient Visit", fieldname="case_no")


def _resolve_patient(file_no: str | None) -> str | None:
	file_no = _clean_oracle_num(file_no)
	if not file_no:
		return None
	if frappe.db.exists("Patient", file_no):
		return file_no
	match = frappe.db.get_value("Patient", {"file_no": file_no}, "name")
	if match:
		return match
	result = ensure_patient_for_legacy_import(file_no)
	return result.get("patient") if result.get("status") != "skip" else None


def upsert_daily_auto_visit_from_row(row: dict, *, default_company: str | None = None) -> dict:
	case_no = _resolve_case_no(row)
	if not case_no:
		return {"status": "skip", "reason": "no_case_no", "key": row.get("_key")}

	patient = _resolve_patient(row.get("patient_num"))
	if not patient:
		return {"status": "skip", "reason": "no_patient", "key": row.get("_key")}

	encounter_date = row.get("encounter_date")
	if not encounter_date:
		return {"status": "skip", "reason": "no_date", "key": row.get("_key")}

	if frappe.db.exists("Patient Visit", case_no):
		doc = frappe.get_doc("Patient Visit", case_no)
		action = "updated"
	else:
		doc = frappe.new_doc("Patient Visit")
		doc.case_no = case_no
		action = "created"

	doc.patient = patient
	doc.patient_name = frappe.db.get_value("Patient", patient, "patient_name")
	doc.visit_type = DAILY_AUTO_VISIT_TYPE
	doc.encounter_date = getdate(encounter_date)
	doc.encounter_time = "00:00:00"
	doc.status = "Completed"

	company = default_company if default_company is not None else _default_company()
	if company:
		doc.company = company

	if row.get("inv_num"):
		doc.inv_num = row.get("inv_num")
	if row.get("trans_num"):
		doc.old_trans_num = row.get("trans_num")
	if row.get("year_month"):
		doc.year_month = row.get("year_month")

	doc.flags.ignore_mandatory = True
	doc.flags.ignore_links = True
	doc.flags.from_legacy_import = True
	doc.flags.ignore_permissions = True
	if doc.get("__islocal"):
		doc.insert()
	else:
		doc.save()

	submitted = False
	if doc.docstatus == 0:
		doc.flags.ignore_validate = True
		doc.flags.from_legacy_import = True
		try:
			doc.submit()
			submitted = True
		except Exception:
			frappe.log_error(title=f"Daily auto visit submit failed: {doc.name}")

	return {
		"status": action,
		"name": doc.name,
		"case_no": case_no,
		"patient": patient,
		"key": row.get("_key"),
		"submitted": submitted,
	}


@frappe.whitelist()
def preview_daily_auto_visit_import(file_url: str) -> dict:
	_require_admin()
	if not (file_url or "").strip():
		frappe.throw(_("Please upload the DAILY_PATIENTS_02 Excel file."))
	return parse_and_cache_excel(file_url)


def parse_and_cache_excel(file_url: str) -> dict:
	if _cache_is_warm(file_url):
		return _build_preview_summary(list(_load_cached_rows().values()))

	rows = _parse_excel_rows(file_url)
	keys = [row["_key"] for row in rows if row.get("_key")]
	rows_map = {
		row["_key"]: _serialize_row_for_cache(row) for row in rows if row.get("_key")
	}

	cache = frappe.cache()
	cache.set_value(CACHE_KEYS["file_url"], file_url, expires_in_sec=CACHE_TTL)
	cache.set_value(CACHE_KEYS["keys"], keys, expires_in_sec=CACHE_TTL)
	cache.set_value(CACHE_KEYS["rows"], json.dumps(rows_map, default=str), expires_in_sec=CACHE_TTL)
	return _build_preview_summary(rows)


def run_daily_auto_visit_import_batch(*, offset: int = 0) -> dict:
	cache = frappe.cache()
	file_url = cache.get_value(CACHE_KEYS["file_url"])
	keys = cache.get_value(CACHE_KEYS["keys"]) or []
	rows_map = _load_cached_rows()

	if not file_url or not keys or not rows_map:
		return {"done": True, "processed": offset, "error": "missing_cache"}

	batch = keys[offset : offset + DAILY_AUTO_VISIT_IMPORT_BATCH_SIZE]
	created = updated = skipped = submitted = 0
	skip_no_patient = skip_no_date = skip_no_case_no = 0
	errors: list[str] = []
	default_company = _default_company()

	for key in batch:
		row = rows_map.get(key)
		if not row:
			skipped += 1
			continue
		try:
			result = upsert_daily_auto_visit_from_row(row, default_company=default_company)
			status = result.get("status")
			if status == "created":
				created += 1
			elif status == "updated":
				updated += 1
			else:
				skipped += 1
				reason = result.get("reason")
				if reason == "no_patient":
					skip_no_patient += 1
				elif reason == "no_date":
					skip_no_date += 1
				elif reason == "no_case_no":
					skip_no_case_no += 1
			if result.get("submitted"):
				submitted += 1
		except Exception:
			skipped += 1
			tb = frappe.get_traceback()
			errors.append(f"{key}: {tb}")
			if len(errors) <= 3:
				frappe.log_error(title=f"Daily auto visit import failed: {key}", message=tb)

	frappe.db.commit()
	processed = offset + len(batch)
	done = processed >= len(keys)
	return {
		"done": done,
		"processed": processed,
		"total": len(keys),
		"created": created,
		"updated": updated,
		"skipped": skipped,
		"skip_no_patient": skip_no_patient,
		"skip_no_date": skip_no_date,
		"skip_no_case_no": skip_no_case_no,
		"submitted": submitted,
		"errors": len(errors),
		"error_samples": errors[:5],
	}
