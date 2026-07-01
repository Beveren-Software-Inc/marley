"""Import Oracle SRV_00_03 (header) + SRV_00_04 (detail) into IP Service records."""

from __future__ import annotations

import json
from datetime import date, datetime, time
from typing import Any

import frappe
from frappe import _
from frappe.utils import cint, flt, getdate

from healthcare.api.legacy_id_normalize import normalize_legacy_id
from healthcare.api.patient_info_import import _cell_text, _clean_oracle_num, _excel_file_path
from healthcare.api.service_request_visit_import import ensure_patient_visit_for_legacy_service_import
from healthcare.api.visit_diagnosis_sync import (
	_resolve_inpatient_admission,
	_resolve_patient_visit,
)

LEGACY_IP_SERVICE_IMPORT_BATCH_SIZE = 50
CACHE_TTL = 7200
CACHE_KEYS = {
	"header_file_url": "healthcare:data_migration:legacy_ip_service_import:header_file_url",
	"detail_file_url": "healthcare:data_migration:legacy_ip_service_import:detail_file_url",
	"batch_id": "healthcare:data_migration:legacy_ip_service_import:batch_id",
	"trans_nums": "healthcare:data_migration:legacy_ip_service_import:trans_nums",
	"headers": "healthcare:data_migration:legacy_ip_service_import:headers",
	"details": "healthcare:data_migration:legacy_ip_service_import:details",
	"failures": "healthcare:data_migration:legacy_ip_service_import:failures",
}
FAILURE_DETAIL_MAX = 800

HEADER_EXCEL_MAP = {
	"TRANS_NUM": "trans_num",
	"TRANS_DATE": "trans_date",
	"DR_GL_CODE": "dr_gl_code",
	"SUB_DR_GL_CODE": "sub_dr_gl_code",
	"TRANS_SOURCE": "trans_source",
	"VISIT_NUM": "visit_num",
	"STA_FLG": "sta_flg",
	"BRANCH_NUM": "branch_num",
	"TRANS_REMARKS": "trans_remarks",
	"NET_AMOUNT": "net_amount",
	"ADMISSION_NUM": "admission_num",
	"TOTAL_AMOUNT": "total_amount",
	"EXTRA_AMOUNT": "extra_amount",
	"DISCOUNT_AMOUNT": "discount_amount",
	"FIELD5": "field5",
	"FIELD6": "field6",
	"FIELD7": "field7",
	"FIELD8": "field8",
	"FIELD9": "field9",
	"FIELD10": "field10",
	"CR_ID": "cr_id",
	"CR_DATE": "cr_date",
	"UP_ID": "up_id",
	"UP_DATE": "up_date",
}

DETAIL_EXCEL_MAP = {
	"TRANS_NUM": "trans_num",
	"SR_NUM": "sr_num",
	"SRV_GROUP_NUM": "srv_group_num",
	"SRV_SUB_NUM": "srv_sub_num",
	"SRV_AMT_BOOK": "srv_amt_book",
	"SRV_AMT_ADD": "srv_amt_add",
	"SRV_AMT_DISC": "srv_amt_disc",
	"SRV_AMT_NET": "srv_amt_net",
	"STA_FLG": "sta_flg",
	"FIELD1": "field1",
	"FIELD2": "field2",
	"FIELD3": "field3",
	"FIELD4": "field4",
	"FIELD5": "field5",
	"FIELD6": "field6",
	"FIELD7": "field7",
	"FIELD8": "field8",
	"FIELD9": "field9",
	"FIELD10": "field10",
	"CR_ID": "cr_id",
	"CR_DATE": "cr_date",
	"UP_ID": "up_id",
	"UP_DATE": "up_date",
}


def _require_admin() -> None:
	frappe.only_for(("System Manager", "Healthcare Administrator"))


def _clean_visit_num(value: Any) -> str:
	"""VISIT_NUM with thousand-separator commas stripped (e.g. 1,415 → 1415)."""
	return _clean_oracle_num(value)


def _clean_service_code(value: Any) -> str:
	return _cell_text(value).replace(",", "").strip()


def _oracle_flag_to_check(value: Any) -> int:
	if value is None or value == "":
		return 0
	if isinstance(value, (int, float)):
		return 1 if float(value) > 0 else 0
	text = str(value).strip().upper()
	return 1 if text in ("Y", "YES", "1", "TRUE", "T", "2") else 0


def _to_currency(value: Any) -> float:
	if value is None or value == "":
		return 0.0
	if isinstance(value, str):
		value = value.strip().replace(",", "")
	try:
		return flt(value)
	except Exception:
		return 0.0


def _format_legacy_datetime(value: Any):
	if value is None or value == "":
		return None
	if isinstance(value, datetime):
		return value
	if isinstance(value, date):
		return datetime.combine(value, time.min)
	text = str(value).strip()
	if not text:
		return None
	try:
		from frappe.utils import get_datetime

		return get_datetime(text)
	except Exception:
		return None


def _format_legacy_date_str(value: Any) -> str:
	dt = _format_legacy_datetime(value)
	if dt:
		return dt.strftime("%Y-%m-%d %H:%M:%S")
	if value is None:
		return ""
	return str(value).strip()


def _line_date(value: Any, fallback: Any = None):
	dt = _format_legacy_datetime(value) or _format_legacy_datetime(fallback)
	if dt:
		return getdate(dt)
	return getdate()


def _load_workbook_sheets(file_url: str):
	try:
		import openpyxl
	except ImportError:
		frappe.throw(
			_(
				"openpyxl is required to read Excel files. Install it in the bench environment: pip install openpyxl"
			)
		)
	path = _excel_file_path(file_url)
	return openpyxl.load_workbook(path, read_only=True, data_only=True)


def _normalize_header_key(value: Any) -> str:
	return str(value or "").strip().upper()


def _parse_sheet_rows(ws, header_map: dict[str, str]) -> list[dict]:
	rows_iter = ws.iter_rows(values_only=True)
	try:
		header_row = next(rows_iter)
	except StopIteration:
		return []

	headers = [
		header_map.get(_normalize_header_key(h), str(h).strip().lower()) if h is not None else ""
		for h in header_row
	]
	parsed: list[dict] = []
	for raw in rows_iter:
		if not raw or all(cell is None or str(cell).strip() == "" for cell in raw):
			continue
		row: dict[str, Any] = {}
		for idx, key in enumerate(headers):
			if not key or idx >= len(raw):
				continue
			row[key] = raw[idx]
		parsed.append(row)
	return parsed


def _parse_header_excel(file_url: str) -> tuple[list[dict], dict[str, int]]:
	wb = _load_workbook_sheets(file_url)
	by_trans: dict[str, dict] = {}
	sheet_counts: dict[str, int] = {}
	try:
		for sheet_name in wb.sheetnames:
			rows = _parse_sheet_rows(wb[sheet_name], HEADER_EXCEL_MAP)
			sheet_counts[sheet_name] = len(rows)
			for row in rows:
				trans_num = _cell_text(row.get("trans_num")).strip()
				if not trans_num:
					continue
				row["trans_num"] = trans_num
				row["visit_num"] = _clean_visit_num(row.get("visit_num"))
				row["admission_num"] = _clean_oracle_num(row.get("admission_num"))
				row["sub_dr_gl_code"] = _clean_oracle_num(row.get("sub_dr_gl_code"))
				row["dr_gl_code"] = _clean_oracle_num(row.get("dr_gl_code"))
				row["cr_id"] = _clean_oracle_num(row.get("cr_id"))
				row["up_id"] = _clean_oracle_num(row.get("up_id"))
				by_trans[trans_num] = row
	finally:
		wb.close()
	return list(by_trans.values()), sheet_counts


def _parse_detail_excel(file_url: str) -> tuple[list[dict], dict[str, int]]:
	wb = _load_workbook_sheets(file_url)
	seen: set[tuple[str, str, str]] = set()
	parsed: list[dict] = []
	sheet_counts: dict[str, int] = {}
	try:
		for sheet_name in wb.sheetnames:
			rows = _parse_sheet_rows(wb[sheet_name], DETAIL_EXCEL_MAP)
			sheet_counts[sheet_name] = len(rows)
			for row in rows:
				trans_num = _cell_text(row.get("trans_num")).strip()
				srv_sub_num = _clean_service_code(row.get("srv_sub_num"))
				if not trans_num or not srv_sub_num:
					continue
				sr_num = _clean_oracle_num(row.get("sr_num"))
				dedupe_key = (trans_num, sr_num, srv_sub_num)
				if dedupe_key in seen:
					continue
				seen.add(dedupe_key)
				row["trans_num"] = trans_num
				row["sr_num"] = sr_num
				row["srv_sub_num"] = srv_sub_num
				row["srv_group_num"] = _cell_text(row.get("srv_group_num")).strip()
				row["cr_id"] = _clean_oracle_num(row.get("cr_id"))
				row["up_id"] = _clean_oracle_num(row.get("up_id"))
				parsed.append(row)
	finally:
		wb.close()
	return parsed, sheet_counts


def _resolve_cost_center(branch_label: Any) -> str | None:
	from healthcare.api.discharge_checklist_import import _resolve_cost_center as resolve_cc

	return resolve_cc(branch_label)


def _resolve_patient_from_sub_dr(sub_dr: str) -> str | None:
	key = _clean_oracle_num(sub_dr)
	if not key:
		return None
	if frappe.db.exists("Patient", key):
		return key
	plain = normalize_legacy_id(key)
	if plain and frappe.db.exists("Patient", plain):
		return plain
	name = frappe.db.get_value("Patient", {"file_no": key}, "name")
	if name:
		return name
	if plain:
		return frappe.db.get_value("Patient", {"file_no": plain}, "name")
	return None


def _build_healthcare_service_template_index() -> dict[str, str]:
	index: dict[str, str] = {}
	for row in frappe.get_all(
		"Healthcare Service Template",
		fields=["name", "service_id", "item_code", "old_no", "service_group"],
		limit_page_length=0,
	):
		name = (row.name or "").strip()
		if not name:
			continue
		for raw in (row.name, row.service_id, row.item_code, row.old_no):
			key = _clean_service_code(raw).upper()
			if key:
				index[key] = name
	return index


def _resolve_service_template(
	code: Any,
	*,
	template_index: dict[str, str] | None = None,
) -> str | None:
	serv_code = _clean_service_code(code)
	if not serv_code:
		return None
	keys = [serv_code.upper(), serv_code]
	if template_index is not None:
		for key in keys:
			if key in template_index:
				return template_index[key]
	else:
		for candidate in keys:
			if frappe.db.exists("Healthcare Service Template", candidate):
				return candidate
			for field in ("service_id", "item_code", "old_no"):
				name = frappe.db.get_value(
					"Healthcare Service Template",
					{field: candidate},
					"name",
				)
				if name:
					return name
	return None


def _resolve_service_group(
	value: Any,
	*,
	template_name: str | None = None,
) -> str | None:
	text = _cell_text(value).strip()
	candidates: list[str] = []
	if text:
		candidates.extend([text, text.title(), text.upper()])
		candidates.append(" ".join(part.capitalize() for part in text.split()))
	for candidate in candidates:
		if candidate and frappe.db.exists("Service Template Group", candidate):
			return candidate
	if template_name:
		group = frappe.db.get_value("Healthcare Service Template", template_name, "service_group")
		if group and frappe.db.exists("Service Template Group", group):
			return group
	return None


def _sort_detail_lines(detail_lines: list[dict]) -> list[dict]:
	def sort_key(line: dict):
		sr = line.get("sr_num") or ""
		try:
			return float(sr)
		except (TypeError, ValueError):
			return sr

	return sorted(detail_lines, key=sort_key)


def _resolve_patient_context(
	header: dict | None,
	*,
	ensure_visit: bool = True,
) -> dict:
	header = header or {}
	sub_dr = header.get("sub_dr_gl_code") or ""
	visit_num = header.get("visit_num") or ""
	admission_num = header.get("admission_num") or ""

	patient = None
	patient_visit = None
	inpatient_admission = None
	visit_created = False
	patient_created = False

	if visit_num:
		patient_visit = _resolve_patient_visit(visit_num, sub_dr or None)
		if patient_visit:
			patient = frappe.db.get_value("Patient Visit", patient_visit, "patient")
		elif ensure_visit:
			visit_result = ensure_patient_visit_for_legacy_service_import(
				visit_num,
				order_date=header.get("trans_date") or header.get("cr_date"),
			)
			patient_visit = visit_result.get("visit")
			visit_created = bool(visit_result.get("created"))
			patient_created = bool(visit_result.get("patient_created"))
			if patient_visit:
				patient = frappe.db.get_value("Patient Visit", patient_visit, "patient") or visit_num

	if admission_num:
		inpatient_admission = _resolve_inpatient_admission(admission_num, sub_dr or None)
		if inpatient_admission and not patient:
			patient = frappe.db.get_value("Inpatient Admission", inpatient_admission, "patient")

	if not patient:
		patient = _resolve_patient_from_sub_dr(sub_dr)

	return {
		"patient": patient,
		"patient_visit": patient_visit,
		"inpatient_admission": inpatient_admission,
		"visit_created": visit_created,
		"patient_created": patient_created,
	}


def _build_service_detail_rows(
	detail_lines: list[dict],
	*,
	template_index: dict[str, str],
	default_date: Any = None,
) -> tuple[list[dict], int]:
	"""SRV_00_04 child lines → IP Service Detail; service_type = SRV_SUB_NUM (Healthcare Service Template)."""
	rows: list[dict] = []
	unresolved_templates = 0
	for line in _sort_detail_lines(detail_lines):
		srv_sub_num = _clean_service_code(line.get("srv_sub_num"))
		if not srv_sub_num:
			continue

		# Service Type on the child row is the Excel SRV_SUB_NUM (e.g. OP-0092).
		service_type = _resolve_service_template(
			srv_sub_num,
			template_index=template_index,
		) or srv_sub_num
		if not frappe.db.exists("Healthcare Service Template", service_type):
			unresolved_templates += 1

		template_doc = None
		if frappe.db.exists("Healthcare Service Template", service_type):
			template_doc = frappe.get_cached_doc("Healthcare Service Template", service_type)

		item_code = (template_doc.item_code if template_doc else None) or srv_sub_num
		service_name = (
			(template_doc.service_name or template_doc.display_name if template_doc else None)
			or srv_sub_num
		)
		book = _to_currency(line.get("srv_amt_book"))
		add = _to_currency(line.get("srv_amt_add"))
		disc = _to_currency(line.get("srv_amt_disc"))
		net = _to_currency(line.get("srv_amt_net"))
		if net <= 0 and book > 0:
			net = max(book + add - disc, 0)

		rows.append(
			{
				"date": _line_date(line.get("cr_date"), default_date),
				"service_group": _resolve_service_group(
					line.get("srv_group_num"),
					template_name=service_type if template_doc else None,
				),
				"service_type": service_type,
				"service_code": item_code,
				"service_name": service_name,
				"sr_no": line.get("sr_num") or "",
				"amount": book,
				"total_amount": book,
				"additional_amount": add,
				"discount_amount": disc,
				"net_amount": net,
			}
		)
	return rows, unresolved_templates


def _sum_child_amounts(service_rows: list[dict], field: str) -> float:
	return flt(sum(_to_currency(row.get(field)) for row in service_rows), 2)


def _apply_header_amounts(doc, header: dict | None, service_rows: list[dict]) -> None:
	header = header or {}
	header_total = _to_currency(header.get("total_amount"))
	header_extra = _to_currency(header.get("extra_amount"))
	header_disc = _to_currency(header.get("discount_amount"))
	header_net = _to_currency(header.get("net_amount"))

	child_total = _sum_child_amounts(service_rows, "total_amount")
	child_extra = _sum_child_amounts(service_rows, "additional_amount")
	child_disc = _sum_child_amounts(service_rows, "discount_amount")
	child_net = _sum_child_amounts(service_rows, "net_amount")

	doc.total_amount = header_total if header_total else child_total
	doc.additional_amount = header_extra if header_extra else child_extra
	doc.discount_amount = header_disc if header_disc else child_disc
	doc.net_amount = header_net if header_net else child_net


def _apply_header_fields(
	doc,
	header: dict | None,
	ctx: dict,
	*,
	trans_num: str,
) -> None:
	header = header or {}
	doc.trans_no = trans_num
	doc.remarks = _cell_text(header.get("trans_remarks")).strip()
	doc.dr_gl_code = header.get("dr_gl_code") or ""
	doc.sub_dr_gl_code = header.get("sub_dr_gl_code") or ""
	doc.trans_source = _cell_text(header.get("trans_source")).strip()
	doc.sta_flg = _oracle_flag_to_check(header.get("sta_flg"))
	doc.cr_id = header.get("cr_id") or ""
	doc.cr_date = _format_legacy_date_str(header.get("cr_date"))
	doc.up_id = header.get("up_id") or ""
	doc.up_date = _format_legacy_date_str(header.get("up_date"))
	doc.ap_date = _format_legacy_date_str(header.get("trans_date"))

	if ctx.get("patient_visit"):
		doc.patient_visit = ctx["patient_visit"]
	if ctx.get("inpatient_admission"):
		doc.admission_no = ctx["inpatient_admission"]
	if ctx.get("patient"):
		doc.file_number = ctx["patient"]

	cc = _resolve_cost_center(header.get("branch_num"))
	if cc:
		doc.cost_center = cc
	elif doc.patient_visit:
		cc = frappe.db.get_value("Patient Visit", doc.patient_visit, "cost_center")
		if cc:
			doc.cost_center = cc
	elif doc.admission_no:
		cc = frappe.db.get_value("Inpatient Admission", doc.admission_no, "cost_center")
		if cc:
			doc.cost_center = cc

	doc.category = doc.category or "Medical Service"


def _finalize_legacy_ip_service(doc) -> bool:
	submitted = False
	try:
		doc.flags.ignore_validate = True
		doc.flags.ignore_mandatory = True
		doc.flags.from_legacy_import = True
		if doc.docstatus == 0:
			doc.submit()
			submitted = True
	except Exception:
		frappe.log_error(
			title=f"Legacy IP Service submit failed: {doc.name}",
			message=frappe.get_traceback(),
		)
	return submitted


def import_legacy_ip_service(
	trans_num: str,
	header: dict | None,
	detail_lines: list[dict],
	*,
	template_index: dict[str, str] | None = None,
	batch_id: str | None = None,
) -> dict:
	if not header and not detail_lines:
		return {"status": "skip_no_data", "trans_num": trans_num}

	template_index = template_index or _build_healthcare_service_template_index()
	standalone = not header
	if standalone:
		header = {"trans_num": trans_num}
		ctx = {
			"patient": None,
			"patient_visit": None,
			"inpatient_admission": None,
			"visit_created": False,
			"patient_created": False,
		}
	else:
		ctx = _resolve_patient_context(header, ensure_visit=True)

	service_rows, unresolved_template_lines = _build_service_detail_rows(
		detail_lines,
		template_index=template_index,
		default_date=header.get("trans_date") or header.get("cr_date"),
	)

	if not service_rows and not standalone and not detail_lines:
		return {"status": "skip_no_lines", "trans_num": trans_num}

	if not service_rows and detail_lines:
		return {"status": "skip_no_lines", "trans_num": trans_num}

	existing = frappe.db.exists("IP Service", trans_num)
	if existing:
		doc = frappe.get_doc("IP Service", trans_num)
		action = "updated"
	else:
		doc = frappe.new_doc("IP Service")
		doc.trans_no = trans_num
		action = "created"

	_apply_header_fields(doc, header if not standalone else None, ctx, trans_num=trans_num)
	_apply_header_amounts(doc, header if not standalone else None, service_rows)

	doc.set("services", [])
	for row in service_rows:
		doc.append("services", row)

	if not doc.get("services") and not standalone:
		_apply_header_amounts(doc, header, [])

	doc.flags.ignore_validate = True
	doc.flags.ignore_mandatory = True
	doc.flags.ignore_links = True
	doc.flags.from_legacy_import = True

	if existing:
		doc.save(ignore_permissions=True)
	else:
		doc.insert(ignore_permissions=True)

	submitted = _finalize_legacy_ip_service(doc)

	return {
		"status": "ok",
		"trans_num": trans_num,
		"action": action,
		"name": doc.name,
		"standalone": standalone,
		"submitted": submitted,
		"visit_created": ctx.get("visit_created"),
		"patient_created": ctx.get("patient_created"),
		"service_lines": len(service_rows),
		"unresolved_template_lines": unresolved_template_lines,
	}


def _clear_import_failures() -> None:
	frappe.cache().delete_value(CACHE_KEYS["failures"])


def _append_import_failure(trans_num: str, reason: str, detail: str = "") -> None:
	raw = frappe.cache().get_value(CACHE_KEYS["failures"]) or []
	failures = list(raw) if isinstance(raw, list) else []
	failures.append(
		{
			"trans_num": trans_num,
			"reason": reason,
			"detail": (detail or "")[:FAILURE_DETAIL_MAX],
		}
	)
	frappe.cache().set_value(CACHE_KEYS["failures"], failures, expires_in_sec=CACHE_TTL)


def parse_and_cache_excel(header_file_url: str, detail_file_url: str) -> dict:
	headers_list, header_sheet_counts = _parse_header_excel(header_file_url)
	details_list, detail_sheet_counts = _parse_detail_excel(detail_file_url)

	headers_by_trans = {row["trans_num"]: row for row in headers_list}
	details_by_trans: dict[str, list[dict]] = {}
	for row in details_list:
		details_by_trans.setdefault(row["trans_num"], []).append(row)

	trans_nums = sorted(set(headers_by_trans.keys()) | set(details_by_trans.keys()))
	batch_id = frappe.generate_hash(length=10)
	_clear_import_failures()

	template_index = _build_healthcare_service_template_index()
	header_trans = set(headers_by_trans.keys())
	detail_only_trans = set(details_by_trans.keys()) - header_trans

	resolvable_visit = 0
	visits_to_create = 0
	with_service_lines = 0
	standalone_with_lines = 0
	matching_templates = 0

	for trans_num in trans_nums:
		header = headers_by_trans.get(trans_num)
		lines = details_by_trans.get(trans_num) or []
		if header:
			visit_num = header.get("visit_num") or ""
			if visit_num:
				if _resolve_patient_visit(visit_num):
					resolvable_visit += 1
				else:
					visits_to_create += 1
		elif lines:
			standalone_with_lines += 1
		if lines:
			with_service_lines += 1
			if any(
				_resolve_service_template(line.get("srv_sub_num"), template_index=template_index)
				for line in lines
			):
				matching_templates += 1

	frappe.cache().set_value(CACHE_KEYS["header_file_url"], header_file_url, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(CACHE_KEYS["detail_file_url"], detail_file_url, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(CACHE_KEYS["batch_id"], batch_id, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(CACHE_KEYS["trans_nums"], trans_nums, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(
		CACHE_KEYS["headers"],
		json.dumps(headers_by_trans, default=str),
		expires_in_sec=CACHE_TTL,
	)
	frappe.cache().set_value(
		CACHE_KEYS["details"],
		json.dumps(details_by_trans, default=str),
		expires_in_sec=CACHE_TTL,
	)

	return {
		"batch_id": batch_id,
		"header_rows": len(headers_list),
		"detail_rows": len(details_list),
		"header_sheet_row_counts": header_sheet_counts,
		"detail_sheet_row_counts": detail_sheet_counts,
		"transactions": len(trans_nums),
		"transactions_with_header": len(header_trans),
		"transactions_with_service_lines": with_service_lines,
		"resolvable_visits": resolvable_visit,
		"visits_to_create": visits_to_create,
		"matching_templates": matching_templates,
		"detail_without_header": len(detail_only_trans),
		"standalone_transactions": standalone_with_lines,
		"unique_serv_sub_nums": len(
			{_clean_service_code(row.get("srv_sub_num")) for row in details_list if row.get("srv_sub_num")}
		),
	}


def _load_cached_headers() -> dict[str, dict]:
	raw = frappe.cache().get_value(CACHE_KEYS["headers"])
	if not raw:
		return {}
	if isinstance(raw, dict):
		return raw
	return json.loads(raw)


def _load_cached_details() -> dict[str, list[dict]]:
	raw = frappe.cache().get_value(CACHE_KEYS["details"])
	if not raw:
		return {}
	if isinstance(raw, dict):
		return raw
	return json.loads(raw)


@frappe.whitelist()
def preview_legacy_ip_service_import(header_file_url: str, detail_file_url: str) -> dict:
	_require_admin()
	if not (header_file_url or "").strip():
		frappe.throw(_("Please upload the SRV_00_03 Excel file (header / parent)."))
	if not (detail_file_url or "").strip():
		frappe.throw(_("Please upload the SRV_00_04 Excel file (detail lines — both sheets)."))
	return parse_and_cache_excel(header_file_url, detail_file_url)


def run_legacy_ip_service_import_batch(offset: int = 0) -> dict:
	trans_nums = frappe.cache().get_value(CACHE_KEYS["trans_nums"]) or []
	headers = _load_cached_headers()
	details = _load_cached_details()
	batch_id = frappe.cache().get_value(CACHE_KEYS["batch_id"]) or ""
	template_index = _build_healthcare_service_template_index()

	if not trans_nums:
		return {"processed": offset, "done": True, "batch_count": 0}

	batch_keys = trans_nums[offset : offset + LEGACY_IP_SERVICE_IMPORT_BATCH_SIZE]
	if not batch_keys:
		return {"processed": offset, "done": True, "batch_count": 0}

	ok = created = updated = standalone_ok = skip_no_data = skip_no_template = skip_no_lines = 0
	visits_created = patients_created = submitted = 0
	errors: list[str] = []

	for trans_num in batch_keys:
		header = headers.get(trans_num)
		lines = details.get(trans_num) or []
		if not header and not lines:
			skip_no_data += 1
			_append_import_failure(trans_num, "skip_no_data")
			continue

		savepoint = f"legacy_ip_srv_{offset}_{trans_num}".replace("/", "_").replace(" ", "_")[:60]
		try:
			frappe.db.savepoint(savepoint)
			result = import_legacy_ip_service(
				trans_num,
				header,
				lines,
				template_index=template_index,
				batch_id=batch_id,
			)
			status = result.get("status")
			if status == "ok":
				ok += 1
				if result.get("action") == "created":
					created += 1
				elif result.get("action") == "updated":
					updated += 1
				if result.get("standalone"):
					standalone_ok += 1
				if result.get("visit_created"):
					visits_created += 1
				if result.get("patient_created"):
					patients_created += 1
				if result.get("submitted"):
					submitted += 1
			elif status == "skip_no_template":
				skip_no_template += 1
				_append_import_failure(trans_num, status)
			elif status == "skip_no_lines":
				skip_no_lines += 1
				_append_import_failure(trans_num, status)
			elif status == "skip_no_data":
				skip_no_data += 1
				_append_import_failure(trans_num, status)
			else:
				errors.append(f"{trans_num}: {status}")
				_append_import_failure(trans_num, status or "unknown")
		except Exception:
			frappe.db.rollback(save_point=savepoint)
			tb = frappe.get_traceback()
			errors.append(f"{trans_num}: {tb}")
			_append_import_failure(trans_num, "exception", tb)
			frappe.log_error(title=f"Legacy IP Service import failed: {trans_num}")

	frappe.db.commit()
	processed = offset + len(batch_keys)
	return {
		"processed": processed,
		"done": len(batch_keys) < LEGACY_IP_SERVICE_IMPORT_BATCH_SIZE,
		"batch_count": len(batch_keys),
		"ok": ok,
		"created": created,
		"updated": updated,
		"standalone_ok": standalone_ok,
		"skip_no_data": skip_no_data,
		"skip_no_template": skip_no_template,
		"skip_no_lines": skip_no_lines,
		"visits_created": visits_created,
		"patients_created": patients_created,
		"submitted": submitted,
		"errors": len(errors),
	}
