"""Import psychiatric max-dose Excel into Item custom fields."""

from __future__ import annotations

from typing import Any

import frappe
from frappe import _

from healthcare.api.patient_info_import import _cell_text, _excel_file_path, _require_admin

EXCEL_HEADER_MAP = {
	"ITEM_CODE": "item_code",
	"ITEM_NAME": "item_name",
	"ACTIVE_SUBSTANCES": "active_substances",
	"DRUG_CATEGORY": "drug_category",
	"ITEM_GROUP": "item_group",
	"ROUTE_OF_ADMINISTRATION": "route_of_administration",
	"MAX_DOSE_PER_SINGLE_DOSE": "max_dose_per_single_dose",
	"MAX_DOSE_PER_DAY": "max_dose_per_day",
	"HIGH_ALERT": "high_alert",
	"CLINICAL_NOTES": "clinical_notes",
}


def _normalize_header(cell: Any) -> str:
	if cell is None:
		return ""
	text = str(cell).strip().upper()
	for ch in ("?", "-", "(", ")"):
		text = text.replace(ch, " ")
	text = "_".join(text.split())
	return EXCEL_HEADER_MAP.get(text, text.lower())


def _is_duplicate_row(item_name: Any) -> bool:
	text = _cell_text(item_name)
	return text.startswith("Duplicated from item code")


def _parse_high_alert(value: Any) -> int:
	text = (_cell_text(value) or "").upper()
	return 1 if text in ("YES", "Y", "1", "TRUE") else 0


def _resolve_item(item_code: Any) -> str | None:
	code = _cell_text(item_code)
	if not code:
		return None
	if frappe.db.exists("Item", code):
		return code
	return frappe.db.get_value("Item", {"item_code": code}, "name")


def _parse_sheet_rows(ws) -> tuple[list[dict], int]:
	rows_iter = ws.iter_rows(values_only=True)
	try:
		header_row = next(rows_iter)
	except StopIteration:
		return [], 0

	headers = [_normalize_header(h) for h in header_row]
	parsed: list[dict] = []
	skipped_duplicates = 0

	for raw in rows_iter:
		if not raw or all(cell is None or str(cell).strip() == "" for cell in raw):
			continue

		row: dict[str, Any] = {}
		for idx, key in enumerate(headers):
			if not key or idx >= len(raw):
				continue
			row[key] = raw[idx]

		item_code = _cell_text(row.get("item_code"))
		if not item_code:
			continue
		if _is_duplicate_row(row.get("item_name")):
			skipped_duplicates += 1
			continue

		row["item_code"] = item_code
		parsed.append(row)

	return parsed, skipped_duplicates


def _parse_excel_rows(file_url: str) -> tuple[list[dict], dict[str, int], int]:
	try:
		import openpyxl
	except ImportError:
		frappe.throw(
			_(
				"openpyxl is required to read Excel files. Install it in the bench environment: pip install openpyxl"
			)
		)

	path = _excel_file_path(file_url)
	wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
	by_code: dict[str, dict] = {}
	sheet_row_counts: dict[str, int] = {}
	skipped_duplicates = 0

	try:
		for sheet_name in wb.sheetnames:
			sheet_rows, sheet_skipped = _parse_sheet_rows(wb[sheet_name])
			sheet_row_counts[sheet_name] = len(sheet_rows)
			skipped_duplicates += sheet_skipped
			for row in sheet_rows:
				by_code[row["item_code"]] = row
	finally:
		wb.close()

	return list(by_code.values()), sheet_row_counts, skipped_duplicates


def _preview_stats(rows: list[dict]) -> dict:
	matched_items = 0
	missing_items = 0
	missing_codes: list[str] = []

	for row in rows:
		if _resolve_item(row.get("item_code")):
			matched_items += 1
		else:
			missing_items += 1
			if len(missing_codes) < 10:
				missing_codes.append(row.get("item_code") or "")

	return {
		"matched_items": matched_items,
		"missing_items": missing_items,
		"sample_item_codes": [row.get("item_code") for row in rows[:5]],
		"sample_missing_codes": missing_codes,
	}


def update_item_from_row(row: dict) -> dict:
	item_name = _resolve_item(row.get("item_code"))
	if not item_name:
		return {"status": "not_found", "item_code": row.get("item_code")}

	doc = frappe.get_doc("Item", item_name)
	doc.custom_max_dose_per_single_dose = _cell_text(row.get("max_dose_per_single_dose")) or None
	doc.custom_max_dose_per_day = _cell_text(row.get("max_dose_per_day")) or None
	doc.custom_drug_category = _cell_text(row.get("drug_category")) or None
	doc.custom_high_alert = _parse_high_alert(row.get("high_alert"))

	clinical_notes = _cell_text(row.get("clinical_notes"))
	if clinical_notes:
		doc.description = clinical_notes

	doc.flags.ignore_validate = True
	doc.save(ignore_permissions=True)
	return {"status": "updated", "item_code": row.get("item_code"), "name": item_name}


@frappe.whitelist()
def preview_psychiatric_item_max_dose_import(file_url: str) -> dict:
	_require_admin()
	if not (file_url or "").strip():
		frappe.throw(_("Please upload the psychiatric max-dose Excel file."))

	rows, sheet_row_counts, skipped_duplicates = _parse_excel_rows(file_url)
	stats = _preview_stats(rows)
	raw_row_total = sum(sheet_row_counts.values()) + skipped_duplicates

	return {
		"excel_rows": len(rows),
		"raw_excel_rows": raw_row_total,
		"skipped_duplicate_rows": skipped_duplicates,
		"sheets": list(sheet_row_counts.keys()),
		"sheet_row_counts": sheet_row_counts,
		**stats,
	}


@frappe.whitelist()
def run_psychiatric_item_max_dose_import(file_url: str) -> dict:
	"""Update Item max-dose fields synchronously from Excel."""
	_require_admin()
	if not (file_url or "").strip():
		frappe.throw(_("Please upload the psychiatric max-dose Excel file."))

	rows, _, skipped_duplicates = _parse_excel_rows(file_url)
	updated = not_found = skipped = 0
	errors: list[str] = []

	for row in rows:
		try:
			result = update_item_from_row(row)
			status = result.get("status")
			if status == "updated":
				updated += 1
			elif status == "not_found":
				not_found += 1
			else:
				skipped += 1
		except Exception:
			errors.append(f"{row.get('item_code')}: {frappe.get_traceback()}")
			frappe.log_error(title=f"Psychiatric max-dose import failed: {row.get('item_code')}")

	frappe.db.commit()
	return {
		"ok": True,
		"total": len(rows),
		"updated": updated,
		"not_found": not_found,
		"skipped": skipped + skipped_duplicates,
		"skipped_duplicate_rows": skipped_duplicates,
		"errors": len(errors),
	}
