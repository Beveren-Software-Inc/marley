"""Import TRICARE / insurance price list Excel files into Health Insurance inclusive items."""

from __future__ import annotations

import re
from contextlib import contextmanager
from typing import Any

import frappe
from frappe import _
from frappe.utils import cint, flt, today

from healthcare.api.discharge_checklist_import import _excel_file_path


@contextmanager
def _mute_frappe_messages():
	previous = getattr(frappe.flags, "mute_messages", False)
	frappe.flags.mute_messages = True
	try:
		yield
	finally:
		frappe.flags.mute_messages = previous


def _load_workbook(file_url: str):
	try:
		import openpyxl
	except ImportError:
		frappe.throw(
			_("openpyxl is required to read Excel files. Install it in the bench environment: pip install openpyxl")
		)

	path = _excel_file_path(file_url)
	# read_only can skip rows/cells on some uploaded workbooks
	return openpyxl.load_workbook(path, data_only=True)


def _norm_name(value: Any) -> str:
	return re.sub(r"\s+", " ", (str(value or "")).strip()).upper()


def _norm_code(value: Any) -> str:
	return (str(value or "")).strip().upper()


def _has_discount(*values) -> bool:
	for value in values:
		if value is None or value == "":
			continue
		if isinstance(value, (int, float)) and flt(value) > 0:
			return True
		text = str(value).strip().upper()
		if text in ("ACTUAL CHARGES", "N/A", "NA", "-"):
			continue
		try:
			if flt(value) > 0:
				return True
		except Exception:
			pass
	return False


def _is_data_row(values: tuple) -> bool:
	if not values:
		return False
	return not all(cell is None or str(cell).strip() == "" for cell in values)


def _is_iop_service_name(name: Any) -> bool:
	text = _norm_name(name)
	if not text:
		return False
	return text.startswith("IOP-") or text.startswith("IOP ")


def _is_ip_service_code(code: Any) -> bool:
	return bool(re.match(r"^IP-\d", _norm_code(code)))


def _default_selling_price_list() -> str:
	return (
		frappe.db.get_single_value("Selling Settings", "selling_price_list")
		or frappe.db.get_value("Price List", {"selling": 1, "enabled": 1})
		or "Standard Selling"
	)


def _parse_price(value: Any) -> float | None:
	if value is None or value == "":
		return None
	text = str(value).strip().upper()
	if text in ("ACTUAL CHARGES", "N/A", "NA", "-"):
		return None
	try:
		amount = flt(value)
		return amount if amount > 0 else None
	except Exception:
		return None


def _apply_item_pricing(item_code: str, rate: Any) -> None:
	"""Set Item standard_rate and selling Item Price from the Excel amount."""
	amount = _parse_price(rate)
	if not item_code or amount is None:
		return

	with _mute_frappe_messages():
		frappe.db.set_value("Item", item_code, "standard_rate", amount, update_modified=False)
		price_list = _default_selling_price_list()
		existing = frappe.db.get_value(
			"Item Price",
			{"item_code": item_code, "price_list": price_list, "selling": 1},
			"name",
		)
		if existing:
			frappe.db.set_value("Item Price", existing, "price_list_rate", amount, update_modified=False)
		else:
			frappe.get_doc(
				{
					"doctype": "Item Price",
					"item_code": item_code,
					"price_list": price_list,
					"selling": 1,
					"price_list_rate": amount,
					"valid_from": today(),
				}
			).insert(ignore_permissions=True)


def _default_lab_item_group() -> str:
	for candidate in ("Laboratory", "Lab Tests", "Lab", "Services"):
		if frappe.db.exists("Item Group", candidate):
			return candidate
	return frappe.db.get_value("Item Group", {}, "name") or "All Item Groups"


def _default_service_item_group(category: str = "Medical Service") -> str:
	mapping = {
		"Medical Service": "Medical Services",
		"Other Service": "Other Services",
	}
	target = mapping.get(category, "Healthcare Services")
	if not frappe.db.exists("Item Group", target):
		with _mute_frappe_messages():
			frappe.get_doc(
				{
					"doctype": "Item Group",
					"item_group_name": target,
					"parent_item_group": "All Item Groups",
				}
			).insert(ignore_permissions=True)
	return target


def _service_id_from_name(service_name: Any, fallback_code: Any = None) -> str:
	if fallback_code:
		code = _norm_code(fallback_code)
		if code:
			return code[:140]

	text = _norm_name(service_name)
	text = text.replace("IOP ", "IOP-")
	text = re.sub(r"[^A-Z0-9\-/]+", "-", text)
	text = re.sub(r"-+", "-", text).strip("-")
	return (text or "SERVICE")[:140]


def _template_payload(name: str, display_name: str | None, item_code: str | None) -> dict:
	return {
		"name": name,
		"lab_test_name": display_name,
		"service_name": display_name,
		"item_code": item_code,
	}


def _item_name(item_code: str | None) -> str | None:
	if not item_code:
		return None
	return frappe.db.get_value("Item", item_code, "item_name")


def _lookup_lab_template_direct(test_code: Any, test_name: Any = None) -> dict | None:
	code = (str(test_code or "")).strip()
	if not code:
		return None

	for candidate in (code, code.upper()):
		if frappe.db.exists("Lab Test Template", candidate):
			row = frappe.db.get_value(
				"Lab Test Template", candidate, ["name", "lab_test_name", "item"], as_dict=True
			)
			return _template_payload(row.name, row.lab_test_name, row.item)

	for field in ("no", "name", "lab_test_code"):
		for candidate in (code, code.upper()):
			row = frappe.db.get_value(
				"Lab Test Template",
				{field: candidate},
				["name", "lab_test_name", "item"],
				as_dict=True,
			)
			if row:
				return _template_payload(row.name, row.lab_test_name, row.item)

	if test_name:
		clean = str(test_name).strip()
		row = frappe.db.get_value(
			"Lab Test Template",
			{"lab_test_name": clean},
			["name", "lab_test_name", "item"],
			as_dict=True,
		)
		if row:
			return _template_payload(row.name, row.lab_test_name, row.item)

	return None


def _build_lab_template_index() -> dict[str, dict]:
	by_code: dict[str, dict] = {}
	by_name: dict[str, dict] = {}
	for row in frappe.get_all(
		"Lab Test Template",
		fields=["name", "no", "lab_test_name", "lab_test_code", "item"],
		limit_page_length=0,
	):
		payload = _template_payload(row.name, row.lab_test_name, row.item)
		for key in (row.no, row.name, row.lab_test_code):
			code_key = (key or "").strip().upper()
			if code_key:
				by_code[code_key] = payload
		name = _norm_name(row.lab_test_name)
		if name:
			by_name[name] = payload
	return {"by_code": by_code, "by_name": by_name}


def _build_service_index() -> dict[str, dict]:
	by_id: dict[str, dict] = {}
	by_name: dict[str, dict] = {}
	for row in frappe.get_all(
		"Healthcare Service Template",
		fields=["name", "service_id", "service_name", "display_name", "item_code"],
		limit_page_length=0,
	):
		payload = _template_payload(row.name, row.service_name or row.display_name, row.item_code)
		for key in (row.service_id, row.name):
			code = (key or "").strip().upper()
			if code:
				by_id[code] = payload
		for field in (row.service_name, row.display_name):
			key = _norm_name(field)
			if key:
				by_name[key] = payload
	return {"by_id": by_id, "by_name": by_name}


def _register_lab_in_index(index: dict, payload: dict, test_code: str, test_name: str | None = None) -> None:
	code = _norm_code(test_code)
	if code:
		index["by_code"][code] = payload
	name = _norm_name(test_name)
	if name:
		index["by_name"][name] = payload


def _register_service_in_index(
	index: dict,
	payload: dict,
	service_id: str | None = None,
	service_name: str | None = None,
) -> None:
	if service_id:
		index["by_id"][_norm_code(service_id)] = payload
	if service_name:
		index["by_name"][_norm_name(service_name)] = payload


def _resolve_lab_template(test_code: Any, test_name: Any, index: dict) -> dict | None:
	direct = _lookup_lab_template_direct(test_code, test_name)
	if direct:
		return direct

	code = _norm_code(test_code)
	name = _norm_name(test_name)
	if code and code in index["by_code"]:
		return index["by_code"][code]
	if name and name in index["by_name"]:
		return index["by_name"][name]
	return None


def _lookup_service_template_direct(
	service_code: Any = None,
	service_name: Any = None,
) -> dict | None:
	code = (str(service_code or "")).strip()
	if code:
		if frappe.db.exists("Healthcare Service Template", code):
			row = frappe.db.get_value(
				"Healthcare Service Template",
				code,
				["name", "service_name", "display_name", "item_code"],
				as_dict=True,
			)
			return _template_payload(row.name, row.service_name or row.display_name, row.item_code)

		row = frappe.db.get_value(
			"Healthcare Service Template",
			{"service_id": code},
			["name", "service_name", "display_name", "item_code"],
			as_dict=True,
		)
		if row:
			return _template_payload(row.name, row.service_name or row.display_name, row.item_code)

	if service_name:
		clean = str(service_name).strip()
		for field in ("service_name", "display_name"):
			row = frappe.db.get_value(
				"Healthcare Service Template",
				{field: clean},
				["name", "service_name", "display_name", "item_code"],
				as_dict=True,
			)
			if row:
				return _template_payload(row.name, row.service_name or row.display_name, row.item_code)

	return None


def _resolve_service_template(
	service_code: Any = None,
	service_name: Any = None,
	index: dict | None = None,
) -> dict | None:
	direct = _lookup_service_template_direct(service_code, service_name)
	if direct:
		return direct

	index = index or _build_service_index()
	code = _norm_code(service_code)
	name = _norm_name(service_name)
	if code and code in index["by_id"]:
		return index["by_id"][code]
	if name and name in index["by_name"]:
		return index["by_name"][name]

	if name:
		alt = name.replace("IOP ", "IOP-")
		if alt in index["by_name"]:
			return index["by_name"][alt]

	return None


def _ensure_lab_item(item_code: str, item_name: str, item_group: str, rate: Any = None) -> str:
	if not frappe.db.exists("Item", item_code):
		uom = frappe.db.exists("UOM", "Unit") or frappe.db.get_single_value("Stock Settings", "stock_uom") or "Nos"
		amount = _parse_price(rate)
		with _mute_frappe_messages():
			item = frappe.get_doc(
				{
					"doctype": "Item",
					"item_code": item_code,
					"item_name": item_name,
					"item_group": item_group,
					"is_sales_item": 1,
					"is_service_item": 1,
					"is_purchase_item": 0,
					"is_stock_item": 0,
					"include_item_in_manufacturing": 0,
					"show_in_website": 0,
					"is_pro_applicable": 0,
					"disabled": 0,
					"stock_uom": uom,
					"standard_rate": amount or 0,
				}
			)
			item.flags.ignore_mandatory = True
			item.insert(ignore_permissions=True)

	_apply_item_pricing(item_code, rate)
	return item_code


def _get_or_create_lab_template(
	test_code: Any,
	test_name: Any,
	index: dict,
	rate: Any = None,
	created: dict | None = None,
) -> dict | None:
	code = (str(test_code or "")).strip()
	if not code:
		return None

	existing = _resolve_lab_template(test_code, test_name, index)
	if existing:
		return existing

	test_name_clean = str(test_name or code).strip()
	item_group = _default_lab_item_group()
	price = _parse_price(rate)

	with _mute_frappe_messages():
		doc = frappe.get_doc(
			{
				"doctype": "Lab Test Template",
				"no": code,
				"lab_test_code": code,
				"lab_test_name": test_name_clean,
				"lab_test_rate": price or 0,
				"price": price or 0,
				"is_billable": 1,
				"lab_item_group": item_group,
				"lab_test_template_type": "Single",
			}
		)
		doc.flags.ignore_mandatory = True
		doc.insert(ignore_permissions=True)

		item_code = doc.item
		if not item_code:
			item_code = _ensure_lab_item(code, test_name_clean, item_group, rate=rate)
			frappe.db.set_value("Lab Test Template", doc.name, "item", item_code, update_modified=False)
			doc.item = item_code
		else:
			_apply_item_pricing(item_code, rate)

		payload = _template_payload(doc.name, doc.lab_test_name or test_name_clean, doc.item)

	_register_lab_in_index(index, payload, code, test_name_clean)
	if created is not None:
		created["lab_templates"] = created.get("lab_templates", 0) + 1
	return payload


def _ensure_service_item(service_id: str, service_name: str, item_group: str, rate: Any = None) -> str:
	if not frappe.db.exists("Item", service_id):
		uom = frappe.db.exists("UOM", "Nos") or frappe.db.get_single_value("Stock Settings", "stock_uom") or "Nos"
		amount = _parse_price(rate)
		with _mute_frappe_messages():
			item = frappe.get_doc(
				{
					"doctype": "Item",
					"item_code": service_id,
					"item_name": service_name,
					"item_group": item_group,
					"description": service_name,
					"is_stock_item": 0,
					"is_sales_item": 1,
					"is_purchase_item": 0,
					"stock_uom": uom,
					"disabled": 0,
					"standard_rate": amount or 0,
				}
			)
			item.flags.ignore_mandatory = True
			item.insert(ignore_permissions=True)

	_apply_item_pricing(service_id, rate)
	return service_id


def _get_or_create_service_template(
	service_code: Any = None,
	service_name: Any = None,
	index: dict | None = None,
	rate: Any = None,
	category: str = "Medical Service",
	created: dict | None = None,
) -> dict | None:
	index = index or _build_service_index()
	existing = _resolve_service_template(service_code, service_name, index)
	if existing:
		return existing

	service_name_clean = str(service_name or service_code or "").strip()
	service_id = _service_id_from_name(service_name_clean, service_code)
	if not service_id:
		return None

	item_group = _default_service_item_group(category)
	price = _parse_price(rate)
	item_code = _ensure_service_item(service_id, service_name_clean, item_group, rate=rate)

	with _mute_frappe_messages():
		doc = frappe.get_doc(
			{
				"doctype": "Healthcare Service Template",
				"service_id": service_id,
				"service_name": service_name_clean,
				"display_name": service_name_clean,
				"category": category,
				"rate": price or 0,
				"item_code": item_code,
				"disabled": 0,
			}
		)
		doc.flags.ignore_mandatory = True
		doc.insert(ignore_permissions=True)

	_apply_item_pricing(doc.item_code or item_code, rate)

	payload = _template_payload(doc.name, doc.service_name or service_name_clean, doc.item_code or item_code)
	_register_service_in_index(index, payload, service_id, service_name_clean)
	if created is not None:
		created["service_templates"] = created.get("service_templates", 0) + 1
	return payload


def _make_inclusive_row(
	*,
	lab_test_template: str | None = None,
	lab_test_name: str | None = None,
	healthcare_service: str | None = None,
	healthcare_service_name: str | None = None,
	item_code: str | None = None,
	discount_apply: int = 0,
) -> dict:
	return {
		"lab_test_template": lab_test_template,
		"lab_test_name": lab_test_name,
		"healthcare_service": healthcare_service,
		"healthcare_service_name": healthcare_service_name,
		"item_code": item_code,
		"item_name": _item_name(item_code),
		"discount_apply": cint(discount_apply),
	}


def _row_identity(row: dict) -> tuple | None:
	if row.get("lab_test_template"):
		return ("lab", row["lab_test_template"])
	if row.get("healthcare_service"):
		return ("svc", row["healthcare_service"])
	if row.get("item_code"):
		return ("item", row["item_code"])
	return None


def _norm_header_cell(cell: Any) -> str:
	return re.sub(r"\s+", " ", str(cell or "").strip().lower().replace("\n", " "))


# Known TRICARE lab export layout: Group No | Group Name | Test Code | Test Name | ...
_DEFAULT_LAB_COLUMNS: dict[str, int] = {
	"test_code": 2,
	"test_name": 3,
	"test_price": 4,
	"ip_disc": 6,
	"op_disc": 7,
}


def _is_lab_test_code(value: Any) -> bool:
	text = (str(value or "")).strip()
	return bool(re.match(r"^LAB-\d", text, re.IGNORECASE))


def _find_lab_header(cells: list) -> dict[str, int] | None:
	normalized = [_norm_header_cell(c) for c in cells]

	test_code_idx = None
	for idx, label in enumerate(normalized):
		if label == "test code" or "test code" in label:
			test_code_idx = idx
			break

	if test_code_idx is None and len(normalized) >= 4:
		if "group no" in normalized[0] and "group name" in normalized[1]:
			test_code_idx = 2

	if test_code_idx is None:
		return None

	test_name_idx = test_code_idx + 1
	test_price_idx = -1
	ip_disc_idx = -1
	op_disc_idx = -1

	for idx, label in enumerate(normalized):
		if "test name" in label:
			test_name_idx = idx
		if "test price" in label:
			test_price_idx = idx
		if "inpatient" in label and "%" in label:
			ip_disc_idx = idx
		if "outpatient" in label and "%" in label:
			op_disc_idx = idx

	return {
		"test_code": test_code_idx,
		"test_name": test_name_idx,
		"test_price": test_price_idx,
		"ip_disc": ip_disc_idx,
		"op_disc": op_disc_idx,
	}


def _resolve_lab_sheet_header(ws) -> dict[str, int] | None:
	for raw in ws.iter_rows(max_row=50, values_only=True):
		if not raw:
			continue
		cells = list(raw)
		header = _find_lab_header(cells)
		if header:
			return header

	for raw in ws.iter_rows(max_row=50, values_only=True):
		if not raw:
			continue
		cells = list(raw)
		if len(cells) > _DEFAULT_LAB_COLUMNS["test_code"] and _is_lab_test_code(
			cells[_DEFAULT_LAB_COLUMNS["test_code"]]
		):
			return dict(_DEFAULT_LAB_COLUMNS)

	return None


def parse_lab_price_list(file_url: str, created: dict | None = None) -> tuple[list[dict], list[str]]:
	wb = _load_workbook(file_url)
	lab_index = _build_lab_template_index()
	rows: list[dict] = []
	missing: list[str] = []
	header_map: dict[str, int] | None = None

	for sheet_name in wb.sheetnames:
		ws = wb[sheet_name]
		sheet_header = _resolve_lab_sheet_header(ws)
		if not sheet_header:
			continue
		header_map = sheet_header

		for raw in ws.iter_rows(values_only=True):
			if not raw:
				continue
			cells = list(raw)

			if _find_lab_header(cells):
				continue

			test_code = (
				cells[sheet_header["test_code"]]
				if len(cells) > sheet_header["test_code"]
				else None
			)
			test_name = (
				cells[sheet_header["test_name"]]
				if len(cells) > sheet_header["test_name"]
				else None
			)
			if not _is_lab_test_code(test_code):
				continue

			test_code = str(test_code).strip()

			rate = None
			if sheet_header.get("test_price", -1) >= 0 and len(cells) > sheet_header["test_price"]:
				rate = cells[sheet_header["test_price"]]

			template = _get_or_create_lab_template(test_code, test_name, lab_index, rate=rate, created=created)
			if not template:
				missing.append(f"Lab: {test_code} — {test_name or ''}".strip())
				continue

			ip_disc = (
				cells[sheet_header["ip_disc"]]
				if sheet_header.get("ip_disc", -1) >= 0 and len(cells) > sheet_header["ip_disc"]
				else None
			)
			op_disc = (
				cells[sheet_header["op_disc"]]
				if sheet_header.get("op_disc", -1) >= 0 and len(cells) > sheet_header["op_disc"]
				else None
			)
			rows.append(
				_make_inclusive_row(
					lab_test_template=template["name"],
					lab_test_name=template.get("lab_test_name") or str(test_name or "").strip() or None,
					item_code=template.get("item_code"),
					discount_apply=1 if _has_discount(ip_disc, op_disc) else 0,
				)
			)

	if not header_map:
		missing.append(_("Lab file: could not find a header row with Test Code / Test Name."))

	wb.close()
	return rows, missing


def parse_ip_price_list(file_url: str, created: dict | None = None) -> tuple[list[dict], list[str]]:
	wb = _load_workbook(file_url)
	service_index = _build_service_index()
	rows_by_code: dict[str, dict] = {}
	missing: list[str] = []

	for sheet_name in wb.sheetnames:
		ws = wb[sheet_name]
		header_found = False
		for raw in ws.iter_rows(values_only=True):
			if not _is_data_row(raw):
				continue
			cells = list(raw)
			if str(cells[0] or "").strip() == "Services Code":
				header_found = True
				continue
			if not header_found:
				continue

			service_code = cells[0] if len(cells) > 0 else None
			service_name = cells[1] if len(cells) > 1 else None
			if not _is_ip_service_code(service_code):
				continue

			code_key = _norm_code(service_code)
			if code_key in rows_by_code:
				continue

			rate = cells[2] if len(cells) > 2 else None
			template = _get_or_create_service_template(
				service_code=service_code,
				service_name=service_name,
				index=service_index,
				rate=rate,
				created=created,
			)
			if not template:
				missing.append(f"IP service: {service_code} — {service_name or ''}".strip())
				continue

			discount_col = cells[3] if len(cells) > 3 else None
			rows_by_code[code_key] = _make_inclusive_row(
				healthcare_service=template["name"],
				healthcare_service_name=template.get("service_name") or str(service_name or "").strip() or None,
				item_code=template.get("item_code"),
				discount_apply=1 if _has_discount(discount_col) else 0,
			)

	wb.close()
	return list(rows_by_code.values()), missing


def parse_iop_price_list(file_url: str, created: dict | None = None) -> tuple[list[dict], list[str]]:
	wb = _load_workbook(file_url)
	service_index = _build_service_index()
	rows_by_name: dict[str, dict] = {}
	missing: list[str] = []

	for sheet_name in wb.sheetnames:
		ws = wb[sheet_name]
		header_found = False
		for raw in ws.iter_rows(values_only=True):
			if not _is_data_row(raw):
				continue
			cells = list(raw)
			if str(cells[0] or "").strip() == "Service Name":
				header_found = True
				continue
			if not header_found:
				continue

			service_name = cells[0] if len(cells) > 0 else None
			if not _is_iop_service_name(service_name):
				continue

			name_key = _norm_name(service_name)
			if name_key in rows_by_name:
				continue

			rate = cells[2] if len(cells) > 2 else None
			template = _get_or_create_service_template(
				service_name=service_name,
				index=service_index,
				rate=rate,
				category="Other Service",
				created=created,
			)
			if not template:
				missing.append(f"IOP service: {service_name}")
				continue

			rows_by_name[name_key] = _make_inclusive_row(
				healthcare_service=template["name"],
				healthcare_service_name=template.get("service_name") or str(service_name or "").strip() or None,
				item_code=template.get("item_code"),
				discount_apply=0,
			)

	wb.close()
	return list(rows_by_name.values()), missing


def _merge_rows_into_doc(doc, parsed_rows: list[dict]) -> tuple[int, int]:
	existing: dict[tuple, int] = {}
	existing_by_item_code: dict[str, int] = {}
	for idx, row in enumerate(doc.get("inclusive_item") or []):
		identity = _row_identity(
			{
				"lab_test_template": row.lab_test_template,
				"healthcare_service": row.healthcare_service,
				"item_code": row.item_code,
			}
		)
		if identity:
			existing[identity] = idx
		if row.item_code:
			existing_by_item_code[row.item_code] = idx

	added = 0
	updated = 0
	for payload in parsed_rows:
		identity = _row_identity(payload)
		child_idx = existing.get(identity) if identity else None
		if child_idx is None and payload.get("item_code"):
			candidate_idx = existing_by_item_code.get(payload["item_code"])
			if candidate_idx is not None:
				existing_row = doc.inclusive_item[candidate_idx]
				if payload.get("lab_test_template") and not existing_row.lab_test_template:
					child_idx = candidate_idx
				elif payload.get("healthcare_service") and not existing_row.healthcare_service:
					child_idx = candidate_idx

		if child_idx is not None:
			child = doc.inclusive_item[child_idx]
			child.discount_apply = payload.get("discount_apply") or 0
			if payload.get("lab_test_template"):
				child.lab_test_template = payload["lab_test_template"]
				child.lab_test_name = payload.get("lab_test_name")
			if payload.get("healthcare_service"):
				child.healthcare_service = payload["healthcare_service"]
				child.healthcare_service_name = payload.get("healthcare_service_name")
			if payload.get("item_code"):
				child.item_code = payload["item_code"]
				child.item_name = payload.get("item_name")
			if identity:
				existing[identity] = child_idx
			updated += 1
		else:
			doc.append("inclusive_item", payload)
			new_idx = len(doc.inclusive_item) - 1
			if identity:
				existing[identity] = new_idx
			if payload.get("item_code"):
				existing_by_item_code[payload["item_code"]] = new_idx
			added += 1
	return added, updated


@frappe.whitelist()
def import_inclusive_items_from_price_lists(
	docname: str,
	lab_file_url: str | None = None,
	ip_file_url: str | None = None,
	iop_file_url: str | None = None,
):
	"""Parse the three agreed price list workbooks and populate inclusive_item rows."""
	if not docname:
		frappe.throw(_("Health Insurance is required"))
	if not lab_file_url and not ip_file_url and not iop_file_url:
		frappe.throw(_("Upload at least one price list file."))

	doc = frappe.get_doc("Health Insurance", docname)
	all_rows: list[dict] = []
	missing: list[str] = []
	summary = {"lab": 0, "ip": 0, "iop": 0}
	created = {"lab_templates": 0, "service_templates": 0}

	if lab_file_url:
		lab_rows, lab_missing = parse_lab_price_list(lab_file_url, created=created)
		all_rows.extend(lab_rows)
		missing.extend(lab_missing)
		summary["lab"] = len(lab_rows)

	if ip_file_url:
		ip_rows, ip_missing = parse_ip_price_list(ip_file_url, created=created)
		all_rows.extend(ip_rows)
		missing.extend(ip_missing)
		summary["ip"] = len(ip_rows)

	if iop_file_url:
		iop_rows, iop_missing = parse_iop_price_list(iop_file_url, created=created)
		all_rows.extend(iop_rows)
		missing.extend(iop_missing)
		summary["iop"] = len(iop_rows)

	if not all_rows:
		frappe.throw(_("No matching lab tests or healthcare services were found in the uploaded files."))

	added, updated = _merge_rows_into_doc(doc, all_rows)
	doc.save(ignore_permissions=True)

	message = _(
		"Imported {0} inclusive item row(s) ({1} lab, {2} IP service, {3} IOP service). "
		"{4} added, {5} updated. Created {6} lab template(s) and {7} service template(s)."
	).format(
		len(all_rows),
		summary["lab"],
		summary["ip"],
		summary["iop"],
		added,
		updated,
		created.get("lab_templates", 0),
		created.get("service_templates", 0),
	)

	return {
		"ok": True,
		"message": message,
		"summary": summary,
		"created": created,
		"added": added,
		"updated": updated,
		"total_parsed": len(all_rows),
		"missing_count": len(missing),
		"missing_sample": missing[:40],
	}
