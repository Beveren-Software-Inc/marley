"""Import Oracle DAILY_PATIENTS_01 Excel into Daily Patient Visit Setup records."""

from __future__ import annotations

import json

import frappe
from frappe import _
from frappe.utils import flt, getdate

from healthcare.api.patient_info_import import (
	_cell_text,
	_clean_oracle_num,
	_excel_file_path,
	_parse_date_value,
	_parse_datetime_value,
	_require_admin,
	_yn_to_check,
)
from healthcare.api.patient_visit_import import (
	_resolve_practitioner_by_doctors_id,
	ensure_patient_for_legacy_import,
)

DAILY_PATIENT_VISIT_SETUP_IMPORT_BATCH_SIZE = 500
CACHE_TTL = 7200
CACHE_KEYS = {
	"file_url": "healthcare:data_migration:daily_patient_visit_setup_import:file_url",
	"trans_nums": "healthcare:data_migration:daily_patient_visit_setup_import:trans_nums",
	"rows": "healthcare:data_migration:daily_patient_visit_setup_import:rows",
}

EXCEL_HEADER_MAP = {
	"TRANS_NUM": "trans_num",
	"TRANS_DATE": "trans_date",
	"PATIENT_NUM": "patient_num",
	"DOC_NUM": "doc_num",
	"START_DATE": "start_date",
	"END_DATE": "end_date",
	"SERVICE_NUM": "service_num",
	"SERVICE_AMT": "service_amt",
	"ACTIVE_YN": "active_yn",
	"SERVICE_NUM_2": "service_num_2",
	"SERVICE_AMT_2": "service_amt_2",
	"SERVICE_NUM_3": "service_num_3",
	"SERVICE_AMT_3": "service_amt_3",
	"BRANCH_NUM": "branch_num",
	"CR_ID": "cr_id",
	"CR_DATE": "cr_date",
	"UP_ID": "up_id",
	"UP_DATE": "up_date",
}


def _normalize_header(cell) -> str:
	if cell is None:
		return ""
	text = str(cell).strip().upper().replace(" ", "_")
	return EXCEL_HEADER_MAP.get(text, text.lower())


def _service_lines_from_row(row: dict) -> list[dict]:
	lines: list[dict] = []
	for num_key, amt_key in (
		("service_num", "service_amt"),
		("service_num_2", "service_amt_2"),
		("service_num_3", "service_amt_3"),
	):
		session = _cell_text(row.get(num_key))
		amount = flt(row.get(amt_key))
		if session or amount:
			lines.append({"session": session, "amount": amount})
	return lines


def _parse_sheet_rows(ws, *, datemode: int = 0) -> list[dict]:
	rows_iter = ws.iter_rows(values_only=True)
	try:
		header_row = next(rows_iter)
	except StopIteration:
		return []

	headers = [_normalize_header(h) for h in header_row]
	parsed: list[dict] = []
	for raw in rows_iter:
		if not raw or all(cell is None or str(cell).strip() == "" for cell in raw):
			continue
		row: dict = {}
		for idx, key in enumerate(headers):
			if not key or idx >= len(raw):
				continue
			row[key] = raw[idx]
		trans_num = _clean_oracle_num(row.get("trans_num"))
		if not trans_num:
			continue
		row["trans_num"] = trans_num
		row["patient_num"] = _clean_oracle_num(row.get("patient_num"))
		for date_key in ("start_date", "end_date", "trans_date"):
			parsed_date = _parse_date_value(row.get(date_key), datemode=datemode)
			if parsed_date:
				row[date_key] = parsed_date
		for dt_key in ("cr_date", "up_date"):
			dt = _parse_datetime_value(row.get(dt_key), datemode=datemode)
			if dt:
				row[dt_key] = dt.isoformat(sep=" ", timespec="seconds")
			elif row.get(dt_key) not in (None, ""):
				row[dt_key] = _cell_text(row.get(dt_key))
		row["services"] = _service_lines_from_row(row)
		parsed.append(row)
	return parsed


def _parse_excel_rows(file_url: str) -> list[dict]:
	try:
		import openpyxl
	except ImportError:
		frappe.throw(
			_("openpyxl is required to read Excel files. Install it in the bench environment: pip install openpyxl")
		)

	path = _excel_file_path(file_url)
	wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
	seen: set[str] = set()
	parsed: list[dict] = []
	try:
		for sheet_name in wb.sheetnames:
			rows = _parse_sheet_rows(wb[sheet_name], datemode=wb.epoch)
			for row in rows:
				trans_num = row.get("trans_num")
				if not trans_num or trans_num in seen:
					continue
				seen.add(trans_num)
				parsed.append(row)
	finally:
		wb.close()
	return parsed


def _resolve_patient(file_no: str | None) -> str | None:
	file_no = _clean_oracle_num(file_no)
	if not file_no:
		return None
	if frappe.db.exists("Patient", file_no):
		return file_no
	match = frappe.db.get_value("Patient", {"file_no": file_no}, "name")
	if match:
		return match
	result = ensure_patient_for_legacy_import(file_no)
	return result.get("patient") if result.get("status") != "skip" else None


def _apply_services(doc, services: list[dict]):
	doc.set("services", [])
	for line in services or []:
		session = _cell_text(line.get("session"))
		amount = flt(line.get("amount"))
		if not session and not amount:
			continue
		doc.append("services", {"session": session, "amount": amount})


def upsert_daily_patient_visit_setup_from_row(row: dict) -> dict:
	trans_num = row.get("trans_num")
	if not trans_num:
		return {"status": "skip", "reason": "no_trans_num"}

	patient = _resolve_patient(row.get("patient_num"))
	if not patient:
		return {"status": "skip", "reason": "no_patient", "trans_num": trans_num}

	existing = frappe.db.get_value("Daily Patient Visit Setup", {"trans_num": trans_num}, "name")
	if existing:
		doc = frappe.get_doc("Daily Patient Visit Setup", existing)
		action = "updated"
	else:
		doc = frappe.new_doc("Daily Patient Visit Setup")
		doc.trans_num = trans_num
		action = "created"

	doc.patient = patient
	practitioner = _resolve_practitioner_by_doctors_id(row.get("doc_num"))
	if practitioner:
		doc.practioner = practitioner
		doc.practitioner_name = frappe.db.get_value(
			"Healthcare Practitioner", practitioner, "practitioner_name"
		)

	if row.get("start_date"):
		doc.from_date = getdate(row.get("start_date"))
	if row.get("end_date"):
		doc.to_date = getdate(row.get("end_date"))

	cr_parsed = _parse_date_value(row.get("cr_date"))
	if cr_parsed:
		doc.cr_date = str(getdate(cr_parsed))
		doc.posting_date = doc.cr_date
	elif row.get("cr_date"):
		doc.cr_date = _cell_text(row.get("cr_date"))
		doc.posting_date = doc.cr_date
	elif row.get("trans_date"):
		doc.posting_date = str(getdate(row.get("trans_date")))

	active_raw = row.get("active_yn")
	if active_raw not in (None, ""):
		doc.is_active = 1 if _yn_to_check(active_raw) else 0

	doc.cr_id = _clean_oracle_num(row.get("cr_id")) or doc.cr_id
	doc.up_id = _clean_oracle_num(row.get("up_id")) or doc.up_id
	if row.get("up_date"):
		doc.up_date = _cell_text(row.get("up_date"))

	services = row.get("services") or []
	if services:
		_apply_services(doc, services)

	doc.flags.ignore_permissions = True
	if doc.get("__islocal"):
		doc.insert()
	else:
		doc.save()

	return {"status": action, "name": doc.name, "trans_num": trans_num, "patient": patient}


@frappe.whitelist()
def preview_daily_patient_visit_setup_import(file_url: str) -> dict:
	_require_admin()
	if not (file_url or "").strip():
		frappe.throw(_("Please upload the DAILY_PATIENTS_01 Excel file."))
	return parse_and_cache_excel(file_url)


def _load_cached_rows() -> dict[str, dict]:
	raw = frappe.cache().get_value(CACHE_KEYS["rows"])
	if not raw:
		return {}
	if isinstance(raw, dict):
		return raw
	return json.loads(raw)


def _cache_is_warm(file_url: str) -> bool:
	cache = frappe.cache()
	return (
		cache.get_value(CACHE_KEYS["file_url"]) == file_url
		and bool(cache.get_value(CACHE_KEYS["trans_nums"]))
		and bool(cache.get_value(CACHE_KEYS["rows"]))
	)


def _build_preview_summary(rows: list[dict]) -> dict:
	trans_nums = [row["trans_num"] for row in rows if row.get("trans_num")]
	existing = 0
	if trans_nums:
		existing = frappe.db.count("Daily Patient Visit Setup", {"trans_num": ["in", trans_nums]})

	patient_nums = {
		_clean_oracle_num(row.get("patient_num"))
		for row in rows
		if _clean_oracle_num(row.get("patient_num"))
	}
	patients_to_create = 0
	if patient_nums:
		found = set(frappe.get_all("Patient", filters={"name": ["in", list(patient_nums)]}, pluck="name"))
		remaining = patient_nums - found
		if remaining:
			found.update(
				frappe.get_all(
					"Patient",
					filters={"file_no": ["in", list(remaining)]},
					pluck="file_no",
				)
			)
		patients_to_create = len(patient_nums - found)

	return {
		"excel_rows": len(rows),
		"setups": len(rows),
		"existing_setups": existing,
		"patients_to_create": patients_to_create,
	}


def parse_and_cache_excel(file_url: str) -> dict:
	if _cache_is_warm(file_url):
		return _build_preview_summary(list(_load_cached_rows().values()))

	rows = _parse_excel_rows(file_url)
	cache = frappe.cache()
	cache.set_value(CACHE_KEYS["file_url"], file_url, expires_in_sec=CACHE_TTL)
	cache.set_value(
		CACHE_KEYS["trans_nums"],
		[row["trans_num"] for row in rows if row.get("trans_num")],
		expires_in_sec=CACHE_TTL,
	)
	cache.set_value(
		CACHE_KEYS["rows"],
		json.dumps(
			{row["trans_num"]: row for row in rows if row.get("trans_num")},
			default=str,
		),
		expires_in_sec=CACHE_TTL,
	)
	return _build_preview_summary(rows)


def run_daily_patient_visit_setup_import_batch(*, offset: int = 0) -> dict:
	cache = frappe.cache()
	file_url = cache.get_value(CACHE_KEYS["file_url"])
	trans_nums = cache.get_value(CACHE_KEYS["trans_nums"]) or []
	rows_map = _load_cached_rows()

	if not file_url or not trans_nums:
		return {"done": True, "processed": offset, "error": "missing_cache"}

	batch = trans_nums[offset : offset + DAILY_PATIENT_VISIT_SETUP_IMPORT_BATCH_SIZE]
	created = updated = skipped = 0
	errors: list[str] = []

	for trans_num in batch:
		row = rows_map.get(trans_num)
		if not row:
			skipped += 1
			continue
		try:
			result = upsert_daily_patient_visit_setup_from_row(row)
			if result.get("status") == "created":
				created += 1
			elif result.get("status") == "updated":
				updated += 1
			else:
				skipped += 1
		except Exception:
			skipped += 1
			errors.append(f"{trans_num}: {frappe.get_traceback()}")

	frappe.db.commit()
	processed = offset + len(batch)
	done = processed >= len(trans_nums)
	return {
		"done": done,
		"processed": processed,
		"total": len(trans_nums),
		"created": created,
		"updated": updated,
		"skipped": skipped,
		"errors": errors[:20],
	}
