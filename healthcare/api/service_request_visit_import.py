"""Import Oracle VISIT_00_02 Excel into legacy Service Request rows (Patient Visit services)."""

from __future__ import annotations

import json
from datetime import date, datetime, time
from typing import Any

import frappe
from frappe import _
from frappe.utils import cint, flt, getdate

from healthcare.api.patient_info_import import (
	_cell_text,
	_clean_oracle_num,
	_excel_file_path,
	_parse_date_value,
	_parse_datetime_value,
	_require_admin,
)
from healthcare.api.patient_visit_import import ensure_patient_for_legacy_import
from healthcare.api.visit_diagnosis_sync import _resolve_patient_visit

SERVICE_REQUEST_VISIT_IMPORT_BATCH_SIZE = 150
CACHE_TTL = 7200
CACHE_KEYS = {
	"file_url": "healthcare:data_migration:service_request_visit_import:file_url",
	"row_keys": "healthcare:data_migration:service_request_visit_import:row_keys",
	"rows": "healthcare:data_migration:service_request_visit_import:rows",
}
LEGACY_SERVICE_REQUEST_STATUS = "completed-Request Status"

EXCEL_HEADER_MAP = {
	"VISIT_NUM": "visit_num",
	"SR_NUM": "sr_num",
	"SERV_GROUP_NUM": "serv_group_num",
	"SERV_NUM": "serv_num",
	"SERV_AMT_BOOK": "serv_amt_book",
	"SERV_AMT_ADD": "serv_amt_add",
	"SERV_AMT_DISC": "serv_amt_disc",
	"SERV_AMT_NET": "serv_amt_net",
	"BRANCH_NUM": "branch_num",
	"CR_ID": "cr_id",
	"CR_DATE": "cr_date",
	"UP_ID": "up_id",
	"UP_DATE": "up_date",
}


def _normalize_header(cell: Any) -> str:
	if cell is None:
		return ""
	text = str(cell).strip().upper().replace(" ", "_")
	return EXCEL_HEADER_MAP.get(text, text.lower())


def _clean_service_code(value: Any) -> str:
	text = _cell_text(value)
	return text.replace(",", "").strip()


def _to_currency(value: Any) -> float:
	if value is None or value == "":
		return 0.0
	if isinstance(value, str):
		value = value.strip().replace(",", "")
	try:
		return flt(value)
	except Exception:
		return 0.0


def _format_legacy_datetime(value: Any) -> str | None:
	dt = _parse_datetime_value(value)
	if dt:
		return dt.strftime("%Y-%m-%d %H:%M:%S")
	parsed_date = _parse_date_value(value)
	if parsed_date:
		return parsed_date.strftime("%Y-%m-%d")
	text = _cell_text(value)
	return text or None


def _order_date_time_from_cr_date(cr_date: Any) -> tuple[Any, str]:
	"""Map Service Request order_date / order_time from CR_DATE."""
	dt = _parse_datetime_value(cr_date)
	if dt:
		return getdate(dt), dt.strftime("%H:%M:%S")
	parsed_date = _parse_date_value(cr_date)
	if parsed_date:
		return parsed_date, "00:00:00"
	return getdate(), "00:00:00"


def _default_company() -> str | None:
	company = frappe.defaults.get_global_default("company")
	if company and frappe.db.exists("Company", company):
		return company
	rows = frappe.get_all("Company", pluck="name", limit=1)
	return rows[0] if rows else None


def _resolve_cost_center(branch_label: Any) -> str | None:
	from healthcare.api.discharge_checklist_import import _resolve_cost_center as resolve_cc

	return resolve_cc(branch_label)


def _build_healthcare_service_template_index() -> dict[str, str]:
	"""Map normalized SERV_NUM / service_id / item_code → Healthcare Service Template name."""
	index: dict[str, str] = {}
	for row in frappe.get_all(
		"Healthcare Service Template",
		fields=["name", "service_id", "item_code", "old_no"],
		limit_page_length=0,
	):
		name = (row.name or "").strip()
		if not name:
			continue
		for raw in (row.name, row.service_id, row.item_code, row.old_no):
			key = _clean_service_code(raw).upper()
			if key:
				index[key] = name
	return index


def _resolve_healthcare_service_template(
	serv_num: Any,
	*,
	template_index: dict[str, str] | None = None,
) -> str | None:
	"""Resolve Excel SERV_NUM to a Healthcare Service Template name (must exist in DB)."""
	code = _clean_service_code(serv_num)
	if not code:
		return None

	keys = [code.upper()]
	if code != code.upper():
		keys.append(code)

	if template_index is not None:
		for key in keys:
			if key in template_index:
				return template_index[key]
	else:
		for candidate in keys:
			if frappe.db.exists("Healthcare Service Template", candidate):
				return candidate
			for field in ("service_id", "item_code", "old_no"):
				name = frappe.db.get_value(
					"Healthcare Service Template",
					{field: candidate},
					"name",
				)
				if name:
					return name
	return None


def _row_key(visit_num: str, sr_num: str) -> str:
	return f"{visit_num}::{sr_num}"


def _parse_sheet_rows(ws) -> list[dict]:
	rows_iter = ws.iter_rows(values_only=True)
	try:
		header_row = next(rows_iter)
	except StopIteration:
		return []

	headers = [_normalize_header(h) for h in header_row]
	parsed: list[dict] = []
	for raw in rows_iter:
		if not raw or all(cell is None or str(cell).strip() == "" for cell in raw):
			continue
		row: dict[str, Any] = {}
		for idx, key in enumerate(headers):
			if not key or idx >= len(raw):
				continue
			row[key] = raw[idx]
		visit_num = _clean_oracle_num(row.get("visit_num"))
		sr_num = _clean_oracle_num(row.get("sr_num"))
		serv_num = _clean_service_code(row.get("serv_num"))
		if not visit_num or not sr_num or not serv_num:
			continue
		row["visit_num"] = visit_num
		row["sr_num"] = sr_num
		row["serv_num"] = serv_num
		parsed.append(row)
	return parsed


def _parse_excel_rows(file_url: str) -> tuple[list[dict], dict[str, int]]:
	"""Parse every worksheet; duplicate VISIT_NUM+SR_NUM keeps the last sheet's row."""
	try:
		import openpyxl
	except ImportError:
		frappe.throw(
			_("openpyxl is required to read Excel files. Install it in the bench environment: pip install openpyxl")
		)

	path = _excel_file_path(file_url)
	wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
	by_key: dict[str, dict] = {}
	sheet_row_counts: dict[str, int] = {}
	try:
		for sheet_name in wb.sheetnames:
			sheet_rows = _parse_sheet_rows(wb[sheet_name])
			sheet_row_counts[sheet_name] = len(sheet_rows)
			for row in sheet_rows:
				by_key[_row_key(row["visit_num"], row["sr_num"])] = row
	finally:
		wb.close()
	return list(by_key.values()), sheet_row_counts


def _default_practitioner() -> str | None:
	rows = frappe.get_all("Healthcare Practitioner", pluck="name", limit=1)
	return rows[0] if rows else None


def _practitioner_for_visit(visit_name: str) -> str | None:
	practitioner = frappe.db.get_value("Patient Visit", visit_name, "practitioner")
	if practitioner:
		return practitioner
	return _default_practitioner()


def ensure_patient_visit_for_legacy_service_import(
	visit_num: str,
	*,
	order_date: Any = None,
) -> dict:
	"""Create a minimal Patient Visit when VISIT_00_02 references an unknown visit."""
	visit_num = _clean_oracle_num(visit_num)
	if not visit_num:
		return {"status": "skip", "visit": None, "created": False}

	existing = _resolve_patient_visit(visit_num)
	if existing:
		return {"status": "existing", "visit": existing, "created": False}

	patient_created = False
	patient_result = ensure_patient_for_legacy_import(visit_num)
	if patient_result.get("status") == "created":
		patient_created = True
	elif patient_result.get("status") not in ("existing", "created"):
		return {"status": "skip_no_patient", "visit": None, "created": False}

	encounter_date, encounter_time = _order_date_time_from_cr_date(order_date)
	doc = frappe.new_doc("Patient Visit")
	doc.case_no = visit_num
	doc.patient = visit_num
	doc.encounter_date = encounter_date
	doc.encounter_time = encounter_time
	doc.status = "Completed"
	company = _default_company()
	if company:
		doc.company = company
	doc.flags.ignore_mandatory = True
	doc.flags.ignore_links = True
	doc.flags.from_legacy_import = True
	doc.insert(ignore_permissions=True)

	submitted = False
	try:
		doc.flags.ignore_validate = True
		doc.flags.from_legacy_import = True
		doc.submit()
		submitted = True
	except Exception:
		frappe.log_error(title=f"Legacy service-request visit submit failed: {visit_num}")

	return {
		"status": "created",
		"visit": doc.name,
		"created": True,
		"patient_created": patient_created,
		"submitted": submitted,
	}


def _find_existing_service_request(patient_visit: str, sr_num: str) -> str | None:
	rows = frappe.get_all(
		"Service Request",
		filters={"legacy": 1, "patient_visit": patient_visit, "sr_num": sr_num},
		pluck="name",
		limit=1,
	)
	return rows[0] if rows else None


def _finalize_legacy_service_request(doc) -> bool:
	"""Submit draft legacy Service Request and set status to Completed."""
	submitted = False
	try:
		doc.flags.ignore_validate = True
		doc.flags.ignore_mandatory = True
		doc.flags.from_legacy_import = True
		if doc.docstatus == 0:
			doc.submit()
			submitted = True
		if doc.docstatus == 1 and doc.status != LEGACY_SERVICE_REQUEST_STATUS:
			doc.db_set("status", LEGACY_SERVICE_REQUEST_STATUS, update_modified=False)
	except Exception:
		frappe.log_error(
			title=f"Legacy service request submit failed: {doc.name}",
			message=frappe.get_traceback(),
		)
	return submitted


def _build_service_request_fields(
	row: dict,
	*,
	patient_visit: str,
	patient: str,
	template_index: dict[str, str] | None = None,
) -> dict[str, Any]:
	"""VISIT_00_02 → Service Request (legacy).

	VISIT_NUM → patient_visit | SERV_NUM → template_dn (Healthcare Service Template)
	SERV_AMT_BOOK → cost/serv_amt_book | SERV_AMT_DISC → discount_amount
	SERV_AMT_NET → grand_total | BRANCH_NUM → cost_center (1=Serene Hospital, 2=Serene Center, …)
	CR_DATE → order_date + order_time
	"""
	serv_num_raw = _clean_service_code(row.get("serv_num"))
	template_dn = _resolve_healthcare_service_template(serv_num_raw, template_index=template_index)
	book = _to_currency(row.get("serv_amt_book"))
	add = _to_currency(row.get("serv_amt_add"))
	discount = _to_currency(row.get("serv_amt_disc"))
	grand_total = _to_currency(row.get("serv_amt_net"))
	if grand_total <= 0 and book > 0:
		grand_total = max(book + add - discount, 0)

	order_date, order_time = _order_date_time_from_cr_date(row.get("cr_date"))

	fields: dict[str, Any] = {
		"legacy": 1,
		"patient": patient,
		"patient_visit": patient_visit,
		"template_dt": "Healthcare Service Template",
		"template_dn": template_dn,
		"sr_num": row.get("sr_num"),
		"serv_group_num": _clean_oracle_num(row.get("serv_group_num")) or _cell_text(row.get("serv_group_num")),
		"serv_amt_book": book,
		"serv_amount_add": _cell_text(row.get("serv_amt_add")) or str(add),
		"cost": book,
		"amount": grand_total,
		"discount_amount": discount,
		"discount_margin": "Amount" if discount else "",
		"grand_total": grand_total,
		"order_date": order_date,
		"order_time": order_time,
		"quantity": 1,
		"practitioner": _practitioner_for_visit(patient_visit),
		"cr_id": _clean_oracle_num(row.get("cr_id")),
		"up_id": _clean_oracle_num(row.get("up_id")),
	}

	if template_dn:
		item_code = frappe.db.get_value("Healthcare Service Template", template_dn, "item_code")
		if item_code:
			fields["item_code"] = item_code
	elif serv_num_raw:
		fields["order_description"] = _("Legacy SERV_NUM (no template): {0}").format(serv_num_raw)

	cr_date = _format_legacy_datetime(row.get("cr_date"))
	if cr_date:
		fields["cr_date"] = cr_date
	up_date = _format_legacy_datetime(row.get("up_date"))
	if up_date:
		fields["up_date"] = up_date

	cc = _resolve_cost_center(row.get("branch_num"))
	if cc:
		fields["cost_center"] = cc

	company = _default_company()
	if company:
		fields["company"] = company

	return {key: value for key, value in fields.items() if value not in (None, "")}


def upsert_service_request_from_row(
	row: dict,
	*,
	template_index: dict[str, str] | None = None,
) -> dict:
	visit_num = row.get("visit_num")
	sr_num = row.get("sr_num")
	visit_result = ensure_patient_visit_for_legacy_service_import(
		visit_num,
		order_date=row.get("cr_date"),
	)
	patient_visit = visit_result.get("visit")
	if not patient_visit:
		return {
			"status": visit_result.get("status") or "skip_no_visit",
			"visit_num": visit_num,
			"sr_num": sr_num,
		}

	patient = frappe.db.get_value("Patient Visit", patient_visit, "patient")
	if not patient:
		patient = visit_num

	fields = _build_service_request_fields(
		row,
		patient_visit=patient_visit,
		patient=patient,
		template_index=template_index,
	)
	if not fields.get("template_dn"):
		return {
			"status": "skip_no_template",
			"visit_num": visit_num,
			"sr_num": sr_num,
			"serv_num": row.get("serv_num"),
		}

	existing = _find_existing_service_request(patient_visit, sr_num)
	if existing:
		doc = frappe.get_doc("Service Request", existing)
		for key, value in fields.items():
			doc.set(key, value)
		doc.flags.ignore_mandatory = True
		doc.flags.ignore_links = True
		doc.flags.from_legacy_import = True
		doc.save(ignore_permissions=True)
		action = "updated"
	else:
		doc = frappe.new_doc("Service Request")
		doc.naming_series = "HSR-"
		doc.update(fields)
		doc.flags.ignore_mandatory = True
		doc.flags.ignore_links = True
		doc.flags.from_legacy_import = True
		doc.insert(ignore_permissions=True)
		action = "created"

	submitted = _finalize_legacy_service_request(doc)

	return {
		"status": action,
		"visit_num": visit_num,
		"sr_num": sr_num,
		"name": doc.name,
		"visit_created": visit_result.get("created"),
		"patient_created": visit_result.get("patient_created"),
		"submitted": submitted,
	}


def _preview_template_matches(
	rows: list[dict],
	*,
	template_index: dict[str, str] | None = None,
) -> dict:
	codes: set[str] = set()
	for row in rows:
		code = _clean_service_code(row.get("serv_num"))
		if code:
			codes.add(code)

	matched: set[str] = set()
	unmatched: list[str] = []
	for code in sorted(codes):
		if _resolve_healthcare_service_template(code, template_index=template_index):
			matched.add(code)
		else:
			unmatched.append(code)

	op_0092 = _resolve_healthcare_service_template("OP-0092", template_index=template_index)

	return {
		"unique_serv_nums": len(codes),
		"matching_templates": len(matched),
		"unmatched_serv_num_count": len(unmatched),
		"unmatched_serv_nums": unmatched[:25],
		"sample_serv_nums": sorted(codes)[:10],
		"op_0092_resolved": op_0092,
	}


def parse_and_cache_excel(file_url: str) -> dict:
	rows, sheet_row_counts = _parse_excel_rows(file_url)
	by_key = {_row_key(row["visit_num"], row["sr_num"]): row for row in rows}
	row_keys = sorted(by_key.keys())
	raw_row_total = sum(sheet_row_counts.values())

	frappe.cache().set_value(CACHE_KEYS["file_url"], file_url, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(CACHE_KEYS["row_keys"], row_keys, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(
		CACHE_KEYS["rows"],
		json.dumps(by_key, default=str),
		expires_in_sec=CACHE_TTL,
	)

	visit_nums = {row["visit_num"] for row in rows}
	existing_visits = sum(1 for visit_num in visit_nums if _resolve_patient_visit(visit_num))
	existing_sr = 0
	for row in rows:
		visit_name = _resolve_patient_visit(row["visit_num"])
		if visit_name and _find_existing_service_request(visit_name, row["sr_num"]):
			existing_sr += 1

	template_index = _build_healthcare_service_template_index()
	template_stats = _preview_template_matches(rows, template_index=template_index)

	return {
		"excel_rows": len(rows),
		"raw_excel_rows": raw_row_total,
		"sheets": list(sheet_row_counts.keys()),
		"sheet_row_counts": sheet_row_counts,
		"service_requests": len(row_keys),
		"unique_visits": len(visit_nums),
		"existing_visits": existing_visits,
		"visits_to_create": len(visit_nums) - existing_visits,
		"existing_service_requests": existing_sr,
		"unique_serv_nums": template_stats["unique_serv_nums"],
		"matching_templates": template_stats["matching_templates"],
		"unmatched_serv_num_count": template_stats["unmatched_serv_num_count"],
		"unmatched_serv_nums": template_stats["unmatched_serv_nums"],
		"sample_serv_nums": template_stats["sample_serv_nums"],
		"op_0092_resolved": template_stats["op_0092_resolved"],
		"healthcare_service_templates_in_db": len(template_index),
		"sample_visit_nums": sorted(visit_nums)[:5],
	}


def _load_cached_rows() -> dict[str, dict]:
	raw = frappe.cache().get_value(CACHE_KEYS["rows"])
	if not raw:
		return {}
	if isinstance(raw, dict):
		return raw
	return json.loads(raw)


@frappe.whitelist()
def preview_service_request_visit_import(file_url: str) -> dict:
	_require_admin()
	if not (file_url or "").strip():
		frappe.throw(_("Please upload the VISIT_00_02 Excel file."))
	return parse_and_cache_excel(file_url)


def run_service_request_visit_import_batch(offset: int = 0) -> dict:
	row_keys = frappe.cache().get_value(CACHE_KEYS["row_keys"]) or []
	rows_by_key = _load_cached_rows()
	if not row_keys or not rows_by_key:
		return {"processed": offset, "done": True, "batch_count": 0}

	batch_keys = row_keys[offset : offset + SERVICE_REQUEST_VISIT_IMPORT_BATCH_SIZE]
	if not batch_keys:
		return {"processed": offset, "done": True, "batch_count": 0}

	created = updated = skip_no_visit = skip_no_template = visits_created = patients_created = submitted = 0
	errors: list[str] = []
	template_index = _build_healthcare_service_template_index()

	for key in batch_keys:
		row = rows_by_key.get(key) or {}
		try:
			result = upsert_service_request_from_row(row, template_index=template_index)
			status = result.get("status")
			if status == "created":
				created += 1
			elif status == "updated":
				updated += 1
			elif status in ("skip_no_visit", "skip_no_patient"):
				skip_no_visit += 1
			elif status == "skip_no_template":
				skip_no_template += 1
			if result.get("visit_created"):
				visits_created += 1
			if result.get("patient_created"):
				patients_created += 1
			if result.get("submitted"):
				submitted += 1
		except Exception:
			errors.append(f"{key}: {frappe.get_traceback()}")
			frappe.log_error(title=f"Service request visit import failed: {key}")

	frappe.db.commit()
	processed = offset + len(batch_keys)
	return {
		"processed": processed,
		"done": len(batch_keys) < SERVICE_REQUEST_VISIT_IMPORT_BATCH_SIZE,
		"batch_count": len(batch_keys),
		"created": created,
		"updated": updated,
		"skip_no_visit": skip_no_visit,
		"skip_no_template": skip_no_template,
		"visits_created": visits_created,
		"patients_created": patients_created,
		"submitted": submitted,
		"errors": len(errors),
	}
