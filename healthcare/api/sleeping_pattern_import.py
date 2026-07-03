"""Import Oracle IP_SLEEPING_PATTERN Excel into standalone Sleeping Pattern rows."""

from __future__ import annotations

import json
from datetime import date, datetime, time, timedelta
from typing import Any

import frappe
from frappe import _
from frappe.utils import cint, flt, get_datetime

from healthcare.api.discharge_checklist_import import _resolve_cost_center
from healthcare.api.legacy_id_normalize import normalize_legacy_id
from healthcare.api.patient_info_import import _clean_oracle_num, _excel_file_path, _require_admin

DOCTYPE = "Sleeping Pattern"
SLEEPING_PATTERN_IMPORT_BATCH_SIZE = 500
CACHE_TTL = 7200
CACHE_KEYS = {
	"file_url": "healthcare:data_migration:sleeping_pattern_import:file_url",
	"row_keys": "healthcare:data_migration:sleeping_pattern_import:row_keys",
	"rows": "healthcare:data_migration:sleeping_pattern_import:rows",
}

EXCEL_HEADER_MAP = {
	"TRANS_NO": "trans_no",
	"TRANS_DATE": "trans_date",
	"ADMISSION_NUM": "admission_num",
	"CR_ID": "cr_id",
	"CR_DATE": "cr_date",
	"UP_ID": "up_id",
	"UP_DATE": "up_date",
	"BRANCH_NUM": "branch_num",
	"MORNING_FROM": "morning_from",
	"MORNING_TO": "morning_to",
	"EVENING_FROM": "evening_from",
	"EVENING_TO": "evening_to",
	"NIGHT_FROM": "night_from",
	"NIGHT_TO": "night_to",
	"ADMISSION_NUM_OLD": "admission_num_old",
}


def _chunked(values: list[str], size: int = 1000):
	for index in range(0, len(values), size):
		yield values[index : index + size]


def _normalize_header(cell: Any) -> str:
	if cell is None:
		return ""
	text = str(cell).strip().upper().replace(" ", "_")
	return EXCEL_HEADER_MAP.get(text, text.lower())


def _parse_date_field(value: Any) -> date | None:
	if value in (None, ""):
		return None
	if isinstance(value, datetime):
		return value.date()
	if isinstance(value, date):
		return value
	try:
		return get_datetime(value).date()
	except Exception:
		return None


def _format_legacy_datetime(value: Any) -> str | None:
	if value in (None, ""):
		return None
	if isinstance(value, datetime):
		return value.strftime("%Y-%m-%d %H:%M:%S")
	if isinstance(value, date):
		return value.strftime("%Y-%m-%d")
	text = str(value).strip()
	if not text:
		return None
	try:
		return get_datetime(text).strftime("%Y-%m-%d %H:%M:%S")
	except Exception:
		return text


def _extract_time(value: Any) -> time | None:
	if value in (None, ""):
		return None
	if isinstance(value, datetime):
		return value.time().replace(tzinfo=None)
	if isinstance(value, time):
		return value.replace(tzinfo=None)
	try:
		return get_datetime(value).time().replace(tzinfo=None)
	except Exception:
		return None


def _hours_between(start, end) -> float | None:
	if not start or not end:
		return None
	try:
		start_dt = get_datetime(start)
		end_dt = get_datetime(end)
		seconds = (end_dt - start_dt).total_seconds()
		if seconds < 0:
			seconds += 24 * 60 * 60
		if seconds <= 0:
			return None
		return flt(seconds / 3600.0, 2)
	except Exception:
		return None


def _build_period_range(base_date: date | None, start_value: Any, end_value: Any) -> tuple[datetime | None, datetime | None]:
	start_time = _extract_time(start_value)
	end_time = _extract_time(end_value)
	if not start_time and not end_time:
		return None, None

	if base_date:
		anchor = base_date
	else:
		fallback_date = _parse_date_field(start_value) or _parse_date_field(end_value)
		anchor = fallback_date or date.today()

	start_dt = datetime.combine(anchor, start_time) if start_time else None
	end_dt = datetime.combine(anchor, end_time) if end_time else None
	if start_dt and end_dt and end_dt < start_dt:
		end_dt += timedelta(days=1)
	return start_dt, end_dt


def _candidate_admission_keys(row: dict) -> list[str]:
	keys: list[str] = []
	for raw in (row.get("admission_num"), row.get("admission_num_old")):
		key = (raw or "").strip()
		if not key:
			continue
		keys.append(key)
		normalized = normalize_legacy_id(key)
		if normalized:
			keys.append(normalized)
	return keys


def _build_admission_lookup(rows: list[dict]) -> dict[str, str]:
	candidates: set[str] = set()
	for row in rows:
		candidates.update(_candidate_admission_keys(row))

	if not candidates:
		return {}

	lookup: dict[str, str] = {}
	candidate_list = list(candidates)
	for fieldname in ("name", "case_no", "admission_no_old"):
		for chunk in _chunked(candidate_list):
			records = frappe.get_all(
				"Inpatient Admission",
				filters={fieldname: ["in", chunk]},
				fields=["name", fieldname],
				limit_page_length=0,
			)
			for record in records:
				key = (record.get(fieldname) or "").strip()
				if key:
					lookup[key] = record["name"]
	return lookup


def _resolve_admission(row: dict, admission_lookup: dict[str, str]) -> str | None:
	for key in _candidate_admission_keys(row):
		if admission_lookup.get(key):
			return admission_lookup[key]
	return None


def _build_patient_maps(admission_names: set[str]) -> tuple[dict[str, str], dict[str, str]]:
	if not admission_names:
		return {}, {}

	admissions = []
	for chunk in _chunked(list(admission_names)):
		admissions.extend(
			frappe.get_all(
				"Inpatient Admission",
				filters={"name": ["in", chunk]},
				fields=["name", "patient"],
				limit_page_length=0,
			)
		)
	patient_by_admission = {
		row["name"]: row["patient"] for row in admissions if (row.get("patient") or "").strip()
	}
	patient_ids = list({patient for patient in patient_by_admission.values() if patient})
	patient_name_by_id: dict[str, str] = {}
	if patient_ids:
		for chunk in _chunked(patient_ids):
			for row in frappe.get_all(
				"Patient",
				filters={"name": ["in", chunk]},
				fields=["name", "patient_name"],
				limit_page_length=0,
			):
				patient_name_by_id[row["name"]] = row.get("patient_name") or row["name"]
	return patient_by_admission, patient_name_by_id


def _parse_sheet_rows(ws) -> list[dict]:
	rows_iter = ws.iter_rows(values_only=True)
	try:
		header_row = next(rows_iter)
	except StopIteration:
		return []

	headers = [_normalize_header(h) for h in header_row]
	parsed: list[dict] = []
	for raw in rows_iter:
		if not raw or all(cell is None or str(cell).strip() == "" for cell in raw):
			continue

		row: dict[str, Any] = {}
		for idx, key in enumerate(headers):
			if not key or idx >= len(raw):
				continue
			row[key] = raw[idx]

		trans_no = _clean_oracle_num(row.get("trans_no"))
		if not trans_no:
			continue

		row["trans_no"] = trans_no
		row["admission_num"] = _clean_oracle_num(row.get("admission_num"))
		row["admission_num_old"] = _clean_oracle_num(row.get("admission_num_old"))
		row["cr_id"] = _clean_oracle_num(row.get("cr_id"))
		row["up_id"] = _clean_oracle_num(row.get("up_id"))
		row["branch_num"] = _clean_oracle_num(row.get("branch_num"))
		parsed.append(row)

	return parsed


def _parse_excel_rows(file_url: str) -> tuple[list[dict], dict[str, int]]:
	try:
		import openpyxl
	except ImportError:
		frappe.throw(
			_("openpyxl is required to read Excel files. Install it in the bench environment: pip install openpyxl")
		)

	path = _excel_file_path(file_url)
	wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
	all_rows: list[dict] = []
	sheet_row_counts: dict[str, int] = {}
	try:
		for sheet_name in wb.sheetnames:
			sheet_rows = _parse_sheet_rows(wb[sheet_name])
			sheet_row_counts[sheet_name] = len(sheet_rows)
			all_rows.extend(sheet_rows)
	finally:
		wb.close()
	return all_rows, sheet_row_counts


def _build_fields(
	row: dict,
	admission_lookup: dict[str, str],
	patient_by_admission: dict[str, str],
	patient_name_by_id: dict[str, str],
) -> tuple[dict[str, Any], dict[str, int]]:
	stats = {"skip_no_admission": 0, "skip_no_patient": 0}

	admission = _resolve_admission(row, admission_lookup)
	if not admission:
		stats["skip_no_admission"] = 1
		return {}, stats

	patient = patient_by_admission.get(admission)
	if not patient:
		stats["skip_no_patient"] = 1
		return {}, stats

	base_date = _parse_date_field(row.get("trans_date"))
	morning_from, morning_to = _build_period_range(base_date, row.get("morning_from"), row.get("morning_to"))
	evening_from, evening_to = _build_period_range(base_date, row.get("evening_from"), row.get("evening_to"))
	night_from, night_to = _build_period_range(base_date, row.get("night_from"), row.get("night_to"))

	fields: dict[str, Any] = {
		"trans_no": row["trans_no"],
		"admission_no": admission,
		"file_no": patient,
		"patient_name": patient_name_by_id.get(patient) or patient,
		"admission_no_old": row.get("admission_num_old") or row.get("admission_num") or None,
		"morning_from": morning_from,
		"morning_to": morning_to,
		"evening_from": evening_from,
		"evening_to": evening_to,
		"night_from": night_from,
		"night_to": night_to,
		"cr_id": row.get("cr_id"),
		"up_id": row.get("up_id"),
	}

	if base_date:
		fields["date"] = base_date.isoformat()

	cr_date = _format_legacy_datetime(row.get("cr_date"))
	if cr_date:
		fields["cr_date"] = cr_date

	up_date = _format_legacy_datetime(row.get("up_date"))
	if up_date:
		fields["up_date"] = up_date

	cost_center = _resolve_cost_center(row.get("branch_num"))
	if cost_center:
		fields["cost_center"] = cost_center

	return {key: value for key, value in fields.items() if value not in (None, "")}, stats


def _preview_counts(rows: list[dict]) -> dict:
	row_keys = [row["trans_no"] for row in rows]
	existing_records = set()
	if row_keys:
		for chunk in _chunked(row_keys):
			existing_records.update(
				frappe.get_all(
					DOCTYPE,
					filters={"trans_no": ["in", chunk]},
					pluck="trans_no",
					limit_page_length=0,
				)
			)

	admission_lookup = _build_admission_lookup(rows)
	patient_by_admission, _patient_name_by_id = _build_patient_maps(set(admission_lookup.values()))

	resolved_admissions = 0
	skip_no_admission = 0
	skip_no_patient = 0
	for row in rows:
		admission = _resolve_admission(row, admission_lookup)
		if not admission:
			skip_no_admission += 1
			continue
		resolved_admissions += 1
		if not patient_by_admission.get(admission):
			skip_no_patient += 1

	return {
		"existing_records": len(existing_records),
		"new_records": len(rows) - len(existing_records),
		"resolved_admissions": resolved_admissions,
		"skip_no_admission": skip_no_admission,
		"skip_no_patient": skip_no_patient,
		"sample_trans_nos": [row.get("trans_no") for row in rows[:5]],
	}


def parse_and_cache_excel(file_url: str) -> dict:
	rows, sheet_row_counts = _parse_excel_rows(file_url)
	by_key = {row["trans_no"]: row for row in rows}
	row_keys = list(by_key.keys())

	frappe.cache().set_value(CACHE_KEYS["file_url"], file_url, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(CACHE_KEYS["row_keys"], row_keys, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(
		CACHE_KEYS["rows"],
		json.dumps(by_key, default=str),
		expires_in_sec=CACHE_TTL,
	)

	preview = _preview_counts(list(by_key.values()))
	raw_row_total = sum(sheet_row_counts.values())
	return {
		"excel_rows": len(by_key),
		"raw_excel_rows": raw_row_total,
		"sheets": list(sheet_row_counts.keys()),
		"sheet_row_counts": sheet_row_counts,
		**preview,
	}


def _load_cached_rows() -> dict[str, dict]:
	raw = frappe.cache().get_value(CACHE_KEYS["rows"])
	if not raw:
		return {}
	if isinstance(raw, dict):
		return raw
	return json.loads(raw)


def upsert_sleeping_pattern(
	row: dict,
	admission_lookup: dict[str, str],
	patient_by_admission: dict[str, str],
	patient_name_by_id: dict[str, str],
	existing_names_by_trans: dict[str, str],
) -> dict:
	fields, stats = _build_fields(row, admission_lookup, patient_by_admission, patient_name_by_id)
	if not fields.get("trans_no"):
		return {"status": "skip", "trans_no": row.get("trans_no"), **stats}
	if stats.get("skip_no_admission") or stats.get("skip_no_patient"):
		return {"status": "skip", "trans_no": row.get("trans_no"), **stats}

	trans_no = fields["trans_no"]
	existing_name = existing_names_by_trans.get(trans_no)
	if existing_name:
		doc = frappe.get_doc(DOCTYPE, existing_name)
		for key, value in fields.items():
			doc.set(key, value)
		doc.flags.ignore_mandatory = True
		doc.flags.ignore_links = True
		doc.flags.ignore_validate = True
		doc.save(ignore_permissions=True)
		action = "updated"
	else:
		doc = frappe.get_doc({"doctype": DOCTYPE, **fields})
		doc.flags.ignore_mandatory = True
		doc.flags.ignore_links = True
		doc.flags.ignore_validate = True
		doc.insert(ignore_permissions=True)
		action = "created"

	return {"status": action, "trans_no": trans_no, **stats}


@frappe.whitelist()
def preview_sleeping_pattern_import(file_url: str) -> dict:
	_require_admin()
	if not (file_url or "").strip():
		frappe.throw(_("Please upload the IP_SLEEPING_PATTERN Excel file."))
	return parse_and_cache_excel(file_url)


def run_sleeping_pattern_import_batch(*, offset: int = 0) -> dict:
	row_keys = frappe.cache().get_value(CACHE_KEYS["row_keys"]) or []
	rows_by_key = _load_cached_rows()
	if not row_keys or not rows_by_key:
		return {"processed": offset, "done": True, "batch_count": 0}

	batch_keys = row_keys[offset : offset + SLEEPING_PATTERN_IMPORT_BATCH_SIZE]
	if not batch_keys:
		return {"processed": offset, "done": True, "batch_count": 0}

	batch_rows = [rows_by_key.get(key) or {} for key in batch_keys]
	admission_lookup = _build_admission_lookup(batch_rows)
	patient_by_admission, patient_name_by_id = _build_patient_maps(set(admission_lookup.values()))
	existing_names_by_trans = {}
	for chunk in _chunked(batch_keys):
		for row in frappe.get_all(
			DOCTYPE,
			filters={"trans_no": ["in", chunk]},
			fields=["name", "trans_no"],
			limit_page_length=0,
		):
			existing_names_by_trans[row["trans_no"]] = row["name"]

	created = updated = skipped = 0
	skip_no_admission = skip_no_patient = 0
	errors: list[str] = []

	for row in batch_rows:
		trans_no = row.get("trans_no")
		try:
			result = upsert_sleeping_pattern(
				row,
				admission_lookup,
				patient_by_admission,
				patient_name_by_id,
				existing_names_by_trans,
			)
			status = result.get("status")
			if status == "created":
				created += 1
			elif status == "updated":
				updated += 1
			else:
				skipped += 1
			skip_no_admission += cint(result.get("skip_no_admission", 0))
			skip_no_patient += cint(result.get("skip_no_patient", 0))
		except Exception:
			errors.append(f"{trans_no}: {frappe.get_traceback()}")
			frappe.log_error(title=f"IP_SLEEPING_PATTERN import failed: {trans_no}")

	frappe.db.commit()
	processed = offset + len(batch_keys)
	return {
		"processed": processed,
		"done": len(batch_keys) < SLEEPING_PATTERN_IMPORT_BATCH_SIZE,
		"batch_count": len(batch_keys),
		"created": created,
		"updated": updated,
		"skipped": skipped,
		"skip_no_admission": skip_no_admission,
		"skip_no_patient": skip_no_patient,
		"errors": len(errors),
	}
