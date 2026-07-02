"""Import Oracle IP_ADMISSION_TRANSFER_BAL Excel into Admission Transfer Balance rows."""

from __future__ import annotations

import json
from datetime import date, datetime
from typing import Any

import frappe
from frappe import _
from frappe.utils import flt, get_datetime, getdate

from healthcare.api.discharge_checklist_import import _resolve_cost_center
from healthcare.api.patient_info_import import (
	_cell_text,
	_clean_oracle_num,
	_excel_file_path,
	_require_admin,
)
from healthcare.api.visit_diagnoses_op_import import _legacy_data_datetime
from healthcare.api.visit_diagnosis_sync import _resolve_inpatient_admission

IP_ADMISSION_TRANSFER_BAL_IMPORT_BATCH_SIZE = 500
CACHE_TTL = 7200
CACHE_KEYS = {
	"file_url": "healthcare:data_migration:ip_admission_transfer_bal_import:file_url",
	"row_keys": "healthcare:data_migration:ip_admission_transfer_bal_import:row_keys",
	"rows": "healthcare:data_migration:ip_admission_transfer_bal_import:rows",
}

EXCEL_HEADER_MAP = {
	"TRANS_NUM": "trans_num",
	"TRANS_DATE": "trans_date",
	"PATIENT_NUM": "patient_num",
	"OLD_ADMISSION_NUM": "old_admission_num",
	"NEW_ADMISSION_NUM": "new_admission_num",
	"OLD_BRANCH_NUM": "old_branch_num",
	"NEW_BRANCH_NUM": "new_branch_num",
	"FROM_BRANCH_NUM": "old_branch_num",
	"TO_BRANCH_NUM": "new_branch_num",
	"BRANCH_NUM": "branch_num",
	"BAL_AMT": "bal_amt",
	"CR_ID": "cr_id",
	"CR_DATE": "cr_date",
	"UP_ID": "up_id",
	"UP_DATE": "up_date",
}


def _normalize_header(cell: Any) -> str:
	if cell is None:
		return ""
	text = str(cell).strip().upper().replace(" ", "_")
	return EXCEL_HEADER_MAP.get(text, text.lower())


def _trans_num_value(value: Any) -> str:
	text = _cell_text(value)
	if text:
		return text
	return _clean_oracle_num(value) or ""


def _parse_trans_date(value: Any):
	if value in (None, ""):
		return None
	if isinstance(value, datetime):
		return value
	if isinstance(value, date):
		return datetime.combine(value, datetime.min.time())
	try:
		return get_datetime(value)
	except Exception:
		return None


def _resolve_company(admission: str | None) -> str | None:
	if admission:
		company = frappe.db.get_value("Inpatient Admission", admission, "company")
		if company:
			return company
	return frappe.defaults.get_global_default("company")


def _patient_matches_admission(patient_num: str | None, admission: str | None) -> bool:
	if not patient_num or not admission:
		return True
	adm_patient = frappe.db.get_value("Inpatient Admission", admission, "patient")
	if not adm_patient:
		return True
	return adm_patient == patient_num


def _branch_value(value: Any) -> str | None:
	if value in (None, ""):
		return None
	cleaned = _clean_oracle_num(value)
	if cleaned:
		return cleaned
	text = _cell_text(value)
	return text or None


def _parse_sheet_rows(ws) -> list[dict]:
	rows_iter = ws.iter_rows(values_only=True)
	try:
		header_row = next(rows_iter)
	except StopIteration:
		return []

	headers = [_normalize_header(h) for h in header_row]
	parsed: list[dict] = []
	for raw in rows_iter:
		if not raw or all(cell is None or str(cell).strip() == "" for cell in raw):
			continue
		row: dict[str, Any] = {}
		for idx, key in enumerate(headers):
			if not key or idx >= len(raw):
				continue
			row[key] = raw[idx]

		trans_num = _trans_num_value(row.get("trans_num"))
		if not trans_num:
			continue

		row["trans_num"] = trans_num
		row["patient_num"] = _clean_oracle_num(row.get("patient_num"))
		row["old_admission_num"] = _clean_oracle_num(row.get("old_admission_num"))
		row["new_admission_num"] = _clean_oracle_num(row.get("new_admission_num"))
		row["old_branch_num"] = _branch_value(row.get("old_branch_num")) or _branch_value(row.get("branch_num"))
		row["new_branch_num"] = _branch_value(row.get("new_branch_num"))
		row["cr_id"] = _clean_oracle_num(row.get("cr_id"))
		row["up_id"] = _clean_oracle_num(row.get("up_id"))
		parsed.append(row)
	return parsed


def _parse_excel_rows(file_url: str) -> tuple[list[dict], dict[str, int]]:
	try:
		import openpyxl
	except ImportError:
		frappe.throw(
			_(
				"openpyxl is required to read Excel files. Install it in the bench environment: pip install openpyxl"
			)
		)

	path = _excel_file_path(file_url)
	wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
	all_rows: list[dict] = []
	sheet_row_counts: dict[str, int] = {}
	try:
		for sheet_name in wb.sheetnames:
			sheet_rows = _parse_sheet_rows(wb[sheet_name])
			sheet_row_counts[sheet_name] = len(sheet_rows)
			all_rows.extend(sheet_rows)
	finally:
		wb.close()
	return all_rows, sheet_row_counts


def _build_transfer_balance_fields(row: dict) -> tuple[dict[str, Any], dict[str, int]]:
	patient_num = row.get("patient_num") or ""

	if patient_num and not frappe.db.exists("Patient", patient_num):
		return {}, {"skip_no_patient": 1}

	new_admission_num = row.get("new_admission_num") or ""
	new_admission = _resolve_inpatient_admission(new_admission_num, patient_num or None)
	if not new_admission:
		return {}, {"skip_no_new_admission": 1}

	if patient_num and not _patient_matches_admission(patient_num, new_admission):
		return {}, {"skip_patient_mismatch": 1}

	from_cost_center = _resolve_cost_center(row.get("old_branch_num"))
	to_cost_center = _resolve_cost_center(row.get("new_branch_num"))
	if not from_cost_center or not to_cost_center:
		return {}, {"skip_no_cost_center": 1}

	transfer_dt = _parse_trans_date(row.get("trans_date"))
	if not transfer_dt:
		return {}, {"skip_no_trans_date": 1}

	inpatient = frappe.db.get_value(
		"Inpatient Admission",
		new_admission,
		["patient", "patient_name", "company"],
		as_dict=True,
	) or {}

	fields: dict[str, Any] = {
		"trans_num": row["trans_num"],
		"inpatient_admission": new_admission,
		"patient": inpatient.get("patient"),
		"patient_name": inpatient.get("patient_name"),
		"old_admission": row.get("old_admission_num") or None,
		"new_admission": new_admission_num or None,
		"from_cost_center": from_cost_center,
		"to_cost_center": to_cost_center,
		"company": inpatient.get("company") or _resolve_company(new_admission),
		"transfer_datetime": transfer_dt,
		"new_admission_date": getdate(transfer_dt),
		"new_admission_time": transfer_dt,
		"bal_amount": flt(row.get("bal_amt")),
		"cr_id": row.get("cr_id") or None,
		"up_id": row.get("up_id") or None,
	}

	cr_date = _legacy_data_datetime(row.get("cr_date"))
	if cr_date:
		fields["cr_date"] = cr_date
	up_date = _legacy_data_datetime(row.get("up_date"))
	if up_date:
		fields["up_date"] = up_date

	return {key: value for key, value in fields.items() if value not in (None, "")}, {}


def _apply_legacy_import_flags(doc) -> None:
	doc.flags.ignore_validate = True
	doc.flags.ignore_mandatory = True
	doc.flags.ignore_links = True
	doc.flags.ignore_permissions = True
	doc.flags.legacy_import = True


def _update_inpatient_admission_from_transfer(
	admission: str,
	*,
	to_cost_center: str,
	transfer_dt: datetime,
	old_admission_num: str | None = None,
	new_admission_num: str | None = None,
) -> None:
	updates: dict[str, Any] = {
		"cost_center": to_cost_center,
		"admission_date": getdate(transfer_dt),
		"admission_time": transfer_dt.strftime("%H:%M:%S"),
	}
	if (
		old_admission_num
		and new_admission_num
		and old_admission_num != new_admission_num
	):
		updates["admission_no_old"] = old_admission_num

	doc = frappe.get_doc("Inpatient Admission", admission)
	for key, value in updates.items():
		doc.set(key, value)
	doc.flags.ignore_validate = True
	doc.flags.ignore_permissions = True
	doc.flags.legacy_import = True
	doc.save(ignore_permissions=True)


def _persist_transfer_balance(doc, *, existing: bool) -> None:
	_apply_legacy_import_flags(doc)
	if existing:
		doc.save(ignore_permissions=True)
	else:
		doc.insert(ignore_permissions=True)


def upsert_admission_transfer_balance_from_row(row: dict) -> dict:
	trans_num = row.get("trans_num")
	if not trans_num:
		return {"status": "skip_no_trans_num"}

	fields, side_stats = _build_transfer_balance_fields(row)
	if side_stats.get("skip_no_patient"):
		return {"status": "skip_no_patient", "trans_num": trans_num}
	if side_stats.get("skip_no_new_admission"):
		return {"status": "skip_no_new_admission", "trans_num": trans_num}
	if side_stats.get("skip_patient_mismatch"):
		return {"status": "skip_patient_mismatch", "trans_num": trans_num}
	if side_stats.get("skip_no_cost_center"):
		return {"status": "skip_no_cost_center", "trans_num": trans_num}
	if side_stats.get("skip_no_trans_date"):
		return {"status": "skip_no_trans_date", "trans_num": trans_num}
	if not fields:
		return {"status": "skip_error", "trans_num": trans_num}

	new_admission = fields["inpatient_admission"]
	to_cost_center = fields["to_cost_center"]
	transfer_dt = fields["transfer_datetime"]

	existing_name = frappe.db.exists("Admission Transfer Balance", trans_num)
	if existing_name:
		doc = frappe.get_doc("Admission Transfer Balance", existing_name)
		for key, value in fields.items():
			if key == "trans_num":
				continue
			doc.set(key, value)
		action = "updated"
		existing = True
	else:
		doc = frappe.new_doc("Admission Transfer Balance")
		doc.update(fields)
		action = "created"
		existing = False

	_persist_transfer_balance(doc, existing=existing)
	_update_inpatient_admission_from_transfer(
		new_admission,
		to_cost_center=to_cost_center,
		transfer_dt=transfer_dt,
		old_admission_num=row.get("old_admission_num") or None,
		new_admission_num=row.get("new_admission_num") or None,
	)

	return {
		"status": action,
		"trans_num": trans_num,
		"name": doc.name,
		"inpatient_admission": new_admission,
	}


def _preview_counts(rows: list[dict]) -> dict:
	existing = sum(
		1 for row in rows if frappe.db.exists("Admission Transfer Balance", row["trans_num"])
	)
	resolved_new = unresolved_new = skip_no_patient = skip_patient_mismatch = skip_no_cost_center = 0
	missing_trans_date = 0

	for row in rows:
		if not _parse_trans_date(row.get("trans_date")):
			missing_trans_date += 1

		patient_num = row.get("patient_num") or ""
		if patient_num and not frappe.db.exists("Patient", patient_num):
			skip_no_patient += 1
			continue

		new_num = row.get("new_admission_num") or ""
		new_admission = _resolve_inpatient_admission(new_num, patient_num or None)
		if not new_admission:
			unresolved_new += 1
			continue

		if patient_num and not _patient_matches_admission(patient_num, new_admission):
			skip_patient_mismatch += 1
			continue

		old_branch = row.get("old_branch_num") or _branch_value(row.get("branch_num"))
		from_cc = _resolve_cost_center(old_branch)
		to_cc = _resolve_cost_center(row.get("new_branch_num"))
		if not from_cc or not to_cc:
			skip_no_cost_center += 1
			continue

		resolved_new += 1

	return {
		"existing_transfers": existing,
		"new_transfers": len(rows) - existing,
		"resolved_new_admissions": resolved_new,
		"unresolved_new_admissions": unresolved_new,
		"skip_no_patient": skip_no_patient,
		"skip_patient_mismatch": skip_patient_mismatch,
		"skip_no_cost_center": skip_no_cost_center,
		"missing_trans_date": missing_trans_date,
		"sample_trans_nums": [row["trans_num"] for row in rows[:5]],
	}


def parse_and_cache_excel(file_url: str) -> dict:
	rows, sheet_row_counts = _parse_excel_rows(file_url)
	by_key = {row["trans_num"]: row for row in rows}
	row_keys = list(by_key.keys())

	frappe.cache().set_value(CACHE_KEYS["file_url"], file_url, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(CACHE_KEYS["row_keys"], row_keys, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(
		CACHE_KEYS["rows"],
		json.dumps(by_key, default=str),
		expires_in_sec=CACHE_TTL,
	)

	preview = _preview_counts(list(by_key.values()))
	raw_row_total = sum(sheet_row_counts.values())
	return {
		"excel_rows": len(by_key),
		"raw_excel_rows": raw_row_total,
		"sheets": list(sheet_row_counts.keys()),
		"sheet_row_counts": sheet_row_counts,
		**preview,
	}


def _load_cached_rows() -> dict[str, dict]:
	raw = frappe.cache().get_value(CACHE_KEYS["rows"])
	if not raw:
		return {}
	if isinstance(raw, dict):
		return raw
	return json.loads(raw)


@frappe.whitelist()
def preview_ip_admission_transfer_bal_import(file_url: str) -> dict:
	_require_admin()
	if not (file_url or "").strip():
		frappe.throw(_("Please upload the IP_ADMISSION_TRANSFER_BAL Excel file."))
	return parse_and_cache_excel(file_url)


def run_ip_admission_transfer_bal_import_batch(*, offset: int = 0) -> dict:
	row_keys = frappe.cache().get_value(CACHE_KEYS["row_keys"]) or []
	rows_by_key = _load_cached_rows()
	if not row_keys or not rows_by_key:
		return {"processed": offset, "done": True, "batch_count": 0}

	batch_keys = row_keys[offset : offset + IP_ADMISSION_TRANSFER_BAL_IMPORT_BATCH_SIZE]
	if not batch_keys:
		return {"processed": offset, "done": True, "batch_count": 0}

	created = updated = 0
	skip_no_patient = skip_no_new_admission = skip_patient_mismatch = skip_no_cost_center = 0
	skip_no_trans_date = 0
	errors: list[str] = []

	for key in batch_keys:
		row = rows_by_key.get(key) or {}
		try:
			result = upsert_admission_transfer_balance_from_row(row)
			status = result.get("status")
			if status == "created":
				created += 1
			elif status == "updated":
				updated += 1
			elif status == "skip_no_patient":
				skip_no_patient += 1
			elif status == "skip_no_new_admission":
				skip_no_new_admission += 1
			elif status == "skip_patient_mismatch":
				skip_patient_mismatch += 1
			elif status == "skip_no_cost_center":
				skip_no_cost_center += 1
			elif status == "skip_no_trans_date":
				skip_no_trans_date += 1
		except Exception:
			errors.append(f"{key}: {frappe.get_traceback()}")
			frappe.log_error(title=f"IP_ADMISSION_TRANSFER_BAL import failed: {key}")

	frappe.db.commit()
	processed = offset + len(batch_keys)
	return {
		"processed": processed,
		"done": len(batch_keys) < IP_ADMISSION_TRANSFER_BAL_IMPORT_BATCH_SIZE,
		"batch_count": len(batch_keys),
		"created": created,
		"updated": updated,
		"skip_no_patient": skip_no_patient,
		"skip_no_new_admission": skip_no_new_admission,
		"skip_patient_mismatch": skip_patient_mismatch,
		"skip_no_cost_center": skip_no_cost_center,
		"skip_no_trans_date": skip_no_trans_date,
		"errors": len(errors),
	}
