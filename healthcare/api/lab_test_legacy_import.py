"""Import Oracle lab header (C LAB_00_03) + detail (C-I LAB_00_04) into Lab Test records."""

from __future__ import annotations

import html
import json
from collections import Counter
from datetime import date, datetime, time
from typing import Any

import frappe
from frappe import _
from frappe.utils import cint, flt, get_datetime, getdate

from healthcare.api.legacy_id_normalize import normalize_legacy_id
from healthcare.api.visit_diagnosis_sync import (
	_resolve_inpatient_admission,
	_resolve_patient_visit,
)

LEGACY_LAB_IMPORT_BATCH_SIZE = 500
CACHE_TTL = 7200
CACHE_KEYS = {
	"header_file_url": "healthcare:data_migration:legacy_lab_import:header_file_url",
	"detail_file_url": "healthcare:data_migration:legacy_lab_import:detail_file_url",
	"batch_id": "healthcare:data_migration:legacy_lab_import:batch_id",
	"trans_nums": "healthcare:data_migration:legacy_lab_import:trans_nums",
	"headers": "healthcare:data_migration:legacy_lab_import:headers",
	"details": "healthcare:data_migration:legacy_lab_import:details",
	"failures": "healthcare:data_migration:legacy_lab_import:failures",
}
FAILURE_DETAIL_MAX = 800
FAILURE_LOG_LIST_MAX = 500

HEADER_EXCEL_MAP = {
	"TRANS_NUM": "trans_num",
	"TRANS_DATE": "trans_date",
	"DR_GL_CODE": "dr_gl_code",
	"SUB_DR_GL_CODE": "sub_dr_gl_code",
	"TRANS_SOURCE": "trans_source",
	"VISIT_NUM": "visit_num",
	"STA_FLG": "sta_flg",
	"BRANCH_NUM": "branch_num",
	"TRANS_REMARKS": "trans_remarks",
	"NET_AMOUNT": "net_amount",
	"DOC_NUM": "doc_num",
	"ADMISSION_NUM": "admission_num",
	"OLD_NUM": "old_num",
	"TOTAL_AMOUNT": "total_amount",
	"EXTRA_AMOUNT": "extra_amount",
	"DISCOUNT_AMOUNT": "discount_amount",
	"FIELD7": "oracle_field7",
	"FIELD8": "oracle_field8",
	"FIELD9": "oracle_field9",
	"FIELD10": "oracle_field10",
	"CR_ID": "cr_id",
	"CR_DATE": "cr_date",
	"UP_ID": "up_id",
	"UP_DATE": "up_date",
	"AP_ID": "ap_id",
	"AP_DATE": "ap_date",
	"IN_OUT": "in_out",
	"LAB_TEST_TYPE": "lab_test_type",
	"CONFIRM": "confirm",
	"FILL_ID": "fill_id",
	"FILL_DATE": "fill_date",
	"SAMPLE_COLLECT_ID": "sample_collect_id",
	"SAMPLE_COLLECT_DATE": "sample_collect_date",
}

DETAIL_EXCEL_MAP = {
	"TRANS_NUM": "trans_num",
	"SR_NUM": "sr_num",
	"LAB_GROUP_NUM": "lab_group_num",
	"GROUP NAME": "group_name",
	"LAB_SUB_NUM": "lab_sub_num",
	"LAB_RESULT_VALUE": "lab_result_value",
	"LAB_AMT_BOOK": "lab_amt_book",
	"LAB_AMT_ADD": "lab_amt_add",
	"LAB_AMT_DISC": "lab_amt_disc",
	"LAB_AMT_NET": "lab_amt_net",
	"STA_FLG": "sta_flg",
	"FIELD1": "field1",
	"FIELD2": "field2",
	"FIELD3": "field3",
	"FIELD4": "field4",
	"FIELD5": "field5",
	"FIELD6": "field6",
	"FIELD7": "field7",
	"FIELD8": "field8",
	"FIELD9": "field9",
	"FIELD10": "field10",
	"CR_ID": "cr_id",
	"CR_DATE": "cr_date",
	"UP_ID": "up_id",
	"UP_DATE": "up_date",
	"LAB_04_REMARKS": "lab_04_remarks",
}


def _require_admin() -> None:
	frappe.only_for(("System Manager", "Healthcare Administrator"))


def _clean_oracle_num(value: Any) -> str:
	if value is None or value == "":
		return ""
	if isinstance(value, float) and value == int(value):
		return str(int(value))
	if isinstance(value, int):
		return str(value)
	return str(value).strip()


def _oracle_flag_to_check(value: Any) -> int:
	if value is None or value == "":
		return 0
	if isinstance(value, (int, float)):
		return 1 if float(value) > 0 else 0
	text = str(value).strip().upper()
	return 1 if text in ("Y", "YES", "1", "TRUE", "T", "2") else 0


def _to_currency(value: Any) -> float:
	if value is None or value == "":
		return 0.0
	try:
		return flt(value)
	except Exception:
		return 0.0


def _format_legacy_datetime(value: Any):
	if value is None or value == "":
		return None
	if isinstance(value, datetime):
		return value
	if isinstance(value, date):
		return datetime.combine(value, time.min)
	text = str(value).strip()
	if not text:
		return None
	try:
		return get_datetime(text)
	except Exception:
		return None


def _format_legacy_date_str(value: Any) -> str:
	dt = _format_legacy_datetime(value)
	if dt:
		return dt.strftime("%Y-%m-%d %H:%M:%S")
	if value is None:
		return ""
	return str(value).strip()


def _excel_file_path(file_url: str) -> str:
	if not file_url:
		frappe.throw(_("File URL is required."))
	file_name = frappe.db.get_value("File", {"file_url": file_url}, "name")
	if not file_name:
		frappe.throw(_("Uploaded file was not found. Please upload the Excel file again."))
	from frappe.utils.file_manager import get_file_path

	return get_file_path(file_name)


def _load_workbook_sheets(file_url: str):
	try:
		import openpyxl
	except ImportError:
		frappe.throw(
			_(
				"openpyxl is required to read Excel files. Install it in the bench environment: pip install openpyxl"
			)
		)
	path = _excel_file_path(file_url)
	return openpyxl.load_workbook(path, read_only=True, data_only=True)


def _normalize_header_key(value: Any) -> str:
	return str(value or "").strip().upper()


def _parse_sheet_rows(ws, header_map: dict[str, str]) -> list[dict]:
	rows_iter = ws.iter_rows(values_only=True)
	try:
		header_row = next(rows_iter)
	except StopIteration:
		return []

	headers = [
		header_map.get(_normalize_header_key(h), str(h).strip().lower()) if h is not None else ""
		for h in header_row
	]
	parsed: list[dict] = []
	for raw in rows_iter:
		if not raw or all(cell is None or str(cell).strip() == "" for cell in raw):
			continue
		row: dict[str, Any] = {}
		for idx, key in enumerate(headers):
			if not key or idx >= len(raw):
				continue
			row[key] = raw[idx]
		parsed.append(row)
	return parsed


def _parse_header_excel(file_url: str) -> list[dict]:
	wb = _load_workbook_sheets(file_url)
	try:
		rows = _parse_sheet_rows(wb.active, HEADER_EXCEL_MAP)
	finally:
		wb.close()

	parsed: list[dict] = []
	for row in rows:
		trans_num = _clean_oracle_num(row.get("trans_num"))
		if not trans_num:
			continue
		row["trans_num"] = trans_num
		row["visit_num"] = _clean_oracle_num(row.get("visit_num"))
		row["admission_num"] = _clean_oracle_num(row.get("admission_num"))
		row["sub_dr_gl_code"] = _clean_oracle_num(row.get("sub_dr_gl_code"))
		row["dr_gl_code"] = _clean_oracle_num(row.get("dr_gl_code"))
		row["doc_num"] = _clean_oracle_num(row.get("doc_num"))
		row["old_num"] = _clean_oracle_num(row.get("old_num"))
		row["fill_id"] = _clean_oracle_num(row.get("fill_id"))
		row["sample_collect_id"] = _clean_oracle_num(row.get("sample_collect_id"))
		row["cr_id"] = _clean_oracle_num(row.get("cr_id"))
		row["up_id"] = _clean_oracle_num(row.get("up_id"))
		row["ap_id"] = _clean_oracle_num(row.get("ap_id"))
		parsed.append(row)
	return parsed


def _parse_detail_excel(file_url: str) -> list[dict]:
	wb = _load_workbook_sheets(file_url)
	seen: set[tuple[str, str, str]] = set()
	parsed: list[dict] = []
	try:
		for sheet_name in wb.sheetnames:
			rows = _parse_sheet_rows(wb[sheet_name], DETAIL_EXCEL_MAP)
			for row in rows:
				trans_num = _clean_oracle_num(row.get("trans_num"))
				lab_sub_num = _clean_oracle_num(row.get("lab_sub_num"))
				if not trans_num:
					continue
				sr_num = _clean_oracle_num(row.get("sr_num"))
				dedupe_key = (trans_num, sr_num, lab_sub_num)
				if dedupe_key in seen:
					continue
				seen.add(dedupe_key)
				row["trans_num"] = trans_num
				row["sr_num"] = sr_num
				row["lab_group_num"] = _clean_oracle_num(row.get("lab_group_num"))
				row["lab_sub_num"] = lab_sub_num
				row["cr_id"] = _clean_oracle_num(row.get("cr_id"))
				row["up_id"] = _clean_oracle_num(row.get("up_id"))
				parsed.append(row)
	finally:
		wb.close()
	return parsed


def _resolve_cost_center(branch_label: Any) -> str | None:
	from healthcare.api.discharge_checklist_import import _resolve_cost_center as resolve_cc

	return resolve_cc(branch_label)


def _resolve_patient_from_sub_dr(sub_dr: str) -> str | None:
	key = _clean_oracle_num(sub_dr)
	if not key:
		return None
	if frappe.db.exists("Patient", key):
		return key
	plain = normalize_legacy_id(key)
	if plain and frappe.db.exists("Patient", plain):
		return plain
	name = frappe.db.get_value("Patient", {"file_no": key}, "name")
	if name:
		return name
	if plain:
		return frappe.db.get_value("Patient", {"file_no": plain}, "name")
	return None


def _ensure_patient_from_sub_dr(sub_dr: str) -> tuple[str | None, bool]:
	"""Use SUB_DR_GL_CODE as Patient id; create a minimal Patient when missing."""
	key = _clean_oracle_num(sub_dr)
	if not key:
		return None, False
	existing = _resolve_patient_from_sub_dr(key)
	if existing:
		return existing, False
	from healthcare.api.patient_visit_import import ensure_patient_for_legacy_import

	result = ensure_patient_for_legacy_import(key)
	patient = result.get("patient")
	return patient, result.get("status") == "created"


def _resolve_patient_context(header: dict, *, ensure_patient: bool = True) -> dict:
	"""Resolve visit/admission/patient links; never block import when they are missing.

	When Patient / Patient Visit / Inpatient Admission cannot be found, fall back to
	SUB_DR_GL_CODE as the patient id (creating the Patient if needed).
	"""
	sub_dr = header.get("sub_dr_gl_code") or ""
	visit_num = header.get("visit_num") or ""
	admission_num = header.get("admission_num") or ""

	patient = None
	patient_visit = None
	inpatient_admission = None
	patient_created = False

	if visit_num:
		patient_visit = _resolve_patient_visit(visit_num, sub_dr or None)
		if patient_visit:
			patient = frappe.db.get_value("Patient Visit", patient_visit, "patient")

	if admission_num:
		inpatient_admission = _resolve_inpatient_admission(admission_num, sub_dr or None)
		if inpatient_admission and not patient:
			patient = frappe.db.get_value("Inpatient Admission", inpatient_admission, "patient")

	if not patient:
		patient = _resolve_patient_from_sub_dr(sub_dr)

	if not patient and ensure_patient and sub_dr:
		patient, patient_created = _ensure_patient_from_sub_dr(sub_dr)

	return {
		"patient": patient,
		"patient_visit": patient_visit,
		"inpatient_admission": inpatient_admission,
		"patient_created": patient_created,
		"legacy_visit_num": visit_num if visit_num and not patient_visit else "",
		"legacy_admission_num": admission_num if admission_num and not inpatient_admission else "",
	}


def _pick_primary_template(detail_lines: list[dict]) -> str | None:
	groups = [line.get("lab_group_num") for line in detail_lines if line.get("lab_group_num")]
	if not groups:
		return None
	primary = Counter(groups).most_common(1)[0][0]
	if frappe.db.exists("Lab Test Template", primary):
		return primary
	for group in Counter(groups).keys():
		if frappe.db.exists("Lab Test Template", group):
			return group
	return None


def _primary_lab_group(detail_lines: list[dict]) -> str:
	groups = [line.get("lab_group_num") for line in detail_lines if line.get("lab_group_num")]
	if not groups:
		return ""
	return Counter(groups).most_common(1)[0][0]


def _sort_detail_lines(detail_lines: list[dict]) -> list[dict]:
	def sort_key(line: dict):
		sr = line.get("sr_num") or ""
		try:
			return float(sr)
		except (TypeError, ValueError):
			return sr

	return sorted(detail_lines, key=sort_key)


def _result_text(value: Any) -> str:
	if value is None:
		return ""
	return str(value).strip()


def _build_custom_result(detail_lines: list[dict]) -> str:
	lines = _sort_detail_lines(detail_lines)
	rows: list[str] = []
	plain_parts: list[str] = []

	for line in lines:
		result = _result_text(line.get("lab_result_value"))
		if not result:
			continue
		panel = html.escape(_result_text(line.get("lab_group_num")))
		group_name = html.escape(_result_text(line.get("group_name")))
		sub = html.escape(_result_text(line.get("lab_sub_num")))
		val = html.escape(result)
		label = group_name or panel or sub
		rows.append(f"<tr><td>{label}</td><td>{sub}</td><td>{val}</td></tr>")
		plain_parts.append(f"{sub or label}: {result}")

	if not rows:
		return ""

	table = (
		"<table class='table table-bordered table-sm'>"
		"<thead><tr><th>Panel</th><th>Test Code</th><th>Result</th></tr></thead>"
		f"<tbody>{''.join(rows)}</tbody></table>"
	)
	return table


def _build_results_summary(detail_lines: list[dict]) -> str:
	parts = []
	for line in _sort_detail_lines(detail_lines):
		result = _result_text(line.get("lab_result_value"))
		if not result:
			continue
		sub = _result_text(line.get("lab_sub_num"))
		parts.append(f"{sub}: {result}" if sub else result)
	return "\n".join(parts)


def _legacy_results_field_value(detail_lines: list[dict], *, max_length: int = 140) -> str:
	"""Plain Results field is Data (140 chars). Panel tests store full lines in lab_test_lines."""
	if not detail_lines:
		return ""
	if len(detail_lines) == 1:
		single = _result_text(detail_lines[0].get("lab_result_value"))
		return single[:max_length] if single else ""
	summary = _build_results_summary(detail_lines)
	if len(summary) <= max_length:
		return summary
	return ""


def _build_lab_test_lines(trans_num: str, detail_lines: list[dict]) -> list[dict]:
	rows: list[dict] = []
	for line in _sort_detail_lines(detail_lines):
		rows.append(
			{
				"trans_num": trans_num,
				"sr_num": line.get("sr_num") or "",
				"lab_group_num": line.get("lab_group_num") or "",
				"group_name": _result_text(line.get("group_name")),
				"lab_sub_num": line.get("lab_sub_num") or "",
				"lab_result_value": _result_text(line.get("lab_result_value")),
				"lab_amt_book": _to_currency(line.get("lab_amt_book")),
				"lab_amt_add": _to_currency(line.get("lab_amt_add")),
				"lab_amt_disc": _to_currency(line.get("lab_amt_disc")),
				"lab_amt_net": _to_currency(line.get("lab_amt_net")),
				"sta_flg": _oracle_flag_to_check(line.get("sta_flg")),
				"field1": _result_text(line.get("field1")),
				"field2": _result_text(line.get("field2")),
				"field3": _result_text(line.get("field3")),
				"field4": _result_text(line.get("field4")),
				"field5": _result_text(line.get("field5")),
				"field6": _result_text(line.get("field6")),
				"field7": _result_text(line.get("field7")),
				"field8": _result_text(line.get("field8")),
				"field9": _result_text(line.get("field9")),
				"field10": _result_text(line.get("field10")),
				"cr_id": line.get("cr_id") or "",
				"cr_date": _format_legacy_date_str(line.get("cr_date")),
				"up_id": line.get("up_id") or "",
				"up_date": _format_legacy_date_str(line.get("up_date")),
				"lab_04_remarks": _result_text(line.get("lab_04_remarks")),
			}
		)
	return rows


def _apply_amount_fields(doc, header: dict, detail_lines: list[dict]) -> None:
	total_amount = _to_currency(header.get("total_amount"))
	net_amount = _to_currency(header.get("net_amount"))
	discount_amount = _to_currency(header.get("discount_amount"))
	extra_amount = _to_currency(header.get("extra_amount"))

	if not total_amount and detail_lines:
		total_amount = sum(_to_currency(line.get("lab_amt_book")) for line in detail_lines)
	if not discount_amount and detail_lines:
		discount_amount = sum(_to_currency(line.get("lab_amt_disc")) for line in detail_lines)
	if not extra_amount and detail_lines:
		extra_amount = sum(_to_currency(line.get("lab_amt_add")) for line in detail_lines)
	if not net_amount and detail_lines:
		net_amount = sum(_to_currency(line.get("lab_amt_net")) for line in detail_lines)

	doc.amount = total_amount
	doc.discount_margin = "Amount"
	doc.discount_amount = discount_amount
	doc.lab_amount_addition = extra_amount

	if net_amount:
		doc.grand_total = net_amount
	elif total_amount:
		doc.grand_total = max(total_amount - discount_amount, 0)
	else:
		doc.grand_total = 0


def _default_company() -> str | None:
	company = frappe.defaults.get_global_default("company")
	if company and frappe.db.exists("Company", company):
		return company
	return frappe.db.get_single_value("Global Defaults", "default_company")


def _legacy_status(header: dict | None, detail_lines: list[dict]) -> str:
	header = header or {}
	if _oracle_flag_to_check(header.get("sta_flg")):
		return "Completed"
	for line in detail_lines:
		if _result_text(line.get("lab_result_value")):
			return "Completed"
	return "Draft"


def _effective_trans_datetime(header: dict | None, detail_lines: list[dict]):
	header = header or {}
	trans_dt = _format_legacy_datetime(header.get("trans_date"))
	if trans_dt:
		return trans_dt
	for line in _sort_detail_lines(detail_lines):
		trans_dt = _format_legacy_datetime(line.get("cr_date"))
		if trans_dt:
			return trans_dt
	return None


def _apply_header_fields(
	doc,
	header: dict | None,
	ctx: dict,
	template_name: str | None,
	detail_lines: list[dict],
	*,
	trans_num: str | None = None,
) -> None:
	header = header or {}
	trans_dt = _effective_trans_datetime(header, detail_lines)
	primary_group = _primary_lab_group(detail_lines)
	doc_trans_num = trans_num or header.get("trans_num") or doc.get("trans_num") or ""

	doc.trans_num = doc_trans_num
	doc.is_legacy_import = 1
	doc.template = template_name
	doc.patient = ctx.get("patient")
	doc.patient_visit = ctx.get("patient_visit")
	doc.inpatient_admission = ctx.get("inpatient_admission")
	doc.legacy_visit_num = ctx.get("legacy_visit_num") or ""
	doc.legacy_admission_num = ctx.get("legacy_admission_num") or ""

	if doc.patient and frappe.db.exists("Patient", doc.patient):
		patient = frappe.get_doc("Patient", doc.patient)
		doc.patient_name = patient.patient_name
		doc.patient_sex = patient.sex or doc.patient_sex
		if patient.dob:
			doc.patient_age = patient.age
	elif doc.patient:
		doc.patient_name = _("Patient {0}").format(doc.patient)

	doc.dr_gl_code = header.get("dr_gl_code") or ""
	doc.sub_dr_gl_code = header.get("sub_dr_gl_code") or ""
	doc.trans_source = (header.get("trans_source") or "").strip()
	doc.in_out = (header.get("in_out") or "").strip()
	doc.sta_flg = _oracle_flag_to_check(header.get("sta_flg"))
	doc.old_no = header.get("old_num") or ""
	doc.doc_no = header.get("doc_num") or ""
	doc.lab_test_type = (header.get("lab_test_type") or "").strip()
	doc.confirm = _oracle_flag_to_check(header.get("confirm"))
	doc.lab_test_comment = (header.get("trans_remarks") or "").strip()
	doc.transfer_remark = (header.get("trans_remarks") or "").strip()

	doc.cr_id = header.get("cr_id") or ""
	doc.cr_date = _format_legacy_date_str(header.get("cr_date"))
	doc.up_id = header.get("up_id") or ""
	doc.up_date = _format_legacy_date_str(header.get("up_date"))
	doc.ap_id = header.get("ap_id") or ""
	doc.ap_date = _format_legacy_date_str(header.get("ap_date"))
	doc.branch_num = _clean_oracle_num(header.get("branch_num")) or _result_text(header.get("branch_num"))
	doc.oracle_field7 = _result_text(header.get("oracle_field7"))
	doc.oracle_field8 = _result_text(header.get("oracle_field8"))
	doc.oracle_field9 = _result_text(header.get("oracle_field9"))
	doc.oracle_field10 = _result_text(header.get("oracle_field10"))

	doc.fill_id = header.get("fill_id") or ""
	doc.fill_date = _format_legacy_date_str(header.get("fill_date"))
	doc.sample_collect_id = header.get("sample_collect_id") or ""
	doc.sample_collected_id = header.get("sample_collect_id") or ""
	doc.sample_collected_date = _format_legacy_date_str(header.get("sample_collect_date"))
	doc.transaction_date = _format_legacy_date_str(header.get("trans_date"))

	if detail_lines:
		doc.sr_num = detail_lines[0].get("sr_num") or ""

	if trans_dt:
		doc.date = getdate(trans_dt)
		doc.time = trans_dt.time()
		doc.result_date = getdate(trans_dt)

	cc = _resolve_cost_center(header.get("branch_num"))
	if cc:
		doc.cost_center = cc

	company = _default_company()
	if company:
		doc.company = company

	if template_name and frappe.db.exists("Lab Test Template", template_name):
		template = frappe.get_doc("Lab Test Template", template_name)
		doc.lab_test_name = template.lab_test_name
		doc.legend_print_position = template.legend_print_position
		doc.result_legend = template.result_legend
		doc.worksheet_instructions = template.worksheet_instructions
	elif primary_group:
		first_name = _result_text(detail_lines[0].get("group_name")) if detail_lines else ""
		doc.lab_test_name = first_name or primary_group

	_apply_amount_fields(doc, header, detail_lines)
	doc.custom_result = _build_custom_result(detail_lines)
	doc.results = _legacy_results_field_value(detail_lines)
	doc.status = _legacy_status(header, detail_lines)


def import_legacy_lab_test(
	trans_num: str,
	header: dict | None,
	detail_lines: list[dict],
	*,
	batch_id: str | None = None,
) -> dict:
	if not header and not detail_lines:
		return {"status": "skip_no_data", "trans_num": trans_num}

	detail_only = not header
	patient_created = False
	if detail_only:
		header = {"trans_num": trans_num}
		ctx = {
			"patient": None,
			"patient_visit": None,
			"inpatient_admission": None,
			"patient_created": False,
			"legacy_visit_num": "",
			"legacy_admission_num": "",
		}
	else:
		ctx = _resolve_patient_context(header, ensure_patient=True)
		patient_created = bool(ctx.get("patient_created"))
		# Missing visit / admission / patient must not block import.
		# Patient is created from SUB_DR_GL_CODE when possible; otherwise Lab Test is stored without link.

	template_name = _pick_primary_template(detail_lines)
	lab_lines = _build_lab_test_lines(trans_num, detail_lines)
	existing = frappe.db.exists("Lab Test", trans_num)

	if existing:
		doc = frappe.get_doc("Lab Test", trans_num)
		if not doc.get("is_legacy_import"):
			return {"status": "skip_existing_non_legacy", "trans_num": trans_num}
	else:
		doc = frappe.new_doc("Lab Test")
		doc.trans_num = trans_num

	_apply_header_fields(
		doc, header, ctx, template_name, detail_lines, trans_num=trans_num
	)
	if batch_id:
		doc.legacy_import_batch = batch_id

	doc.set("lab_test_lines", [])
	for row in lab_lines:
		doc.append("lab_test_lines", row)

	doc.flags.ignore_validate = True
	doc.flags.ignore_mandatory = True
	doc.flags.ignore_links = True
	doc.flags.legacy_import = True

	if existing:
		doc.save(ignore_permissions=True)
		action = "updated"
	else:
		doc.insert(ignore_permissions=True)
		action = "created"

	if doc.docstatus == 0 and doc.status == "Completed":
		try:
			doc.reload()
			doc.flags.ignore_validate = True
			doc.flags.ignore_mandatory = True
			doc.flags.ignore_links = True
			doc.flags.legacy_import = True
			doc.submit()
		except Exception:
			# Keep the imported Lab Test even when submit fails (still in database as Draft/Completed).
			frappe.log_error(
				title=f"Legacy lab import submit failed: {trans_num}",
				message=frappe.get_traceback(),
			)

	return {
		"status": "ok",
		"trans_num": trans_num,
		"action": action,
		"lab_test": doc.name,
		"patient": doc.patient,
		"patient_created": patient_created,
		"template": template_name,
		"result_rows": len(lab_lines),
		"standalone": detail_only,
	}


def _clear_import_failures() -> None:
	frappe.cache().delete_value(CACHE_KEYS["failures"])


def _get_import_failures() -> list[dict]:
	raw = frappe.cache().get_value(CACHE_KEYS["failures"]) or []
	return list(raw) if isinstance(raw, list) else []


def _append_import_failure(trans_num: str, reason: str, detail: str = "") -> None:
	failures = _get_import_failures()
	failures.append(
		{
			"trans_num": trans_num,
			"reason": reason,
			"detail": (detail or "")[:FAILURE_DETAIL_MAX],
		}
	)
	frappe.cache().set_value(CACHE_KEYS["failures"], failures, expires_in_sec=CACHE_TTL)


def get_missing_legacy_lab_trans_nums(trans_nums: list[str] | None = None) -> list[str]:
	"""Excel TRANS_NUM values with no legacy Lab Test in the database."""
	if trans_nums is None:
		trans_nums = frappe.cache().get_value(CACHE_KEYS["trans_nums"]) or []
	if not trans_nums:
		return []
	existing = set(frappe.db.sql_list("SELECT name FROM `tabLab Test` WHERE is_legacy_import = 1"))
	return sorted(set(trans_nums) - existing)


def build_legacy_lab_import_summary(progress: dict | None = None) -> dict:
	"""Final import report: counts, failures, and TRANS_NUM not in database."""
	progress = progress or {}
	trans_nums = frappe.cache().get_value(CACHE_KEYS["trans_nums"]) or []
	failures = _get_import_failures()
	missing = get_missing_legacy_lab_trans_nums(trans_nums)
	in_db = frappe.db.count("Lab Test", {"is_legacy_import": 1})
	total = len(trans_nums)
	ok = cint(progress.get("ok", 0))
	errors = cint(progress.get("errors", 0))
	skipped = (
		cint(progress.get("skip_no_patient", 0))
		+ cint(progress.get("skip_no_header", 0))
		+ cint(progress.get("skip_existing_non_legacy", 0))
	)
	return {
		"total_transactions": total,
		"processed": cint(progress.get("processed", 0)),
		"in_database": in_db,
		"ok": ok,
		"errors": errors,
		"skipped": skipped,
		"skip_no_patient": cint(progress.get("skip_no_patient", 0)),
		"skip_no_header": cint(progress.get("skip_no_header", 0)),
		"skip_existing_non_legacy": cint(progress.get("skip_existing_non_legacy", 0)),
		"missing_from_database": len(missing),
		"missing_trans_nums": missing[:FAILURE_LOG_LIST_MAX],
		"failure_count": len(failures),
		"failures": failures[:FAILURE_LOG_LIST_MAX],
		"batch_id": progress.get("batch_id") or frappe.cache().get_value(CACHE_KEYS["batch_id"]),
		"done": progress.get("done"),
		"updated_at": progress.get("updated_at"),
	}


def log_legacy_lab_import_completion(progress: dict | None = None) -> dict:
	"""Write a detailed Error Log entry when the background import finishes."""
	summary = build_legacy_lab_import_summary(progress)
	message_lines = [
		"Legacy lab import summary",
		f"Excel transactions: {summary['total_transactions']}",
		f"Processed by job: {summary['processed']}",
		f"Imported OK (created/updated): {summary['ok']}",
		f"Standalone (detail only): {cint(progress.get('standalone_ok', 0))}",
		f"Patients created from SUB_DR_GL_CODE: {cint(progress.get('patients_created', 0))}",
		f"Skipped: {summary['skipped']} "
		f"(no patient: {summary['skip_no_patient']}, "
		f"no header: {summary['skip_no_header']}, "
		f"existing non-legacy: {summary['skip_existing_non_legacy']})",
		f"Errors during import: {summary['errors']}",
		f"Legacy Lab Tests in database: {summary['in_database']}",
		f"Missing from database: {summary['missing_from_database']}",
	]
	if summary["missing_trans_nums"]:
		message_lines.append("")
		message_lines.append("Missing TRANS_NUM (not in database):")
		message_lines.extend(summary["missing_trans_nums"])
	if summary["failures"]:
		message_lines.append("")
		message_lines.append("Recorded failures:")
		for row in summary["failures"]:
			line = f"- {row.get('trans_num')}: {row.get('reason')}"
			if row.get("detail"):
				line += f" — {row['detail'][:200]}"
			message_lines.append(line)

	frappe.log_error(
		title="Legacy lab import complete",
		message="\n".join(message_lines),
	)
	# One Error Log row per failed TRANS_NUM (easy to search in Error Log list).
	failure_trans = {f.get("trans_num") for f in failures if f.get("trans_num")}
	for row in summary["failures"]:
		trans_num = row.get("trans_num") or "?"
		frappe.log_error(
			title=f"Legacy lab import skipped/failed: {trans_num}",
			message=row.get("detail") or row.get("reason") or "unknown",
		)
	for trans_num in summary["missing_trans_nums"]:
		if trans_num in failure_trans:
			continue
		frappe.log_error(
			title=f"Legacy lab import missing in database: {trans_num}",
			message=_(
				"TRANS_NUM was in the Excel header file but no legacy Lab Test exists after import completed."
			),
		)
	return summary


def parse_and_cache_excel(header_file_url: str, detail_file_url: str) -> dict:
	headers_list = _parse_header_excel(header_file_url)
	details_list = _parse_detail_excel(detail_file_url)

	headers_by_trans = {row["trans_num"]: row for row in headers_list}
	details_by_trans: dict[str, list[dict]] = {}
	for row in details_list:
		key = row["trans_num"]
		details_by_trans.setdefault(key, []).append(row)

	trans_nums = sorted(set(headers_by_trans.keys()) | set(details_by_trans.keys()))
	batch_id = frappe.generate_hash(length=10)
	_clear_import_failures()

	header_trans = set(headers_by_trans.keys())
	detail_only_trans = set(details_by_trans.keys()) - header_trans

	resolvable_patient = 0
	will_create_patient = 0
	with_results = 0
	with_template = 0
	standalone_with_lines = 0
	for trans_num in trans_nums:
		header = headers_by_trans.get(trans_num)
		lines = details_by_trans.get(trans_num) or []
		if header:
			ctx = _resolve_patient_context(header, ensure_patient=False)
			if ctx.get("patient"):
				resolvable_patient += 1
			elif header.get("sub_dr_gl_code"):
				will_create_patient += 1
		elif lines:
			standalone_with_lines += 1
		if _pick_primary_template(lines):
			with_template += 1
		if lines:
			with_results += 1

	frappe.cache().set_value(CACHE_KEYS["header_file_url"], header_file_url, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(CACHE_KEYS["detail_file_url"], detail_file_url, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(CACHE_KEYS["batch_id"], batch_id, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(CACHE_KEYS["trans_nums"], trans_nums, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(
		CACHE_KEYS["headers"],
		json.dumps(headers_by_trans, default=str),
		expires_in_sec=CACHE_TTL,
	)
	frappe.cache().set_value(
		CACHE_KEYS["details"],
		json.dumps(details_by_trans, default=str),
		expires_in_sec=CACHE_TTL,
	)

	detail_only = len(detail_only_trans)
	return {
		"batch_id": batch_id,
		"header_rows": len(headers_list),
		"detail_rows": len(details_list),
		"transactions": len(trans_nums),
		"transactions_with_header": len(header_trans),
		"transactions_with_results": with_results,
		"resolvable_patient": resolvable_patient,
		"will_create_patient_from_sub_dr": will_create_patient,
		"resolvable_template": with_template,
		"detail_without_header": detail_only,
		"standalone_transactions": standalone_with_lines,
	}


def _load_cached_headers() -> dict[str, dict]:
	raw = frappe.cache().get_value(CACHE_KEYS["headers"])
	if not raw:
		return {}
	if isinstance(raw, dict):
		return raw
	return json.loads(raw)


def _load_cached_details() -> dict[str, list[dict]]:
	raw = frappe.cache().get_value(CACHE_KEYS["details"])
	if not raw:
		return {}
	if isinstance(raw, dict):
		return raw
	return json.loads(raw)


@frappe.whitelist()
def get_legacy_lab_import_report() -> dict:
	"""Current import summary + TRANS_NUM missing from the database (admin)."""
	_require_admin()
	try:
		from healthcare.api.data_migration_jobs import _job_progress_key

		progress = frappe.cache().get_value(_job_progress_key("legacy_lab_import")) or {}
	except Exception:
		progress = {}
	return build_legacy_lab_import_summary(progress)


@frappe.whitelist()
def preview_legacy_lab_bundle_import(header_file_url: str, detail_file_url: str) -> dict:
	"""Direct Upload: LAB_00_03 header + LAB_00_04 detail (all sheets)."""
	return preview_legacy_lab_import(header_file_url, detail_file_url)


@frappe.whitelist()
def preview_legacy_lab_import(header_file_url: str, detail_file_url: str) -> dict:
	_require_admin()
	if not (header_file_url or "").strip():
		frappe.throw(_("Please upload the lab header Excel file (C LAB_00_03)."))
	if not (detail_file_url or "").strip():
		frappe.throw(_("Please upload the lab detail Excel file (C-I LAB_00_04)."))
	return parse_and_cache_excel(header_file_url, detail_file_url)


def run_legacy_lab_import_batch(offset: int = 0) -> dict:
	trans_nums = frappe.cache().get_value(CACHE_KEYS["trans_nums"]) or []
	headers = _load_cached_headers()
	details = _load_cached_details()
	batch_id = frappe.cache().get_value(CACHE_KEYS["batch_id"]) or ""

	if not trans_nums:
		return {"processed": offset, "done": True, "batch_count": 0}

	batch_keys = trans_nums[offset : offset + LEGACY_LAB_IMPORT_BATCH_SIZE]
	if not batch_keys:
		return {"processed": offset, "done": True, "batch_count": 0}

	ok = skip_no_patient = skip_no_header = skip_existing = standalone_ok = patients_created = 0
	errors: list[str] = []

	for trans_num in batch_keys:
		header = headers.get(trans_num)
		lines = details.get(trans_num) or []
		if not header and not lines:
			skip_no_header += 1
			_append_import_failure(trans_num, "skip_no_data")
			continue
		savepoint = f"legacy_lab_{offset}_{trans_num}".replace("/", "_").replace(" ", "_")[:60]
		try:
			frappe.db.savepoint(savepoint)
			result = import_legacy_lab_test(
				trans_num,
				header,
				lines,
				batch_id=batch_id,
			)
			status = result.get("status")
			if status == "ok":
				ok += 1
				if result.get("standalone"):
					standalone_ok += 1
				if result.get("patient_created"):
					patients_created += 1
			elif status == "skip_no_patient":
				# Kept for older cached jobs; current importer no longer skips for missing patient.
				skip_no_patient += 1
				_append_import_failure(trans_num, status)
			elif status == "skip_no_header":
				skip_no_header += 1
				_append_import_failure(trans_num, status)
			elif status == "skip_no_data":
				skip_no_header += 1
				_append_import_failure(trans_num, status)
			elif status == "skip_existing_non_legacy":
				skip_existing += 1
				_append_import_failure(trans_num, status)
			else:
				errors.append(f"{trans_num}: {status}")
				_append_import_failure(trans_num, status or "unknown")
		except Exception:
			frappe.db.rollback(save_point=savepoint)
			tb = frappe.get_traceback()
			errors.append(f"{trans_num}: {tb}")
			_append_import_failure(trans_num, "exception", tb)

	frappe.db.commit()

	processed = offset + len(batch_keys)
	return {
		"processed": processed,
		"done": processed >= len(trans_nums),
		"batch_count": len(batch_keys),
		"ok": ok,
		"standalone_ok": standalone_ok,
		"patients_created": patients_created,
		"skip_no_patient": skip_no_patient,
		"skip_no_header": skip_no_header,
		"skip_existing_non_legacy": skip_existing,
		"errors": len(errors),
		"error_samples": errors[:5],
	}
