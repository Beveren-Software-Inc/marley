"""Import Oracle IP_RISK_ANALYSIS Excel into IP Risk Analysis rows (one per admission)."""

from __future__ import annotations

import json
from typing import Any

import frappe
from frappe import _

from healthcare.api.discharge_checklist_import import _resolve_cost_center
from healthcare.api.patient_info_import import (
	_cell_text,
	_clean_oracle_num,
	_excel_file_path,
	_require_admin,
)
from healthcare.api.visit_diagnoses_op_import import _legacy_data_datetime
from healthcare.api.visit_diagnosis_sync import _resolve_inpatient_admission

IP_RISK_ANALYSIS_IMPORT_BATCH_SIZE = 500
CACHE_TTL = 7200
CACHE_KEYS = {
	"file_url": "healthcare:data_migration:ip_risk_analysis_import:file_url",
	"row_keys": "healthcare:data_migration:ip_risk_analysis_import:row_keys",
	"rows": "healthcare:data_migration:ip_risk_analysis_import:rows",
}

EXCEL_HEADER_MAP = {
	"ADMISSION_NUM": "admission_num",
	"BRANCH_NUM": "branch_num",
	"RTS_SLF_HARM_PARASUI_YN": "rts_slf_harm_parasui_yn",
	"RTS_SLF_HARM_PARASUI_DTL": "rts_slf_harm_parasui_dtl",
	"RTS_SUICIDAL_BEHAV_IN_PAST_YN": "rts_suicidal_behav_in_past_yn",
	"RTS_SUICIDAL_BEHAV_IN_PAST_DTL": "rts_suicidal_behav_in_past_dtl",
	"RTS_SUICIDAL_THOUGHTS_YN": "rts_suicidal_thoughts_yn",
	"RTS_SUICIDAL_THOUGHTS_DTL": "rts_suicidal_thoughts_dtl",
	"RTS_COMPL_TO_PSHYC_FOLLO_YN": "rts_compl_to_pshyc_follo_yn",
	"RTS_COMPL_TO_PSHYC_MEDIC_YN": "rts_compl_to_pshyc_medic_yn",
	"RTS_DRIVING_YN": "rts_driving_yn",
	"RTS_SELF_NEGLECT_YN": "rts_self_neglect_yn",
	"RTS_MALNUTRITION_DEHYDR_YN": "rts_malnutrition_dehydr_yn",
	"RTS_FALLS_YN": "rts_falls_yn",
	"RTS_SETTING_FIRE_OVEN_YN": "rts_setting_fire_oven_yn",
	"RTS_1_FREE": "rts_1_free",
	"RTS_2_FREE": "rts_2_free",
	"RTS_3_FREE": "rts_3_free",
	"RTS_4_FREE": "rts_4_free",
	"RTS_5_FREE": "rts_5_free",
	"RTO_BULLYING_YN": "rto_bullying_yn",
	"RTO_EXPLOITATION_LIST": "rto_exploitation_list",
	"RTO_EXPLOITATION_DTL": "rto_exploitation_dtl",
	"RTO_NEGLECT_LIST": "rto_neglect_list",
	"RTO_NEGLECT_DTL": "rto_neglect_dtl",
	"RTO_VIOLENCE_AGGRESSION_YN": "rto_violence_aggression_yn",
	"RTO_VIOLENCE_AGGRESSION_DTL": "rto_violence_aggression_dtl",
	"RTO_1_FREE": "rto_1_free",
	"RTO_2_FREE": "rto_2_free",
	"RTO_3_FREE": "rto_3_free",
	"RTO_4_FREE": "rto_4_free",
	"RTO_5_FREE": "rto_5_free",
	"RFO_BULLYING_YN": "rfo_bullying_yn",
	"RFO_EXPLOITATION_LIST": "rfo_exploitation_list",
	"RFO_EXPLOITATION_DTL": "rfo_exploitation_dtl",
	"RFO_NEGLECT_LIST": "rfo_neglect_list",
	"RFO_VIOLENCE_AGGRESSION_YN": "rfo_violence_aggression_yn",
	"RFO_VIOLENCE_AGGRESSION_DTL": "rfo_violence_aggression_dtl",
	"RFO_1_FREE": "rfo_1_free",
	"RFO_2_FREE": "rfo_2_free",
	"RFO_3_FREE": "rfo_3_free",
	"RFO_4_FREE": "rfo_4_free",
	"RFO_5_FREE": "rfo_5_free",
	"CR_ID": "cr_id",
	"CR_DATE": "cr_date",
	"UP_ID": "up_id",
	"UP_DATE": "up_date",
	"ADMISSION_NUM_OLD": "admission_num_old",
}

EXPLOITATION_MAP = {
	"F": "Financial",
	"S": "Sexual",
	"O": "Others",
	"FINANCIAL": "Financial",
	"SEXUAL": "Sexual",
	"OTHERS": "Others",
}

NEGLECT_MAP = {
	"C": "Children",
	"P": "Pets",
	"O": "Others",
	"CHILDREN": "Children",
	"PETS": "Pets",
	"OTHERS": "Others",
}


def _normalize_header(cell: Any) -> str:
	if cell is None:
		return ""
	text = str(cell).strip().upper().replace(" ", "_")
	return EXCEL_HEADER_MAP.get(text, text.lower())


def _yn_to_check(value: Any) -> int:
	text = _cell_text(value).upper()
	return 1 if text in ("Y", "YES", "1", "TRUE", "T") else 0


def _yn_to_yes_no(value: Any) -> str | None:
	text = _cell_text(value).upper()
	if text in ("Y", "YES", "1", "TRUE", "T"):
		return "Yes"
	if text in ("N", "NO", "0", "FALSE", "F"):
		return "No"
	return None


def _text_value(value: Any) -> str | None:
	text = _cell_text(value)
	return text or None


def _map_exploitation(value: Any) -> str | None:
	text = _cell_text(value).upper()
	if not text:
		return None
	return EXPLOITATION_MAP.get(text, _cell_text(value))


def _map_neglect(value: Any) -> str | None:
	text = _cell_text(value).upper()
	if not text:
		return None
	return NEGLECT_MAP.get(text, _cell_text(value))


def _resolve_admission(row: dict) -> str | None:
	admission_num = row.get("admission_num") or ""
	old_num = row.get("admission_num_old") or ""
	for candidate in (admission_num, old_num):
		if not candidate:
			continue
		resolved = _resolve_inpatient_admission(candidate)
		if resolved:
			return resolved
	return None


def _existing_ip_risk_analysis_name(admission: str) -> str | None:
	return frappe.db.get_value(
		"IP Risk Analysis",
		{"admission_no": admission},
		"name",
		order_by="creation asc",
	)


def _free_child_rows(row: dict, keys: list[str], child_field: str) -> list[dict]:
	rows: list[dict] = []
	for key in keys:
		value = _text_value(row.get(key))
		if value:
			rows.append({child_field: value})
	return rows


def _build_ip_risk_analysis_fields(row: dict, admission: str) -> dict[str, Any]:
	patient = frappe.db.get_value("Inpatient Admission", admission, "patient")
	patient_name = frappe.db.get_value("Patient", patient, "patient_name") if patient else None

	fields: dict[str, Any] = {
		"admission_no": admission,
		"self_harming": _yn_to_check(row.get("rts_slf_harm_parasui_yn")),
		"suicidal_behaviourattempt": _yn_to_check(row.get("rts_suicidal_behav_in_past_yn")),
		"suicidal_thoughts_plans_currently": _yn_to_check(row.get("rts_suicidal_thoughts_yn")),
		"rto_bullying": _yn_to_check(row.get("rto_bullying_yn")),
		"rto_violence_aggression": _yn_to_check(row.get("rto_violence_aggression_yn")),
		"rfo_bullying": _yn_to_check(row.get("rfo_bullying_yn")),
		"rfo_violence": _yn_to_check(row.get("rfo_violence_aggression_yn")),
	}

	if patient:
		fields["file_number"] = patient
	if patient_name:
		fields["patient_name"] = patient_name

	text_map = {
		"self_harming_description": row.get("rts_slf_harm_parasui_dtl"),
		"suicide_behaviour_description": row.get("rts_suicidal_behav_in_past_dtl"),
		"suicidal_thoughts_description": row.get("rts_suicidal_thoughts_dtl"),
		"rto_exploitation_description": row.get("rto_exploitation_dtl"),
		"rto_violence_aggression_reason": row.get("rto_violence_aggression_dtl"),
		"rto_neglect_description": row.get("rto_neglect_dtl"),
		"rfo_exploitation_description": row.get("rfo_exploitation_dtl"),
		"rfo_violence_reason": row.get("rfo_violence_aggression_dtl"),
	}
	for field, value in text_map.items():
		text = _text_value(value)
		if text:
			fields[field] = text

	yes_no_map = {
		"compliance_to_psychiatric_followup": row.get("rts_compl_to_pshyc_follo_yn"),
		"compliance_to_psychiatric_medication": row.get("rts_compl_to_pshyc_medic_yn"),
		"driving": row.get("rts_driving_yn"),
		"self_neglect": row.get("rts_self_neglect_yn"),
		"malnutrition__dehydration": row.get("rts_malnutrition_dehydr_yn"),
		"falls": row.get("rts_falls_yn"),
		"setting_fire_overelectricity_etc": row.get("rts_setting_fire_oven_yn"),
	}
	for field, value in yes_no_map.items():
		mapped = _yn_to_yes_no(value)
		if mapped:
			fields[field] = mapped

	exploitation = _map_exploitation(row.get("rto_exploitation_list"))
	if exploitation:
		fields["rto_exploitation"] = exploitation
	neglect = _map_neglect(row.get("rto_neglect_list"))
	if neglect:
		fields["rto_neglect"] = neglect

	rfo_exploitation = _map_exploitation(row.get("rfo_exploitation_list"))
	if rfo_exploitation:
		fields["rfo_exploitation"] = rfo_exploitation
	rfo_neglect = _map_neglect(row.get("rfo_neglect_list"))
	if rfo_neglect:
		fields["rfo_neglect"] = rfo_neglect

	if row.get("admission_num_old"):
		fields["admission_no_old"] = row.get("admission_num_old")

	cost_center = _resolve_cost_center(row.get("branch_num"))
	if cost_center:
		fields["cost_center"] = cost_center
	elif admission:
		cc = frappe.db.get_value("Inpatient Admission", admission, "cost_center")
		if cc:
			fields["cost_center"] = cc

	if row.get("cr_id"):
		fields["cr_id"] = row["cr_id"]
	if row.get("up_id"):
		fields["up_id"] = row["up_id"]
	cr_date = _legacy_data_datetime(row.get("cr_date"))
	if cr_date:
		fields["cr_date"] = cr_date
	up_date = _legacy_data_datetime(row.get("up_date"))
	if up_date:
		fields["up_date"] = up_date

	fields["risk_to_others_free"] = _free_child_rows(
		row,
		["rto_1_free", "rto_2_free", "rto_3_free", "rto_4_free", "rto_5_free"],
		"rto_free",
	)
	fields["risk_from_others_free"] = _free_child_rows(
		row,
		["rfo_1_free", "rfo_2_free", "rfo_3_free", "rfo_4_free", "rfo_5_free"],
		"rfo_free",
	)

	return fields


def _apply_legacy_import_flags(doc) -> None:
	doc.flags.ignore_validate = True
	doc.flags.ignore_mandatory = True
	doc.flags.ignore_links = True
	doc.flags.ignore_permissions = True
	doc.flags.legacy_import = True


def _set_child_table(doc, fieldname: str, rows: list[dict]) -> None:
	doc.set(fieldname, [])
	for row in rows:
		doc.append(fieldname, row)


def upsert_ip_risk_analysis_from_row(row: dict) -> dict:
	admission = _resolve_admission(row)
	if not admission:
		admission_key = row.get("admission_num") or row.get("admission_num_old") or ""
		return {"status": "skip_no_admission", "admission_key": admission_key}

	fields = _build_ip_risk_analysis_fields(row, admission)
	rto_free = fields.pop("risk_to_others_free", [])
	rfo_free = fields.pop("risk_from_others_free", [])

	existing_name = _existing_ip_risk_analysis_name(admission)
	if existing_name:
		doc = frappe.get_doc("IP Risk Analysis", existing_name)
		action = "updated"
	else:
		doc = frappe.get_doc({"doctype": "IP Risk Analysis", **fields})
		action = "created"

	if existing_name:
		for key, value in fields.items():
			doc.set(key, value)

	_set_child_table(doc, "risk_to_others_free", rto_free)
	_set_child_table(doc, "risk_from_others_free", rfo_free)

	_apply_legacy_import_flags(doc)
	if existing_name:
		doc.save(ignore_permissions=True)
	else:
		doc.insert(ignore_permissions=True)

	return {
		"status": action,
		"admission_no": admission,
		"name": doc.name,
	}


def _parse_sheet_rows(ws) -> list[dict]:
	rows_iter = ws.iter_rows(values_only=True)
	try:
		header_row = next(rows_iter)
	except StopIteration:
		return []

	headers = [_normalize_header(h) for h in header_row]
	parsed: list[dict] = []
	for raw in rows_iter:
		if not raw or all(cell is None or str(cell).strip() == "" for cell in raw):
			continue
		row: dict[str, Any] = {}
		for idx, key in enumerate(headers):
			if not key or idx >= len(raw):
				continue
			row[key] = raw[idx]

		admission_num = _clean_oracle_num(row.get("admission_num"))
		if not admission_num:
			continue

		row["admission_num"] = admission_num
		row["admission_num_old"] = _clean_oracle_num(row.get("admission_num_old"))
		row["branch_num"] = _clean_oracle_num(row.get("branch_num"))
		row["cr_id"] = _clean_oracle_num(row.get("cr_id"))
		row["up_id"] = _clean_oracle_num(row.get("up_id"))
		parsed.append(row)
	return parsed


def _parse_excel_rows(file_url: str) -> tuple[list[dict], dict[str, int]]:
	try:
		import openpyxl
	except ImportError:
		frappe.throw(
			_(
				"openpyxl is required to read Excel files. Install it in the bench environment: pip install openpyxl"
			)
		)

	path = _excel_file_path(file_url)
	wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
	all_rows: list[dict] = []
	sheet_row_counts: dict[str, int] = {}
	try:
		for sheet_name in wb.sheetnames:
			sheet_rows = _parse_sheet_rows(wb[sheet_name])
			sheet_row_counts[sheet_name] = len(sheet_rows)
			all_rows.extend(sheet_rows)
	finally:
		wb.close()
	return all_rows, sheet_row_counts


def _preview_counts(rows: list[dict]) -> dict:
	existing = 0
	resolved = 0
	unresolved = 0
	for row in rows:
		admission = _resolve_admission(row)
		if not admission:
			unresolved += 1
			continue
		resolved += 1
		if _existing_ip_risk_analysis_name(admission):
			existing += 1

	return {
		"existing_analyses": existing,
		"new_analyses": resolved - existing,
		"resolved_admissions": resolved,
		"unresolved_admissions": unresolved,
		"sample_admissions": [row.get("admission_num") for row in rows[:5]],
	}


def parse_and_cache_excel(file_url: str) -> dict:
	rows, sheet_row_counts = _parse_excel_rows(file_url)
	by_key = {row["admission_num"]: row for row in rows}
	row_keys = list(by_key.keys())

	frappe.cache().set_value(CACHE_KEYS["file_url"], file_url, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(CACHE_KEYS["row_keys"], row_keys, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(
		CACHE_KEYS["rows"],
		json.dumps(by_key, default=str),
		expires_in_sec=CACHE_TTL,
	)

	preview = _preview_counts(list(by_key.values()))
	raw_row_total = sum(sheet_row_counts.values())
	return {
		"excel_rows": len(by_key),
		"raw_excel_rows": raw_row_total,
		"sheets": list(sheet_row_counts.keys()),
		"sheet_row_counts": sheet_row_counts,
		**preview,
	}


def _load_cached_rows() -> dict[str, dict]:
	raw = frappe.cache().get_value(CACHE_KEYS["rows"])
	if not raw:
		return {}
	if isinstance(raw, dict):
		return raw
	return json.loads(raw)


@frappe.whitelist()
def preview_ip_risk_analysis_import(file_url: str) -> dict:
	_require_admin()
	if not (file_url or "").strip():
		frappe.throw(_("Please upload the IP_RISK_ANALYSIS Excel file."))
	return parse_and_cache_excel(file_url)


def run_ip_risk_analysis_import_batch(*, offset: int = 0) -> dict:
	row_keys = frappe.cache().get_value(CACHE_KEYS["row_keys"]) or []
	rows_by_key = _load_cached_rows()
	if not row_keys or not rows_by_key:
		return {"processed": offset, "done": True, "batch_count": 0}

	batch_keys = row_keys[offset : offset + IP_RISK_ANALYSIS_IMPORT_BATCH_SIZE]
	if not batch_keys:
		return {"processed": offset, "done": True, "batch_count": 0}

	created = updated = skip_no_admission = 0
	errors: list[str] = []

	for key in batch_keys:
		row = rows_by_key.get(key) or {}
		try:
			result = upsert_ip_risk_analysis_from_row(row)
			status = result.get("status")
			if status == "created":
				created += 1
			elif status == "updated":
				updated += 1
			elif status == "skip_no_admission":
				skip_no_admission += 1
			else:
				errors.append(f"{key}: {status}")
		except Exception:
			errors.append(f"{key}: {frappe.get_traceback()}")
			frappe.log_error(title=f"IP_RISK_ANALYSIS import failed: {key}")

	frappe.db.commit()
	processed = offset + len(batch_keys)
	return {
		"processed": processed,
		"done": len(batch_keys) < IP_RISK_ANALYSIS_IMPORT_BATCH_SIZE,
		"batch_count": len(batch_keys),
		"created": created,
		"updated": updated,
		"skip_no_admission": skip_no_admission,
		"errors": len(errors),
	}
