"""Import Oracle PATIENT_VISIT_PRESCRIPTION Excel into Patient Medication Order (OP)."""

from __future__ import annotations

import json
from typing import Any

import frappe
from frappe import _
from frappe.utils import cint, nowdate

from healthcare.api.discharge_checklist_import import _resolve_cost_center
from healthcare.api.patient_info_import import (
	_clean_oracle_num,
	_excel_file_path,
	_parse_date_value,
	_require_admin,
)
from healthcare.api.patient_medication_order_import import (
	_append_child_line,
	_ensure_prescription_frequency_label,
	_patient_display_fields,
)
from healthcare.api.patient_visit_prescription_common import (
	existing_pmo_for_legacy_visit,
	preview_counts_for_legacy_visit_pmo,
	submit_and_complete_legacy_visit_pmo,
)
from healthcare.api.utils.api_utility import get_next_transaction_number
from healthcare.api.visit_diagnosis_sync import _resolve_patient_visit

DOCTYPE = "Patient Medication Order"
VISIT_PRESCRIPTION_IMPORT_BATCH_SIZE = 200
CACHE_TTL = 7200
CACHE_KEYS = {
	"file_url": "healthcare:data_migration:patient_visit_prescription_import:file_url",
	"visit_keys": "healthcare:data_migration:patient_visit_prescription_import:visit_keys",
	"visits": "healthcare:data_migration:patient_visit_prescription_import:visits",
}

EXCEL_HEADER_MAP = {
	"PRESCRIP_ID": "prescrip_id",
	"PRESCRIP_DURATION": "prescrip_duration",
	"PRESCRIP_FREQ": "prescrip_freq",
	"NOTE": "note",
	"PATIENT_FILE_NO": "patient_file_no",
	"VISIT_CD": "visit_cd",
	"HIST_ORD": "hist_ord",
	"MEDICINE_CD": "medicine_cd",
	"IF_NEEDED": "if_needed",
	"QUANTITY": "quantity",
	"PERIOD": "period",
	"CREATE_USER_ID": "create_user_id",
	"CREATE_DATE": "create_date",
	"ROUTE": "route",
	"STRENGTH": "strength",
	"UNIT": "unit",
	"FREQUENCY": "frequency",
	"DURATION": "duration",
	"DURATION_TYPE": "duration_type",
	"QTY": "qty",
	"START_DATE": "start_date",
	"END_DATE": "end_date",
	"UPDATE_USER_ID": "update_user_id",
	"UPDATE_DATE": "update_date",
	"BRANCH_NUM": "branch_num",
	"CR_ID": "cr_id",
	"CR_DATE": "cr_date",
	"UP_ID": "up_id",
	"UP_DATE": "up_date",
}


def _normalize_header(cell: Any) -> str:
	if cell is None:
		return ""
	text = str(cell).strip().upper().replace(" ", "_")
	return EXCEL_HEADER_MAP.get(text, text.lower())


def _resolve_patient(patient_file_no: Any) -> str | None:
	patient = _clean_oracle_num(patient_file_no)
	if not patient:
		return None
	if frappe.db.exists("Patient", patient):
		return patient
	return None


def _resolve_patient_visit_link(visit_cd: Any, patient_file_no: Any = None) -> str | None:
	visit_key = _clean_oracle_num(visit_cd)
	if not visit_key:
		return None
	patient = _clean_oracle_num(patient_file_no) if patient_file_no not in (None, "") else None
	resolved = _resolve_patient_visit(visit_key, patient)
	if resolved and frappe.db.exists("Patient Visit", resolved):
		return resolved
	return None


def _default_company() -> str | None:
	company = frappe.defaults.get_global_default("company")
	if company and frappe.db.exists("Company", company):
		return company
	rows = frappe.get_all("Company", pluck="name", limit=1)
	return rows[0] if rows else None


def _prepare_line_row(row: dict[str, Any]) -> dict[str, Any] | None:
	visit_cd = _clean_oracle_num(row.get("visit_cd"))
	patient_file_no = row.get("patient_file_no")
	if not visit_cd or patient_file_no in (None, ""):
		return None
	row["visit_cd"] = visit_cd
	row["patient_file_no"] = _clean_oracle_num(patient_file_no)
	row["prescrip_id"] = _clean_oracle_num(row.get("prescrip_id")) or None
	row["medicine_cd"] = _clean_oracle_num(row.get("medicine_cd")) or None
	row["create_user_id"] = _clean_oracle_num(row.get("create_user_id")) or None
	row["update_user_id"] = _clean_oracle_num(row.get("update_user_id")) or None
	row["cr_id"] = _clean_oracle_num(row.get("cr_id")) or None
	row["up_id"] = _clean_oracle_num(row.get("up_id")) or None
	if row.get("hist_ord") not in (None, ""):
		row["hist_ord"] = cint(row.get("hist_ord"))
	return row


def _parse_sheet_rows(ws) -> list[dict]:
	rows_iter = ws.iter_rows(values_only=True)
	try:
		header_row = next(rows_iter)
	except StopIteration:
		return []

	headers = [_normalize_header(h) for h in header_row]
	parsed: list[dict] = []
	for raw in rows_iter:
		if not raw or all(cell is None or str(cell).strip() == "" for cell in raw):
			continue
		row = {headers[idx]: raw[idx] for idx in range(min(len(headers), len(raw))) if headers[idx]}
		prepared = _prepare_line_row(row)
		if prepared:
			parsed.append(prepared)
	return parsed


def _group_rows_by_visit(rows: list[dict]) -> dict[str, dict]:
	grouped: dict[str, dict] = {}
	for row in rows:
		visit_cd = row["visit_cd"]
		bucket = grouped.setdefault(
			visit_cd,
			{
				"visit_cd": visit_cd,
				"patient_visit": row.get("patient_visit"),
				"patient_file_no": row.get("patient_file_no"),
				"branch_num": row.get("branch_num"),
				"lines": [],
			},
		)
		if not bucket.get("patient_visit") and row.get("patient_visit"):
			bucket["patient_visit"] = row.get("patient_visit")
		if not bucket.get("patient_file_no") and row.get("patient_file_no"):
			bucket["patient_file_no"] = row.get("patient_file_no")
		if bucket.get("branch_num") in (None, "") and row.get("branch_num") not in (None, ""):
			bucket["branch_num"] = row.get("branch_num")
		bucket["lines"].append(row)
	return grouped


def _parse_excel_rows(file_url: str) -> tuple[dict[str, dict], dict[str, int], int]:
	try:
		import openpyxl
	except ImportError:
		frappe.throw(
			_(
				"openpyxl is required to read Excel files. Install it in the bench environment: pip install openpyxl"
			)
		)

	path = _excel_file_path(file_url)
	wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
	all_rows: list[dict] = []
	sheet_row_counts: dict[str, int] = {}
	try:
		for sheet_name in wb.sheetnames:
			sheet_rows = _parse_sheet_rows(wb[sheet_name])
			sheet_row_counts[sheet_name] = len(sheet_rows)
			all_rows.extend(sheet_rows)
	finally:
		wb.close()
	grouped = _group_rows_by_visit(all_rows)
	return grouped, sheet_row_counts, len(all_rows)


def _pmo_line_from_row(row: dict) -> dict[str, Any]:
	line = {
		"prescrip_id": row.get("prescrip_id"),
		"medicine_num": row.get("medicine_cd"),
		"frequency": row.get("frequency"),
		"start_date": row.get("start_date"),
		"end_date": row.get("end_date"),
		"route": row.get("route"),
		"strength": row.get("strength"),
		"unit": row.get("unit"),
		"qty": row.get("qty"),
		"quantity": row.get("quantity"),
		"duration": row.get("duration"),
		"duration_type": row.get("duration_type"),
		"instructions": row.get("note"),
		"username": row.get("create_user_id"),
		"cr_id": row.get("cr_id"),
		"cr_date": row.get("cr_date"),
		"up_id": row.get("up_id"),
		"up_date": row.get("update_date") or row.get("up_date"),
		"trans_date": row.get("create_date"),
	}
	_ensure_prescription_frequency_label(row.get("frequency"))
	return line


def import_patient_medication_order_for_visit(visit_payload: dict) -> dict:
	visit_cd = visit_payload.get("visit_cd")
	lines = visit_payload.get("lines") or []
	if not visit_cd or not lines:
		return {"status": "skip_empty", "visit_cd": visit_cd}

	first = lines[0]
	patient = _resolve_patient(visit_payload.get("patient_file_no"))
	patient_visit = visit_payload.get("patient_visit") or _resolve_patient_visit_link(
		visit_cd, visit_payload.get("patient_file_no")
	)

	visit_data = None
	if patient_visit:
		visit_data = frappe.db.get_value(
			"Patient Visit",
			patient_visit,
			["patient", "patient_name", "company", "practitioner", "cost_center", "encounter_date"],
			as_dict=True,
		)
		if visit_data and visit_data.patient:
			patient = visit_data.patient

	if not patient:
		return {"status": "skip_no_patient", "visit_cd": visit_cd}

	existing_name = existing_pmo_for_legacy_visit(
		visit_cd,
		patient=patient,
		patient_visit=patient_visit,
		source_flag="patient_visit_prescription",
	)
	created = False
	if existing_name:
		doc = frappe.get_doc(DOCTYPE, existing_name)
	else:
		doc = frappe.new_doc(DOCTYPE)
		doc.trans_no = get_next_transaction_number("Patient Medication Order", fieldname="trans_no")
		created = True

	doc.care_context = "Patient Visit"
	doc.patient_visit_prescription = 1
	doc.patient_visit_prescription_his = 0
	doc.legacy = 1
	doc.visit_cd = visit_cd
	if patient_visit:
		doc.patient_encounter = patient_visit
	doc.patient = patient
	doc.patient_name = (visit_data and visit_data.patient_name) or frappe.db.get_value(
		"Patient", patient, "patient_name"
	)
	doc.company = (visit_data and visit_data.company) or _default_company()
	doc.practitioner = visit_data.practitioner if visit_data else None

	patient_fields = _patient_display_fields(patient)
	if patient_fields.get("nationality") is not None:
		doc.nationality = patient_fields["nationality"]
	if patient_fields.get("patient_age") is not None:
		doc.patient_age = patient_fields["patient_age"]

	doc.posting_date = _parse_date_value(first.get("start_date")) or _parse_date_value(
		first.get("create_date")
	) or (visit_data and visit_data.encounter_date) or nowdate()
	doc.start_date = _parse_date_value(first.get("start_date")) or doc.posting_date
	doc.end_date = _parse_date_value(first.get("end_date"))

	branch = _resolve_cost_center(visit_payload.get("branch_num"))
	if not branch and first.get("branch_num") not in (None, ""):
		branch = _resolve_cost_center(first.get("branch_num"))
	doc.cost_center = branch or (visit_data and visit_data.cost_center)

	doc.set("medication_orders", [])
	for row in lines:
		_append_child_line(doc, _pmo_line_from_row(row))

	if not doc.get("medication_orders"):
		return {"status": "skip_no_lines", "visit_cd": visit_cd}

	doc.flags.ignore_mandatory = True
	doc.flags.ignore_links = True
	doc.flags.ignore_validate = True
	submit_and_complete_legacy_visit_pmo(doc)

	return {
		"status": "created" if created else "updated",
		"visit_cd": visit_cd,
		"name": doc.name,
		"lines": len(doc.get("medication_orders") or []),
		"patient": patient,
		"patient_visit": patient_visit,
	}


def _preview_counts(grouped: dict[str, dict], raw_line_count: int) -> dict:
	return preview_counts_for_legacy_visit_pmo(
		grouped,
		raw_line_count,
		source_flag="patient_visit_prescription",
	)


def parse_and_cache_excel(file_url: str) -> dict:
	grouped, sheet_row_counts, raw_line_count = _parse_excel_rows(file_url)
	visit_keys = sorted(grouped.keys(), key=lambda key: cint(key) if str(key).isdigit() else key)

	frappe.cache().set_value(CACHE_KEYS["file_url"], file_url, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(CACHE_KEYS["visit_keys"], visit_keys, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(
		CACHE_KEYS["visits"],
		json.dumps(grouped, default=str),
		expires_in_sec=CACHE_TTL,
	)

	preview = _preview_counts(grouped, raw_line_count)
	raw_row_total = sum(sheet_row_counts.values())
	return {
		"excel_rows": raw_line_count,
		"raw_excel_rows": raw_row_total,
		"sheets": list(sheet_row_counts.keys()),
		"sheet_row_counts": sheet_row_counts,
		**preview,
	}


def _load_cached_visits() -> dict[str, dict]:
	raw = frappe.cache().get_value(CACHE_KEYS["visits"])
	if not raw:
		return {}
	if isinstance(raw, dict):
		return raw
	return json.loads(raw)


@frappe.whitelist()
def preview_patient_visit_prescription_import(file_url: str) -> dict:
	_require_admin()
	if not (file_url or "").strip():
		frappe.throw(_("Please upload the PATIENT_VISIT_PRESCRIPTION Excel file."))
	return parse_and_cache_excel(file_url)


def run_patient_visit_prescription_import_batch(*, offset: int = 0) -> dict:
	visit_keys = frappe.cache().get_value(CACHE_KEYS["visit_keys"]) or []
	visits_by_key = _load_cached_visits()
	if not visit_keys or not visits_by_key:
		return {"processed": offset, "done": True, "batch_count": 0}

	batch_keys = visit_keys[offset : offset + VISIT_PRESCRIPTION_IMPORT_BATCH_SIZE]
	if not batch_keys:
		return {"processed": offset, "done": True, "batch_count": 0}

	created = updated = skipped = lines_imported = 0
	errors: list[str] = []

	for key in batch_keys:
		payload = visits_by_key.get(key) or {}
		try:
			result = import_patient_medication_order_for_visit(payload)
			status = result.get("status")
			if status == "created":
				created += 1
				lines_imported += cint(result.get("lines", 0))
			elif status == "updated":
				updated += 1
				lines_imported += cint(result.get("lines", 0))
			else:
				skipped += 1
		except Exception:
			errors.append(f"{key}: {frappe.get_traceback()}")
			frappe.log_error(title=f"PATIENT_VISIT_PRESCRIPTION PMO import failed: {key}")

	frappe.db.commit()
	processed = offset + len(batch_keys)
	return {
		"processed": processed,
		"done": len(batch_keys) < VISIT_PRESCRIPTION_IMPORT_BATCH_SIZE,
		"batch_count": len(batch_keys),
		"created": created,
		"updated": updated,
		"skipped": skipped,
		"lines_imported": lines_imported,
		"errors": len(errors),
	}
