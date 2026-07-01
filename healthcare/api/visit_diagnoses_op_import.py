"""Import Oracle VISIT_DIAGNOSES_01 Excel into Medical Diagnosis Entry rows (OP/IP)."""

from __future__ import annotations

import json
from datetime import date, datetime
from typing import Any

import frappe
from frappe import _
from frappe.utils import cint, get_datetime, getdate

from healthcare.api.discharge_checklist_import import _resolve_cost_center
from healthcare.api.lab_test_legacy_import import _format_legacy_date_str
from healthcare.api.patient_info_import import (
	_cell_text,
	_clean_oracle_num,
	_excel_file_path,
	_require_admin,
)
from healthcare.api.patient_visit_import import ensure_patient_for_legacy_import
from healthcare.api.visit_diagnosis_sync import (
	_resolve_inpatient_admission,
	_resolve_patient_visit,
)

VISIT_DIAGNOSES_OP_IMPORT_BATCH_SIZE = 100
DATA_FIELD_MAX_LEN = 140
CACHE_TTL = 7200
CACHE_KEYS = {
	"file_url": "healthcare:data_migration:visit_diagnoses_op_import:file_url",
	"row_keys": "healthcare:data_migration:visit_diagnoses_op_import:row_keys",
	"rows": "healthcare:data_migration:visit_diagnoses_op_import:rows",
}

EXCEL_HEADER_MAP = {
	"PATIENT": "patient",
	"SR_NO": "sr_no",
	"SR NUM": "sr_no",
	"DIAGNOSIS": "diagnosis",
	"DETAILS": "details",
	"CR_ID": "cr_id",
	"CD_DATE": "cd_date",
	"CR_DATE": "cd_date",
	"UP_ID": "up_id",
	"UP_DATE": "up_date",
	"PATIENT_VISIT": "patient_visit",
	"VISIT_NUM": "patient_visit",
	"COST_CENTER": "cost_center",
	"BRANCH_NUM": "cost_center",
	"INPATIENT_ADMISSION": "inpatient_admission",
	"GROUP_CODE": "group_code",
	"SOURCE": "source",
}


def _normalize_header(cell: Any) -> str:
	if cell is None:
		return ""
	text = str(cell).strip().upper().replace(" ", "_")
	return EXCEL_HEADER_MAP.get(text, text.lower())


def _truncate_data(value: Any, max_len: int = DATA_FIELD_MAX_LEN) -> str:
	text = _cell_text(value)
	if len(text) <= max_len:
		return text
	return text[:max_len]


def _legacy_data_datetime(value: Any) -> str:
	if value is None or value == "":
		return ""
	if isinstance(value, datetime):
		return value.strftime("%Y-%m-%d %H:%M:%S")
	if isinstance(value, date):
		return datetime.combine(value, datetime.min.time()).strftime("%Y-%m-%d %H:%M:%S")
	return _format_legacy_date_str(value)


def _parse_posting_datetime(value: Any):
	if value is None or value == "":
		return None
	if isinstance(value, datetime):
		return value
	if isinstance(value, date):
		return datetime.combine(value, datetime.min.time())
	try:
		return get_datetime(value)
	except Exception:
		return None


def _resolve_diagnosis(code: Any) -> str | None:
	text = _cell_text(code)
	if not text:
		return None
	if frappe.db.exists("Diagnosis", text):
		return text
	for field in ("disease_no", "old_no", "name"):
		name = frappe.db.get_value("Diagnosis", {field: text}, "name")
		if name:
			return name
	return None


def _resolve_diagnosis_group(code: Any) -> str | None:
	text = _cell_text(code)
	if not text:
		return None
	if frappe.db.exists("Diagnosis Group", text):
		return text
	name = frappe.db.get_value("Diagnosis Group", {"disease_no": text}, "name")
	if name:
		return name
	return frappe.db.get_value("Diagnosis Group", {"old_no": text}, "name")


def _apply_patient_fetch_fields(fields: dict[str, Any], patient: str | None) -> None:
	if not patient or not frappe.db.exists("Patient", patient):
		return
	patient_row = frappe.db.get_value(
		"Patient",
		patient,
		["patient_name"],
		as_dict=True,
	)
	if patient_row and patient_row.patient_name:
		fields["patient_name"] = _truncate_data(patient_row.patient_name)


def _apply_diagnosis_fetch_fields(fields: dict[str, Any], diagnosis: str | None) -> None:
	if not diagnosis or not frappe.db.exists("Diagnosis", diagnosis):
		return
	diagnosis_name = frappe.db.get_value("Diagnosis", diagnosis, "diagnosis")
	if diagnosis_name:
		fields["diagnosis_name"] = _truncate_data(diagnosis_name)


def _legacy_trans_num(row: dict) -> str:
	patient = row.get("patient") or "0"
	sr_no = row.get("sr_no") or "0"
	visit = row.get("patient_visit_raw") or "0"
	admission = row.get("inpatient_admission_raw") or "0"
	diagnosis = row.get("diagnosis_raw") or "-"
	return f"VD01/{patient}/{sr_no}/{visit}/{admission}/{diagnosis}"


def _row_key(trans_num: str) -> str:
	return trans_num


def _parse_sheet_rows(ws) -> list[dict]:
	rows_iter = ws.iter_rows(values_only=True)
	try:
		header_row = next(rows_iter)
	except StopIteration:
		return []

	headers = [_normalize_header(h) for h in header_row]
	parsed: list[dict] = []
	for raw in rows_iter:
		if not raw or all(cell is None or str(cell).strip() == "" for cell in raw):
			continue
		row: dict[str, Any] = {}
		for idx, key in enumerate(headers):
			if not key or idx >= len(raw):
				continue
			row[key] = raw[idx]

		patient = _clean_oracle_num(row.get("patient"))
		if not patient:
			continue

		row["patient"] = patient
		row["sr_no"] = _clean_oracle_num(row.get("sr_no"))
		row["diagnosis_raw"] = _cell_text(row.get("diagnosis"))
		row["patient_visit_raw"] = _clean_oracle_num(row.get("patient_visit"))
		row["inpatient_admission_raw"] = _clean_oracle_num(row.get("inpatient_admission"))
		row["details"] = _cell_text(row.get("details"))
		row["cr_id"] = _clean_oracle_num(row.get("cr_id"))
		row["up_id"] = _clean_oracle_num(row.get("up_id"))
		row["cd_date"] = row.get("cd_date")
		row["up_date"] = row.get("up_date")
		row["group_code_raw"] = _cell_text(row.get("group_code"))
		row["source"] = _cell_text(row.get("source")).upper()
		row["cost_center_raw"] = row.get("cost_center")
		row["trans_num"] = _legacy_trans_num(row)
		parsed.append(row)
	return parsed


def _parse_excel_rows(file_url: str) -> tuple[list[dict], dict[str, int]]:
	try:
		import openpyxl
	except ImportError:
		frappe.throw(
			_("openpyxl is required to read Excel files. Install it in the bench environment: pip install openpyxl")
		)

	path = _excel_file_path(file_url)
	wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
	by_key: dict[str, dict] = {}
	sheet_row_counts: dict[str, int] = {}
	try:
		for sheet_name in wb.sheetnames:
			sheet_rows = _parse_sheet_rows(wb[sheet_name])
			sheet_row_counts[sheet_name] = len(sheet_rows)
			for row in sheet_rows:
				by_key[_row_key(row["trans_num"])] = row
	finally:
		wb.close()
	return list(by_key.values()), sheet_row_counts


def _build_entry_fields(row: dict) -> tuple[dict, dict]:
	stats = {"patient_created": 0}
	patient = row["patient"]
	if not frappe.db.exists("Patient", patient):
		result = ensure_patient_for_legacy_import(patient)
		if result.get("status") == "created":
			stats["patient_created"] = 1
		elif result.get("status") != "existing":
			return {}, stats

	diagnosis = _resolve_diagnosis(row.get("diagnosis_raw"))
	if not diagnosis:
		return {}, stats

	visit_num = None
	if row.get("patient_visit_raw"):
		visit_num = _resolve_patient_visit(row["patient_visit_raw"], patient)

	admission = None
	if row.get("inpatient_admission_raw"):
		admission = _resolve_inpatient_admission(row["inpatient_admission_raw"], patient)

	cost_center = _resolve_cost_center(row.get("cost_center_raw"))
	group_code = _resolve_diagnosis_group(row.get("group_code_raw"))

	posting_dt = _parse_posting_datetime(row.get("cd_date"))

	fields: dict[str, Any] = {
		"trans_num": row["trans_num"],
		"patient": patient,
		"sr_no": row.get("sr_no") or None,
		"diagnosis": diagnosis,
		"details": row.get("details") or None,
		"source": row.get("source") or None,
		"visit_num": visit_num,
		"inpatient_admission": admission,
		"cost_center": cost_center,
		"group_code": group_code,
		"cr_id": row.get("cr_id") or None,
		"up_id": row.get("up_id") or None,
	}
	cd_date = _legacy_data_datetime(row.get("cd_date"))
	if cd_date:
		fields["cd_date"] = cd_date
	up_date = _legacy_data_datetime(row.get("up_date"))
	if up_date:
		fields["up_date"] = up_date
	if posting_dt:
		fields["posting_date"] = posting_dt
		fields["diagnoses_time"] = posting_dt

	_apply_patient_fetch_fields(fields, patient)
	_apply_diagnosis_fetch_fields(fields, diagnosis)

	return {key: value for key, value in fields.items() if value not in (None, "")}, stats


def _find_existing_entry(trans_num: str) -> str | None:
	rows = frappe.get_all(
		"Medical Diagnosis Entry",
		filters={"trans_num": trans_num},
		pluck="name",
		limit=1,
	)
	return rows[0] if rows else None


def upsert_medical_diagnosis_entry_from_row(row: dict) -> dict:
	trans_num = row.get("trans_num")
	if not trans_num:
		return {"status": "skip_no_key"}

	if not row.get("diagnosis_raw"):
		return {"status": "skip_no_diagnosis", "trans_num": trans_num}

	fields, side_stats = _build_entry_fields(row)
	if not fields.get("diagnosis"):
		return {"status": "skip_unresolved_diagnosis", "trans_num": trans_num, **side_stats}

	existing = _find_existing_entry(trans_num)
	if existing:
		doc = frappe.get_doc("Medical Diagnosis Entry", existing)
		for key, value in fields.items():
			if key == "trans_num":
				continue
			doc.set(key, value)
		action = "updated"
	else:
		doc = frappe.new_doc("Medical Diagnosis Entry")
		doc.update(fields)
		action = "created"

	doc.flags.ignore_validate = True
	doc.flags.ignore_mandatory = True
	doc.flags.ignore_links = True
	doc.flags.legacy_import = True

	if existing:
		doc.save(ignore_permissions=True)
	else:
		doc.insert(ignore_permissions=True)

	return {
		"status": action,
		"trans_num": trans_num,
		"name": doc.name,
		**side_stats,
	}


def _preview_counts(rows: list[dict]) -> dict:
	patient_nums = {row["patient"] for row in rows}
	missing_patients = [pn for pn in patient_nums if not frappe.db.exists("Patient", pn)]
	existing = 0
	matched_diagnosis = 0
	unresolved_diagnosis: list[str] = []
	resolved_visits = 0
	resolved_admissions = 0
	skip_no_diagnosis = 0
	sources: dict[str, int] = {}

	for row in rows:
		src = (row.get("source") or "").strip().upper() or "(empty)"
		sources[src] = sources.get(src, 0) + 1
		if not row.get("diagnosis_raw"):
			skip_no_diagnosis += 1
			continue
		if _resolve_diagnosis(row.get("diagnosis_raw")):
			matched_diagnosis += 1
		else:
			code = row.get("diagnosis_raw")
			if code and code not in unresolved_diagnosis:
				unresolved_diagnosis.append(code)
		if row.get("patient_visit_raw") and _resolve_patient_visit(row["patient_visit_raw"], row["patient"]):
			resolved_visits += 1
		if row.get("inpatient_admission_raw") and _resolve_inpatient_admission(
			row["inpatient_admission_raw"], row["patient"]
		):
			resolved_admissions += 1
		if _find_existing_entry(row["trans_num"]):
			existing += 1

	return {
		"unique_patients": len(patient_nums),
		"patients_to_create": len(missing_patients),
		"sample_patients_to_create": missing_patients[:10],
		"existing_entries": existing,
		"new_entries": len(rows) - existing,
		"skip_no_diagnosis": skip_no_diagnosis,
		"matched_diagnosis": matched_diagnosis,
		"unresolved_diagnosis_count": len(rows) - skip_no_diagnosis - matched_diagnosis,
		"unresolved_diagnosis_codes": unresolved_diagnosis[:15],
		"resolved_visits": resolved_visits,
		"resolved_admissions": resolved_admissions,
		"with_visit_num": sum(1 for row in rows if row.get("patient_visit_raw")),
		"with_admission": sum(1 for row in rows if row.get("inpatient_admission_raw")),
		"sources": sources,
	}


def parse_and_cache_excel(file_url: str) -> dict:
	rows, sheet_row_counts = _parse_excel_rows(file_url)
	row_keys = [_row_key(row["trans_num"]) for row in rows]
	by_key = {_row_key(row["trans_num"]): row for row in rows}

	frappe.cache().set_value(CACHE_KEYS["file_url"], file_url, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(CACHE_KEYS["row_keys"], row_keys, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(
		CACHE_KEYS["rows"],
		json.dumps(by_key, default=str),
		expires_in_sec=CACHE_TTL,
	)

	preview = _preview_counts(rows)
	raw_row_total = sum(sheet_row_counts.values())
	return {
		"excel_rows": len(rows),
		"raw_excel_rows": raw_row_total,
		"sheets": list(sheet_row_counts.keys()),
		"sheet_row_counts": sheet_row_counts,
		**preview,
		"sample_trans_nums": row_keys[:5],
	}


def _load_cached_rows() -> dict[str, dict]:
	raw = frappe.cache().get_value(CACHE_KEYS["rows"])
	if not raw:
		return {}
	if isinstance(raw, dict):
		return raw
	return json.loads(raw)


@frappe.whitelist()
def preview_visit_diagnoses_op_import(file_url: str) -> dict:
	_require_admin()
	if not (file_url or "").strip():
		frappe.throw(_("Please upload the VISIT_DIAGNOSES_01 Excel file."))
	return parse_and_cache_excel(file_url)


def run_visit_diagnoses_op_import_batch(*, offset: int = 0) -> dict:
	row_keys = frappe.cache().get_value(CACHE_KEYS["row_keys"]) or []
	rows_by_key = _load_cached_rows()
	if not row_keys or not rows_by_key:
		return {"processed": offset, "done": True, "batch_count": 0}

	batch_keys = row_keys[offset : offset + VISIT_DIAGNOSES_OP_IMPORT_BATCH_SIZE]
	if not batch_keys:
		return {"processed": offset, "done": True, "batch_count": 0}

	created = updated = skip_no_diagnosis = skip_unresolved = 0
	patients_created = 0
	errors: list[str] = []

	for key in batch_keys:
		row = rows_by_key.get(key) or {}
		try:
			result = upsert_medical_diagnosis_entry_from_row(row)
			status = result.get("status")
			if status == "created":
				created += 1
			elif status == "updated":
				updated += 1
			elif status == "skip_no_diagnosis":
				skip_no_diagnosis += 1
			elif status == "skip_unresolved_diagnosis":
				skip_unresolved += 1
			patients_created += cint(result.get("patient_created"))
		except Exception:
			errors.append(f"{key}: {frappe.get_traceback()}")
			frappe.log_error(title=f"VISIT_DIAGNOSES_01 import failed: {key}")

	frappe.db.commit()
	processed = offset + len(batch_keys)
	return {
		"processed": processed,
		"done": len(batch_keys) < VISIT_DIAGNOSES_OP_IMPORT_BATCH_SIZE,
		"batch_count": len(batch_keys),
		"created": created,
		"updated": updated,
		"skip_no_diagnosis": skip_no_diagnosis,
		"skip_unresolved_diagnosis": skip_unresolved,
		"patients_created": patients_created,
		"errors": len(errors),
	}
