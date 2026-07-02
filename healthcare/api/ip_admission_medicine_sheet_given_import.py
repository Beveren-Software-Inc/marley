"""Direct-upload IP_ADMISSION_MEDICINE_SHEET Excel into Admission Detail given medicine rows."""

from __future__ import annotations

import json
from datetime import date, datetime
from typing import Any

import frappe
from frappe import _
from frappe.utils import cint, getdate, nowdate

from healthcare.api.discharge_checklist_import import _resolve_cost_center
from healthcare.api.patient_info_import import (
	_clean_oracle_num,
	_excel_file_path,
	_require_admin,
)

IP_ADMISSION_MEDICINE_SHEET_GIVEN_IMPORT_BATCH_SIZE = 500
CACHE_TTL = 7200
CACHE_KEYS = {
	"file_url": "healthcare:data_migration:ip_admission_medicine_sheet_given:file_url",
	"row_keys": "healthcare:data_migration:ip_admission_medicine_sheet_given:row_keys",
	"rows": "healthcare:data_migration:ip_admission_medicine_sheet_given:rows",
}

EXCEL_HEADER_MAP = {
	"TRANS_NUM": "trans_num",
	"MEDI_TRANS_NUM": "medi_trans_num",
	"ADMISSION_NUM": "admission_num",
	"PATIENT_NUM": "patient_num",
	"GIVEN_YN": "given_yn",
	"GIVEN_DATE": "given_date",
	"REMARKS": "remarks",
	"CR_ID": "cr_id",
	"CR_DATE": "cr_date",
	"UP_ID": "up_id",
	"UP_DATE": "up_date",
	"BRANCH_NUM": "branch_num",
	"ADMISSION_NUM_OLD": "admission_num_old",
}


def _normalize_header(cell: Any) -> str:
	if cell is None:
		return ""
	text = str(cell).strip().upper().replace(" ", "_")
	return EXCEL_HEADER_MAP.get(text, text.lower())


def _parse_datetime(value: Any) -> datetime | None:
	if value in (None, ""):
		return None
	if isinstance(value, datetime):
		return value
	if isinstance(value, date):
		return datetime.combine(value, datetime.min.time())
	if isinstance(value, (int, float)) and not isinstance(value, bool):
		try:
			import openpyxl.utils.datetime

			return openpyxl.utils.datetime.from_excel(float(value))
		except Exception:
			return None
	text = str(value).strip()
	if not text:
		return None
	try:
		from frappe.utils import get_datetime

		return get_datetime(text)
	except Exception:
		return None


def _parse_sheet_rows(ws) -> list[dict]:
	rows_iter = ws.iter_rows(values_only=True)
	try:
		header_row = next(rows_iter)
	except StopIteration:
		return []

	headers = [_normalize_header(h) for h in header_row]
	parsed: list[dict] = []
	for raw in rows_iter:
		if not raw or all(cell is None or str(cell).strip() == "" for cell in raw):
			continue
		row: dict[str, Any] = {}
		for idx, key in enumerate(headers):
			if not key or idx >= len(raw):
				continue
			row[key] = raw[idx]

		trans_num = _clean_oracle_num(row.get("trans_num"))
		if not trans_num:
			continue
		row["trans_num"] = trans_num
		row["medi_trans_num"] = _clean_oracle_num(row.get("medi_trans_num"))
		row["admission_num"] = _clean_oracle_num(row.get("admission_num"))
		row["patient_num"] = _clean_oracle_num(row.get("patient_num"))
		row["admission_num_old"] = _clean_oracle_num(row.get("admission_num_old"))
		row["cr_id"] = _clean_oracle_num(row.get("cr_id"))
		row["up_id"] = _clean_oracle_num(row.get("up_id"))
		row["given_yn"] = (row.get("given_yn") or "").strip().upper()
		row["remarks"] = (row.get("remarks") or "").strip()
		parsed.append(row)
	return parsed


def _parse_excel_rows(file_url: str) -> tuple[list[dict], dict[str, int]]:
	try:
		import openpyxl
	except ImportError:
		frappe.throw(
			_(
				"openpyxl is required to read Excel files. Install it in the bench environment: pip install openpyxl"
			)
		)

	path = _excel_file_path(file_url)
	wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
	by_key: dict[str, dict] = {}
	sheet_row_counts: dict[str, int] = {}
	try:
		for sheet_name in wb.sheetnames:
			sheet_rows = _parse_sheet_rows(wb[sheet_name])
			sheet_row_counts[sheet_name] = len(sheet_rows)
			for row in sheet_rows:
				by_key[row["trans_num"]] = row
	finally:
		wb.close()
	return list(by_key.values()), sheet_row_counts


def _existing_sheet_doc_name(trans_num: str) -> str | None:
	return frappe.db.get_value("IP Admission Medicine Sheet", {"trans_num": trans_num}, "name")


def _preview_counts(rows: list[dict]) -> dict[str, Any]:
	given_rows = 0
	not_given_rows = 0
	existing_rows = 0
	admissions = set()
	for row in rows:
		if row.get("admission_num"):
			admissions.add(row.get("admission_num"))
		if (row.get("given_yn") or "").strip().upper() == "Y":
			given_rows += 1
		else:
			not_given_rows += 1
		if _existing_sheet_doc_name(row.get("trans_num") or ""):
			existing_rows += 1

	return {
		"excel_rows": len(rows),
		"given_rows": given_rows,
		"not_given_rows": not_given_rows,
		"admissions": len(admissions),
		"existing_rows": existing_rows,
		"new_rows": len(rows) - existing_rows,
		"sample_trans_nums": [row.get("trans_num") for row in rows[:5]],
	}


def parse_and_cache_excel(file_url: str) -> dict:
	rows, sheet_row_counts = _parse_excel_rows(file_url)
	row_keys = [row["trans_num"] for row in rows]
	by_key = {row["trans_num"]: row for row in rows}

	frappe.cache().set_value(CACHE_KEYS["file_url"], file_url, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(CACHE_KEYS["row_keys"], row_keys, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(
		CACHE_KEYS["rows"],
		json.dumps(by_key, default=str),
		expires_in_sec=CACHE_TTL,
	)

	preview = _preview_counts(rows)
	raw_row_total = sum(sheet_row_counts.values())
	return {
		"raw_excel_rows": raw_row_total,
		"sheets": list(sheet_row_counts.keys()),
		"sheet_row_counts": sheet_row_counts,
		**preview,
	}


def _load_cached_rows() -> dict[str, dict]:
	raw = frappe.cache().get_value(CACHE_KEYS["rows"])
	if not raw:
		return {}
	if isinstance(raw, dict):
		return raw
	return json.loads(raw)


def _upsert_sheet_doc(row: dict) -> tuple[str, str]:
	trans_num = row["trans_num"]
	existing_name = _existing_sheet_doc_name(trans_num)
	values = {
		"trans_num": cint(trans_num),
		"medi_trans_num": cint(row.get("medi_trans_num") or 0) or None,
		"admission_num": row.get("admission_num") or None,
		"patient_num": row.get("patient_num") or None,
		"given_yn": row.get("given_yn") or None,
		"given_date": _parse_datetime(row.get("given_date")),
		"remarks": row.get("remarks") or None,
		"cr_id": cint(row.get("cr_id") or 0) or None,
		"cr_date": row.get("cr_date") or None,
		"up_id": cint(row.get("up_id") or 0) or None,
		"up_date": row.get("up_date") or None,
		"cost_center": _resolve_cost_center(row.get("branch_num")) or None,
		"admission_num_old": cint(row.get("admission_num_old") or 0) or None,
	}

	if existing_name:
		doc = frappe.get_doc("IP Admission Medicine Sheet", existing_name)
		for key, value in values.items():
			doc.set(key, value)
		action = "updated"
	else:
		doc = frappe.get_doc({"doctype": "IP Admission Medicine Sheet", **values})
		action = "created"

	doc.flags.ignore_validate = True
	doc.flags.ignore_mandatory = True
	doc.flags.ignore_links = True
	if existing_name:
		doc.save(ignore_permissions=True)
	else:
		doc.insert(ignore_permissions=True)
	return doc.name, action


def _row_already_mapped(ad_detail, sheet_doc_name: str) -> bool:
	return any(
		(r.ip_admission_medicine_sheet or "") == sheet_doc_name for r in (ad_detail.get("table_yrwe") or [])
	)


@frappe.whitelist()
def preview_ip_admission_medicine_sheet_given_import(file_url: str) -> dict:
	_require_admin()
	if not (file_url or "").strip():
		frappe.throw(_("Please upload the IP_ADMISSION_MEDICINE_SHEET Excel file."))
	return parse_and_cache_excel(file_url)


def run_ip_admission_medicine_sheet_given_import_batch(*, offset: int = 0) -> dict:
	from healthcare.api.data_migration_jobs import (
		_get_admission_detail_for_sheet_row,
		_normalize_time_value,
		_resolve_item_00_01_name,
		_resolve_item_code_for_given_tables,
	)

	row_keys = frappe.cache().get_value(CACHE_KEYS["row_keys"]) or []
	rows_by_key = _load_cached_rows()
	if not row_keys or not rows_by_key:
		return {"processed": offset, "done": True, "batch_count": 0}

	batch_keys = row_keys[offset : offset + IP_ADMISSION_MEDICINE_SHEET_GIVEN_IMPORT_BATCH_SIZE]
	if not batch_keys:
		return {"processed": offset, "done": True, "batch_count": 0}

	staging_created = staging_updated = created_given = 0
	skip_not_given = skip_no_admission_detail = skip_already_mapped = errors = 0
	created_admission_detail = 0
	admission_cache: dict[str, object] = {}

	for key in batch_keys:
		row = rows_by_key.get(key) or {}
		try:
			sheet_doc_name, staging_action = _upsert_sheet_doc(row)
			if staging_action == "created":
				staging_created += 1
			else:
				staging_updated += 1

			if (row.get("given_yn") or "").strip().upper() != "Y":
				skip_not_given += 1
				continue

			admission = row.get("admission_num") or ""
			patient_num = row.get("patient_num") or ""
			cache_key = f"{admission}|{patient_num}"
			ad_detail = admission_cache.get(cache_key)
			if ad_detail is None:
				existed_before = bool(admission and frappe.db.exists("Admission Detail", {"admission": admission}))
				ad_detail = _get_admission_detail_for_sheet_row(admission, patient_num)
				if not ad_detail:
					skip_no_admission_detail += 1
					continue
				if not existed_before and admission and (ad_detail.get("admission") == admission):
					created_admission_detail += 1
				admission_cache[cache_key] = ad_detail

			if _row_already_mapped(ad_detail, sheet_doc_name):
				skip_already_mapped += 1
				continue

			old_code = _resolve_item_00_01_name(row.get("medi_trans_num"))
			old_name = frappe.db.get_value("ITEM_00_01", old_code, "item_nam") if old_code else None
			item_code = _resolve_item_code_for_given_tables(row.get("medi_trans_num"))
			item_name = (
				frappe.db.get_value("Item", item_code, "item_name")
				if item_code
				else (old_name or (row.get("medi_trans_num") or ""))
			)

			given_dt = _parse_datetime(row.get("given_date"))
			ad_detail.append(
				"table_yrwe",
				{
					"date": getdate(given_dt) if given_dt else nowdate(),
					"time": _normalize_time_value(given_dt if given_dt else "00:00:00"),
					"medicine_code": item_code,
					"medicine_name": item_name,
					"qty": 0,
					"dose_notes": row.get("remarks") or None,
					"user": frappe.session.user,
					"old_medicine_code": old_code,
					"old_medicine_name": old_name or item_name,
					"ip_admission_medicine_sheet": sheet_doc_name,
				},
			)
			created_given += 1
		except Exception:
			errors += 1
			frappe.log_error(
				title=f"IP_ADMISSION_MEDICINE_SHEET given import failed: {key}",
				message=frappe.get_traceback(),
			)

	for ad_doc in admission_cache.values():
		ad_doc.flags.ignore_mandatory = True
		ad_doc.save(ignore_permissions=True)

	frappe.db.commit()
	processed = offset + len(batch_keys)
	return {
		"processed": processed,
		"done": len(batch_keys) < IP_ADMISSION_MEDICINE_SHEET_GIVEN_IMPORT_BATCH_SIZE,
		"batch_count": len(batch_keys),
		"staging_created": staging_created,
		"staging_updated": staging_updated,
		"created_given": created_given,
		"created_admission_detail": created_admission_detail,
		"skip_not_given": skip_not_given,
		"skip_no_admission_detail": skip_no_admission_detail,
		"skip_already_mapped": skip_already_mapped,
		"errors": errors,
	}
