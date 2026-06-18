"""Backfill Patient Appointment doc_code, practitioner, and status from Oracle data."""

from __future__ import annotations

import json
from typing import Any

import frappe
from frappe import _
from frappe.utils import cint, getdate, nowdate

from healthcare.api.discharge_checklist_import import _clean_oracle_num, _excel_file_path

PATIENT_APPOINTMENT_OLD_STATUS_BATCH_SIZE = 2000
PATIENT_APPOINTMENT_DOC_CODE_BATCH_SIZE = 500
CACHE_TTL = 7200
CACHE_KEYS = {
	"file_url": "healthcare:data_migration:appointment_oracle_backfill:file_url",
	"excel_map": "healthcare:data_migration:appointment_oracle_backfill:excel_map",
	"pending_doc_code": "healthcare:data_migration:appointment_oracle_backfill:pending_doc_code",
}

EXCEL_HEADER_MAP = {
	"APP_NUM": "app_num",
	"DOC_CODE": "doc_code",
	"APP_STATUS": "app_status",
}


def _require_admin() -> None:
	frappe.only_for(("System Manager", "Healthcare Administrator"))


def _normalize_header(cell: Any) -> str:
	if cell is None:
		return ""
	text = str(cell).strip().upper().replace(" ", "_")
	return EXCEL_HEADER_MAP.get(text, text.lower())


def _parse_appointments_excel(file_url: str) -> dict[str, dict[str, str]]:
	"""Read all sheets and return ``app_num`` -> ``{doc_code, app_status}``."""
	try:
		import openpyxl
	except ImportError:
		frappe.throw(
			_(
				"openpyxl is required to read Excel files. Install it in the bench environment: pip install openpyxl"
			)
		)

	path = _excel_file_path(file_url)
	wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
	excel_map: dict[str, dict[str, str]] = {}

	for sheet_name in wb.sheetnames:
		ws = wb[sheet_name]
		rows_iter = ws.iter_rows(values_only=True)
		try:
			header_row = next(rows_iter)
		except StopIteration:
			continue

		headers = [_normalize_header(h) for h in header_row]
		for raw in rows_iter:
			if not raw or all(cell is None or str(cell).strip() == "" for cell in raw):
				continue
			row: dict[str, Any] = {}
			for idx, key in enumerate(headers):
				if not key or idx >= len(raw):
					continue
				row[key] = raw[idx]

			app_num = _clean_oracle_num(row.get("app_num"))
			if not app_num:
				continue

			doc_code = _clean_oracle_num(row.get("doc_code"))
			app_status = (row.get("app_status") or "").strip().upper()
			excel_map[app_num] = {
				"doc_code": doc_code,
				"app_status": app_status,
			}

	wb.close()
	return excel_map


def cache_appointments_excel_for_migration(file_url: str) -> dict[str, dict[str, str]]:
	"""Parse Excel and cache map + pending doc_code queue for background batches."""
	excel_map = _parse_appointments_excel(file_url)
	pending = build_pending_doc_code_queue(excel_map)
	cache = frappe.cache()
	cache.set_value(CACHE_KEYS["file_url"], file_url, expires_in_sec=CACHE_TTL)
	cache.set_value(CACHE_KEYS["excel_map"], json.dumps(excel_map), expires_in_sec=CACHE_TTL)
	cache.set_value(CACHE_KEYS["pending_doc_code"], json.dumps(pending), expires_in_sec=CACHE_TTL)
	return excel_map


def get_cached_excel_map() -> dict[str, dict[str, str]]:
	raw = frappe.cache().get_value(CACHE_KEYS["excel_map"])
	if not raw:
		return {}
	try:
		return json.loads(raw)
	except Exception:
		return {}


def get_cached_pending_doc_code_queue() -> list[str]:
	raw = frappe.cache().get_value(CACHE_KEYS["pending_doc_code"])
	if not raw:
		return []
	try:
		return json.loads(raw)
	except Exception:
		return []


def clear_appointments_excel_cache() -> None:
	cache = frappe.cache()
	for key in CACHE_KEYS.values():
		cache.delete_value(key)


def build_pending_doc_code_queue(excel_map: dict[str, dict[str, str]]) -> list[str]:
	"""Appointments in Excel that exist in ERP and need doc_code and/or practitioner."""
	if not excel_map:
		return []

	app_nums = list(excel_map.keys())
	existing: dict[str, dict[str, str | None]] = {}
	chunk_size = 500
	for start in range(0, len(app_nums), chunk_size):
		chunk = app_nums[start : start + chunk_size]
		for row in frappe.get_all(
			"Patient Appointment",
			filters={"name": ["in", chunk]},
			fields=["name", "doc_code", "practitioner", "old_status"],
		):
			existing[row.name] = row

	pending: list[str] = []
	for app_num, excel_row in excel_map.items():
		pa = existing.get(app_num)
		if not pa:
			continue

		doc_code = (excel_row.get("doc_code") or "").strip()
		if not doc_code:
			continue

		needs_doc_code = (pa.get("doc_code") or "").strip() != doc_code
		needs_practitioner = not (pa.get("practitioner") or "").strip()
		needs_old_status = not (pa.get("old_status") or "").strip() and bool(
			(excel_row.get("app_status") or "").strip()
		)
		if needs_doc_code or needs_practitioner or needs_old_status:
			pending.append(app_num)

	return sorted(pending)


def count_doc_code_backfill_preview(excel_map: dict[str, dict[str, str]]) -> dict:
	pending = build_pending_doc_code_queue(excel_map)
	matched = 0
	with_doc_code = 0
	for app_num, row in excel_map.items():
		if not frappe.db.exists("Patient Appointment", app_num):
			continue
		matched += 1
		if (row.get("doc_code") or "").strip():
			with_doc_code += 1

	practitioner_to_set = 0
	doc_code_to_set = 0
	for app_num in pending:
		excel_row = excel_map.get(app_num) or {}
		pa = frappe.db.get_value(
			"Patient Appointment",
			app_num,
			["doc_code", "practitioner"],
			as_dict=True,
		)
		if not pa:
			continue
		doc_code = (excel_row.get("doc_code") or "").strip()
		if doc_code and (pa.doc_code or "").strip() != doc_code:
			doc_code_to_set += 1
		if doc_code and not (pa.practitioner or "").strip():
			practitioner_to_set += 1

	sample: list[dict] = []
	for app_num in pending[:10]:
		excel_row = excel_map.get(app_num) or {}
		pa = frappe.db.get_value(
			"Patient Appointment",
			app_num,
			["doc_code", "practitioner", "old_status"],
			as_dict=True,
		)
		if not pa:
			continue
		doc_code = (excel_row.get("doc_code") or "").strip()
		sample.append(
			{
				"app_num": app_num,
				"current_doc_code": pa.doc_code,
				"target_doc_code": doc_code,
				"current_practitioner": pa.practitioner,
				"will_set_practitioner": bool(doc_code and not (pa.practitioner or "").strip()),
				"app_status": excel_row.get("app_status"),
			}
		)

	return {
		"excel_rows": len(excel_map),
		"matched_appointments": matched,
		"with_doc_code_in_excel": with_doc_code,
		"pending_doc_code_updates": len(pending),
		"doc_code_to_set": doc_code_to_set,
		"practitioner_to_set": practitioner_to_set,
		"sample_doc_code": sample,
	}


def ensure_healthcare_practitioner(doc_code: str) -> tuple[str, bool]:
	"""Return Healthcare Practitioner name and whether a new record was created."""
	doc_code = (doc_code or "").strip()
	if not doc_code:
		return "", False

	name = frappe.db.get_value("Healthcare Practitioner", {"doctors_id": doc_code}, "name")
	if name:
		return name, False
	if frappe.db.exists("Healthcare Practitioner", doc_code):
		return doc_code, False

	doc = frappe.new_doc("Healthcare Practitioner")
	doc.doctors_id = doc_code
	doc.first_name = doc_code
	doc.practitioner_name = doc_code
	doc.status = "Active"
	doc.insert(ignore_permissions=True)
	return doc.name, True


def apply_excel_row_to_appointment(app_num: str, excel_row: dict[str, str]) -> dict[str, int]:
	"""Update doc_code, optional old_status, and practitioner (only when empty)."""
	stats = {"doc_code": 0, "practitioner": 0, "old_status": 0, "practitioner_created": 0}
	doc_code = (excel_row.get("doc_code") or "").strip()
	app_status = (excel_row.get("app_status") or "").strip().upper()

	if not frappe.db.exists("Patient Appointment", app_num):
		return stats

	pa = frappe.db.get_value(
		"Patient Appointment",
		app_num,
		["doc_code", "practitioner", "old_status"],
		as_dict=True,
	)
	if not pa:
		return stats

	updates: dict[str, str] = {}

	if doc_code and (pa.doc_code or "").strip() != doc_code:
		updates["doc_code"] = doc_code
		stats["doc_code"] = 1

	if app_status and not (pa.old_status or "").strip():
		updates["old_status"] = app_status
		stats["old_status"] = 1

	if doc_code and not (pa.practitioner or "").strip():
		practitioner, created = ensure_healthcare_practitioner(doc_code)
		if created:
			stats["practitioner_created"] = 1
		if practitioner:
			updates["practitioner"] = practitioner
			stats["practitioner"] = 1

	if updates:
		frappe.db.set_value("Patient Appointment", app_num, updates, update_modified=False)

	return stats


def target_status_from_old_status(old_status: str | None, appointment_date) -> str | None:
	"""Map Oracle old_status to ERPNext Patient Appointment status."""
	code = (old_status or "").strip().upper()
	if code == "V":
		return "Closed"
	if code == "S":
		appt_date = getdate(appointment_date) if appointment_date else None
		if appt_date and appt_date < getdate(nowdate()):
			return "No Show"
		return "Scheduled"
	return None


def _needs_update_filters_sql() -> tuple[str, tuple]:
	today = getdate(nowdate())
	clause = """
		IFNULL(old_status, '') != ''
		AND IFNULL(status, '') != 'Cancelled'
		AND (
			(UPPER(TRIM(old_status)) = 'V' AND IFNULL(status, '') != 'Closed')
			OR (
				UPPER(TRIM(old_status)) = 'S'
				AND appointment_date IS NOT NULL
				AND appointment_date < %s
				AND IFNULL(status, '') != 'No Show'
			)
			OR (
				UPPER(TRIM(old_status)) = 'S'
				AND (appointment_date IS NULL OR appointment_date >= %s)
				AND IFNULL(status, '') != 'Scheduled'
			)
		)
	"""
	return clause, (today, today)


def count_appointments_needing_old_status_update() -> dict:
	today = getdate(nowdate())
	row = frappe.db.sql(
		"""
		SELECT
			SUM(CASE
				WHEN UPPER(TRIM(old_status)) = 'V' AND IFNULL(status, '') != 'Closed' THEN 1
				ELSE 0
			END) AS to_closed,
			SUM(CASE
				WHEN UPPER(TRIM(old_status)) = 'S'
					AND appointment_date IS NOT NULL
					AND appointment_date < %s
					AND IFNULL(status, '') != 'No Show' THEN 1
				ELSE 0
			END) AS to_no_show,
			SUM(CASE
				WHEN UPPER(TRIM(old_status)) = 'S'
					AND (appointment_date IS NULL OR appointment_date >= %s)
					AND IFNULL(status, '') != 'Scheduled' THEN 1
				ELSE 0
			END) AS to_scheduled
		FROM `tabPatient Appointment`
		WHERE IFNULL(old_status, '') != ''
			AND IFNULL(status, '') != 'Cancelled'
			AND UPPER(TRIM(old_status)) IN ('S', 'V')
		""",
		(today, today),
		as_dict=True,
	)
	data = row[0] if row else {}
	to_closed = cint(data.get("to_closed"))
	to_no_show = cint(data.get("to_no_show"))
	to_scheduled = cint(data.get("to_scheduled"))
	return {
		"to_closed": to_closed,
		"to_no_show": to_no_show,
		"to_scheduled": to_scheduled,
		"total_needing_update": to_closed + to_no_show + to_scheduled,
	}


def list_appointments_needing_old_status_update(*, limit: int) -> list[dict]:
	where, params = _needs_update_filters_sql()
	return frappe.db.sql(
		f"""
		SELECT name, old_status, appointment_date, status
		FROM `tabPatient Appointment`
		WHERE {where}
		ORDER BY name
		LIMIT %s
		""",
		(*params, limit),
		as_dict=True,
	)


@frappe.whitelist()
def run_patient_appointment_old_status_backfill_preview(file_url: str | None = None) -> dict:
	"""Preview doc_code/practitioner backfill from Excel and status updates from old_status."""
	_require_admin()
	status_counts = count_appointments_needing_old_status_update()
	status_sample = list_appointments_needing_old_status_update(limit=10)

	result: dict[str, Any] = {
		**status_counts,
		"sample": [
			{
				"name": row.name,
				"old_status": row.old_status,
				"appointment_date": str(row.appointment_date) if row.appointment_date else None,
				"current_status": row.status,
				"target_status": target_status_from_old_status(
					row.old_status, row.appointment_date
				),
			}
			for row in status_sample
		],
	}

	if file_url:
		excel_map = _parse_appointments_excel(file_url)
		result.update(count_doc_code_backfill_preview(excel_map))
	else:
		result.update(
			{
				"excel_rows": 0,
				"matched_appointments": 0,
				"pending_doc_code_updates": 0,
				"doc_code_to_set": 0,
				"practitioner_to_set": 0,
				"sample_doc_code": [],
			}
		)

	return result


def run_patient_appointment_doc_code_backfill_batch(*, offset: int = 0) -> dict:
	"""Update one batch of appointments from cached Excel (doc_code + practitioner)."""
	excel_map = get_cached_excel_map()
	pending = get_cached_pending_doc_code_queue()
	batch = pending[offset : offset + PATIENT_APPOINTMENT_DOC_CODE_BATCH_SIZE]

	updated = 0
	errors = 0
	error_names: list[str] = []
	totals = {
		"doc_code": 0,
		"practitioner": 0,
		"old_status": 0,
		"practitioner_created": 0,
	}

	for app_num in batch:
		excel_row = excel_map.get(app_num) or {}
		try:
			stats = apply_excel_row_to_appointment(app_num, excel_row)
			if any(stats.values()):
				updated += 1
			for key, value in stats.items():
				totals[key] = totals.get(key, 0) + value
		except Exception:
			errors += 1
			error_names.append(app_num)
			frappe.log_error(
				title=f"Patient Appointment doc_code backfill failed: {app_num}",
				message=frappe.get_traceback(),
			)

	frappe.db.commit()

	next_offset = offset + len(batch)
	done = next_offset >= len(pending) or not batch

	return {
		"processed": next_offset,
		"updated": updated,
		"errors": errors,
		"error_names": error_names,
		"totals": totals,
		"remaining": max(len(pending) - next_offset, 0),
		"done": done,
		"stats": {
			"doc_code_phase_processed": next_offset,
			"doc_code_phase_total": len(pending),
			"errors": errors,
			**totals,
		},
	}


def run_patient_appointment_old_status_backfill_batch(*, offset: int = 0) -> dict:
	"""Update one batch of appointments from old_status."""
	rows = list_appointments_needing_old_status_update(
		limit=PATIENT_APPOINTMENT_OLD_STATUS_BATCH_SIZE
	)
	updated = 0
	skipped = 0
	errors = 0
	error_names: list[str] = []
	by_target: dict[str, int] = {"Closed": 0, "No Show": 0, "Scheduled": 0}

	for row in rows:
		target = target_status_from_old_status(row.old_status, row.appointment_date)
		if not target:
			skipped += 1
			continue
		if (row.status or "").strip() == target:
			skipped += 1
			continue
		try:
			frappe.db.set_value(
				"Patient Appointment",
				row.name,
				"status",
				target,
				update_modified=False,
			)
			updated += 1
			by_target[target] = by_target.get(target, 0) + 1
		except Exception:
			errors += 1
			error_names.append(row.name)
			frappe.log_error(
				title=f"Patient Appointment old_status backfill failed: {row.name}",
				message=frappe.get_traceback(),
			)

	frappe.db.commit()

	remaining = count_appointments_needing_old_status_update()
	done = not rows or cint(remaining.get("total_needing_update")) == 0

	return {
		"processed": offset + updated,
		"updated": updated,
		"skipped": skipped,
		"errors": errors,
		"error_names": error_names,
		"by_target": by_target,
		"remaining": remaining,
		"done": done,
		"stats": {
			"updated": offset + updated,
			"errors": errors,
			"remaining": remaining.get("total_needing_update"),
			"by_target": by_target,
		},
	}
