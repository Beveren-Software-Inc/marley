"""Import Oracle IP_PATIENT_SHORT_LEAVE Excel into IP_PATIENT_SHORT_LEAVE rows."""

from __future__ import annotations

from datetime import date, datetime, time
from typing import Any

import frappe
from frappe import _
from frappe.utils import getdate

from healthcare.api.discharge_checklist_import import _resolve_cost_center
from healthcare.api.patient_info_import import (
	_cell_text,
	_clean_oracle_num,
	_excel_file_path,
	_parse_date_value,
	_parse_datetime_value,
	_require_admin,
)
from healthcare.api.visit_diagnosis_sync import _resolve_inpatient_admission

DOCTYPE = "IP_PATIENT_SHORT_LEAVE"

EXCEL_HEADER_MAP = {
	"TRANS_NUM": "trans_num",
	"TRANS_DATE": "trans_date",
	"ADMISSION_NUM": "admission_num",
	"PATIENT_NUM": "patient_num",
	"DATE_FROM": "date_from",
	"TIME_FROM": "time_from",
	"DATE_TO": "date_to",
	"TIME_TO": "time_to",
	"FINAL_COME_DATE": "final_come_date",
	"CARE_OF": "care_of",
	"RELATIONSHIP": "relationship",
	"PURPOSE_OF_THE_LEAVE": "purpose_of_leave",
	"CONTACT_NUM": "contact_num",
	"DOC_CR_ID": "doc_cr_id",
	"DOC_CR_DATE": "doc_cr_date",
	"NUR_CR_ID": "nur_cr_id",
	"NUR_CR_DATE": "nur_cr_date",
	"LEAVE_STATUS": "leave_status",
	"BRANCH_NUM": "branch_num",
	"CR_ID": "cr_id",
	"CR_DATE": "cr_date",
	"UP_ID": "up_id",
	"UP_DATE": "up_date",
	"ADMISSION_NUM_OLD": "admission_num_old",
}


def _normalize_header(cell: Any) -> str:
	if cell is None:
		return ""
	text = str(cell).strip().upper().replace(" ", "_")
	return EXCEL_HEADER_MAP.get(text, text.lower())


def _format_legacy_datetime(value: Any) -> str | None:
	dt = _parse_datetime_value(value)
	if dt:
		return dt.strftime("%Y-%m-%d %H:%M:%S")
	parsed_date = _parse_date_value(value)
	if parsed_date:
		return parsed_date.strftime("%Y-%m-%d")
	text = _cell_text(value)
	return text or None


def _parse_date_field(value: Any):
	dt = _parse_datetime_value(value)
	if dt:
		return getdate(dt)
	parsed_date = _parse_date_value(value)
	if parsed_date:
		return parsed_date
	return None


def _legacy_time_text(value: Any) -> str | None:
	if value in (None, ""):
		return None
	if isinstance(value, datetime):
		return value.strftime("%H:%M:%S")
	if isinstance(value, time):
		return value.strftime("%H:%M:%S")
	if isinstance(value, date):
		return "00:00:00"
	if isinstance(value, (int, float)) and not isinstance(value, bool):
		try:
			import openpyxl.utils.datetime

			dt = openpyxl.utils.datetime.from_excel(float(value))
			return dt.strftime("%H:%M:%S")
		except Exception:
			frac = float(value) % 1
			if frac:
				seconds = int(round(frac * 24 * 3600))
				hour, rem = divmod(seconds, 3600)
				minute, second = divmod(rem, 60)
				return f"{hour:02d}:{minute:02d}:{second:02d}"
	text = _cell_text(value)
	return text or None


def _resolve_patient(patient_num: str | None) -> str | None:
	patient = _clean_oracle_num(patient_num)
	if not patient:
		return None
	if frappe.db.exists("Patient", patient):
		return patient
	return None


def _resolve_admission(row: dict) -> str | None:
	admission_num = row.get("admission_num") or ""
	old_num = row.get("admission_num_old") or ""
	for candidate in (admission_num, old_num):
		if not candidate:
			continue
		if frappe.db.exists("Inpatient Admission", candidate):
			return candidate
		resolved = _resolve_inpatient_admission(candidate)
		if resolved:
			return resolved
	return None


def _parse_sheet_rows(ws) -> list[dict]:
	rows_iter = ws.iter_rows(values_only=True)
	try:
		header_row = next(rows_iter)
	except StopIteration:
		return []

	headers = [_normalize_header(h) for h in header_row]
	parsed: list[dict] = []
	for raw in rows_iter:
		if not raw or all(cell is None or str(cell).strip() == "" for cell in raw):
			continue
		row: dict[str, Any] = {}
		for idx, key in enumerate(headers):
			if not key or idx >= len(raw):
				continue
			row[key] = raw[idx]

		trans_num = _clean_oracle_num(row.get("trans_num"))
		if not trans_num:
			continue
		row["trans_num"] = trans_num
		row["admission_num"] = _clean_oracle_num(row.get("admission_num"))
		row["patient_num"] = _clean_oracle_num(row.get("patient_num"))
		row["admission_num_old"] = _clean_oracle_num(row.get("admission_num_old"))
		row["doc_cr_id"] = _clean_oracle_num(row.get("doc_cr_id"))
		row["nur_cr_id"] = _clean_oracle_num(row.get("nur_cr_id"))
		row["cr_id"] = _clean_oracle_num(row.get("cr_id"))
		row["up_id"] = _clean_oracle_num(row.get("up_id"))
		parsed.append(row)
	return parsed


def _parse_excel_rows(file_url: str) -> tuple[list[dict], dict[str, int]]:
	try:
		import openpyxl
	except ImportError:
		frappe.throw(
			_(
				"openpyxl is required to read Excel files. Install it in the bench environment: pip install openpyxl"
			)
		)

	path = _excel_file_path(file_url)
	wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
	by_key: dict[str, dict] = {}
	sheet_row_counts: dict[str, int] = {}
	try:
		for sheet_name in wb.sheetnames:
			sheet_rows = _parse_sheet_rows(wb[sheet_name])
			sheet_row_counts[sheet_name] = len(sheet_rows)
			for row in sheet_rows:
				by_key[row["trans_num"]] = row
	finally:
		wb.close()
	return list(by_key.values()), sheet_row_counts


def _build_fields(row: dict) -> dict[str, Any]:
	admission = _resolve_admission(row)
	patient = _resolve_patient(row.get("patient_num"))

	fields: dict[str, Any] = {
		"trans_num": row["trans_num"],
		"care_of": _cell_text(row.get("care_of")) or None,
		"relationship": _cell_text(row.get("relationship")) or None,
		"purpose_of_leave": _cell_text(row.get("purpose_of_leave")) or None,
		"contact_num": _cell_text(row.get("contact_num")) or None,
		"leave_status": _cell_text(row.get("leave_status")) or None,
		"doc_cr_id": row.get("doc_cr_id") or None,
		"nur_cr_id": row.get("nur_cr_id") or None,
		"cr_id": row.get("cr_id") or None,
		"up_id": row.get("up_id") or None,
	}

	trans_date = _parse_date_field(row.get("trans_date"))
	if trans_date:
		fields["trans_date"] = trans_date
	date_from = _parse_date_field(row.get("date_from"))
	if date_from:
		fields["date_from"] = date_from
	date_to = _parse_date_field(row.get("date_to"))
	if date_to:
		fields["date_to"] = date_to
	final_come_date = _parse_date_field(row.get("final_come_date"))
	if final_come_date:
		fields["final_come_date"] = final_come_date

	time_from = _legacy_time_text(row.get("time_from"))
	if time_from:
		fields["time_from"] = time_from
	time_to = _legacy_time_text(row.get("time_to"))
	if time_to:
		fields["time_to"] = time_to

	cr_date = _format_legacy_datetime(row.get("cr_date"))
	if cr_date:
		fields["cr_date"] = cr_date
	up_date = _format_legacy_datetime(row.get("up_date"))
	if up_date:
		fields["up_date"] = up_date

	if admission:
		fields["admission"] = admission
	if patient:
		fields["patient_no"] = patient

	branch = _resolve_cost_center(row.get("branch_num"))
	if branch:
		fields["branch"] = branch

	return {key: value for key, value in fields.items() if value not in (None, "")}


def _preview_stats(rows: list[dict]) -> dict:
	existing = 0
	resolvable_admissions = 0
	resolvable_patients = 0
	for row in rows:
		if frappe.db.exists(DOCTYPE, row["trans_num"]):
			existing += 1
		if _resolve_admission(row):
			resolvable_admissions += 1
		if _resolve_patient(row.get("patient_num")):
			resolvable_patients += 1

	return {
		"existing_records": existing,
		"resolvable_admissions": resolvable_admissions,
		"resolvable_patients": resolvable_patients,
		"sample_trans_nums": [row.get("trans_num") for row in rows[:5]],
	}


def upsert_ip_patient_short_leave(row: dict) -> dict:
	fields = _build_fields(row)
	if not fields.get("trans_num"):
		return {"status": "skip", "trans_num": row.get("trans_num")}

	trans_num = fields["trans_num"]
	if frappe.db.exists(DOCTYPE, trans_num):
		doc = frappe.get_doc(DOCTYPE, trans_num)
		for key, value in fields.items():
			doc.set(key, value)
		doc.flags.ignore_mandatory = True
		doc.flags.ignore_links = True
		doc.flags.ignore_validate = True
		doc.save(ignore_permissions=True)
		action = "updated"
	else:
		doc = frappe.get_doc({"doctype": DOCTYPE, **fields})
		doc.flags.ignore_mandatory = True
		doc.flags.ignore_links = True
		doc.flags.ignore_validate = True
		doc.insert(ignore_permissions=True)
		action = "created"

	return {"status": action, "trans_num": trans_num, "name": doc.name}


@frappe.whitelist()
def preview_ip_patient_short_leave_import(file_url: str) -> dict:
	_require_admin()
	if not (file_url or "").strip():
		frappe.throw(_("Please upload the IP_PATIENT_SHORT_LEAVE Excel file."))

	rows, sheet_row_counts = _parse_excel_rows(file_url)
	stats = _preview_stats(rows)
	raw_row_total = sum(sheet_row_counts.values())
	return {
		"excel_rows": len(rows),
		"raw_excel_rows": raw_row_total,
		"sheets": list(sheet_row_counts.keys()),
		"sheet_row_counts": sheet_row_counts,
		**stats,
	}


@frappe.whitelist()
def run_ip_patient_short_leave_import(file_url: str) -> dict:
	"""Import all rows synchronously (small file — no background job)."""
	_require_admin()
	if not (file_url or "").strip():
		frappe.throw(_("Please upload the IP_PATIENT_SHORT_LEAVE Excel file."))

	rows, _ = _parse_excel_rows(file_url)
	created = updated = skipped = 0
	errors: list[str] = []

	for row in rows:
		try:
			result = upsert_ip_patient_short_leave(row)
			status = result.get("status")
			if status == "created":
				created += 1
			elif status == "updated":
				updated += 1
			else:
				skipped += 1
		except Exception:
			errors.append(f"{row.get('trans_num')}: {frappe.get_traceback()}")
			frappe.log_error(title=f"IP_PATIENT_SHORT_LEAVE import failed: {row.get('trans_num')}")

	frappe.db.commit()
	return {
		"ok": True,
		"total": len(rows),
		"created": created,
		"updated": updated,
		"skipped": skipped,
		"errors": len(errors),
	}
