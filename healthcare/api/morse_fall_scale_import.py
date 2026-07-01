"""Import Oracle MORSE_FALL_SCALE_01 Excel directly into Morse Fall Scale rows."""

from __future__ import annotations

import json
from datetime import date, datetime
from typing import Any

import frappe
from frappe import _
from frappe.utils import cint, getdate

from healthcare.api.discharge_checklist_import import _resolve_cost_center
from healthcare.api.morse_fall_scale_detail_import import (
	_detail_rows_from_staging,
	replace_morse_fall_scale_detail_from_staging,
)
from healthcare.api.patient_info_import import (
	_clean_oracle_num,
	_excel_file_path,
	_require_admin,
)
from healthcare.api.patient_visit_import import ensure_patient_for_legacy_import
from healthcare.api.visit_diagnoses_op_import import _legacy_data_datetime
from healthcare.api.visit_diagnosis_sync import _resolve_inpatient_admission

MORSE_FALL_SCALE_IMPORT_BATCH_SIZE = 200
CACHE_TTL = 7200
CACHE_KEYS = {
	"file_url": "healthcare:data_migration:morse_fall_scale_import:file_url",
	"row_keys": "healthcare:data_migration:morse_fall_scale_import:row_keys",
	"rows": "healthcare:data_migration:morse_fall_scale_import:rows",
}

EXCEL_HEADER_MAP = {
	"TRANS_NUM": "trans_num",
	"TRANS_DATE": "trans_date",
	"ADMISSION_NUM": "admission_num",
	"PATIENT_NUM": "patient_num",
	"ORDER_NUM": "order_num",
	"TEXT_MESSAGE_1": "text_message_1",
	"GET_POINTS_1": "get_points_1",
	"TEXT_MESSAGE_2": "text_message_2",
	"GET_POINTS_2": "get_points_2",
	"TEXT_MESSAGE_3": "text_message_3",
	"GET_POINTS_3": "get_points_3",
	"TEXT_MESSAGE_4": "text_message_4",
	"GET_POINTS_4": "get_points_4",
	"TEXT_MESSAGE_5": "text_message_5",
	"GET_POINTS_5": "get_points_5",
	"TEXT_MESSAGE_6": "text_message_6",
	"GET_POINTS_6": "get_points_6",
	"TEXT_MESSAGE_7": "text_message_7",
	"GET_POINTS_7": "get_points_7",
	"TOTAL_POINTS": "total_points",
	"BRANCH_NUM": "branch_num",
	"BRANCH": "branch_num",
	"CR_ID": "cr_id",
	"CR_DATE": "cr_date",
	"UP_ID": "up_id",
	"UP_DATE": "up_date",
}


def _normalize_header(cell: Any) -> str:
	if cell is None:
		return ""
	text = str(cell).strip().upper().replace(" ", "_")
	return EXCEL_HEADER_MAP.get(text, text.lower())


def _parse_trans_date(value: Any):
	if value in (None, ""):
		return None
	if isinstance(value, datetime):
		return value.date()
	if isinstance(value, date):
		return value
	try:
		return getdate(value)
	except Exception:
		return None


def _resolve_company(admission: str | None) -> str | None:
	if admission:
		company = frappe.db.get_value("Inpatient Admission", admission, "company")
		if company:
			return company
	return frappe.defaults.get_global_default("company")


def _resolve_patient(row: dict, admission: str) -> tuple[str | None, int]:
	"""Return (patient name, patients_created count)."""
	patient = _clean_oracle_num(row.get("patient_num")) or None
	patients_created = 0
	if patient and not frappe.db.exists("Patient", patient):
		result = ensure_patient_for_legacy_import(patient)
		if result.get("status") == "created":
			patients_created = 1
		elif result.get("status") != "existing":
			patient = None
	if not patient and admission:
		patient = frappe.db.get_value("Inpatient Admission", admission, "patient")
	return patient, patients_created


def _parse_sheet_rows(ws) -> list[dict]:
	rows_iter = ws.iter_rows(values_only=True)
	try:
		header_row = next(rows_iter)
	except StopIteration:
		return []

	headers = [_normalize_header(h) for h in header_row]
	parsed: list[dict] = []
	for raw in rows_iter:
		if not raw or all(cell is None or str(cell).strip() == "" for cell in raw):
			continue
		row: dict[str, Any] = {}
		for idx, key in enumerate(headers):
			if not key or idx >= len(raw):
				continue
			row[key] = raw[idx]

		trans_num = _clean_oracle_num(row.get("trans_num"))
		if not trans_num:
			continue

		row["trans_num"] = trans_num
		row["admission_num"] = _clean_oracle_num(row.get("admission_num"))
		row["patient_num"] = _clean_oracle_num(row.get("patient_num"))
		row["order_num"] = _clean_oracle_num(row.get("order_num"))
		row["cr_id"] = _clean_oracle_num(row.get("cr_id"))
		row["up_id"] = _clean_oracle_num(row.get("up_id"))
		parsed.append(row)
	return parsed


def _parse_excel_rows(file_url: str) -> tuple[list[dict], dict[str, int]]:
	try:
		import openpyxl
	except ImportError:
		frappe.throw(
			_("openpyxl is required to read Excel files. Install it in the bench environment: pip install openpyxl")
		)

	path = _excel_file_path(file_url)
	wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
	all_rows: list[dict] = []
	sheet_row_counts: dict[str, int] = {}
	try:
		for sheet_name in wb.sheetnames:
			sheet_rows = _parse_sheet_rows(wb[sheet_name])
			sheet_row_counts[sheet_name] = len(sheet_rows)
			all_rows.extend(sheet_rows)
	finally:
		wb.close()
	return all_rows, sheet_row_counts


def _apply_detail_rows(doc, row: dict) -> int:
	return replace_morse_fall_scale_detail_from_staging(doc, row)


def _build_morse_fall_scale_fields(row: dict) -> tuple[dict[str, Any], dict[str, int]]:
	stats = {"patients_created": 0}
	admission_num = row.get("admission_num") or ""
	admission = _resolve_inpatient_admission(admission_num, row.get("patient_num"))
	if not admission:
		return {}, {"skip_no_admission": 1}

	patient, patients_created = _resolve_patient(row, admission)
	stats["patients_created"] = patients_created
	if not patient:
		return {}, {"skip_no_patient": 1}

	fields: dict[str, Any] = {
		"trans_no": row["trans_num"],
		"admission_no": admission,
		"patient_no": patient,
		"written_admission": admission_num or None,
		"date": _parse_trans_date(row.get("trans_date")),
		"orderer_number": row.get("order_num") or None,
		"company": _resolve_company(admission),
		"cost_center": _resolve_cost_center(row.get("branch_num")),
		"cr_id": row.get("cr_id") or None,
		"up_id": row.get("up_id") or None,
	}
	cr_date = _legacy_data_datetime(row.get("cr_date"))
	if cr_date:
		fields["cr_date"] = cr_date
	up_date = _legacy_data_datetime(row.get("up_date"))
	if up_date:
		fields["up_date"] = up_date

	return {key: value for key, value in fields.items() if value not in (None, "")}, stats


def _persist_morse_fall_scale(doc, *, existing: bool) -> None:
	doc.flags.ignore_validate = True
	doc.flags.ignore_mandatory = True
	doc.flags.ignore_links = True
	doc.flags.ignore_permissions = True
	doc.flags.legacy_import = True
	if existing:
		doc.save(ignore_permissions=True)
	else:
		doc.insert(ignore_permissions=True)


def upsert_morse_fall_scale_from_row(row: dict) -> dict:
	trans_num = row.get("trans_num")
	if not trans_num:
		return {"status": "skip_no_trans_num"}

	fields, side_stats = _build_morse_fall_scale_fields(row)
	if side_stats.get("skip_no_admission"):
		return {"status": "skip_no_admission", "trans_num": trans_num}
	if side_stats.get("skip_no_patient"):
		return {"status": "skip_no_patient", "trans_num": trans_num}
	if not fields:
		return {"status": "skip_error", "trans_num": trans_num}

	existing_name = frappe.db.exists("Morse Fall Scale", trans_num)
	if existing_name:
		doc = frappe.get_doc("Morse Fall Scale", existing_name)
		for key, value in fields.items():
			if key == "trans_no":
				continue
			doc.set(key, value)
		action = "updated"
		existing = True
	else:
		doc = frappe.new_doc("Morse Fall Scale")
		doc.update(fields)
		action = "created"
		existing = False

	detail_count = _apply_detail_rows(doc, row)
	if detail_count == 0:
		return {"status": "skip_empty_details", "trans_num": trans_num}

	_persist_morse_fall_scale(doc, existing=existing)
	return {
		"status": action,
		"trans_num": trans_num,
		"name": doc.name,
		"detail_rows": detail_count,
		"patients_created": side_stats.get("patients_created", 0),
	}


def _preview_counts(rows: list[dict]) -> dict:
	existing = sum(1 for row in rows if frappe.db.exists("Morse Fall Scale", row["trans_num"]))
	resolved_admissions = 0
	unresolved_admissions = 0
	empty_details = 0
	patient_nums = {row["patient_num"] for row in rows if row.get("patient_num")}
	missing_patients = [pn for pn in patient_nums if not frappe.db.exists("Patient", pn)]

	for row in rows:
		if not _detail_rows_from_staging(row):
			empty_details += 1
		admission_num = row.get("admission_num") or ""
		if not admission_num:
			continue
		if _resolve_inpatient_admission(admission_num, row.get("patient_num")):
			resolved_admissions += 1
		else:
			unresolved_admissions += 1

	return {
		"existing_scales": existing,
		"new_scales": len(rows) - existing,
		"resolved_admissions": resolved_admissions,
		"unresolved_admissions": unresolved_admissions,
		"empty_detail_rows": empty_details,
		"unique_patients": len(patient_nums),
		"patients_to_create": len(missing_patients),
		"sample_patients_to_create": missing_patients[:10],
	}


def parse_and_cache_excel(file_url: str) -> dict:
	rows, sheet_row_counts = _parse_excel_rows(file_url)
	by_key = {row["trans_num"]: row for row in rows}
	row_keys = list(by_key.keys())

	frappe.cache().set_value(CACHE_KEYS["file_url"], file_url, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(CACHE_KEYS["row_keys"], row_keys, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(
		CACHE_KEYS["rows"],
		json.dumps(by_key, default=str),
		expires_in_sec=CACHE_TTL,
	)

	preview = _preview_counts(list(by_key.values()))
	raw_row_total = sum(sheet_row_counts.values())
	return {
		"excel_rows": len(by_key),
		"raw_excel_rows": raw_row_total,
		"sheets": list(sheet_row_counts.keys()),
		"sheet_row_counts": sheet_row_counts,
		**preview,
		"sample_trans_nums": row_keys[:5],
	}


def _load_cached_rows() -> dict[str, dict]:
	raw = frappe.cache().get_value(CACHE_KEYS["rows"])
	if not raw:
		return {}
	if isinstance(raw, dict):
		return raw
	return json.loads(raw)


@frappe.whitelist()
def preview_morse_fall_scale_excel_import(file_url: str) -> dict:
	_require_admin()
	if not (file_url or "").strip():
		frappe.throw(_("Please upload the MORSE_FALL_SCALE_01 Excel file."))
	return parse_and_cache_excel(file_url)


def run_morse_fall_scale_import_batch(*, offset: int = 0) -> dict:
	row_keys = frappe.cache().get_value(CACHE_KEYS["row_keys"]) or []
	rows_by_key = _load_cached_rows()
	if not row_keys or not rows_by_key:
		return {"processed": offset, "done": True, "batch_count": 0}

	batch_keys = row_keys[offset : offset + MORSE_FALL_SCALE_IMPORT_BATCH_SIZE]
	if not batch_keys:
		return {"processed": offset, "done": True, "batch_count": 0}

	created = updated = skip_no_admission = skip_no_patient = skip_empty_details = patients_created = 0
	errors: list[str] = []

	for key in batch_keys:
		row = rows_by_key.get(key) or {}
		try:
			result = upsert_morse_fall_scale_from_row(row)
			status = result.get("status")
			if status == "created":
				created += 1
			elif status == "updated":
				updated += 1
			elif status == "skip_no_admission":
				skip_no_admission += 1
			elif status == "skip_no_patient":
				skip_no_patient += 1
			elif status == "skip_empty_details":
				skip_empty_details += 1
			patients_created += cint(result.get("patients_created"))
		except Exception:
			errors.append(f"{key}: {frappe.get_traceback()}")
			frappe.log_error(title=f"MORSE_FALL_SCALE_01 import failed: {key}")

	frappe.db.commit()
	processed = offset + len(batch_keys)
	return {
		"processed": processed,
		"done": len(batch_keys) < MORSE_FALL_SCALE_IMPORT_BATCH_SIZE,
		"batch_count": len(batch_keys),
		"created": created,
		"updated": updated,
		"skip_no_admission": skip_no_admission,
		"skip_no_patient": skip_no_patient,
		"skip_empty_details": skip_empty_details,
		"patients_created": patients_created,
		"errors": len(errors),
	}
