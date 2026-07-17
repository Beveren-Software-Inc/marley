"""Import Oracle SALES_DATA_MASTER Excel into Legacy Sales Transactions (headers only)."""

from __future__ import annotations

import json
import os
from typing import Any

import frappe
from frappe import _
from frappe.utils import cint, flt, get_datetime, getdate

from healthcare.api.discharge_checklist_import import _resolve_cost_center
from healthcare.api.patient_info_import import (
	_cell_text,
	_clean_oracle_num,
	_excel_file_path,
	_parse_date_value,
	_parse_datetime_value,
	_require_admin,
)
from healthcare.api.visit_diagnosis_sync import (
	_resolve_inpatient_admission,
	_resolve_patient_visit,
)

DOCTYPE = "Legacy Sales Transactions"
LEGACY_SALES_MASTER_IMPORT_BATCH_SIZE = 500
CACHE_TTL = 7200
CACHE_KEYS = {
	"file_url": "healthcare:data_migration:legacy_sales_master_import:file_url",
	"cache_path": "healthcare:data_migration:legacy_sales_master_import:cache_path",
	"row_keys": "healthcare:data_migration:legacy_sales_master_import:row_keys",
	"total_rows": "healthcare:data_migration:legacy_sales_master_import:total_rows",
}

EXCEL_HEADER_MAP = {
	"TRANS_NUM": "trans_num",
	"TRANS_TYPE_NUM": "trans_type_num",
	"TRANS_DATE": "trans_date",
	"TRANS_YEAR": "trans_year",
	"TRANS_MONTH": "trans_month",
	"TRANS_PERIOD": "trans_period",
	"ST_NUM": "st_num",
	"INVOICE_NUM": "invoice_num",
	"INVOICE_DATE": "invoice_date",
	"TRANS_SOURCE": "trans_source",
	"TRANS_SUB_SOURCE": "trans_sub_source",
	"TOTAL_BILL_AMOUNT": "total_bill_amount",
	"DISCOUNT_PERCENTAGE": "discount_percentage",
	"TOTAL_DISCOUNT_AMOUNT": "total_discount_amount",
	"TOTAL_TAX_AMOUNT": "total_tax_amount",
	"FREIGHT_AMOUNT": "freight_amount",
	"OTHER_AMOUNT_1": "other_amount_1",
	"OTHER_AMOUNT_2": "other_amount_2",
	"OTHER_AMOUNT_3": "other_amount_3",
	"NET_BILL_AMOUNT": "net_bill_amount",
	"TRANS_REMARKS": "trans_remarks",
	"RETURN_TRANS_NUM": "return_trans_num",
	"FROM_BRANCH": "from_branch",
	"TO_BRANCH": "to_branch",
	"BRANCH_NUM": "branch_num",
	"VCH_STATUS": "vch_status",
	"CR_ID": "cr_id",
	"CR_DATE": "cr_date",
	"UP_ID": "up_id",
	"UP_DATE": "up_date",
	"CK_ID": "ck_id",
	"CK_DATE": "ck_date",
	"AP_ID": "ap_id",
	"AP_DATE": "ap_date",
	"VISIT_NUM": "visit_num",
	"TOTAL_PROFIT_AMT": "total_profit_amt",
	"ADMISSION_NUM": "admission_num",
	"CHARGE_YN": "charge_yn",
	"TOTAL_SERVICES_CHARGES": "total_services_charges",
	"ONLY_SERVICES_CHARGES": "only_services_charges",
	"SERVICES_CHARGES_PERCENTAGE": "services_charges_percentage",
	"LINK_GRN_NUM": "link_grn_num",
	"IS_RETURN_SLR_SRV_CHARGES": "is_return_slr_srv_charges",
	"TO_BRANCH_NUM": "to_branch_num",
	"PINK_PRESC_NUM": "pink_presc_num",
	"IS_PINK_PRESC_CHARGE": "is_pink_presc_charge",
	"PINK_PRESC_AMT": "pink_presc_amt",
}


def _normalize_header(cell: Any) -> str:
	if cell is None:
		return ""
	text = str(cell).strip().upper().replace(" ", "_")
	return EXCEL_HEADER_MAP.get(text, text.lower())


def _format_legacy_datetime(value: Any, *, datemode: int = 0) -> str | None:
	dt = _parse_datetime_value(value, datemode=datemode)
	if dt:
		return dt.strftime("%Y-%m-%d %H:%M:%S")
	parsed_date = _parse_date_value(value, datemode=datemode)
	if parsed_date:
		return parsed_date.strftime("%Y-%m-%d")
	text = _cell_text(value)
	return text or None


def _parse_date_field(value: Any, *, datemode: int = 0):
	dt = _parse_datetime_value(value, datemode=datemode)
	if dt:
		return getdate(dt)
	return _parse_date_value(value, datemode=datemode)


def _parse_datetime_field(value: Any, *, datemode: int = 0):
	dt = _parse_datetime_value(value, datemode=datemode)
	if dt:
		return dt
	parsed_date = _parse_date_value(value, datemode=datemode)
	if parsed_date:
		return get_datetime(parsed_date)
	return None


def _amount(value: Any) -> float | None:
	if value in (None, ""):
		return None
	if isinstance(value, str):
		value = value.strip().replace(",", "")
		if not value:
			return None
	return flt(value)


def _int_or_none(value: Any) -> int | None:
	text = _clean_oracle_num(value)
	if not text:
		return None
	try:
		return cint(text)
	except Exception:
		return None


def _parse_sheet_raw_rows(headers: list[str], rows_iter, datemode: int = 0) -> list[dict]:
	parsed: list[dict] = []
	for raw in rows_iter:
		if not raw or all(cell is None or str(cell).strip() == "" for cell in raw):
			continue
		row: dict[str, Any] = {}
		for idx, key in enumerate(headers):
			if not key or idx >= len(raw):
				continue
			row[key] = raw[idx]

		trans_no = _cell_text(row.get("trans_num")).strip()
		if not trans_no:
			# Sheet 2+ often repeats the header row — skip it.
			if _normalize_header(raw[0] if raw else None) == "trans_num":
				continue
			continue

		row["trans_no"] = trans_no
		row["visit_num"] = _clean_oracle_num(row.get("visit_num")) or None
		row["admission_num"] = _clean_oracle_num(row.get("admission_num")) or None
		row["branch_num"] = _clean_oracle_num(row.get("branch_num")) or None
		row["from_branch"] = _clean_oracle_num(row.get("from_branch")) or None
		row["to_branch"] = _clean_oracle_num(row.get("to_branch")) or None
		row["to_branch_num"] = _clean_oracle_num(row.get("to_branch_num")) or None
		row["cr_id"] = _clean_oracle_num(row.get("cr_id")) or None
		row["up_id"] = _clean_oracle_num(row.get("up_id")) or None
		row["ap_id"] = _clean_oracle_num(row.get("ap_id")) or None
		row["ck_id"] = _clean_oracle_num(row.get("ck_id")) or None
		row["_datemode"] = datemode
		parsed.append(row)
	return parsed


def _parse_excel_rows(file_url: str) -> tuple[list[dict], dict[str, int]]:
	"""Parse every worksheet. Old .xls exports often split at ~65k rows across Sheet 1/Sheet 2."""
	import os

	path = _excel_file_path(file_url)
	ext = os.path.splitext(path)[1].lower()
	by_key: dict[str, dict] = {}
	sheet_row_counts: dict[str, int] = {}

	if ext == ".xls":
		try:
			import xlrd
		except ImportError:
			frappe.throw(
				_("xlrd is required to read .xls files. Install it in the bench environment: pip install xlrd")
			)
		wb = xlrd.open_workbook(path, formatting_info=False)
		datemode = wb.datemode
		for sheet_idx in range(wb.nsheets):
			ws = wb.sheet_by_index(sheet_idx)
			sheet_name = wb.sheet_names()[sheet_idx] or f"Sheet {sheet_idx + 1}"
			if ws.nrows < 1:
				sheet_row_counts[sheet_name] = 0
				continue
			headers = [_normalize_header(ws.cell_value(0, c)) for c in range(ws.ncols)]
			raw_rows = (
				tuple(ws.cell_value(r, c) for c in range(ws.ncols)) for r in range(1, ws.nrows)
			)
			sheet_rows = _parse_sheet_raw_rows(headers, raw_rows, datemode=datemode)
			sheet_row_counts[sheet_name] = len(sheet_rows)
			for row in sheet_rows:
				by_key[row["trans_no"]] = row
		return list(by_key.values()), sheet_row_counts

	try:
		import openpyxl
	except ImportError:
		frappe.throw(
			_(
				"openpyxl is required to read Excel files. Install it in the bench environment: pip install openpyxl"
			)
		)

	wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
	try:
		for sheet_name in wb.sheetnames:
			ws = wb[sheet_name]
			rows_iter = ws.iter_rows(values_only=True)
			try:
				header_row = next(rows_iter)
			except StopIteration:
				sheet_row_counts[sheet_name] = 0
				continue
			headers = [_normalize_header(h) for h in header_row]
			sheet_rows = _parse_sheet_raw_rows(headers, rows_iter, datemode=0)
			sheet_row_counts[sheet_name] = len(sheet_rows)
			for row in sheet_rows:
				by_key[row["trans_no"]] = row
	finally:
		wb.close()

	return list(by_key.values()), sheet_row_counts


def _build_fields(row: dict) -> dict[str, Any]:
	datemode = cint(row.get("_datemode") or 0)
	fields: dict[str, Any] = {
		"trans_no": row["trans_no"],
		"trans_type_num": _cell_text(row.get("trans_type_num")) or None,
		"trans_source": _cell_text(row.get("trans_source")) or None,
		"trans_sub_source": _cell_text(row.get("trans_sub_source")) or None,
		"vch_status": _cell_text(row.get("vch_status")) or None,
		"invoice_num": _cell_text(row.get("invoice_num")) or None,
		"st_num": _clean_oracle_num(row.get("st_num")) or None,
		"trans_remarks": _cell_text(row.get("trans_remarks")) or None,
		"return_trans_num": _cell_text(row.get("return_trans_num")) or None,
		"link_grn_num": _cell_text(row.get("link_grn_num")) or None,
		"charge_yn": _cell_text(row.get("charge_yn")) or None,
		"is_return_slr_srv_charges": _cell_text(row.get("is_return_slr_srv_charges")) or None,
		"pink_presc_num": _cell_text(row.get("pink_presc_num")) or None,
		"is_pink_presc_charge": _cell_text(row.get("is_pink_presc_charge")) or None,
		"visit_num": row.get("visit_num"),
		"admission_num": row.get("admission_num"),
		"to_branch_num": row.get("to_branch_num"),
		"cr_id": row.get("cr_id"),
		"up_id": row.get("up_id"),
		"ap_id": row.get("ap_id"),
		"ck_id": row.get("ck_id"),
		"trans_period": _clean_oracle_num(row.get("trans_period"))
		or _cell_text(row.get("trans_period"))
		or None,
	}

	trans_date = _parse_date_field(row.get("trans_date"), datemode=datemode)
	if trans_date:
		fields["trans_date"] = trans_date
	invoice_date = _parse_date_field(row.get("invoice_date"), datemode=datemode)
	if invoice_date:
		fields["invoice_date"] = invoice_date

	trans_year = _int_or_none(row.get("trans_year"))
	if trans_year is not None:
		fields["trans_year"] = trans_year
	trans_month = _int_or_none(row.get("trans_month"))
	if trans_month is not None:
		fields["trans_month"] = trans_month

	for amount_field in (
		"total_bill_amount",
		"discount_percentage",
		"total_discount_amount",
		"total_tax_amount",
		"freight_amount",
		"other_amount_1",
		"other_amount_2",
		"other_amount_3",
		"net_bill_amount",
		"total_profit_amt",
		"total_services_charges",
		"only_services_charges",
		"services_charges_percentage",
		"pink_presc_amt",
	):
		amount = _amount(row.get(amount_field))
		if amount is not None:
			fields[amount_field] = amount

	for date_field in ("cr_date", "up_date", "ap_date", "ck_date"):
		legacy_date = _format_legacy_datetime(row.get(date_field), datemode=datemode)
		if legacy_date:
			fields[date_field] = legacy_date

	date_created = _parse_datetime_field(row.get("cr_date"), datemode=datemode)
	if date_created:
		fields["date_created"] = date_created

	branch = _resolve_cost_center(row.get("branch_num"))
	if branch:
		fields["branch"] = branch
	from_branch = _resolve_cost_center(row.get("from_branch"))
	if from_branch:
		fields["from_branch"] = from_branch
	to_branch = _resolve_cost_center(row.get("to_branch")) or _resolve_cost_center(
		row.get("to_branch_num")
	)
	if to_branch:
		fields["to_branch"] = to_branch

	visit_num = row.get("visit_num") or ""
	patient_visit = _resolve_patient_visit(visit_num) if visit_num else None
	if patient_visit:
		fields["patient_visit"] = patient_visit
		patient = frappe.db.get_value("Patient Visit", patient_visit, "patient")
		if patient:
			fields["patient"] = patient
			patient_name = frappe.db.get_value("Patient", patient, "patient_name")
			if patient_name:
				fields["patient_name"] = patient_name

	admission_num = row.get("admission_num") or ""
	admission = _resolve_inpatient_admission(admission_num) if admission_num else None
	if admission:
		fields["admission"] = admission
		if not fields.get("patient"):
			patient = frappe.db.get_value("Inpatient Admission", admission, "patient")
			if patient:
				fields["patient"] = patient
				patient_name = frappe.db.get_value("Patient", patient, "patient_name")
				if patient_name:
					fields["patient_name"] = patient_name

	return {key: value for key, value in fields.items() if value not in (None, "")}


def _apply_legacy_flags(doc) -> None:
	doc.flags.ignore_validate = True
	doc.flags.ignore_mandatory = True
	doc.flags.ignore_links = True
	doc.flags.ignore_permissions = True
	doc.flags.legacy_import = True


def upsert_legacy_sales_master(row: dict) -> dict:
	fields = _build_fields(row)
	trans_no = fields.get("trans_no")
	if not trans_no:
		return {"status": "skip", "trans_no": row.get("trans_no")}

	if frappe.db.exists(DOCTYPE, trans_no):
		doc = frappe.get_doc(DOCTYPE, trans_no)
		for key, value in fields.items():
			doc.set(key, value)
		_apply_legacy_flags(doc)
		doc.save(ignore_permissions=True)
		action = "updated"
	else:
		doc = frappe.get_doc({"doctype": DOCTYPE, **fields})
		_apply_legacy_flags(doc)
		doc.insert(ignore_permissions=True)
		action = "created"

	return {"status": action, "trans_no": trans_no, "name": doc.name}


def _count_existing(trans_nos: list[str]) -> int:
	existing = 0
	chunk_size = 1000
	for i in range(0, len(trans_nos), chunk_size):
		chunk = trans_nos[i : i + chunk_size]
		existing += len(frappe.get_all(DOCTYPE, filters={"name": ["in", chunk]}, pluck="name"))
	return existing


def _preview_stats(rows: list[dict]) -> dict:
	trans_nos = [row["trans_no"] for row in rows]
	existing = _count_existing(trans_nos)

	# Resolve rates from a sample so preview stays fast on large files.
	sample = rows[:500]
	resolved_visits = 0
	resolved_patients = 0
	resolved_admissions = 0
	resolved_branches = 0
	for row in sample:
		visit_num = row.get("visit_num") or ""
		if visit_num:
			visit = _resolve_patient_visit(visit_num)
			if visit:
				resolved_visits += 1
				if frappe.db.get_value("Patient Visit", visit, "patient"):
					resolved_patients += 1
		admission_num = row.get("admission_num") or ""
		if admission_num and _resolve_inpatient_admission(admission_num):
			resolved_admissions += 1
		if row.get("branch_num") and _resolve_cost_center(row.get("branch_num")):
			resolved_branches += 1

	return {
		"existing_records": existing,
		"new_records": len(rows) - existing,
		"sample_size": len(sample),
		"resolved_visits": resolved_visits,
		"resolved_patients": resolved_patients,
		"resolved_admissions": resolved_admissions,
		"resolved_branches": resolved_branches,
		"sample_trans_nos": [row.get("trans_no") for row in rows[:5]],
	}


def _cache_file_path() -> str:
	folder = frappe.get_site_path("private", "files", "data_migration_cache")
	os.makedirs(folder, exist_ok=True)
	return os.path.join(folder, "legacy_sales_master_import.json")


def parse_and_cache_excel(file_url: str) -> dict:
	"""Parse all sheets and cache rows on disk (Redis truncates ~65k of this payload)."""
	rows, sheet_row_counts = _parse_excel_rows(file_url)
	by_key = {row["trans_no"]: row for row in rows}
	row_keys = list(by_key.keys())

	# Keep only fields needed for upsert.
	slim: dict[str, dict] = {}
	keep = set(EXCEL_HEADER_MAP.values()) | {
		"trans_no",
		"visit_num",
		"admission_num",
		"branch_num",
		"from_branch",
		"to_branch",
		"to_branch_num",
		"cr_id",
		"up_id",
		"ap_id",
		"ck_id",
		"_datemode",
	}
	for key, row in by_key.items():
		slim[key] = {k: v for k, v in row.items() if k in keep}

	cache_path = _cache_file_path()
	with open(cache_path, "w", encoding="utf-8") as handle:
		json.dump({"row_keys": row_keys, "rows": slim}, handle, default=str)

	frappe.cache().set_value(CACHE_KEYS["file_url"], file_url, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(CACHE_KEYS["cache_path"], cache_path, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(CACHE_KEYS["total_rows"], len(row_keys), expires_in_sec=CACHE_TTL)
	# Keep a short sample of keys in Redis for progress UI only — full list lives on disk.
	frappe.cache().set_value(
		CACHE_KEYS["row_keys"],
		row_keys[:20],
		expires_in_sec=CACHE_TTL,
	)

	preview = _preview_stats(list(by_key.values()))
	return {
		"excel_rows": len(by_key),
		"raw_excel_rows": sum(sheet_row_counts.values()),
		"sheets": list(sheet_row_counts.keys()),
		"sheet_row_counts": sheet_row_counts,
		**preview,
	}


def _load_cached_payload() -> tuple[list[str], dict[str, dict]]:
	cache_path = frappe.cache().get_value(CACHE_KEYS["cache_path"]) or _cache_file_path()
	if not cache_path or not os.path.exists(cache_path):
		return [], {}
	with open(cache_path, encoding="utf-8") as handle:
		payload = json.load(handle)
	if not isinstance(payload, dict):
		return [], {}
	row_keys = payload.get("row_keys") or []
	rows = payload.get("rows") or {}
	if not isinstance(row_keys, list):
		row_keys = list(rows.keys()) if isinstance(rows, dict) else []
	if not isinstance(rows, dict):
		rows = {}
	return row_keys, rows


@frappe.whitelist()
def preview_legacy_sales_master_import(file_url: str) -> dict:
	_require_admin()
	if not (file_url or "").strip():
		frappe.throw(_("Please upload the SALES_DATA_MASTER Excel file."))
	return parse_and_cache_excel(file_url)


def run_legacy_sales_master_import_batch(*, offset: int = 0) -> dict:
	row_keys, rows_by_key = _load_cached_payload()
	if not row_keys or not rows_by_key:
		return {"processed": offset, "done": True, "batch_count": 0}

	batch_keys = row_keys[offset : offset + LEGACY_SALES_MASTER_IMPORT_BATCH_SIZE]
	if not batch_keys:
		return {"processed": offset, "done": True, "batch_count": 0}

	created = updated = skipped = 0
	errors: list[str] = []

	for key in batch_keys:
		row = rows_by_key.get(key) or {}
		try:
			result = upsert_legacy_sales_master(row)
			status = result.get("status")
			if status == "created":
				created += 1
			elif status == "updated":
				updated += 1
			else:
				skipped += 1
		except Exception:
			errors.append(f"{key}: {frappe.get_traceback()}")
			frappe.log_error(title=f"SALES_DATA_MASTER import failed: {key}")

	frappe.db.commit()
	processed = offset + len(batch_keys)
	return {
		"processed": processed,
		"done": len(batch_keys) < LEGACY_SALES_MASTER_IMPORT_BATCH_SIZE,
		"batch_count": len(batch_keys),
		"created": created,
		"updated": updated,
		"skipped": skipped,
		"errors": len(errors),
		"total_rows": len(row_keys),
	}
