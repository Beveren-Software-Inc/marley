"""Import Oracle PATIENT_WARNING_MESSAGES Excel into Warning Message records."""

from __future__ import annotations

import json
from datetime import date, datetime
from typing import Any

import frappe
from frappe import _
from frappe.utils import cint, get_datetime

from healthcare.api.discharge_checklist_import import _resolve_cost_center
from healthcare.api.patient_info_import import (
	_cell_text,
	_clean_oracle_num,
	_excel_file_path,
	_require_admin,
)

PATIENT_WARNING_MESSAGE_IMPORT_BATCH_SIZE = 500
CACHE_TTL = 7200
CACHE_KEYS = {
	"file_url": "healthcare:data_migration:patient_warning_message_import:file_url",
	"row_keys": "healthcare:data_migration:patient_warning_message_import:row_keys",
	"rows": "healthcare:data_migration:patient_warning_message_import:rows",
}

EXCEL_HEADER_MAP = {
	"WARNING_MESSAGE_NUM": "warning_message_num",
	"PATIENT_NUM": "patient_num",
	"WARNING_MESSAGE": "warning_message",
	"HIGH_RISK_TEXT": "high_risk_text",
	"WARNING_MESSAGE_TYPE": "warning_message_type",
	"WARNING_MESSAGE_CLASS": "warning_message_class",
	"BRANCH_NUM": "branch_num",
	"CR_ID": "cr_id",
	"CR_DATE": "cr_date",
	"UP_ID": "up_id",
	"UP_DATE": "up_date",
	"STA_FLG": "sta_flg",
}


def _normalize_header(cell: Any) -> str:
	if cell is None:
		return ""
	text = str(cell).strip().upper().replace(" ", "_")
	return EXCEL_HEADER_MAP.get(text, text.lower())


def _oracle_text(value: Any) -> str | None:
	if value in (None, ""):
		return None
	cleaned = _clean_oracle_num(value)
	if cleaned:
		return cleaned
	text = _cell_text(value)
	return text or None


def _warning_html(value: Any) -> str | None:
	text = _cell_text(value)
	if not text:
		return None
	if "<" in text and ">" in text:
		return text
	return text.replace("\n", "<br>")


def _parse_posting_datetime(value: Any) -> datetime | None:
	if value in (None, ""):
		return None
	if isinstance(value, datetime):
		return value
	if isinstance(value, date):
		return datetime.combine(value, datetime.min.time())
	text = str(value).strip()
	if not text:
		return None
	try:
		return get_datetime(text)
	except Exception:
		return None


def _yn_to_check(value: Any) -> int:
	if value is None or value == "":
		return 0
	if isinstance(value, (int, float)):
		return 1 if int(value) != 0 else 0
	text = str(value).strip().upper()
	return 1 if text in ("Y", "YES", "1", "TRUE", "T") else 0


def _resolve_patient(file_no: Any) -> str | None:
	patient = _clean_oracle_num(file_no)
	if not patient:
		return None
	if frappe.db.exists("Patient", patient):
		return patient
	return None


def _existing_warning_message_name(trans_id: str) -> str | None:
	if frappe.db.exists("Warning Message", trans_id):
		return trans_id
	return frappe.db.get_value("Warning Message", {"trans_id": trans_id}, "name")


def _build_warning_message_fields(row: dict) -> dict[str, Any]:
	trans_id = row["warning_message_num"]
	patient = _resolve_patient(row.get("patient_num"))
	warning_html = _warning_html(row.get("warning_message"))

	fields: dict[str, Any] = {
		"trans_id": trans_id,
		"warning": warning_html,
		"sta_flg": _yn_to_check(row.get("sta_flg")),
	}

	if patient:
		fields["patient"] = patient
		fields["type_of_warning"] = "Medical"
		fields["from_patient"] = 1
		fields["reference_doc"] = "Patient"
		fields["reference_name"] = patient
	else:
		fields["type_of_warning"] = "Organisation"

	warning_type = _oracle_text(row.get("warning_message_type"))
	if warning_type:
		fields["warning_message_type"] = warning_type
		if warning_type.lower() == "allergy":
			fields["is_allergy"] = 1

	warning_class = _oracle_text(row.get("warning_message_class"))
	if warning_class:
		fields["warning_message_class"] = warning_class

	high_risk = _cell_text(row.get("high_risk_text"))
	if high_risk:
		fields["high_risk_text"] = high_risk

	posting_dt = _parse_posting_datetime(row.get("cr_date"))
	if posting_dt:
		fields["posting_date"] = posting_dt

	cost_center = _resolve_cost_center(row.get("branch_num"))
	if cost_center:
		fields["cost_center"] = cost_center

	return {key: value for key, value in fields.items() if value not in (None, "")}


def upsert_warning_message_from_row(row: dict) -> dict:
	trans_id = row.get("warning_message_num") or ""
	if not trans_id:
		return {"status": "skip_no_trans_id"}

	fields = _build_warning_message_fields(row)

	existing_name = _existing_warning_message_name(trans_id)
	if existing_name:
		doc = frappe.get_doc("Warning Message", existing_name)
		action = "updated"
	else:
		doc = frappe.get_doc({"doctype": "Warning Message", "name": trans_id, **fields})
		action = "created"

	if existing_name:
		for key, value in fields.items():
			if key == "trans_id":
				continue
			doc.set(key, value)

	doc.flags.ignore_validate = True
	doc.flags.ignore_mandatory = True
	doc.flags.ignore_links = True
	doc.flags.ignore_permissions = True
	doc.flags.legacy_import = True

	if existing_name:
		doc.save(ignore_permissions=True)
	else:
		doc.insert(ignore_permissions=True)

	return {
		"status": action,
		"trans_id": trans_id,
		"name": doc.name,
		"patient": fields.get("patient"),
		"type_of_warning": fields.get("type_of_warning"),
	}


def _parse_sheet_rows(ws) -> list[dict]:
	rows_iter = ws.iter_rows(values_only=True)
	try:
		header_row = next(rows_iter)
	except StopIteration:
		return []

	headers = [_normalize_header(h) for h in header_row]
	parsed: list[dict] = []
	for raw in rows_iter:
		if not raw or all(cell is None or str(cell).strip() == "" for cell in raw):
			continue
		row: dict[str, Any] = {}
		for idx, key in enumerate(headers):
			if not key or idx >= len(raw):
				continue
			row[key] = raw[idx]

		trans_id = _clean_oracle_num(row.get("warning_message_num"))
		if not trans_id:
			continue
		row["warning_message_num"] = trans_id
		parsed.append(row)
	return parsed


def _parse_excel_rows(file_url: str) -> tuple[list[dict], dict[str, int]]:
	try:
		import openpyxl
	except ImportError:
		frappe.throw(
			_(
				"openpyxl is required to read Excel files. Install it in the bench environment: pip install openpyxl"
			)
		)

	path = _excel_file_path(file_url)
	wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
	by_key: dict[str, dict] = {}
	sheet_row_counts: dict[str, int] = {}
	try:
		for sheet_name in wb.sheetnames:
			sheet_rows = _parse_sheet_rows(wb[sheet_name])
			sheet_row_counts[sheet_name] = len(sheet_rows)
			for row in sheet_rows:
				by_key[row["warning_message_num"]] = row
	finally:
		wb.close()
	return list(by_key.values()), sheet_row_counts


def _preview_counts(rows: list[dict]) -> dict:
	existing = 0
	resolved_patients = 0
	unresolved_patients = 0
	empty_warning_rows = 0
	organisation_rows = 0

	for row in rows:
		if not _cell_text(row.get("warning_message")):
			empty_warning_rows += 1
		patient_num = _clean_oracle_num(row.get("patient_num"))
		if not patient_num:
			organisation_rows += 1
		elif _resolve_patient(patient_num):
			resolved_patients += 1
		else:
			unresolved_patients += 1
		if _existing_warning_message_name(row.get("warning_message_num") or ""):
			existing += 1

	return {
		"existing_warnings": existing,
		"new_warnings": len(rows) - existing,
		"empty_warning_rows": empty_warning_rows,
		"resolved_patients": resolved_patients,
		"unresolved_patients": unresolved_patients,
		"organisation_rows": organisation_rows,
		"sample_trans_ids": [row.get("warning_message_num") for row in rows[:5]],
	}


def parse_and_cache_excel(file_url: str) -> dict:
	rows, sheet_row_counts = _parse_excel_rows(file_url)
	row_keys = [row["warning_message_num"] for row in rows]
	by_key = {row["warning_message_num"]: row for row in rows}

	frappe.cache().set_value(CACHE_KEYS["file_url"], file_url, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(CACHE_KEYS["row_keys"], row_keys, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(
		CACHE_KEYS["rows"],
		json.dumps(by_key, default=str),
		expires_in_sec=CACHE_TTL,
	)

	preview = _preview_counts(rows)
	raw_row_total = sum(sheet_row_counts.values())
	return {
		"excel_rows": len(rows),
		"raw_excel_rows": raw_row_total,
		"sheets": list(sheet_row_counts.keys()),
		"sheet_row_counts": sheet_row_counts,
		**preview,
	}


def _load_cached_rows() -> dict[str, dict]:
	raw = frappe.cache().get_value(CACHE_KEYS["rows"])
	if not raw:
		return {}
	if isinstance(raw, dict):
		return raw
	return json.loads(raw)


@frappe.whitelist()
def preview_patient_warning_message_import(file_url: str) -> dict:
	_require_admin()
	if not (file_url or "").strip():
		frappe.throw(_("Please upload the PATIENT_WARNING_MESSAGES Excel file."))
	return parse_and_cache_excel(file_url)


def run_patient_warning_message_import_batch(*, offset: int = 0) -> dict:
	row_keys = frappe.cache().get_value(CACHE_KEYS["row_keys"]) or []
	rows_by_key = _load_cached_rows()
	if not row_keys or not rows_by_key:
		return {"processed": offset, "done": True, "batch_count": 0}

	batch_keys = row_keys[offset : offset + PATIENT_WARNING_MESSAGE_IMPORT_BATCH_SIZE]
	if not batch_keys:
		return {"processed": offset, "done": True, "batch_count": 0}

	created = updated = 0
	errors: list[str] = []

	for key in batch_keys:
		row = rows_by_key.get(key) or {}
		try:
			result = upsert_warning_message_from_row(row)
			status = result.get("status")
			if status == "created":
				created += 1
			elif status == "updated":
				updated += 1
		except Exception:
			errors.append(f"{key}: {frappe.get_traceback()}")
			frappe.log_error(title=f"PATIENT_WARNING_MESSAGES import failed: {key}")

	frappe.db.commit()
	processed = offset + len(batch_keys)
	return {
		"processed": processed,
		"done": len(batch_keys) < PATIENT_WARNING_MESSAGE_IMPORT_BATCH_SIZE,
		"batch_count": len(batch_keys),
		"created": created,
		"updated": updated,
		"errors": len(errors),
	}
