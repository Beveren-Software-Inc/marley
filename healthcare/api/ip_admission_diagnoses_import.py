"""Import Oracle IP_ADMISSION_DIAGNOSES Excel into Medical Diagnosis Entry rows (IP)."""

from __future__ import annotations

import json
import re
from datetime import date, datetime, time
from typing import Any

import frappe
from frappe import _
from frappe.utils import cint

from healthcare.api.discharge_checklist_import import _resolve_cost_center
from healthcare.api.lab_test_legacy_import import _format_legacy_date_str
from healthcare.api.patient_info_import import (
	_cell_text,
	_clean_oracle_num,
	_excel_file_path,
	_require_admin,
)
from healthcare.api.visit_diagnoses_op_import import (
	_apply_patient_fetch_fields,
	_legacy_data_datetime,
	_parse_posting_datetime,
	_truncate_data,
)
from healthcare.api.visit_diagnosis_sync import _resolve_inpatient_admission

IP_ADMISSION_DIAGNOSES_IMPORT_BATCH_SIZE = 200
CACHE_TTL = 7200
CACHE_KEYS = {
	"file_url": "healthcare:data_migration:ip_admission_diagnoses_import:file_url",
	"row_keys": "healthcare:data_migration:ip_admission_diagnoses_import:row_keys",
	"rows": "healthcare:data_migration:ip_admission_diagnoses_import:rows",
}

EXCEL_HEADER_MAP = {
	"TRANS_NUM": "trans_num",
	"ADMISSION_NUM": "admission_num",
	"DIAGNOSES_FLAG": "diagnoses_flag",
	"DIAGNOSES_DESC": "diagnoses_desc",
	"DIAGNOSIS_DATE": "diagnoses_date",
	"DIAGNOSES_DATE": "diagnoses_date",
	"FROM_TIME": "diagnoses_time_from",
	"DIAGNOSES_TIME": "diagnoses_time_from",
	"TO_TIME": "diagnoses_time_to",
	"DIAGNOSES_TIME_TO": "diagnoses_time_to",
	"USER_NAME": "user_name",
	"COST_CENTER": "cost_center",
	"BRANCH_NUM": "cost_center",
	"CR_ID": "cr_id",
	"CR_DATE": "cr_date",
	"UP_ID": "up_id",
	"UP_DATE": "up_date",
	"OLD_NO": "old_no",
	"ADMISSION_NUM_OLD": "admission_num_old",
}


def _normalize_header(cell: Any) -> str:
	if cell is None:
		return ""
	text = str(cell).strip().upper().replace(" ", "_")
	return EXCEL_HEADER_MAP.get(text, text.lower())


def _check_flag(value: Any) -> int:
	if value is None or value == "":
		return 0
	if isinstance(value, (int, float)):
		return 1 if int(value) != 0 else 0
	text = str(value).strip().upper()
	return 1 if text in ("1", "Y", "YES", "TRUE", "T") else 0


def _parse_time_text(value: Any) -> time | None:
	text = _cell_text(value).replace(" ", "").lower()
	if not text:
		return None
	for fmt in ("%I:%M%p", "%I:%M:%S%p", "%H:%M:%S", "%H:%M"):
		try:
			return datetime.strptime(text, fmt).time()
		except ValueError:
			continue
	match = re.match(r"^(\d{1,2}):(\d{2})$", text)
	if match:
		return time(int(match.group(1)), int(match.group(2)))
	return None


def _merge_date_time(date_value: Any, time_value: Any) -> datetime | None:
	base = _parse_posting_datetime(date_value)
	if not base:
		return None
	parsed_time = _parse_time_text(time_value)
	if not parsed_time:
		return base
	return datetime.combine(base.date(), parsed_time)


def _writing_diagnosis_time(from_value: Any, to_value: Any) -> str | None:
	from_text = _cell_text(from_value)
	to_text = _cell_text(to_value)
	if from_text and to_text:
		combined = f"{from_text} - {to_text}"
	elif from_text:
		combined = from_text
	elif to_text:
		combined = to_text
	else:
		return None
	return _truncate_data(combined)


def _legacy_trans_num(trans_num: str) -> str:
	return f"IPDX/{trans_num}"


def _parse_sheet_rows(ws) -> list[dict]:
	rows_iter = ws.iter_rows(values_only=True)
	try:
		header_row = next(rows_iter)
	except StopIteration:
		return []

	headers = [_normalize_header(h) for h in header_row]
	parsed: list[dict] = []
	for raw in rows_iter:
		if not raw or all(cell is None or str(cell).strip() == "" for cell in raw):
			continue
		row: dict[str, Any] = {}
		for idx, key in enumerate(headers):
			if not key or idx >= len(raw):
				continue
			row[key] = raw[idx]

		trans_num = _clean_oracle_num(row.get("trans_num"))
		if not trans_num:
			continue

		row["trans_num"] = trans_num
		row["legacy_trans_num"] = _legacy_trans_num(trans_num)
		row["admission_num"] = _clean_oracle_num(row.get("admission_num"))
		row["admission_num_old"] = _clean_oracle_num(row.get("admission_num_old"))
		row["diagnoses_desc"] = _cell_text(row.get("diagnoses_desc"))
		row["diagnoses_date"] = row.get("diagnoses_date")
		row["diagnoses_time_from"] = row.get("diagnoses_time_from")
		row["diagnoses_time_to"] = row.get("diagnoses_time_to")
		row["user_name"] = _cell_text(row.get("user_name"))
		row["cost_center_raw"] = row.get("cost_center")
		row["cr_id"] = _clean_oracle_num(row.get("cr_id"))
		row["up_id"] = _clean_oracle_num(row.get("up_id"))
		row["cr_date"] = row.get("cr_date")
		row["up_date"] = row.get("up_date")
		row["old_no"] = _clean_oracle_num(row.get("old_no"))
		parsed.append(row)
	return parsed


def _parse_excel_rows(file_url: str) -> tuple[list[dict], dict[str, int]]:
	try:
		import openpyxl
	except ImportError:
		frappe.throw(
			_("openpyxl is required to read Excel files. Install it in the bench environment: pip install openpyxl")
		)

	path = _excel_file_path(file_url)
	wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
	by_key: dict[str, dict] = {}
	sheet_row_counts: dict[str, int] = {}
	try:
		for sheet_name in wb.sheetnames:
			sheet_rows = _parse_sheet_rows(wb[sheet_name])
			sheet_row_counts[sheet_name] = len(sheet_rows)
			for row in sheet_rows:
				by_key[row["legacy_trans_num"]] = row
	finally:
		wb.close()
	return list(by_key.values()), sheet_row_counts


def _patient_from_admission(admission_name: str | None) -> str | None:
	if not admission_name:
		return None
	return frappe.db.get_value("Inpatient Admission", admission_name, "patient")


def _build_entry_fields(row: dict) -> dict:
	admission_num = row.get("admission_num") or ""
	admission = _resolve_inpatient_admission(admission_num) if admission_num else None
	if not admission and row.get("admission_num_old"):
		admission = _resolve_inpatient_admission(row["admission_num_old"])

	patient = _patient_from_admission(admission)
	cost_center = _resolve_cost_center(row.get("cost_center_raw"))

	posting_dt = _merge_date_time(row.get("diagnoses_date"), row.get("diagnoses_time_from"))
	if not posting_dt:
		posting_dt = _parse_posting_datetime(row.get("cr_date"))

	fields: dict[str, Any] = {
		"trans_num": row["legacy_trans_num"],
		"source": "IP",
		"inpatient_admission": admission,
		"patient": patient,
		"diagnoses_flag": _check_flag(row.get("diagnoses_flag")),
		"details": row.get("diagnoses_desc") or None,
		"cost_center": cost_center,
		"cr_id": row.get("cr_id") or None,
		"up_id": row.get("up_id") or None,
	}
	writing_time = _writing_diagnosis_time(
		row.get("diagnoses_time_from"), row.get("diagnoses_time_to")
	)
	if writing_time:
		fields["writing_diagnosis_time"] = writing_time
	if row.get("user_name"):
		fields["practitioner_name"] = _truncate_data(row.get("user_name"))

	cr_date = _legacy_data_datetime(row.get("cr_date"))
	if cr_date:
		fields["cd_date"] = cr_date
	up_date = _legacy_data_datetime(row.get("up_date"))
	if up_date:
		fields["up_date"] = up_date
	if posting_dt:
		fields["posting_date"] = posting_dt
		fields["diagnoses_time"] = posting_dt

	if patient:
		_apply_patient_fetch_fields(fields, patient)

	return {key: value for key, value in fields.items() if value not in (None, "")}


def _find_existing_entry(legacy_trans_num: str) -> str | None:
	rows = frappe.get_all(
		"Medical Diagnosis Entry",
		filters={"trans_num": legacy_trans_num},
		pluck="name",
		limit=1,
	)
	return rows[0] if rows else None


def upsert_medical_diagnosis_entry_from_row(row: dict) -> dict:
	legacy_trans_num = row.get("legacy_trans_num")
	if not legacy_trans_num:
		return {"status": "skip_no_trans_num"}

	if not row.get("diagnoses_desc"):
		return {"status": "skip_no_details", "trans_num": legacy_trans_num}

	fields = _build_entry_fields(row)
	existing = _find_existing_entry(legacy_trans_num)
	if existing:
		doc = frappe.get_doc("Medical Diagnosis Entry", existing)
		for key, value in fields.items():
			if key == "trans_num":
				continue
			doc.set(key, value)
		action = "updated"
	else:
		doc = frappe.new_doc("Medical Diagnosis Entry")
		doc.update(fields)
		action = "created"

	doc.flags.ignore_validate = True
	doc.flags.ignore_mandatory = True
	doc.flags.ignore_links = True
	doc.flags.legacy_import = True

	if existing:
		doc.save(ignore_permissions=True)
	else:
		doc.insert(ignore_permissions=True)

	return {
		"status": action,
		"trans_num": legacy_trans_num,
		"name": doc.name,
		"admission_resolved": bool(fields.get("inpatient_admission")),
	}


def _preview_counts(rows: list[dict]) -> dict:
	existing = 0
	resolved_admissions = 0
	skip_no_details = 0
	unresolved_admissions = 0

	for row in rows:
		if not row.get("diagnoses_desc"):
			skip_no_details += 1
		admission_num = row.get("admission_num") or ""
		admission = _resolve_inpatient_admission(admission_num) if admission_num else None
		if not admission and row.get("admission_num_old"):
			admission = _resolve_inpatient_admission(row["admission_num_old"])
		if admission:
			resolved_admissions += 1
		elif admission_num:
			unresolved_admissions += 1
		if _find_existing_entry(row["legacy_trans_num"]):
			existing += 1

	return {
		"existing_entries": existing,
		"new_entries": len(rows) - existing,
		"skip_no_details": skip_no_details,
		"resolved_admissions": resolved_admissions,
		"unresolved_admissions": unresolved_admissions,
		"with_admission_num": sum(1 for row in rows if row.get("admission_num")),
	}


def parse_and_cache_excel(file_url: str) -> dict:
	rows, sheet_row_counts = _parse_excel_rows(file_url)
	row_keys = [row["legacy_trans_num"] for row in rows]
	by_key = {row["legacy_trans_num"]: row for row in rows}

	frappe.cache().set_value(CACHE_KEYS["file_url"], file_url, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(CACHE_KEYS["row_keys"], row_keys, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(
		CACHE_KEYS["rows"],
		json.dumps(by_key, default=str),
		expires_in_sec=CACHE_TTL,
	)

	preview = _preview_counts(rows)
	raw_row_total = sum(sheet_row_counts.values())
	return {
		"excel_rows": len(rows),
		"raw_excel_rows": raw_row_total,
		"sheets": list(sheet_row_counts.keys()),
		"sheet_row_counts": sheet_row_counts,
		**preview,
		"sample_trans_nums": row_keys[:5],
	}


def _load_cached_rows() -> dict[str, dict]:
	raw = frappe.cache().get_value(CACHE_KEYS["rows"])
	if not raw:
		return {}
	if isinstance(raw, dict):
		return raw
	return json.loads(raw)


@frappe.whitelist()
def preview_ip_admission_diagnoses_import(file_url: str) -> dict:
	_require_admin()
	if not (file_url or "").strip():
		frappe.throw(_("Please upload the IP_ADMISSION_DIAGNOSES Excel file."))
	return parse_and_cache_excel(file_url)


def run_ip_admission_diagnoses_import_batch(*, offset: int = 0) -> dict:
	row_keys = frappe.cache().get_value(CACHE_KEYS["row_keys"]) or []
	rows_by_key = _load_cached_rows()
	if not row_keys or not rows_by_key:
		return {"processed": offset, "done": True, "batch_count": 0}

	batch_keys = row_keys[offset : offset + IP_ADMISSION_DIAGNOSES_IMPORT_BATCH_SIZE]
	if not batch_keys:
		return {"processed": offset, "done": True, "batch_count": 0}

	created = updated = skip_no_details = 0
	admissions_resolved = 0
	errors: list[str] = []

	for key in batch_keys:
		row = rows_by_key.get(key) or {}
		try:
			result = upsert_medical_diagnosis_entry_from_row(row)
			status = result.get("status")
			if status == "created":
				created += 1
			elif status == "updated":
				updated += 1
			elif status == "skip_no_details":
				skip_no_details += 1
			if result.get("admission_resolved"):
				admissions_resolved += 1
		except Exception:
			errors.append(f"{key}: {frappe.get_traceback()}")
			frappe.log_error(title=f"IP_ADMISSION_DIAGNOSES import failed: {key}")

	frappe.db.commit()
	processed = offset + len(batch_keys)
	return {
		"processed": processed,
		"done": len(batch_keys) < IP_ADMISSION_DIAGNOSES_IMPORT_BATCH_SIZE,
		"batch_count": len(batch_keys),
		"created": created,
		"updated": updated,
		"skip_no_details": skip_no_details,
		"admissions_resolved": admissions_resolved,
		"errors": len(errors),
	}
