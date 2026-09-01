"""Import Oracle ECT_00_01 Excel into ECT Details rows."""

from __future__ import annotations

import json
from datetime import datetime, time
from typing import Any

import frappe
from frappe import _
from frappe.utils import get_time

from healthcare.api.discharge_checklist_import import _resolve_cost_center
from healthcare.api.ect_details import _default_ect_template
from healthcare.api.patient_info_import import (
	_cell_text,
	_clean_oracle_num,
	_excel_file_path,
	_parse_date_value,
	_parse_datetime_value,
	_require_admin,
)
from healthcare.api.patient_visit_import import (
	_resolve_practitioner_by_doctors_id,
	ensure_patient_for_legacy_import,
)
from healthcare.api.visit_diagnosis_sync import _resolve_inpatient_admission, _resolve_patient_visit

DOCTYPE = "ECT Details"
ECT_DETAILS_IMPORT_BATCH_SIZE = 500
CACHE_TTL = 7200
CACHE_KEYS = {
	"file_url": "healthcare:data_migration:ect_details_import:file_url",
	"row_keys": "healthcare:data_migration:ect_details_import:row_keys",
	"rows": "healthcare:data_migration:ect_details_import:rows",
}

EXCEL_HEADER_MAP = {
	"TRANS_NUM": "trans_num",
	"SR_NUM": "sr_num",
	"TRANS_DATE": "trans_date",
	"TRANS_TIME": "trans_time",
	"PATIENT_NUM": "patient_num",
	"INV_NUM": "inv_num",
	"IP_OP_SOURCE": "ip_op_source",
	"VISIT_NUM": "visit_num",
	"ADMISSION_NUM": "admission_num",
	"GTCS_SECONDS": "gtcs_seconds",
	"ENERGY_VALUE": "energy_value",
	"PERCENTAGE": "percentage",
	"IS_SUCCESS": "is_success",
	"IS_REPEATED": "is_repeated",
	"IS_VITALS": "is_vitals",
	"IS_ECG": "is_ecg",
	"BP_1": "bp_1",
	"BP_2": "bp_2",
	"PSYCH_DOC": "psych_doc",
	"ANAEST_DOC": "anaest_doc",
	"EXTRA_REMARKS": "extra_remarks",
	"IS_LOCK": "is_lock",
	"BRANCH_NUM": "branch_num",
	"CR_ID": "cr_id",
	"CR_DATE": "cr_date",
	"UP_ID": "up_id",
	"UP_DATE": "up_date",
	"NOTES_CR_ID": "notes_cr_id",
	"NOTES_CR_DATE": "notes_cr_date",
	"NOTES_REMARKS": "notes_remarks",
	"REMARKS_CR_ID": "remarks_cr_id",
	"REMARKS_CR_DATE": "remarks_cr_date",
	"GTCS_SECONDS2": "gtcs_seconds2",
	"GTCS_SECONDS3": "gtcs_seconds3",
	"ENERGY_VALUE2": "energy_value2",
	"ENERGY_VALUE3": "energy_value3",
	"ASSIST_DOC": "assist_doc",
	"BP_1_2": "bp_1_2",
	"BP_2_2": "bp_2_2",
	"STRENGTH1": "strength1",
	"STRENGTH2": "strength2",
	"STRENGTH3": "strength3",
	"NEXT_PLAN_DATE": "next_plan_date",
	"RECEP_REMARKS": "recep_remarks",
	"P_DETAIL": "p_detail",
	"S_DETAIL": "s_detail",
}


def _normalize_header(cell: Any) -> str:
	if cell is None:
		return ""
	text = str(cell).strip().upper().replace(" ", "_")
	return EXCEL_HEADER_MAP.get(text, text.lower())


def _format_legacy_datetime(value: Any) -> str | None:
	dt = _parse_datetime_value(value)
	if dt:
		return dt.strftime("%Y-%m-%d %H:%M:%S")
	parsed_date = _parse_date_value(value)
	if parsed_date:
		return parsed_date.strftime("%Y-%m-%d")
	text = _cell_text(value)
	return text or None


def _parse_date_field(value: Any):
	dt = _parse_datetime_value(value)
	if dt:
		return dt.date()
	return _parse_date_value(value)


def _parse_time_field(value: Any) -> str | None:
	if value in (None, ""):
		return None
	if isinstance(value, datetime):
		return value.strftime("%H:%M:%S")
	if isinstance(value, time):
		return value.strftime("%H:%M:%S")
	text = _cell_text(value)
	if not text:
		return None
	try:
		return get_time(text).strftime("%H:%M:%S")
	except Exception:
		return text


def _normalize_yes_no(value: Any) -> str | None:
	text = _cell_text(value).upper()
	return text or None


def _to_check(value: Any) -> int:
	text = _cell_text(value).strip().upper()
	return 1 if text in {"1", "Y", "YES", "TRUE"} else 0


def _resolve_patient_existing(patient_num: Any) -> str | None:
	patient = _clean_oracle_num(patient_num)
	if not patient:
		return None
	if frappe.db.exists("Patient", patient):
		return patient
	return None


def _ensure_patient(patient_num: Any) -> str | None:
	patient = _clean_oracle_num(patient_num)
	if not patient:
		return None
	if frappe.db.exists("Patient", patient):
		return patient
	result = ensure_patient_for_legacy_import(patient)
	if result.get("status") in ("existing", "created"):
		return patient
	return None


def _resolve_reference(row: dict[str, Any], patient: str | None) -> tuple[str | None, str | None]:
	visit_num = _clean_oracle_num(row.get("visit_num"))
	if visit_num:
		patient_visit = _resolve_patient_visit(visit_num, patient)
		if patient_visit:
			return "Patient Visit", patient_visit

	admission_num = _clean_oracle_num(row.get("admission_num"))
	if admission_num:
		admission = _resolve_inpatient_admission(admission_num, patient)
		if admission:
			return "Inpatient Admission", admission

	return None, None


def _practitioner_name(practitioner: str | None) -> str | None:
	if not practitioner:
		return None
	return frappe.db.get_value("Healthcare Practitioner", practitioner, "practitioner_name")


def _parse_sheet_rows(ws) -> list[dict]:
	rows_iter = ws.iter_rows(values_only=True)
	try:
		header_row = next(rows_iter)
	except StopIteration:
		return []

	headers = [_normalize_header(h) for h in header_row]
	parsed: list[dict] = []
	for raw in rows_iter:
		if not raw or all(cell is None or str(cell).strip() == "" for cell in raw):
			continue
		row: dict[str, Any] = {}
		for idx, key in enumerate(headers):
			if not key or idx >= len(raw):
				continue
			row[key] = raw[idx]

		trans_num = _clean_oracle_num(row.get("trans_num"))
		if not trans_num:
			continue

		row["trans_num"] = trans_num
		row["sr_num"] = _clean_oracle_num(row.get("sr_num")) or None
		row["patient_num"] = _clean_oracle_num(row.get("patient_num")) or None
		row["visit_num"] = _clean_oracle_num(row.get("visit_num")) or None
		row["admission_num"] = _clean_oracle_num(row.get("admission_num")) or None
		row["cr_id"] = _clean_oracle_num(row.get("cr_id")) or None
		row["up_id"] = _clean_oracle_num(row.get("up_id")) or None
		row["notes_cr_id"] = _clean_oracle_num(row.get("notes_cr_id")) or None
		row["remarks_cr_id"] = _clean_oracle_num(row.get("remarks_cr_id")) or None
		parsed.append(row)
	return parsed


def _parse_excel_rows(file_url: str) -> tuple[list[dict], dict[str, int]]:
	try:
		import openpyxl
	except ImportError:
		frappe.throw(
			_(
				"openpyxl is required to read Excel files. Install it in the bench environment: pip install openpyxl"
			)
		)

	path = _excel_file_path(file_url)
	wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
	by_key: dict[str, dict] = {}
	sheet_row_counts: dict[str, int] = {}
	try:
		for sheet_name in wb.sheetnames:
			sheet_rows = _parse_sheet_rows(wb[sheet_name])
			sheet_row_counts[sheet_name] = len(sheet_rows)
			for row in sheet_rows:
				by_key[row["trans_num"]] = row
	finally:
		wb.close()
	return list(by_key.values()), sheet_row_counts


def _build_fields(row: dict[str, Any]) -> dict[str, Any]:
	patient = _ensure_patient(row.get("patient_num"))
	if not patient:
		return {}

	reference_doctype, reference_name = _resolve_reference(row, patient)
	cost_center = _resolve_cost_center(row.get("branch_num"))
	psychology_doctor = _resolve_practitioner_by_doctors_id(row.get("psych_doc"))
	anaesthetic_doctor = _resolve_practitioner_by_doctors_id(row.get("anaest_doc"))

	fields: dict[str, Any] = {
		"trans_num": row["trans_num"],
		"sr_num": row.get("sr_num"),
		"patient": patient,
		"template": _default_ect_template(),
		"inv_num": _cell_text(row.get("inv_num")) or None,
		"ip_op_source": _cell_text(row.get("ip_op_source")) or None,
		"visit_num": row.get("visit_num"),
		"admission_num": row.get("admission_num"),
		"duration": row.get("gtcs_seconds"),
		"energy": _cell_text(row.get("energy_value")) or None,
		"_age": row.get("percentage"),
		"success": _normalize_yes_no(row.get("is_success")),
		"repeated": _normalize_yes_no(row.get("is_repeated")),
		"vitals": _normalize_yes_no(row.get("is_vitals")),
		"ecg": _normalize_yes_no(row.get("is_ecg")),
		"bp_1": _cell_text(row.get("bp_1")) or None,
		"bp_2": _cell_text(row.get("bp_2")) or None,
		"psychiatrist": _cell_text(row.get("psych_doc")) or None,
		"anathesiologist": _cell_text(row.get("anaest_doc")) or None,
		"assist_doctor": _cell_text(row.get("assist_doc")) or None,
		"extra_remarks": _cell_text(row.get("extra_remarks")) or None,
		"is_lock": _to_check(row.get("is_lock")),
		"cr_id": row.get("cr_id"),
		"up_id": row.get("up_id"),
		"notes_cr_id": row.get("notes_cr_id"),
		"remarks_cr_id": row.get("remarks_cr_id"),
		"note_remarks": _cell_text(row.get("notes_remarks")) or None,
		"gtcs_seconds2": row.get("gtcs_seconds2"),
		"gtcs_seconds3": row.get("gtcs_seconds3"),
		"energy_value2": _cell_text(row.get("energy_value2")) or None,
		"energy_value3": _cell_text(row.get("energy_value3")) or None,
		"bp_1_2": _cell_text(row.get("bp_1_2")) or None,
		"bp_2_2": _cell_text(row.get("bp_2_2")) or None,
		"strength1": _cell_text(row.get("strength1")) or None,
		"strength2": _cell_text(row.get("strength2")) or None,
		"strength3": _cell_text(row.get("strength3")) or None,
		"recep_remarks": _cell_text(row.get("recep_remarks")) or None,
		"p_detail": _cell_text(row.get("p_detail")) or None,
		"s_detail": _cell_text(row.get("s_detail")) or None,
	}

	trans_date = _parse_date_field(row.get("trans_date"))
	if trans_date:
		fields["date"] = trans_date

	trans_time = _parse_time_field(row.get("trans_time"))
	if trans_time:
		fields["time"] = trans_time

	next_plan_date = _parse_date_field(row.get("next_plan_date"))
	if next_plan_date:
		fields["next_plan_date"] = next_plan_date

	cr_date = _format_legacy_datetime(row.get("cr_date"))
	if cr_date:
		fields["cr_date"] = cr_date

	up_date = _format_legacy_datetime(row.get("up_date"))
	if up_date:
		fields["up_date"] = up_date

	notes_cr_date = _format_legacy_datetime(row.get("notes_cr_date"))
	if notes_cr_date:
		fields["notes_cr_date"] = notes_cr_date

	remarks_cr_date = _format_legacy_datetime(row.get("remarks_cr_date"))
	if remarks_cr_date:
		fields["remarks_cr_date"] = remarks_cr_date

	notes_dt = _parse_datetime_value(row.get("notes_cr_date"))
	if notes_dt:
		fields["n_date_and_time"] = notes_dt
		fields["ect_nurse_notes"] = _cell_text(row.get("notes_remarks")) or None

	remarks_dt = _parse_datetime_value(row.get("remarks_cr_date"))
	if remarks_dt:
		fields["date_and_time"] = remarks_dt

	if row.get("extra_remarks") not in (None, ""):
		fields["ect_doctors_notes"] = _cell_text(row.get("extra_remarks")) or None

	if psychology_doctor:
		fields["psychology_doctor"] = psychology_doctor
		fields["doctors_name"] = _practitioner_name(psychology_doctor)

	if anaesthetic_doctor:
		fields["anaesthetic_doctor"] = anaesthetic_doctor

	if cost_center:
		fields["cost_center"] = cost_center
		if frappe.get_meta("ECT Details").has_field("custom_cost_center"):
			fields["custom_cost_center"] = cost_center

	if reference_doctype and reference_name:
		fields["reference_doctype"] = reference_doctype
		fields["reference_name"] = reference_name

	return {key: value for key, value in fields.items() if value not in (None, "")}


def _preview_counts(rows: list[dict]) -> dict:
	existing = 0
	resolved_patients = 0
	resolved_admissions = 0
	resolved_visits = 0
	resolved_practitioners = 0
	op_rows = 0
	ip_rows = 0

	for row in rows:
		if frappe.db.exists(DOCTYPE, {"trans_num": row["trans_num"]}):
			existing += 1

		patient = _resolve_patient_existing(row.get("patient_num"))
		if patient:
			resolved_patients += 1

		if _clean_oracle_num(row.get("visit_num")):
			if _resolve_patient_visit(row.get("visit_num"), patient):
				resolved_visits += 1

		if _clean_oracle_num(row.get("admission_num")):
			if _resolve_inpatient_admission(row.get("admission_num"), patient):
				resolved_admissions += 1

		if (
			_resolve_practitioner_by_doctors_id(row.get("psych_doc"))
			or _resolve_practitioner_by_doctors_id(row.get("anaest_doc"))
			or _resolve_practitioner_by_doctors_id(row.get("assist_doc"))
		):
			resolved_practitioners += 1

		source = (_cell_text(row.get("ip_op_source")) or "").upper()
		if source == "OP":
			op_rows += 1
		elif source == "IP":
			ip_rows += 1

	return {
		"existing_records": existing,
		"resolved_patients": resolved_patients,
		"resolved_admissions": resolved_admissions,
		"resolved_visits": resolved_visits,
		"resolved_practitioners": resolved_practitioners,
		"op_rows": op_rows,
		"ip_rows": ip_rows,
		"sample_trans_nums": [row.get("trans_num") for row in rows[:5]],
	}


def parse_and_cache_excel(file_url: str) -> dict:
	rows, sheet_row_counts = _parse_excel_rows(file_url)
	row_keys = [row["trans_num"] for row in rows]
	by_key = {row["trans_num"]: row for row in rows}

	frappe.cache().set_value(CACHE_KEYS["file_url"], file_url, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(CACHE_KEYS["row_keys"], row_keys, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(
		CACHE_KEYS["rows"],
		json.dumps(by_key, default=str),
		expires_in_sec=CACHE_TTL,
	)

	preview = _preview_counts(rows)
	raw_row_total = sum(sheet_row_counts.values())
	return {
		"excel_rows": len(rows),
		"raw_excel_rows": raw_row_total,
		"sheets": list(sheet_row_counts.keys()),
		"sheet_row_counts": sheet_row_counts,
		**preview,
	}


def _load_cached_rows() -> dict[str, dict]:
	raw = frappe.cache().get_value(CACHE_KEYS["rows"])
	if not raw:
		return {}
	if isinstance(raw, dict):
		return raw
	return json.loads(raw)


def upsert_ect_detail(row: dict[str, Any]) -> dict:
	fields = _build_fields(row)
	if not fields.get("trans_num"):
		return {"status": "skip", "trans_num": row.get("trans_num")}

	trans_num = fields["trans_num"]
	existing_name = frappe.db.get_value(DOCTYPE, {"trans_num": trans_num}, "name")
	if existing_name:
		doc = frappe.get_doc(DOCTYPE, existing_name)
		for key, value in fields.items():
			doc.set(key, value)
		doc.flags.ignore_mandatory = True
		doc.flags.ignore_links = True
		doc.flags.ignore_validate = True
		doc.save(ignore_permissions=True)
		action = "updated"
	else:
		doc = frappe.get_doc({"doctype": DOCTYPE, **fields})
		doc.flags.ignore_mandatory = True
		doc.flags.ignore_links = True
		doc.flags.ignore_validate = True
		doc.insert(ignore_permissions=True)
		action = "created"

	return {"status": action, "trans_num": trans_num, "name": doc.name}


@frappe.whitelist()
def preview_ect_details_import(file_url: str) -> dict:
	_require_admin()
	if not (file_url or "").strip():
		frappe.throw(_("Please upload the ECT_00_01 Excel file."))
	return parse_and_cache_excel(file_url)


def run_ect_details_import_batch(*, offset: int = 0) -> dict:
	row_keys = frappe.cache().get_value(CACHE_KEYS["row_keys"]) or []
	rows_by_key = _load_cached_rows()
	if not row_keys or not rows_by_key:
		return {"processed": offset, "done": True, "batch_count": 0}

	batch_keys = row_keys[offset : offset + ECT_DETAILS_IMPORT_BATCH_SIZE]
	if not batch_keys:
		return {"processed": offset, "done": True, "batch_count": 0}

	created = updated = skipped = 0
	errors: list[str] = []

	for key in batch_keys:
		row = rows_by_key.get(key) or {}
		try:
			result = upsert_ect_detail(row)
			status = result.get("status")
			if status == "created":
				created += 1
			elif status == "updated":
				updated += 1
			else:
				skipped += 1
		except Exception:
			errors.append(f"{key}: {frappe.get_traceback()}")
			frappe.log_error(title=f"ECT_00_01 import failed: {key}")

	frappe.db.commit()
	processed = offset + len(batch_keys)
	return {
		"processed": processed,
		"done": len(batch_keys) < ECT_DETAILS_IMPORT_BATCH_SIZE,
		"batch_count": len(batch_keys),
		"created": created,
		"updated": updated,
		"skipped": skipped,
		"errors": len(errors),
	}
