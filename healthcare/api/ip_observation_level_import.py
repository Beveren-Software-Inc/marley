"""Import Oracle IP_OBSERVATION_LEVEL Excel directly into Observation rows."""

from __future__ import annotations

import json
from datetime import date, datetime
from typing import Any

import frappe
from frappe import _
from frappe.utils import cint, flt, getdate, get_datetime

from healthcare.api.discharge_checklist_import import _resolve_cost_center
from healthcare.api.patient_info_import import (
	_cell_text,
	_clean_oracle_num,
	_excel_file_path,
	_require_admin,
)
from healthcare.api.visit_diagnoses_op_import import _legacy_data_datetime
from healthcare.api.visit_diagnosis_sync import _resolve_inpatient_admission
from healthcare.healthcare.doctype.observation.observation import fill_patient_from_admission

IP_OBSERVATION_LEVEL_IMPORT_BATCH_SIZE = 200
CACHE_TTL = 7200
CACHE_KEYS = {
	"file_url": "healthcare:data_migration:ip_observation_level_import:file_url",
	"row_keys": "healthcare:data_migration:ip_observation_level_import:row_keys",
	"rows": "healthcare:data_migration:ip_observation_level_import:rows",
}

LEGACY_OBS_CODE_TO_LEVEL = {
	"1": "General Observation",
	"2": "Intermittent Observation",
	"3": "One to One (Within Eye Sight)",
	"4": "One to One (Within In Arm's Length)",
}

EXCEL_HEADER_MAP = {
	"TRANS_NUM": "trans_num",
	"ADMISSION_NUM": "admission_num",
	"OBS_CODE": "obs_code",
	"OBS_LEVEL": "obs_level",
	"NOTES": "notes",
	"START_DATE": "start_date",
	"DC_DATE": "dc_date",
	"CR_ID": "cr_id",
	"CR_DATE": "cr_date",
	"UP_ID": "up_id",
	"UP_DATE": "up_date",
	"BRANCH_NUM": "branch_num",
	"SERV_AMT": "serv_amt",
}


def _normalize_header(cell: Any) -> str:
	if cell is None:
		return ""
	text = str(cell).strip().upper().replace(" ", "_")
	return EXCEL_HEADER_MAP.get(text, text.lower())


def _normalize_obs_code(value: Any) -> str | None:
	code = _clean_oracle_num(value)
	if not code:
		return None
	if code.endswith(".0"):
		code = code[:-2]
	return code


def _resolve_observation_level(obs_code: Any) -> tuple[str | None, str | None]:
	code = _normalize_obs_code(obs_code)
	if not code:
		return None, None
	level_name = LEGACY_OBS_CODE_TO_LEVEL.get(code)
	if level_name and frappe.db.exists("Observation Level", level_name):
		return code, level_name
	return code, level_name


def _parse_start_date(value: Any):
	if value in (None, ""):
		return None
	if isinstance(value, datetime):
		return value.date()
	if isinstance(value, date):
		return value
	try:
		return getdate(value)
	except Exception:
		return None


def _parse_dc_datetime(value: Any):
	if value in (None, ""):
		return None
	if isinstance(value, datetime):
		return value
	try:
		return get_datetime(value)
	except Exception:
		return None


def _oracle_currency(value: Any) -> float | None:
	if value in (None, ""):
		return None
	text = str(value).strip().replace(",", "")
	try:
		return flt(text)
	except (TypeError, ValueError):
		return None


def _resolve_company(admission: str | None) -> str | None:
	if admission:
		company = frappe.db.get_value("Inpatient Admission", admission, "company")
		if company:
			return company
	return frappe.defaults.get_global_default("company")


def _level_defaults(level_name: str | None) -> dict[str, Any]:
	if not level_name or not frappe.db.exists("Observation Level", level_name):
		return {}
	row = frappe.db.get_value(
		"Observation Level",
		level_name,
		["rate", "interval"],
		as_dict=True,
	) or {}
	out: dict[str, Any] = {}
	if row.get("interval"):
		out["duration"] = str(row.interval).strip()
	if flt(row.get("rate")):
		out["amount"] = flt(row.rate)
	return out


def _parse_sheet_rows(ws) -> list[dict]:
	rows_iter = ws.iter_rows(values_only=True)
	try:
		header_row = next(rows_iter)
	except StopIteration:
		return []

	headers = [_normalize_header(h) for h in header_row]
	parsed: list[dict] = []
	for raw in rows_iter:
		if not raw or all(cell is None or str(cell).strip() == "" for cell in raw):
			continue
		row: dict[str, Any] = {}
		for idx, key in enumerate(headers):
			if not key or idx >= len(raw):
				continue
			row[key] = raw[idx]

		trans_num = _clean_oracle_num(row.get("trans_num"))
		if not trans_num:
			continue

		row["trans_num"] = trans_num
		row["admission_num"] = _clean_oracle_num(row.get("admission_num"))
		row["notes"] = _cell_text(row.get("notes"))
		row["obs_level"] = _cell_text(row.get("obs_level"))
		row["cr_id"] = _clean_oracle_num(row.get("cr_id"))
		row["up_id"] = _clean_oracle_num(row.get("up_id"))
		parsed.append(row)
	return parsed


def _parse_excel_rows(file_url: str) -> tuple[list[dict], dict[str, int]]:
	try:
		import openpyxl
	except ImportError:
		frappe.throw(
			_("openpyxl is required to read Excel files. Install it in the bench environment: pip install openpyxl")
		)

	path = _excel_file_path(file_url)
	wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
	all_rows: list[dict] = []
	sheet_row_counts: dict[str, int] = {}
	try:
		for sheet_name in wb.sheetnames:
			sheet_rows = _parse_sheet_rows(wb[sheet_name])
			sheet_row_counts[sheet_name] = len(sheet_rows)
			all_rows.extend(sheet_rows)
	finally:
		wb.close()
	return all_rows, sheet_row_counts


def _build_observation_fields(row: dict) -> tuple[dict[str, Any], dict[str, int]]:
	admission_num = row.get("admission_num") or ""
	admission = _resolve_inpatient_admission(admission_num)
	if not admission:
		return {}, {"skip_no_admission": 1}

	obs_code, observation_level = _resolve_observation_level(row.get("obs_code"))
	start_date = _parse_start_date(row.get("start_date"))
	dc_date = _parse_dc_datetime(row.get("dc_date"))
	serv_amount = _oracle_currency(row.get("serv_amt"))

	fields: dict[str, Any] = {
		"trans_no": row["trans_num"],
		"admission_no": admission,
		"company": _resolve_company(admission),
		"naming_series": "HLC-OBS-.YYYY.-",
		"status": "Registered",
		"obs_code": obs_code,
		"observation_level": observation_level or "",
		"obs_level": row.get("obs_level") or observation_level or "",
		"note": row.get("notes") or None,
		"start_date": start_date,
		"dc_date": dc_date,
		"posting_date": start_date,
		"cost_center": _resolve_cost_center(row.get("branch_num")),
		"serv_amount": serv_amount,
		"cr_id": row.get("cr_id") or None,
		"up_id": row.get("up_id") or None,
	}
	cr_date = _legacy_data_datetime(row.get("cr_date"))
	if cr_date:
		fields["cr_date"] = cr_date
	up_date = _legacy_data_datetime(row.get("up_date"))
	if up_date:
		fields["up_date"] = up_date

	level_defaults = _level_defaults(observation_level)
	for key, value in level_defaults.items():
		fields.setdefault(key, value)
	if serv_amount is not None:
		fields["amount"] = serv_amount
	elif "amount" not in fields:
		fields.pop("amount", None)

	return {
		key: value
		for key, value in fields.items()
		if value not in (None, "") or key in {"obs_code", "observation_level", "obs_level"}
	}, {}


def _apply_legacy_import_flags(doc) -> None:
	doc.flags.ignore_validate = True
	doc.flags.ignore_mandatory = True
	doc.flags.ignore_links = True
	doc.flags.ignore_permissions = True
	doc.flags.legacy_import = True
	doc.flags.skip_care_episode_guard = True


def _persist_and_submit_observation(doc, *, existing: bool) -> bool:
	_apply_legacy_import_flags(doc)

	if existing:
		doc.save(ignore_permissions=True)
	else:
		doc.insert(ignore_permissions=True)

	fill_patient_from_admission(doc)
	if not doc.patient:
		frappe.log_error(
			title=f"Observation submit skipped (no patient): {doc.name}",
			message=frappe.as_json({"admission_no": doc.get("admission_no"), "trans_no": doc.get("trans_no")}),
		)
		return False

	if doc.docstatus != 0:
		doc.reload()
	if doc.docstatus != 0:
		return False

	try:
		_apply_legacy_import_flags(doc)
		doc.submit()
		doc.reload()
		return doc.docstatus == 1
	except Exception:
		frappe.log_error(
			title=f"Observation submit failed: {doc.name}",
			message=frappe.get_traceback(),
		)
		return False


def upsert_observation_from_row(row: dict) -> dict:
	trans_num = row.get("trans_num")
	if not trans_num:
		return {"status": "skip_no_trans_num"}

	fields, side_stats = _build_observation_fields(row)
	if side_stats.get("skip_no_admission"):
		return {"status": "skip_no_admission", "trans_num": trans_num}
	if not fields:
		return {"status": "skip_error", "trans_num": trans_num}

	existing_name = frappe.db.exists("Observation", trans_num)
	if existing_name:
		doc = frappe.get_doc("Observation", existing_name)
		if doc.docstatus != 0:
			frappe.delete_doc("Observation", existing_name, force=True, ignore_permissions=True)
			existing_name = None

	obs_code, observation_level = _resolve_observation_level(row.get("obs_code"))
	level_fields = {
		"obs_code": obs_code or "",
		"observation_level": observation_level or "",
		"obs_level": row.get("obs_level") or observation_level or "",
	}

	if existing_name:
		doc = frappe.get_doc("Observation", existing_name)
		for key, value in fields.items():
			if key == "trans_no":
				continue
			doc.set(key, value)
		for key, value in level_fields.items():
			doc.set(key, value)
		action = "updated"
		existing = True
	else:
		doc = frappe.new_doc("Observation")
		doc.update({**fields, **level_fields})
		action = "created"
		existing = False

	submitted = _persist_and_submit_observation(doc, existing=existing)
	return {
		"status": action,
		"trans_num": trans_num,
		"name": doc.name,
		"submitted": submitted,
	}


def _preview_counts(rows: list[dict]) -> dict:
	existing = sum(1 for row in rows if frappe.db.exists("Observation", row["trans_num"]))
	resolved_admissions = 0
	unresolved_admissions = 0
	obs_code_counts: dict[str, int] = {}
	unknown_obs_codes: dict[str, int] = {}

	for row in rows:
		code = _normalize_obs_code(row.get("obs_code")) or "?"
		obs_code_counts[code] = obs_code_counts.get(code, 0) + 1
		_, level = _resolve_observation_level(row.get("obs_code"))
		if not level:
			unknown_obs_codes[code] = unknown_obs_codes.get(code, 0) + 1

		admission_num = row.get("admission_num") or ""
		if not admission_num:
			continue
		if _resolve_inpatient_admission(admission_num):
			resolved_admissions += 1
		else:
			unresolved_admissions += 1

	return {
		"existing_observations": existing,
		"new_observations": len(rows) - existing,
		"resolved_admissions": resolved_admissions,
		"unresolved_admissions": unresolved_admissions,
		"obs_code_counts": obs_code_counts,
		"unknown_obs_code_rows": sum(unknown_obs_codes.values()),
		"unknown_obs_codes": list(unknown_obs_codes.keys())[:10],
	}


def parse_and_cache_excel(file_url: str) -> dict:
	rows, sheet_row_counts = _parse_excel_rows(file_url)
	by_key = {row["trans_num"]: row for row in rows}
	row_keys = list(by_key.keys())

	frappe.cache().set_value(CACHE_KEYS["file_url"], file_url, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(CACHE_KEYS["row_keys"], row_keys, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(
		CACHE_KEYS["rows"],
		json.dumps(by_key, default=str),
		expires_in_sec=CACHE_TTL,
	)

	preview = _preview_counts(list(by_key.values()))
	raw_row_total = sum(sheet_row_counts.values())
	return {
		"excel_rows": len(by_key),
		"raw_excel_rows": raw_row_total,
		"sheets": list(sheet_row_counts.keys()),
		"sheet_row_counts": sheet_row_counts,
		**preview,
		"sample_trans_nums": row_keys[:5],
	}


def _load_cached_rows() -> dict[str, dict]:
	raw = frappe.cache().get_value(CACHE_KEYS["rows"])
	if not raw:
		return {}
	if isinstance(raw, dict):
		return raw
	return json.loads(raw)


@frappe.whitelist()
def preview_ip_observation_level_import(file_url: str) -> dict:
	_require_admin()
	if not (file_url or "").strip():
		frappe.throw(_("Please upload the IP_OBSERVATION_LEVEL Excel file."))
	return parse_and_cache_excel(file_url)


def run_ip_observation_level_import_batch(*, offset: int = 0) -> dict:
	row_keys = frappe.cache().get_value(CACHE_KEYS["row_keys"]) or []
	rows_by_key = _load_cached_rows()
	if not row_keys or not rows_by_key:
		return {"processed": offset, "done": True, "batch_count": 0}

	batch_keys = row_keys[offset : offset + IP_OBSERVATION_LEVEL_IMPORT_BATCH_SIZE]
	if not batch_keys:
		return {"processed": offset, "done": True, "batch_count": 0}

	created = updated = submitted = skip_no_admission = 0
	errors: list[str] = []

	for key in batch_keys:
		row = rows_by_key.get(key) or {}
		try:
			result = upsert_observation_from_row(row)
			status = result.get("status")
			if status == "created":
				created += 1
			elif status == "updated":
				updated += 1
			elif status == "skip_no_admission":
				skip_no_admission += 1
			if result.get("submitted"):
				submitted += 1
		except Exception:
			errors.append(f"{key}: {frappe.get_traceback()}")
			frappe.log_error(title=f"IP_OBSERVATION_LEVEL import failed: {key}")

	frappe.db.commit()
	processed = offset + len(batch_keys)
	return {
		"processed": processed,
		"done": len(batch_keys) < IP_OBSERVATION_LEVEL_IMPORT_BATCH_SIZE,
		"batch_count": len(batch_keys),
		"created": created,
		"updated": updated,
		"submitted": submitted,
		"skip_no_admission": skip_no_admission,
		"errors": len(errors),
	}
