"""Import Oracle PATIENT_SICK_LEAVE Excel into Patient Sick Leave rows.

This is separate from Sick Leave (PATIENT_SICK_LEAVE_01) imported via patient_sick_leave_import.py.
"""

from __future__ import annotations

from datetime import datetime
from typing import Any

import frappe
from frappe import _
from frappe.utils import getdate

from healthcare.api.discharge_checklist_import import _resolve_cost_center
from healthcare.api.ip_admission_discharge_import import _resolve_practitioner_by_doctors_id
from healthcare.api.patient_info_import import (
	_cell_text,
	_clean_oracle_num,
	_excel_file_path,
	_parse_date_value,
	_parse_datetime_value,
	_require_admin,
)
from healthcare.api.patient_visit_practitioner import practitioner_name_from_link
from healthcare.api.visit_diagnosis_sync import _resolve_inpatient_admission

DOCTYPE = "Patient Sick Leave"
LEGACY_NOTE = (
	"Legacy Oracle PATIENT_SICK_LEAVE export. Separate from Sick Leave (PATIENT_SICK_LEAVE_01)."
)

EXCEL_HEADER_MAP = {
	"TRANS_NO": "trans_no",
	"START_DATE": "start_date",
	"END_DATE": "end_date",
	"PATIENT_FILE_NO": "patient_file_no",
	"CREATE_USER_ID": "create_user_id",
	"CREATE_DATE": "create_date",
	"SICK_FLAG": "sick_flag",
	"ACC_PATIENT": "acc_patient",
	"ATTEND_TIME": "attend_time",
	"CONSULT_TIME": "consult_time",
	"ATTEND_DATE": "attend_date",
	"FIT_FLAG": "fit_flag",
	"UNFIT_FLAG": "unfit_flag",
	"LIGHT_DUTY": "light_duty",
	"NEEDS_FLAG": "needs_flag",
	"DIAGNOSIS": "diagnosis",
	"ADMITTE_DATE": "admitted_date",
	"DISCHARGE_DATE": "discharge_date",
	"DOCTOR_CD": "doctor_cd",
	"LEAVE_DATE": "leave_date",
	"TRANS_SOURCE": "trans_source",
	"VISIT_NUM": "visit_num",
	"BRANCH_NUM": "branch_num",
	"ADMISSION_NUM": "admission_num",
}


def _normalize_header(cell: Any) -> str:
	if cell is None:
		return ""
	text = str(cell).strip().upper().replace(" ", "_")
	return EXCEL_HEADER_MAP.get(text, text.lower())


def _trans_no_from_cell(value: Any) -> str:
	if value in (None, ""):
		return ""
	if isinstance(value, datetime):
		# Excel exports legacy TRANS_NO values as 1900-era dates.
		base = datetime(1899, 12, 30)
		return str((value - base).days)
	return _clean_oracle_num(value) or _cell_text(value)


def _legacy_flag(value: Any) -> str | None:
	if value in (None, ""):
		return None
	if isinstance(value, float) and value == int(value):
		return str(int(value))
	return _cell_text(value) or None


def _legacy_time(value: Any) -> str | None:
	if value in (None, ""):
		return None
	if isinstance(value, datetime):
		return value.strftime("%H:%M:%S")
	if isinstance(value, float):
		# Excel time fraction of day
		total_seconds = int(round(value * 24 * 60 * 60))
		hours = (total_seconds // 3600) % 24
		minutes = (total_seconds % 3600) // 60
		seconds = total_seconds % 60
		return f"{hours:02d}:{minutes:02d}:{seconds:02d}"
	return _cell_text(value) or None


def _format_legacy_datetime(value: Any) -> str | None:
	dt = _parse_datetime_value(value)
	if dt:
		return dt.strftime("%Y-%m-%d %H:%M:%S")
	parsed_date = _parse_date_value(value)
	if parsed_date:
		return parsed_date.strftime("%Y-%m-%d")
	text = _cell_text(value)
	return text or None


def _parse_date_field(value: Any):
	dt = _parse_datetime_value(value)
	if dt:
		return getdate(dt)
	parsed_date = _parse_date_value(value)
	if parsed_date:
		return parsed_date
	return None


def _resolve_patient(patient_file_no: Any) -> str | None:
	patient = _clean_oracle_num(patient_file_no)
	if not patient:
		return None
	if frappe.db.exists("Patient", patient):
		return patient
	return None


def _resolve_admission(admission_num: Any) -> str | None:
	admission = _clean_oracle_num(admission_num)
	if not admission:
		return None
	if frappe.db.exists("Inpatient Admission", admission):
		return admission
	return _resolve_inpatient_admission(admission)


def _resolve_patient_visit(visit_num: Any) -> str | None:
	visit = _clean_oracle_num(visit_num)
	if not visit:
		return None
	if frappe.db.exists("Patient Visit", visit):
		return visit
	return None


def _parse_sheet_rows(ws) -> list[dict]:
	rows_iter = ws.iter_rows(values_only=True)
	try:
		header_row = next(rows_iter)
	except StopIteration:
		return []

	headers = [_normalize_header(h) for h in header_row]
	parsed: list[dict] = []
	for raw in rows_iter:
		if not raw or all(cell is None or str(cell).strip() == "" for cell in raw):
			continue
		row: dict[str, Any] = {}
		for idx, key in enumerate(headers):
			if not key or idx >= len(raw):
				continue
			row[key] = raw[idx]

		trans_no = _trans_no_from_cell(row.get("trans_no"))
		patient_file_no = row.get("patient_file_no")
		if not trans_no or patient_file_no in (None, ""):
			continue
		row["trans_no"] = trans_no
		parsed.append(row)
	return parsed


def _parse_excel_rows(file_url: str) -> tuple[list[dict], dict[str, int]]:
	try:
		import openpyxl
	except ImportError:
		frappe.throw(
			_(
				"openpyxl is required to read Excel files. Install it in the bench environment: pip install openpyxl"
			)
		)

	path = _excel_file_path(file_url)
	wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
	by_key: dict[str, dict] = {}
	sheet_row_counts: dict[str, int] = {}
	try:
		for sheet_name in wb.sheetnames:
			sheet_rows = _parse_sheet_rows(wb[sheet_name])
			sheet_row_counts[sheet_name] = len(sheet_rows)
			for row in sheet_rows:
				by_key[row["trans_no"]] = row
	finally:
		wb.close()
	return list(by_key.values()), sheet_row_counts


def _build_fields(row: dict) -> dict[str, Any]:
	fields: dict[str, Any] = {
		"trans_no": row["trans_no"],
		"legacy_note": LEGACY_NOTE,
		"legacy": 1,
		"trans_source": _cell_text(row.get("trans_source")) or None,
		"diagnosis": _cell_text(row.get("diagnosis")) or None,
		"acc_patient": _legacy_flag(row.get("acc_patient")),
		"create_user_id": _clean_oracle_num(row.get("create_user_id")) or None,
	}

	for flag_field in ("sick_flag", "fit_flag", "unfit_flag", "light_duty", "needs_flag"):
		flag_value = _legacy_flag(row.get(flag_field))
		if flag_value is not None:
			fields[flag_field] = flag_value

	for date_field in ("start_date", "end_date", "leave_date", "admitted_date", "discharge_date", "attend_date"):
		parsed_date = _parse_date_field(row.get(date_field))
		if parsed_date:
			fields[date_field] = parsed_date

	for time_field in ("attend_time", "consult_time"):
		time_value = _legacy_time(row.get(time_field))
		if time_value:
			fields[time_field] = time_value

	create_date = _format_legacy_datetime(row.get("create_date"))
	if create_date:
		fields["create_date"] = create_date

	patient = _resolve_patient(row.get("patient_file_no"))
	if patient:
		fields["patient"] = patient
		patient_name = frappe.db.get_value("Patient", patient, "patient_name")
		if patient_name:
			fields["patient_name"] = patient_name

	admission = _resolve_admission(row.get("admission_num"))
	if admission:
		fields["admission"] = admission

	patient_visit = _resolve_patient_visit(row.get("visit_num"))
	if patient_visit:
		fields["patient_visit"] = patient_visit

	practitioner = _resolve_practitioner_by_doctors_id(row.get("doctor_cd"))
	if practitioner:
		fields["practitioner"] = practitioner
		practitioner_name = practitioner_name_from_link(practitioner)
		if practitioner_name:
			fields["practitioner_name"] = practitioner_name

	branch = _resolve_cost_center(row.get("branch_num"))
	if branch:
		fields["branch"] = branch

	return {key: value for key, value in fields.items() if value not in (None, "")}


def _preview_stats(rows: list[dict]) -> dict:
	existing = 0
	resolvable_patients = 0
	resolvable_admissions = 0
	resolvable_visits = 0
	resolvable_practitioners = 0
	for row in rows:
		if frappe.db.exists(DOCTYPE, row["trans_no"]):
			existing += 1
		if _resolve_patient(row.get("patient_file_no")):
			resolvable_patients += 1
		if _resolve_admission(row.get("admission_num")):
			resolvable_admissions += 1
		if _resolve_patient_visit(row.get("visit_num")):
			resolvable_visits += 1
		if _resolve_practitioner_by_doctors_id(row.get("doctor_cd")):
			resolvable_practitioners += 1

	return {
		"existing_records": existing,
		"resolvable_patients": resolvable_patients,
		"resolvable_admissions": resolvable_admissions,
		"resolvable_visits": resolvable_visits,
		"resolvable_practitioners": resolvable_practitioners,
		"sample_trans_nos": [row.get("trans_no") for row in rows[:5]],
	}


def upsert_patient_sick_leave(row: dict) -> dict:
	fields = _build_fields(row)
	if not fields.get("trans_no"):
		return {"status": "skip", "trans_no": row.get("trans_no")}

	trans_no = fields["trans_no"]
	if frappe.db.exists(DOCTYPE, trans_no):
		doc = frappe.get_doc(DOCTYPE, trans_no)
		for key, value in fields.items():
			doc.set(key, value)
		doc.flags.ignore_mandatory = True
		doc.flags.ignore_links = True
		doc.flags.ignore_validate = True
		doc.save(ignore_permissions=True)
		action = "updated"
	else:
		doc = frappe.get_doc({"doctype": DOCTYPE, **fields})
		doc.flags.ignore_mandatory = True
		doc.flags.ignore_links = True
		doc.flags.ignore_validate = True
		doc.insert(ignore_permissions=True)
		action = "created"

	return {"status": action, "trans_no": trans_no, "name": doc.name}


@frappe.whitelist()
def preview_patient_sick_leave_record_import(file_url: str) -> dict:
	_require_admin()
	if not (file_url or "").strip():
		frappe.throw(_("Please upload the PATIENT_SICK_LEAVE Excel file."))

	rows, sheet_row_counts = _parse_excel_rows(file_url)
	stats = _preview_stats(rows)
	raw_row_total = sum(sheet_row_counts.values())
	return {
		"excel_rows": len(rows),
		"raw_excel_rows": raw_row_total,
		"sheets": list(sheet_row_counts.keys()),
		"sheet_row_counts": sheet_row_counts,
		**stats,
	}


@frappe.whitelist()
def run_patient_sick_leave_record_import(file_url: str) -> dict:
	"""Import all rows synchronously into Patient Sick Leave (not Sick Leave 01)."""
	_require_admin()
	if not (file_url or "").strip():
		frappe.throw(_("Please upload the PATIENT_SICK_LEAVE Excel file."))

	rows, _ = _parse_excel_rows(file_url)
	created = updated = skipped = 0
	errors: list[str] = []

	for row in rows:
		try:
			result = upsert_patient_sick_leave(row)
			status = result.get("status")
			if status == "created":
				created += 1
			elif status == "updated":
				updated += 1
			else:
				skipped += 1
		except Exception:
			errors.append(f"{row.get('trans_no')}: {frappe.get_traceback()}")
			frappe.log_error(title=f"PATIENT_SICK_LEAVE import failed: {row.get('trans_no')}")

	frappe.db.commit()
	return {
		"ok": True,
		"total": len(rows),
		"created": created,
		"updated": updated,
		"skipped": skipped,
		"errors": len(errors),
	}
