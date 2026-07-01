"""Import Oracle IP_ADMISSION_PHY_EXAM Excel directly into Physical Examination rows."""

from __future__ import annotations

import json
from typing import Any

import frappe
from frappe import _
from frappe.utils import cint

from healthcare.api.discharge_checklist_import import _resolve_cost_center
from healthcare.api.patient_info_import import (
	_cell_text,
	_clean_oracle_num,
	_excel_file_path,
	_require_admin,
)
from healthcare.api.patient_visit_import import ensure_patient_for_legacy_import
from healthcare.api.visit_diagnoses_op_import import _legacy_data_datetime
from healthcare.api.visit_diagnosis_sync import _resolve_inpatient_admission

IP_ADMISSION_PHY_EXAM_IMPORT_BATCH_SIZE = 200
CACHE_TTL = 7200
CACHE_KEYS = {
	"file_url": "healthcare:data_migration:ip_admission_phy_exam_import:file_url",
	"row_keys": "healthcare:data_migration:ip_admission_phy_exam_import:row_keys",
	"rows": "healthcare:data_migration:ip_admission_phy_exam_import:rows",
}

EXCEL_HEADER_MAP = {
	"TRANS_NO": "trans_no",
	"TRANS_NUM": "trans_no",
	"PATIENT_NUM": "patient_num",
	"ADMISSION_NUM": "admission_num",
	"ADMISSION_NUM_OLD": "admission_num_old",
	"WEIGHT": "weight",
	"HEIGHT": "height",
	"BLOOD_PRESSURE": "blood_pressure",
	"PULSE": "pulse",
	"TEMPRESSURE": "temp_pressure",
	"TEMPERATURE": "temp_pressure",
	"RESP_RATE": "resp_rate",
	"PHYS_SIGNS": "phys_signs",
	"CVS_RESP": "cvs_resp",
	"CNS": "cns",
	"GIT": "git",
	"OTHERS": "others",
	"CR_ID": "cr_id",
	"CR_DATE_TIME": "cr_date_time",
	"CR_DATE": "cr_date",
	"UP_ID": "up_id",
	"UP_DATE": "up_date",
	"BRANCH_NUM": "branch_num",
	"BRANCH": "branch_num",
}


def _normalize_header(cell: Any) -> str:
	if cell is None:
		return ""
	text = str(cell).strip().upper().replace(" ", "_")
	return EXCEL_HEADER_MAP.get(text, text.lower())


def _oracle_data_value(value: Any) -> str | None:
	if value in (None, ""):
		return None
	if isinstance(value, float) and value == int(value):
		return str(int(value))
	if isinstance(value, int):
		return str(value)
	text = str(value).strip().replace(",", "")
	return text or None


def _text_value(value: Any) -> str | None:
	text = _cell_text(value)
	return text or None


def _resolve_patient(row: dict, admission: str) -> tuple[str | None, int]:
	patient = _clean_oracle_num(row.get("patient_num")) or None
	patients_created = 0
	if patient and not frappe.db.exists("Patient", patient):
		result = ensure_patient_for_legacy_import(patient)
		if result.get("status") == "created":
			patients_created = 1
		elif result.get("status") != "existing":
			patient = None
	if not patient and admission:
		patient = frappe.db.get_value("Inpatient Admission", admission, "patient")
	return patient, patients_created


def _legacy_cr_date(row: dict) -> str:
	for key in ("cr_date_time", "cr_date"):
		value = row.get(key)
		if value in (None, ""):
			continue
		formatted = _legacy_data_datetime(value)
		if formatted:
			return formatted
	return ""


def _parse_sheet_rows(ws) -> list[dict]:
	rows_iter = ws.iter_rows(values_only=True)
	try:
		header_row = next(rows_iter)
	except StopIteration:
		return []

	headers = [_normalize_header(h) for h in header_row]
	parsed: list[dict] = []
	for raw in rows_iter:
		if not raw or all(cell is None or str(cell).strip() == "" for cell in raw):
			continue
		row: dict[str, Any] = {}
		for idx, key in enumerate(headers):
			if not key or idx >= len(raw):
				continue
			row[key] = raw[idx]

		trans_no = _clean_oracle_num(row.get("trans_no"))
		if not trans_no:
			continue

		row["trans_no"] = trans_no
		row["admission_num"] = _clean_oracle_num(row.get("admission_num"))
		row["admission_num_old"] = _clean_oracle_num(row.get("admission_num_old"))
		row["patient_num"] = _clean_oracle_num(row.get("patient_num"))
		row["cr_id"] = _clean_oracle_num(row.get("cr_id"))
		row["up_id"] = _clean_oracle_num(row.get("up_id"))
		parsed.append(row)
	return parsed


def _parse_excel_rows(file_url: str) -> tuple[list[dict], dict[str, int]]:
	try:
		import openpyxl
	except ImportError:
		frappe.throw(
			_("openpyxl is required to read Excel files. Install it in the bench environment: pip install openpyxl")
		)

	path = _excel_file_path(file_url)
	wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
	all_rows: list[dict] = []
	sheet_row_counts: dict[str, int] = {}
	try:
		for sheet_name in wb.sheetnames:
			sheet_rows = _parse_sheet_rows(wb[sheet_name])
			sheet_row_counts[sheet_name] = len(sheet_rows)
			all_rows.extend(sheet_rows)
	finally:
		wb.close()
	return all_rows, sheet_row_counts


def _build_physical_examination_fields(row: dict) -> tuple[dict[str, Any], dict[str, int]]:
	stats = {"patients_created": 0}
	admission_num = row.get("admission_num") or row.get("admission_num_old") or ""
	admission = _resolve_inpatient_admission(admission_num, row.get("patient_num"))
	if not admission:
		return {}, {"skip_no_admission": 1}

	patient, patients_created = _resolve_patient(row, admission)
	stats["patients_created"] = patients_created
	if not patient:
		return {}, {"skip_no_patient": 1}

	patient_name = frappe.db.get_value("Patient", patient, "patient_name")

	fields: dict[str, Any] = {
		"trans_no": row["trans_no"],
		"patient": patient,
		"patient_name": patient_name,
		"inpatient_admission": admission,
		"cost_center": _resolve_cost_center(row.get("branch_num")),
		"weight": _oracle_data_value(row.get("weight")),
		"height": _oracle_data_value(row.get("height")),
		"blood_pressure": _oracle_data_value(row.get("blood_pressure")),
		"pulse": _oracle_data_value(row.get("pulse")),
		"temp_pressure": _oracle_data_value(row.get("temp_pressure")),
		"resp_rate": _oracle_data_value(row.get("resp_rate")),
		"skin_": _text_value(row.get("phys_signs")),
		"cvsresp": _text_value(row.get("cvs_resp")),
		"cnc": _text_value(row.get("cns")),
		"git": _text_value(row.get("git")),
		"others": _text_value(row.get("others")),
		"cr_id": row.get("cr_id") or None,
		"up_id": row.get("up_id") or None,
	}
	cr_date = _legacy_cr_date(row)
	if cr_date:
		fields["cr_date"] = cr_date
	up_date = _legacy_data_datetime(row.get("up_date"))
	if up_date:
		fields["up_date"] = up_date

	return {key: value for key, value in fields.items() if value not in (None, "")}, stats


def _apply_legacy_import_flags(doc) -> None:
	doc.flags.ignore_validate = True
	doc.flags.ignore_mandatory = True
	doc.flags.ignore_links = True
	doc.flags.ignore_permissions = True
	doc.flags.legacy_import = True
	doc.flags.skip_care_episode_guard = True


def _persist_physical_examination(doc, *, existing: bool) -> None:
	_apply_legacy_import_flags(doc)
	if existing:
		doc.save(ignore_permissions=True)
	else:
		doc.insert(ignore_permissions=True)


def upsert_physical_examination_from_row(row: dict) -> dict:
	trans_no = row.get("trans_no")
	if not trans_no:
		return {"status": "skip_no_trans_no"}

	fields, side_stats = _build_physical_examination_fields(row)
	if side_stats.get("skip_no_admission"):
		return {"status": "skip_no_admission", "trans_no": trans_no}
	if side_stats.get("skip_no_patient"):
		return {"status": "skip_no_patient", "trans_no": trans_no}
	if not fields:
		return {"status": "skip_error", "trans_no": trans_no}

	existing_name = frappe.db.exists("Physical Examination", trans_no)
	if existing_name:
		doc = frappe.get_doc("Physical Examination", existing_name)
		for key, value in fields.items():
			if key == "trans_no":
				continue
			doc.set(key, value)
		action = "updated"
		existing = True
	else:
		doc = frappe.new_doc("Physical Examination")
		doc.update(fields)
		action = "created"
		existing = False

	_persist_physical_examination(doc, existing=existing)
	return {
		"status": action,
		"trans_no": trans_no,
		"name": doc.name,
		"patients_created": side_stats.get("patients_created", 0),
	}


def _preview_counts(rows: list[dict]) -> dict:
	existing = sum(1 for row in rows if frappe.db.exists("Physical Examination", row["trans_no"]))
	resolved_admissions = 0
	unresolved_admissions = 0
	patient_nums = {row["patient_num"] for row in rows if row.get("patient_num")}
	missing_patients = [pn for pn in patient_nums if not frappe.db.exists("Patient", pn)]

	for row in rows:
		admission_num = row.get("admission_num") or row.get("admission_num_old") or ""
		if not admission_num:
			continue
		if _resolve_inpatient_admission(admission_num, row.get("patient_num")):
			resolved_admissions += 1
		else:
			unresolved_admissions += 1

	return {
		"existing_examinations": existing,
		"new_examinations": len(rows) - existing,
		"resolved_admissions": resolved_admissions,
		"unresolved_admissions": unresolved_admissions,
		"unique_patients": len(patient_nums),
		"patients_to_create": len(missing_patients),
		"sample_patients_to_create": missing_patients[:10],
	}


def parse_and_cache_excel(file_url: str) -> dict:
	rows, sheet_row_counts = _parse_excel_rows(file_url)
	by_key = {row["trans_no"]: row for row in rows}
	row_keys = list(by_key.keys())

	frappe.cache().set_value(CACHE_KEYS["file_url"], file_url, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(CACHE_KEYS["row_keys"], row_keys, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(
		CACHE_KEYS["rows"],
		json.dumps(by_key, default=str),
		expires_in_sec=CACHE_TTL,
	)

	preview = _preview_counts(list(by_key.values()))
	raw_row_total = sum(sheet_row_counts.values())
	return {
		"excel_rows": len(by_key),
		"raw_excel_rows": raw_row_total,
		"sheets": list(sheet_row_counts.keys()),
		"sheet_row_counts": sheet_row_counts,
		**preview,
		"sample_trans_nos": row_keys[:5],
	}


def _load_cached_rows() -> dict[str, dict]:
	raw = frappe.cache().get_value(CACHE_KEYS["rows"])
	if not raw:
		return {}
	if isinstance(raw, dict):
		return raw
	return json.loads(raw)


@frappe.whitelist()
def preview_ip_admission_phy_exam_import(file_url: str) -> dict:
	_require_admin()
	if not (file_url or "").strip():
		frappe.throw(_("Please upload the IP_ADMISSION_PHY_EXAM Excel file."))
	return parse_and_cache_excel(file_url)


def run_ip_admission_phy_exam_import_batch(*, offset: int = 0) -> dict:
	row_keys = frappe.cache().get_value(CACHE_KEYS["row_keys"]) or []
	rows_by_key = _load_cached_rows()
	if not row_keys or not rows_by_key:
		return {"processed": offset, "done": True, "batch_count": 0}

	batch_keys = row_keys[offset : offset + IP_ADMISSION_PHY_EXAM_IMPORT_BATCH_SIZE]
	if not batch_keys:
		return {"processed": offset, "done": True, "batch_count": 0}

	created = updated = skip_no_admission = skip_no_patient = patients_created = 0
	errors: list[str] = []

	for key in batch_keys:
		row = rows_by_key.get(key) or {}
		try:
			result = upsert_physical_examination_from_row(row)
			status = result.get("status")
			if status == "created":
				created += 1
			elif status == "updated":
				updated += 1
			elif status == "skip_no_admission":
				skip_no_admission += 1
			elif status == "skip_no_patient":
				skip_no_patient += 1
			patients_created += cint(result.get("patients_created"))
		except Exception:
			errors.append(f"{key}: {frappe.get_traceback()}")
			frappe.log_error(title=f"IP_ADMISSION_PHY_EXAM import failed: {key}")

	frappe.db.commit()
	processed = offset + len(batch_keys)
	return {
		"processed": processed,
		"done": len(batch_keys) < IP_ADMISSION_PHY_EXAM_IMPORT_BATCH_SIZE,
		"batch_count": len(batch_keys),
		"created": created,
		"updated": updated,
		"skip_no_admission": skip_no_admission,
		"skip_no_patient": skip_no_patient,
		"patients_created": patients_created,
		"errors": len(errors),
	}
