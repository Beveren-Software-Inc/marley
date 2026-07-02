"""Import Oracle IP_PATIENT_RELATIVES Excel into Inpatient Admission patient_relatives."""

from __future__ import annotations

import json
from datetime import date, datetime
from typing import Any

import frappe
from frappe import _
from frappe.utils import cint, getdate

from healthcare.api.patient_info_import import (
	_cell_text,
	_clean_oracle_num,
	_excel_file_path,
	_require_admin,
)
from healthcare.api.visit_diagnosis_sync import _resolve_inpatient_admission

IP_PATIENT_RELATIVES_IMPORT_BATCH_SIZE = 500
CACHE_TTL = 7200
CACHE_KEYS = {
	"file_url": "healthcare:data_migration:ip_patient_relatives_import:file_url",
	"admissions": "healthcare:data_migration:ip_patient_relatives_import:admissions",
	"grouped": "healthcare:data_migration:ip_patient_relatives_import:grouped",
}

EXCEL_HEADER_MAP = {
	"TRANS_NUM": "trans_num",
	"ADMISSION_NUM": "admission_num",
	"RELATIVE_RELATION": "relative_relation",
	"RELATIVE_NAME": "relative_name",
	"RELATIVE_ID_NUM": "relative_id_num",
	"ANY_REMARKS": "any_remarks",
	"FIELD_1": "field_1",
	"FIELD_2": "field_2",
	"FIELD_3": "field_3",
	"FIELD_4": "field_4",
	"FIELD_5": "field_5",
	"BRANCH_NUM": "branch_num",
	"CR_ID": "cr_id",
	"CR_DATE": "cr_date",
	"UP_ID": "up_id",
	"UP_DATE": "up_date",
}

RELATION_ALIASES = {
	"FATHER": "Father",
	"MOTHER": "Mother",
	"MONTHER": "Mother",
	"BROTHER": "Brother",
	"BROTHETR": "Brother",
	"SISTER": "Sister",
	"COUSIN": "Cousin",
	"HUSBAND": "Husband",
	"WIFE": "Wife",
	"SON": "Son",
	"DAUGHTER": "Daughter",
	"AUNT": "aunt",
	"UNCLE": "uncle",
	"FRIEND": "friend",
	"HOUSEMAID": "Housemaid",
	"HOUSE MAID": "House Maid",
}


def _normalize_header(cell: Any) -> str:
	if cell is None:
		return ""
	text = str(cell).strip().upper().replace(" ", "_")
	return EXCEL_HEADER_MAP.get(text, text.lower())


def _parse_date_value(value: Any):
	if value in (None, ""):
		return None
	if isinstance(value, datetime):
		return getdate(value)
	if isinstance(value, date):
		return value
	text = str(value).strip()
	if not text:
		return None
	try:
		return getdate(text)
	except Exception:
		return None


def _resolve_relationship(value: Any) -> str | None:
	text = _cell_text(value)
	if not text:
		return None

	upper = text.upper().replace(" ", "_")
	if upper in RELATION_ALIASES:
		candidate = RELATION_ALIASES[upper]
		if frappe.db.exists("Patient Relative Relationship", candidate):
			return candidate

	for candidate in (text, text.title(), upper.title()):
		if frappe.db.exists("Patient Relative Relationship", candidate):
			return candidate

	name = frappe.db.get_value(
		"Patient Relative Relationship",
		{"relationship": ["like", f"%{text}%"]},
		"name",
	)
	return name or text.title()


def _build_relative_fields(row: dict) -> dict[str, Any]:
	relative_id = _cell_text(row.get("relative_id_num"))
	remarks = _cell_text(row.get("any_remarks"))
	relationship = _resolve_relationship(row.get("relative_relation"))

	fields: dict[str, Any] = {
		"trans_no": row["trans_num"],
		"relative_name": _cell_text(row.get("relative_name")),
	}

	if relationship:
		fields["relationship_with_patient"] = relationship
	if relative_id:
		fields["relative_id_no"] = relative_id
		fields["cpr__id_no"] = relative_id
	if remarks:
		fields["any_remarks"] = remarks

	entered_date = _parse_date_value(row.get("cr_date"))
	if entered_date:
		fields["entered_date"] = entered_date

	return {key: value for key, value in fields.items() if value not in (None, "")}


def _parse_sheet_rows(ws) -> list[dict]:
	rows_iter = ws.iter_rows(values_only=True)
	try:
		header_row = next(rows_iter)
	except StopIteration:
		return []

	headers = [_normalize_header(h) for h in header_row]
	parsed: list[dict] = []
	for raw in rows_iter:
		if not raw or all(cell is None or str(cell).strip() == "" for cell in raw):
			continue
		row: dict[str, Any] = {}
		for idx, key in enumerate(headers):
			if not key or idx >= len(raw):
				continue
			row[key] = raw[idx]

		trans_num = _clean_oracle_num(row.get("trans_num"))
		admission_num = _clean_oracle_num(row.get("admission_num"))
		if not trans_num or not admission_num:
			continue

		row["trans_num"] = trans_num
		row["admission_num"] = admission_num
		parsed.append(row)
	return parsed


def _parse_excel_rows(file_url: str) -> tuple[list[dict], dict[str, int]]:
	try:
		import openpyxl
	except ImportError:
		frappe.throw(
			_(
				"openpyxl is required to read Excel files. Install it in the bench environment: pip install openpyxl"
			)
		)

	path = _excel_file_path(file_url)
	wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
	parsed: list[dict] = []
	sheet_row_counts: dict[str, int] = {}
	try:
		for sheet_name in wb.sheetnames:
			sheet_rows = _parse_sheet_rows(wb[sheet_name])
			sheet_row_counts[sheet_name] = len(sheet_rows)
			parsed.extend(sheet_rows)
	finally:
		wb.close()
	return parsed, sheet_row_counts


def _group_rows_by_admission(rows: list[dict]) -> dict[str, list[dict]]:
	grouped: dict[str, list[dict]] = {}
	for row in rows:
		key = row.get("admission_num") or ""
		if not key:
			continue
		grouped.setdefault(key, []).append(row)
	return grouped


def import_relatives_for_admission(admission_key: str, lines: list[dict]) -> dict:
	if not lines:
		return {"status": "skip_empty", "admission_key": admission_key}

	admission_name = _resolve_inpatient_admission(admission_key)
	if not admission_name:
		return {"status": "skip_no_admission", "admission_key": admission_key}

	doc = frappe.get_doc("Inpatient Admission", admission_name)
	existing_by_trans = {
		(row.trans_no or "").strip(): row for row in (doc.get("patient_relatives") or []) if row.trans_no
	}

	added = updated = skipped = 0
	for row in lines:
		fields = _build_relative_fields(row)
		if not fields.get("relative_name"):
			skipped += 1
			continue

		trans_no = fields.get("trans_no") or ""
		existing_row = existing_by_trans.get(trans_no)
		if existing_row:
			for key, value in fields.items():
				existing_row.set(key, value)
			updated += 1
		else:
			child = doc.append("patient_relatives", {})
			for key, value in fields.items():
				child.set(key, value)
			existing_by_trans[trans_no] = child
			added += 1

	if added == 0 and updated == 0:
		return {
			"status": "skip_no_lines",
			"admission_key": admission_key,
			"admission": admission_name,
			"skipped": skipped,
		}

	doc.flags.ignore_validate = True
	doc.flags.ignore_links = True
	doc.flags.ignore_permissions = True
	doc.flags.legacy_import = True
	doc.save(ignore_permissions=True)

	return {
		"status": "ok",
		"admission_key": admission_key,
		"admission": admission_name,
		"added": added,
		"updated": updated,
		"skipped": skipped,
	}


def parse_and_cache_excel(file_url: str) -> dict:
	rows, sheet_row_counts = _parse_excel_rows(file_url)
	grouped = _group_rows_by_admission(rows)
	admission_keys = sorted(grouped.keys())

	frappe.cache().set_value(CACHE_KEYS["file_url"], file_url, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(CACHE_KEYS["admissions"], admission_keys, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(
		CACHE_KEYS["grouped"],
		json.dumps(grouped, default=str),
		expires_in_sec=CACHE_TTL,
	)

	resolvable = sum(1 for key in admission_keys if _resolve_inpatient_admission(key))
	relative_lines = len(rows)
	return {
		"excel_rows": relative_lines,
		"raw_excel_rows": relative_lines,
		"sheets": list(sheet_row_counts.keys()),
		"sheet_row_counts": sheet_row_counts,
		"admissions": len(admission_keys),
		"relative_lines": relative_lines,
		"resolvable_admissions": resolvable,
		"unresolved_admissions": len(admission_keys) - resolvable,
		"sample_admissions": admission_keys[:5],
	}


def _load_cached_grouped() -> dict[str, list[dict]]:
	raw = frappe.cache().get_value(CACHE_KEYS["grouped"])
	if not raw:
		return {}
	if isinstance(raw, dict):
		return raw
	return json.loads(raw)


@frappe.whitelist()
def preview_ip_patient_relatives_import(file_url: str) -> dict:
	_require_admin()
	if not (file_url or "").strip():
		frappe.throw(_("Please upload the IP_PATIENT_RELATIVES Excel file."))
	return parse_and_cache_excel(file_url)


def run_ip_patient_relatives_import_batch(*, offset: int = 0) -> dict:
	admission_keys = frappe.cache().get_value(CACHE_KEYS["admissions"]) or []
	grouped = _load_cached_grouped()
	if not admission_keys or not grouped:
		return {"processed": offset, "done": True, "batch_count": 0}

	batch_keys = admission_keys[offset : offset + IP_PATIENT_RELATIVES_IMPORT_BATCH_SIZE]
	if not batch_keys:
		return {"processed": offset, "done": True, "batch_count": 0}

	ok = skip_no_admission = skip_no_lines = 0
	relatives_added = relatives_updated = relatives_skipped = 0
	errors: list[str] = []

	for key in batch_keys:
		lines = grouped.get(key) or []
		try:
			result = import_relatives_for_admission(key, lines)
			status = result.get("status")
			if status == "ok":
				ok += 1
				relatives_added += cint(result.get("added", 0))
				relatives_updated += cint(result.get("updated", 0))
				relatives_skipped += cint(result.get("skipped", 0))
			elif status == "skip_no_admission":
				skip_no_admission += 1
			elif status == "skip_no_lines":
				skip_no_lines += 1
				relatives_skipped += cint(result.get("skipped", 0))
		except Exception:
			errors.append(f"{key}: {frappe.get_traceback()}")
			frappe.log_error(title=f"IP_PATIENT_RELATIVES import failed: {key}")

	frappe.db.commit()
	processed = offset + len(batch_keys)
	return {
		"processed": processed,
		"done": len(batch_keys) < IP_PATIENT_RELATIVES_IMPORT_BATCH_SIZE,
		"batch_count": len(batch_keys),
		"ok": ok,
		"skip_no_admission": skip_no_admission,
		"skip_no_lines": skip_no_lines,
		"relatives_added": relatives_added,
		"relatives_updated": relatives_updated,
		"relatives_skipped": relatives_skipped,
		"errors": len(errors),
	}
