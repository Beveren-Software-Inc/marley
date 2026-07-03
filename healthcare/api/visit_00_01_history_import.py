"""Import Oracle VISIT_00_01_HISTORY Excel into VISIT_00_01_HISTORY rows."""

from __future__ import annotations

import json
from datetime import datetime
from typing import Any

import frappe
from frappe import _
from frappe.utils import getdate

from healthcare.api.patient_info_import import (
	_cell_text,
	_clean_oracle_num,
	_excel_file_path,
	_parse_date_value,
	_parse_datetime_value,
	_require_admin,
)
from healthcare.api.visit_diagnosis_sync import _resolve_patient_visit

DOCTYPE = "VISIT_00_01_HISTORY"
VISIT_00_01_HISTORY_IMPORT_BATCH_SIZE = 500
CACHE_TTL = 7200
CACHE_KEYS = {
	"file_url": "healthcare:data_migration:visit_00_01_history_import:file_url",
	"row_keys": "healthcare:data_migration:visit_00_01_history_import:row_keys",
	"rows": "healthcare:data_migration:visit_00_01_history_import:rows",
}

EXCEL_HEADER_MAP = {
	"VISIT_NUM": "visit_num",
	"VISIT_DATE": "visit_date",
	"VISIT_PATIENT_NUM": "visit_patient_num",
	"VISIT_DIAGNOSIS_DETAIL": "visit_diagnosis_detail",
	"CR_ID": "cr_id",
	"CR_DATE": "cr_date",
	"UP_ID": "up_id",
	"UP_DATE": "up_date",
}


def _normalize_header(cell: Any) -> str:
	if cell is None:
		return ""
	text = str(cell).strip().upper().replace(" ", "_")
	return EXCEL_HEADER_MAP.get(text, text.lower())


def _normalize_visit_num(value: Any) -> str:
	if value in (None, ""):
		return ""
	if isinstance(value, datetime):
		try:
			import openpyxl.utils.datetime

			return str(int(openpyxl.utils.datetime.to_excel(value)))
		except Exception:
			return value.strftime("%Y-%m-%d")
	if isinstance(value, float) and value == int(value):
		return str(int(value))
	return _clean_oracle_num(value)


def _format_legacy_datetime(value: Any) -> str | None:
	dt = _parse_datetime_value(value)
	if dt:
		return dt.strftime("%Y-%m-%d %H:%M:%S")
	parsed_date = _parse_date_value(value)
	if parsed_date:
		return parsed_date.strftime("%Y-%m-%d")
	text = _cell_text(value)
	return text or None


def _parse_date_field(value: Any):
	dt = _parse_datetime_value(value)
	if dt:
		return getdate(dt)
	parsed_date = _parse_date_value(value)
	if parsed_date:
		return parsed_date
	return None


def _resolve_patient(patient_num: Any) -> str | None:
	patient = _clean_oracle_num(patient_num)
	if not patient:
		return None
	if frappe.db.exists("Patient", patient):
		return patient
	return None


def _record_key_base(visit_num: str, cr_date: Any) -> str:
	cr_text = _format_legacy_datetime(cr_date) or "unknown"
	return f"{visit_num}-{cr_text.replace(' ', '-').replace(':', '-')}"


def _build_record_key(
	visit_num: str,
	cr_date: Any,
	*,
	duplicate_index: int = 0,
) -> str:
	base = _record_key_base(visit_num, cr_date)
	if duplicate_index <= 0:
		return base
	return f"{base}-{duplicate_index + 1}"


def _parse_sheet_rows(ws) -> list[dict]:
	rows_iter = ws.iter_rows(values_only=True)
	try:
		header_row = next(rows_iter)
	except StopIteration:
		return []

	headers = [_normalize_header(h) for h in header_row]
	parsed: list[dict] = []
	duplicate_counts: dict[str, int] = {}
	for raw in rows_iter:
		if not raw or all(cell is None or str(cell).strip() == "" for cell in raw):
			continue
		row: dict[str, Any] = {}
		for idx, key in enumerate(headers):
			if not key or idx >= len(raw):
				continue
			row[key] = raw[idx]

		visit_num = _normalize_visit_num(row.get("visit_num"))
		if not visit_num:
			continue
		row["visit_num"] = visit_num
		row["visit_patient_num"] = _clean_oracle_num(row.get("visit_patient_num")) or None
		row["cr_id"] = _clean_oracle_num(row.get("cr_id")) or None
		row["up_id"] = _clean_oracle_num(row.get("up_id")) or None

		base = _record_key_base(visit_num, row.get("cr_date"))
		dup_index = duplicate_counts.get(base, 0)
		duplicate_counts[base] = dup_index + 1
		row["record_key"] = _build_record_key(
			visit_num,
			row.get("cr_date"),
			duplicate_index=dup_index,
		)
		parsed.append(row)
	return parsed


def _parse_excel_rows(file_url: str) -> tuple[list[dict], dict[str, int]]:
	try:
		import openpyxl
	except ImportError:
		frappe.throw(
			_(
				"openpyxl is required to read Excel files. Install it in the bench environment: pip install openpyxl"
			)
		)

	path = _excel_file_path(file_url)
	wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
	by_key: dict[str, dict] = {}
	sheet_row_counts: dict[str, int] = {}
	try:
		for sheet_name in wb.sheetnames:
			sheet_rows = _parse_sheet_rows(wb[sheet_name])
			sheet_row_counts[sheet_name] = len(sheet_rows)
			for row in sheet_rows:
				by_key[row["record_key"]] = row
	finally:
		wb.close()
	return list(by_key.values()), sheet_row_counts


def _build_fields(row: dict) -> dict[str, Any]:
	fields: dict[str, Any] = {
		"record_key": row["record_key"],
		"visit_num": row.get("visit_num"),
		"legacy": 1,
		"cr_id": row.get("cr_id") or None,
		"up_id": row.get("up_id") or None,
	}

	visit_date = _parse_date_field(row.get("visit_date"))
	if visit_date:
		fields["visit_date"] = visit_date

	diagnosis = _cell_text(row.get("visit_diagnosis_detail"))
	if diagnosis:
		fields["visit_diagnosis_detail"] = diagnosis

	for date_field in ("cr_date", "up_date"):
		legacy_date = _format_legacy_datetime(row.get(date_field))
		if legacy_date:
			fields[date_field] = legacy_date

	patient = _resolve_patient(row.get("visit_patient_num"))
	if patient:
		fields["patient"] = patient

	patient_visit = _resolve_patient_visit(row.get("visit_num"), patient)
	if patient_visit:
		fields["patient_visit"] = patient_visit

	return {key: value for key, value in fields.items() if value not in (None, "")}


def _preview_counts(rows: list[dict]) -> dict:
	existing = 0
	resolved_patients = 0
	resolved_visits = 0
	with_diagnosis = 0
	for row in rows:
		if frappe.db.exists(DOCTYPE, row["record_key"]):
			existing += 1
		if _resolve_patient(row.get("visit_patient_num")):
			resolved_patients += 1
		patient = _resolve_patient(row.get("visit_patient_num"))
		if _resolve_patient_visit(row.get("visit_num"), patient):
			resolved_visits += 1
		if _cell_text(row.get("visit_diagnosis_detail")):
			with_diagnosis += 1

	return {
		"existing_records": existing,
		"resolved_patients": resolved_patients,
		"resolved_visits": resolved_visits,
		"with_diagnosis": with_diagnosis,
		"sample_record_keys": [row.get("record_key") for row in rows[:5]],
	}


def parse_and_cache_excel(file_url: str) -> dict:
	rows, sheet_row_counts = _parse_excel_rows(file_url)
	row_keys = [row["record_key"] for row in rows]
	by_key = {row["record_key"]: row for row in rows}

	frappe.cache().set_value(CACHE_KEYS["file_url"], file_url, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(CACHE_KEYS["row_keys"], row_keys, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(
		CACHE_KEYS["rows"],
		json.dumps(by_key, default=str),
		expires_in_sec=CACHE_TTL,
	)

	preview = _preview_counts(rows)
	raw_row_total = sum(sheet_row_counts.values())
	return {
		"excel_rows": len(rows),
		"raw_excel_rows": raw_row_total,
		"sheets": list(sheet_row_counts.keys()),
		"sheet_row_counts": sheet_row_counts,
		**preview,
	}


def _load_cached_rows() -> dict[str, dict]:
	raw = frappe.cache().get_value(CACHE_KEYS["rows"])
	if not raw:
		return {}
	if isinstance(raw, dict):
		return raw
	return json.loads(raw)


def upsert_visit_00_01_history(row: dict) -> dict:
	fields = _build_fields(row)
	if not fields.get("record_key"):
		return {"status": "skip", "record_key": row.get("record_key")}

	record_key = fields["record_key"]
	if frappe.db.exists(DOCTYPE, record_key):
		doc = frappe.get_doc(DOCTYPE, record_key)
		for key, value in fields.items():
			doc.set(key, value)
		doc.flags.ignore_mandatory = True
		doc.flags.ignore_links = True
		doc.flags.ignore_validate = True
		doc.save(ignore_permissions=True)
		action = "updated"
	else:
		doc = frappe.get_doc({"doctype": DOCTYPE, **fields})
		doc.flags.ignore_mandatory = True
		doc.flags.ignore_links = True
		doc.flags.ignore_validate = True
		doc.insert(ignore_permissions=True)
		action = "created"

	return {"status": action, "record_key": record_key, "name": doc.name}


@frappe.whitelist()
def preview_visit_00_01_history_import(file_url: str) -> dict:
	_require_admin()
	if not (file_url or "").strip():
		frappe.throw(_("Please upload the VISIT_00_01_HISTORY Excel file."))
	return parse_and_cache_excel(file_url)


def run_visit_00_01_history_import_batch(*, offset: int = 0) -> dict:
	row_keys = frappe.cache().get_value(CACHE_KEYS["row_keys"]) or []
	rows_by_key = _load_cached_rows()
	if not row_keys or not rows_by_key:
		return {"processed": offset, "done": True, "batch_count": 0}

	batch_keys = row_keys[offset : offset + VISIT_00_01_HISTORY_IMPORT_BATCH_SIZE]
	if not batch_keys:
		return {"processed": offset, "done": True, "batch_count": 0}

	created = updated = 0
	errors: list[str] = []

	for key in batch_keys:
		row = rows_by_key.get(key) or {}
		try:
			result = upsert_visit_00_01_history(row)
			status = result.get("status")
			if status == "created":
				created += 1
			elif status == "updated":
				updated += 1
		except Exception:
			errors.append(f"{key}: {frappe.get_traceback()}")
			frappe.log_error(title=f"VISIT_00_01_HISTORY import failed: {key}")

	frappe.db.commit()
	processed = offset + len(batch_keys)
	return {
		"processed": processed,
		"done": len(batch_keys) < VISIT_00_01_HISTORY_IMPORT_BATCH_SIZE,
		"batch_count": len(batch_keys),
		"created": created,
		"updated": updated,
		"errors": len(errors),
	}
