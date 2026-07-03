"""Import Oracle ECT_00_02 Excel into ECT Details Attribute child rows."""

from __future__ import annotations

import json
from typing import Any

import frappe
from frappe import _
from frappe.utils import cint

from healthcare.api.ect_details import _default_ect_template
from healthcare.api.patient_info_import import (
	_cell_text,
	_clean_oracle_num,
	_excel_file_path,
	_parse_date_value,
	_parse_datetime_value,
	_require_admin,
)

DOCTYPE = "ECT Details"
CHILD_FIELD = "ect_details_attributes"
ECT_DETAILS_ATTRIBUTE_IMPORT_BATCH_SIZE = 200
CACHE_TTL = 7200
CACHE_KEYS = {
	"file_url": "healthcare:data_migration:ect_details_attribute_import:file_url",
	"parent_keys": "healthcare:data_migration:ect_details_attribute_import:parent_keys",
	"rows": "healthcare:data_migration:ect_details_attribute_import:rows",
}

EXCEL_HEADER_MAP = {
	"TRANS_NUM": "trans_num",
	"ATTRIB_NUM": "attrib_num",
	"ORDER_OF_ATTRIB": "order_of_attrib",
	"ATT_NOTES": "att_notes",
	"CR_ID": "cr_id",
	"CR_DATE": "cr_date",
	"UP_ID": "up_id",
	"UP_DATE": "up_date",
}


def _normalize_header(cell: Any) -> str:
	if cell is None:
		return ""
	text = str(cell).strip().upper().replace(" ", "_")
	return EXCEL_HEADER_MAP.get(text, text.lower())


def _format_legacy_datetime(value: Any) -> str | None:
	dt = _parse_datetime_value(value)
	if dt:
		return dt.strftime("%Y-%m-%d %H:%M:%S")
	parsed_date = _parse_date_value(value)
	if parsed_date:
		return parsed_date.strftime("%Y-%m-%d")
	text = _cell_text(value)
	return text or None


def _template_attribute_map(template_name: str | None) -> dict[int, dict[str, Any]]:
	template_name = (template_name or "").strip()
	if not template_name or not frappe.db.exists("Patient History Template", template_name):
		return {}
	template = frappe.get_doc("Patient History Template", template_name)
	return {
		cint(row.attrib_num): {
			"attribute": row.attribute or "",
			"order_of_attrib": cint(row.order_no),
		}
		for row in (template.history_detail or [])
		if cint(row.attrib_num)
	}


def _child_index(children) -> dict[int, int]:
	idx: dict[int, int] = {}
	for i, child in enumerate(children or []):
		num = cint(getattr(child, "attrib_num", 0))
		if num:
			idx[num] = i
	return idx


def _parse_sheet_rows(ws) -> list[dict]:
	rows_iter = ws.iter_rows(values_only=True)
	try:
		header_row = next(rows_iter)
	except StopIteration:
		return []

	headers = [_normalize_header(h) for h in header_row]
	parsed: list[dict] = []
	for raw in rows_iter:
		if not raw or all(cell is None or str(cell).strip() == "" for cell in raw):
			continue
		row: dict[str, Any] = {}
		for idx, key in enumerate(headers):
			if not key or idx >= len(raw):
				continue
			row[key] = raw[idx]

		trans_num = _clean_oracle_num(row.get("trans_num"))
		attrib_num = cint(row.get("attrib_num")) if row.get("attrib_num") not in (None, "") else 0
		if not trans_num or not attrib_num:
			continue

		row["trans_num"] = trans_num
		row["attrib_num"] = attrib_num
		row["order_of_attrib"] = cint(row.get("order_of_attrib")) if row.get("order_of_attrib") not in (None, "") else 0
		row["cr_id"] = _clean_oracle_num(row.get("cr_id")) or None
		row["up_id"] = _clean_oracle_num(row.get("up_id")) or None
		parsed.append(row)
	return parsed


def _parse_excel_rows(file_url: str) -> tuple[dict[str, list[dict]], dict[str, int], int]:
	try:
		import openpyxl
	except ImportError:
		frappe.throw(
			_(
				"openpyxl is required to read Excel files. Install it in the bench environment: pip install openpyxl"
			)
		)

	path = _excel_file_path(file_url)
	wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
	grouped: dict[str, list[dict]] = {}
	sheet_row_counts: dict[str, int] = {}
	total_rows = 0
	try:
		for sheet_name in wb.sheetnames:
			sheet_rows = _parse_sheet_rows(wb[sheet_name])
			sheet_row_counts[sheet_name] = len(sheet_rows)
			total_rows += len(sheet_rows)
			for row in sheet_rows:
				grouped.setdefault(row["trans_num"], []).append(row)
	finally:
		wb.close()
	return grouped, sheet_row_counts, total_rows


def _preview_counts(grouped: dict[str, list[dict]], raw_row_count: int) -> dict:
	trans_nums = list(grouped.keys())
	existing_parents = 0
	template_matches = 0
	missing_parents = 0
	template_name = _default_ect_template()
	template_map = _template_attribute_map(template_name)

	for trans_num, rows in grouped.items():
		if frappe.db.exists(DOCTYPE, {"trans_num": trans_num}):
			existing_parents += 1
		else:
			missing_parents += 1
		for row in rows:
			if cint(row.get("attrib_num")) in template_map:
				template_matches += 1

	return {
		"parents": len(grouped),
		"attribute_rows": raw_row_count,
		"existing_parents": existing_parents,
		"missing_parents": missing_parents,
		"matching_template_rows": template_matches,
		"sample_trans_nums": trans_nums[:5],
	}


def parse_and_cache_excel(file_url: str) -> dict:
	grouped, sheet_row_counts, raw_row_count = _parse_excel_rows(file_url)
	parent_keys = sorted(grouped.keys(), key=lambda x: cint(x) if str(x).isdigit() else x)

	frappe.cache().set_value(CACHE_KEYS["file_url"], file_url, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(CACHE_KEYS["parent_keys"], parent_keys, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(
		CACHE_KEYS["rows"],
		json.dumps(grouped, default=str),
		expires_in_sec=CACHE_TTL,
	)

	preview = _preview_counts(grouped, raw_row_count)
	raw_row_total = sum(sheet_row_counts.values())
	return {
		"excel_rows": raw_row_count,
		"raw_excel_rows": raw_row_total,
		"sheets": list(sheet_row_counts.keys()),
		"sheet_row_counts": sheet_row_counts,
		**preview,
	}


def _load_cached_rows() -> dict[str, list[dict]]:
	raw = frappe.cache().get_value(CACHE_KEYS["rows"])
	if not raw:
		return {}
	if isinstance(raw, dict):
		return raw
	return json.loads(raw)


def import_ect_attributes_for_trans(trans_num: str, rows: list[dict]) -> dict:
	if not trans_num or not rows:
		return {"status": "skip_empty", "trans_num": trans_num}

	parent_name = frappe.db.get_value(DOCTYPE, {"trans_num": trans_num}, "name")
	if not parent_name:
		return {"status": "skip_no_parent", "trans_num": trans_num}

	doc = frappe.get_doc(DOCTYPE, parent_name)
	if not (doc.template or "").strip():
		doc.template = _default_ect_template()

	template_map = _template_attribute_map(doc.template)
	doc.flags.ignore_mandatory = True
	doc.flags.ignore_links = True
	doc.flags.ignore_validate = False
	doc.save(ignore_permissions=True)

	idx = _child_index(doc.get(CHILD_FIELD))
	updated = appended = 0

	for row in rows:
		attrib_num = cint(row.get("attrib_num"))
		if not attrib_num:
			continue

		if attrib_num not in idx:
			meta = template_map.get(attrib_num) or {}
			doc.append(
				CHILD_FIELD,
				{
					"attribute": meta.get("attribute") or f"Attribute {attrib_num}",
					"attrib_num": attrib_num,
					"order_of_attrib": cint(row.get("order_of_attrib")) or cint(meta.get("order_of_attrib")),
				},
			)
			idx = _child_index(doc.get(CHILD_FIELD))
			appended += 1

		child = doc.get(CHILD_FIELD)[idx[attrib_num]]
		meta = template_map.get(attrib_num) or {}
		if meta.get("attribute") and not child.attribute:
			child.attribute = meta["attribute"]
		if cint(row.get("order_of_attrib")):
			child.order_of_attrib = cint(row.get("order_of_attrib"))
		elif meta.get("order_of_attrib") and not cint(child.order_of_attrib):
			child.order_of_attrib = cint(meta["order_of_attrib"])

		child.att_notes = _cell_text(row.get("att_notes")) or child.att_notes
		child.cr_id = row.get("cr_id") or child.cr_id
		child.up_id = row.get("up_id") or child.up_id

		cr_date = _format_legacy_datetime(row.get("cr_date"))
		if cr_date:
			child.cr_date = cr_date
		up_date = _format_legacy_datetime(row.get("up_date"))
		if up_date:
			child.up_date = up_date
		updated += 1

	doc.flags.ignore_mandatory = True
	doc.flags.ignore_links = True
	doc.save(ignore_permissions=True)

	return {
		"status": "updated",
		"trans_num": trans_num,
		"name": doc.name,
		"updated_rows": updated,
		"appended_rows": appended,
	}


@frappe.whitelist()
def preview_ect_details_attribute_import(file_url: str) -> dict:
	_require_admin()
	if not (file_url or "").strip():
		frappe.throw(_("Please upload the ECT_00_02 Excel file."))
	return parse_and_cache_excel(file_url)


def run_ect_details_attribute_import_batch(*, offset: int = 0) -> dict:
	parent_keys = frappe.cache().get_value(CACHE_KEYS["parent_keys"]) or []
	rows_by_parent = _load_cached_rows()
	if not parent_keys or not rows_by_parent:
		return {"processed": offset, "done": True, "batch_count": 0}

	batch_keys = parent_keys[offset : offset + ECT_DETAILS_ATTRIBUTE_IMPORT_BATCH_SIZE]
	if not batch_keys:
		return {"processed": offset, "done": True, "batch_count": 0}

	updated = skipped = appended_rows = updated_rows = 0
	errors: list[str] = []

	for key in batch_keys:
		rows = rows_by_parent.get(key) or []
		try:
			result = import_ect_attributes_for_trans(key, rows)
			status = result.get("status")
			if status == "updated":
				updated += 1
				appended_rows += cint(result.get("appended_rows"))
				updated_rows += cint(result.get("updated_rows"))
			else:
				skipped += 1
		except Exception:
			errors.append(f"{key}: {frappe.get_traceback()}")
			frappe.log_error(title=f"ECT_00_02 attribute import failed: {key}")

	frappe.db.commit()
	processed = offset + len(batch_keys)
	return {
		"processed": processed,
		"done": len(batch_keys) < ECT_DETAILS_ATTRIBUTE_IMPORT_BATCH_SIZE,
		"batch_count": len(batch_keys),
		"updated": updated,
		"skipped": skipped,
		"appended_rows": appended_rows,
		"updated_rows": updated_rows,
		"errors": len(errors),
	}
