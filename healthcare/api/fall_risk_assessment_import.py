"""Import Oracle FALL_RISK_ASSESSMENT Excel directly into Fall Risk Assessment rows."""

from __future__ import annotations

import json
from datetime import date, datetime
from typing import Any

import frappe
from frappe import _
from frappe.utils import cint, get_datetime

from healthcare.api.discharge_checklist_import import _resolve_cost_center
from healthcare.api.patient_info_import import (
	_cell_text,
	_clean_oracle_num,
	_excel_file_path,
	_excel_serial_to_datetime,
	_require_admin,
)
from healthcare.api.visit_diagnoses_op_import import _legacy_data_datetime
from healthcare.api.visit_diagnosis_sync import _resolve_inpatient_admission

FALL_RISK_ASSESSMENT_IMPORT_BATCH_SIZE = 500
CACHE_TTL = 7200
CACHE_KEYS = {
	"file_url": "healthcare:data_migration:fall_risk_assessment_import:file_url",
	"row_keys": "healthcare:data_migration:fall_risk_assessment_import:row_keys",
	"rows": "healthcare:data_migration:fall_risk_assessment_import:rows",
}

SELECT_FIELDS = (
	"history",
	"medications",
	"alcohol",
	"activities",
	"balance",
	"disease",
	"poor_eye",
)

REMARKS_FIELDS = (
	"history_remarks",
	"medications_remarks",
	"alcohol_remarks",
	"activities_remarks",
	"balance_remarks",
	"disease_remarks",
	"poor_eye_remarks",
)

EXCEL_HEADER_MAP = {
	"TRANS_NUM": "trans_num",
	"TRANS_DATE": "trans_date",
	"ADMISSION_NUM": "admission_num",
	"HISTORY": "history",
	"MEDICATIONS": "medications",
	"ALCOHOL": "alcohol",
	"ACTIVITIES": "activities",
	"BALANCE": "balance",
	"DISEASE": "disease",
	"POOR_EYE": "poor_eye",
	"CR_ID": "cr_id",
	"CR_DATE": "cr_date",
	"UP_ID": "up_id",
	"UP_DATE": "up_date",
	"BRANCH_NUM": "branch_num",
	"BRANCH": "branch_num",
	"HISTORY_REMARKS": "history_remarks",
	"MEDICATIONS_REMARKS": "medications_remarks",
	"ALCOHOL_REMARKS": "alcohol_remarks",
	"ACTIVITIES_REMARKS": "activities_remarks",
	"BALANCE_REMARKS": "balance_remarks",
	"DISEASE_REMARKS": "disease_remarks",
	"POOR_EYE_REMARKS": "poor_eye_remarks",
}


def _normalize_header(cell: Any) -> str:
	if cell is None:
		return ""
	text = str(cell).strip().upper().replace(" ", "_")
	return EXCEL_HEADER_MAP.get(text, text.lower())


def _select_risk_value(value: Any) -> str | None:
	if value in (None, ""):
		return None
	text = _clean_oracle_num(value) or str(value).strip()
	if text in ("1", "2", "3"):
		return text
	try:
		num = int(float(text))
	except (TypeError, ValueError):
		return None
	if 1 <= num <= 3:
		return str(num)
	return None


def _parse_trans_date(value: Any):
	if value in (None, ""):
		return None
	if isinstance(value, datetime):
		return value
	if isinstance(value, date):
		return datetime.combine(value, datetime.min.time())
	dt = _excel_serial_to_datetime(value)
	if dt:
		return dt
	try:
		return get_datetime(value)
	except Exception:
		return None


def _remarks_value(value: Any) -> str | None:
	text = _cell_text(value)
	return text or None


def _existing_fall_risk_assessment_name(trans_num: int) -> str | None:
	name = str(trans_num)
	if frappe.db.exists("Fall Risk Assessment", name):
		return name
	return frappe.db.get_value("Fall Risk Assessment", {"trans_num": trans_num}, "name")


def _parse_sheet_rows(ws) -> list[dict]:
	rows_iter = ws.iter_rows(values_only=True)
	try:
		header_row = next(rows_iter)
	except StopIteration:
		return []

	headers = [_normalize_header(h) for h in header_row]
	parsed: list[dict] = []
	for raw in rows_iter:
		if not raw or all(cell is None or str(cell).strip() == "" for cell in raw):
			continue
		row: dict[str, Any] = {}
		for idx, key in enumerate(headers):
			if not key or idx >= len(raw):
				continue
			row[key] = raw[idx]

		trans_num_raw = _clean_oracle_num(row.get("trans_num"))
		if not trans_num_raw:
			continue

		row["trans_num"] = cint(trans_num_raw)
		row["admission_num"] = _clean_oracle_num(row.get("admission_num"))
		row["cr_id"] = _clean_oracle_num(row.get("cr_id"))
		row["up_id"] = _clean_oracle_num(row.get("up_id"))
		parsed.append(row)
	return parsed


def _parse_excel_rows(file_url: str) -> tuple[list[dict], dict[str, int]]:
	try:
		import openpyxl
	except ImportError:
		frappe.throw(
			_("openpyxl is required to read Excel files. Install it in the bench environment: pip install openpyxl")
		)

	path = _excel_file_path(file_url)
	wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
	all_rows: list[dict] = []
	sheet_row_counts: dict[str, int] = {}
	try:
		for sheet_name in wb.sheetnames:
			sheet_rows = _parse_sheet_rows(wb[sheet_name])
			sheet_row_counts[sheet_name] = len(sheet_rows)
			all_rows.extend(sheet_rows)
	finally:
		wb.close()
	return all_rows, sheet_row_counts


def _build_fall_risk_assessment_fields(row: dict) -> tuple[dict[str, Any], dict[str, int]]:
	admission_num = row.get("admission_num") or ""
	admission = _resolve_inpatient_admission(admission_num)
	if not admission:
		return {}, {"skip_no_admission": 1}

	trans_date = _parse_trans_date(row.get("trans_date"))
	if not trans_date:
		return {}, {"skip_no_trans_date": 1}

	fields: dict[str, Any] = {
		"trans_num": row["trans_num"],
		"trans_date": trans_date,
		"admission_num": admission,
	}

	cost_center = _resolve_cost_center(row.get("branch_num"))
	if cost_center:
		fields["cost_center"] = cost_center

	for field in SELECT_FIELDS:
		value = _select_risk_value(row.get(field))
		if value is not None:
			fields[field] = value

	for field in REMARKS_FIELDS:
		value = _remarks_value(row.get(field))
		if value:
			fields[field] = value

	if row.get("cr_id"):
		fields["cr_id"] = row["cr_id"]
	if row.get("up_id"):
		fields["up_id"] = row["up_id"]

	cr_date = _legacy_data_datetime(row.get("cr_date"))
	if cr_date:
		fields["cr_date"] = cr_date
	up_date = _legacy_data_datetime(row.get("up_date"))
	if up_date:
		fields["up_date"] = up_date

	return fields, {}


def _apply_legacy_import_flags(doc) -> None:
	doc.flags.ignore_validate = True
	doc.flags.ignore_mandatory = True
	doc.flags.ignore_links = True
	doc.flags.ignore_permissions = True
	doc.flags.legacy_import = True


def _persist_fall_risk_assessment(doc, *, existing: bool) -> None:
	_apply_legacy_import_flags(doc)
	if existing:
		doc.save(ignore_permissions=True)
	else:
		doc.insert(ignore_permissions=True)


def upsert_fall_risk_assessment_from_row(row: dict) -> dict:
	trans_num = row.get("trans_num")
	if not trans_num:
		return {"status": "skip_no_trans_num"}

	fields, side_stats = _build_fall_risk_assessment_fields(row)
	if side_stats.get("skip_no_admission"):
		return {"status": "skip_no_admission", "trans_num": trans_num}
	if side_stats.get("skip_no_trans_date"):
		return {"status": "skip_no_trans_date", "trans_num": trans_num}
	if not fields:
		return {"status": "skip_error", "trans_num": trans_num}

	existing_name = _existing_fall_risk_assessment_name(cint(trans_num))
	if existing_name:
		doc = frappe.get_doc("Fall Risk Assessment", existing_name)
		for key, value in fields.items():
			if key == "trans_num":
				continue
			doc.set(key, value)
		action = "updated"
		existing = True
	else:
		doc = frappe.new_doc("Fall Risk Assessment")
		doc.update(fields)
		action = "created"
		existing = False

	_persist_fall_risk_assessment(doc, existing=existing)
	return {
		"status": action,
		"trans_num": trans_num,
		"name": doc.name,
	}


def _preview_counts(rows: list[dict]) -> dict:
	existing = sum(
		1 for row in rows if _existing_fall_risk_assessment_name(cint(row["trans_num"]))
	)
	resolved_admissions = 0
	unresolved_admissions = 0
	missing_trans_date = 0

	for row in rows:
		if not _parse_trans_date(row.get("trans_date")):
			missing_trans_date += 1
		admission_num = row.get("admission_num") or ""
		if not admission_num:
			unresolved_admissions += 1
			continue
		if _resolve_inpatient_admission(admission_num):
			resolved_admissions += 1
		else:
			unresolved_admissions += 1

	return {
		"existing_assessments": existing,
		"new_assessments": len(rows) - existing,
		"resolved_admissions": resolved_admissions,
		"unresolved_admissions": unresolved_admissions,
		"missing_trans_date": missing_trans_date,
		"sample_trans_nums": [str(row["trans_num"]) for row in rows[:5]],
	}


def parse_and_cache_excel(file_url: str) -> dict:
	rows, sheet_row_counts = _parse_excel_rows(file_url)
	by_key = {str(row["trans_num"]): row for row in rows}
	row_keys = list(by_key.keys())

	frappe.cache().set_value(CACHE_KEYS["file_url"], file_url, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(CACHE_KEYS["row_keys"], row_keys, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(
		CACHE_KEYS["rows"],
		json.dumps(by_key, default=str),
		expires_in_sec=CACHE_TTL,
	)

	preview = _preview_counts(list(by_key.values()))
	raw_row_total = sum(sheet_row_counts.values())
	return {
		"excel_rows": len(by_key),
		"raw_excel_rows": raw_row_total,
		"sheets": list(sheet_row_counts.keys()),
		"sheet_row_counts": sheet_row_counts,
		**preview,
	}


def _load_cached_rows() -> dict[str, dict]:
	raw = frappe.cache().get_value(CACHE_KEYS["rows"])
	if not raw:
		return {}
	if isinstance(raw, dict):
		return raw
	return json.loads(raw)


@frappe.whitelist()
def preview_fall_risk_assessment_import(file_url: str) -> dict:
	_require_admin()
	if not (file_url or "").strip():
		frappe.throw(_("Please upload the FALL_RISK_ASSESSMENT Excel file."))
	return parse_and_cache_excel(file_url)


def run_fall_risk_assessment_import_batch(*, offset: int = 0) -> dict:
	row_keys = frappe.cache().get_value(CACHE_KEYS["row_keys"]) or []
	rows_by_key = _load_cached_rows()
	if not row_keys or not rows_by_key:
		return {"processed": offset, "done": True, "batch_count": 0}

	batch_keys = row_keys[offset : offset + FALL_RISK_ASSESSMENT_IMPORT_BATCH_SIZE]
	if not batch_keys:
		return {"processed": offset, "done": True, "batch_count": 0}

	created = updated = skip_no_admission = skip_no_trans_date = 0
	errors: list[str] = []

	for key in batch_keys:
		row = rows_by_key.get(key) or {}
		try:
			result = upsert_fall_risk_assessment_from_row(row)
			status = result.get("status")
			if status == "created":
				created += 1
			elif status == "updated":
				updated += 1
			elif status == "skip_no_admission":
				skip_no_admission += 1
			elif status == "skip_no_trans_date":
				skip_no_trans_date += 1
		except Exception:
			errors.append(f"{key}: {frappe.get_traceback()}")
			frappe.log_error(title=f"FALL_RISK_ASSESSMENT import failed: {key}")

	frappe.db.commit()
	processed = offset + len(batch_keys)
	return {
		"processed": processed,
		"done": len(batch_keys) < FALL_RISK_ASSESSMENT_IMPORT_BATCH_SIZE,
		"batch_count": len(batch_keys),
		"created": created,
		"updated": updated,
		"skip_no_admission": skip_no_admission,
		"skip_no_trans_date": skip_no_trans_date,
		"errors": len(errors),
	}
