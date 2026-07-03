"""Import SERVICE_RETURN_01 + SERVICE_RETURN_02 into IP Service return records."""

from __future__ import annotations

from datetime import datetime
from typing import Any

import frappe
from frappe import _
from frappe.utils import flt, getdate

from healthcare.api.discharge_checklist_import import _resolve_cost_center
from healthcare.api.patient_info_import import (
	_cell_text,
	_clean_oracle_num,
	_excel_file_path,
	_parse_date_value,
	_parse_datetime_value,
	_require_admin,
)
from healthcare.api.visit_diagnosis_sync import _resolve_patient_visit

DOCTYPE = "IP Service"

HEADER_EXCEL_MAP = {
	"TRANS_NUM": "trans_num",
	"TRANS_DATE": "trans_date",
	"PATIENT_NUM": "patient_num",
	"OP_IP_SOURCE": "op_ip_source",
	"TOTAL_AMOUNT": "total_amount",
	"REMARKS_MASTER": "remarks_master",
	"VCH_STATUS": "vch_status",
	"BRANCH_NUM": "branch_num",
	"CR_ID": "cr_id",
	"CR_DATE": "cr_date",
	"UP_ID": "up_id",
	"UP_DATE": "up_date",
	"AP_ID": "ap_id",
	"AP_DATE": "ap_date",
}

DETAIL_EXCEL_MAP = {
	"TRANS_NUM": "trans_num",
	"SR_NUM": "sr_num",
	"VISIT_NUM": "visit_num",
	"INV_TYPE": "inv_type",
	"INV_NUM": "inv_num",
	"INV_DATE": "inv_date",
	"TOTAL_AMT": "total_amt",
	"ADJUST_AMT": "adjust_amt",
	"DR_GL_CODE": "dr_gl_code",
	"REMARKS_DETAIL": "remarks_detail",
	"CR_ID": "cr_id",
	"CR_DATE": "cr_date",
	"UP_ID": "up_id",
	"UP_DATE": "up_date",
}


def _normalize_header(cell: Any, mapping: dict[str, str]) -> str:
	if cell is None:
		return ""
	text = str(cell).strip().upper().replace(" ", "_")
	return mapping.get(text, text.lower())


def _format_legacy_datetime(value: Any) -> str | None:
	dt = _parse_datetime_value(value)
	if dt:
		return dt.strftime("%Y-%m-%d %H:%M:%S")
	parsed_date = _parse_date_value(value)
	if parsed_date:
		return parsed_date.strftime("%Y-%m-%d")
	text = _cell_text(value)
	return text or None


def _parse_date_field(value: Any):
	dt = _parse_datetime_value(value)
	if dt:
		return getdate(dt)
	parsed_date = _parse_date_value(value)
	if parsed_date:
		return parsed_date
	return None


def _amount(value: Any) -> float:
	if value in (None, ""):
		return 0.0
	return flt(value)


def _resolve_patient(patient_num: Any) -> str | None:
	patient = _clean_oracle_num(patient_num)
	if not patient:
		return None
	if frappe.db.exists("Patient", patient):
		return patient
	return frappe.db.get_value("Patient", {"file_no": patient}, "name")


def _resolve_return_ip_service(inv_num: Any) -> str | None:
	trans_no = _cell_text(inv_num).strip()
	if not trans_no:
		return None
	if frappe.db.exists("IP Service", trans_no):
		return trans_no
	return frappe.db.get_value("IP Service", {"trans_no": trans_no}, "name")


def _sort_detail_lines(detail_lines: list[dict]) -> list[dict]:
	def sort_key(line: dict):
		sr = line.get("sr_num") or ""
		try:
			return float(sr)
		except (TypeError, ValueError):
			return sr

	return sorted(detail_lines, key=sort_key)


def _parse_sheet_rows(ws, mapping: dict[str, str]) -> list[dict]:
	rows_iter = ws.iter_rows(values_only=True)
	try:
		header_row = next(rows_iter)
	except StopIteration:
		return []

	headers = [_normalize_header(h, mapping) for h in header_row]
	parsed: list[dict] = []
	for raw in rows_iter:
		if not raw or all(cell is None or str(cell).strip() == "" for cell in raw):
			continue
		row: dict[str, Any] = {}
		for idx, key in enumerate(headers):
			if not key or idx >= len(raw):
				continue
			row[key] = raw[idx]
		parsed.append(row)
	return parsed


def _load_workbook_rows(file_url: str, mapping: dict[str, str]) -> tuple[list[dict], dict[str, int]]:
	try:
		import openpyxl
	except ImportError:
		frappe.throw(
			_(
				"openpyxl is required to read Excel files. Install it in the bench environment: pip install openpyxl"
			)
		)

	path = _excel_file_path(file_url)
	wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
	all_rows: list[dict] = []
	sheet_row_counts: dict[str, int] = {}
	try:
		for sheet_name in wb.sheetnames:
			sheet_rows = _parse_sheet_rows(wb[sheet_name], mapping)
			sheet_row_counts[sheet_name] = len(sheet_rows)
			all_rows.extend(sheet_rows)
	finally:
		wb.close()
	return all_rows, sheet_row_counts


def _parse_header_excel(file_url: str) -> tuple[list[dict], dict[str, int]]:
	rows, sheet_row_counts = _load_workbook_rows(file_url, HEADER_EXCEL_MAP)
	by_trans: dict[str, dict] = {}
	for row in rows:
		trans_no = _cell_text(row.get("trans_num")).strip()
		if not trans_no:
			continue
		row["trans_no"] = trans_no
		row["patient_num"] = _clean_oracle_num(row.get("patient_num"))
		row["branch_num"] = _clean_oracle_num(row.get("branch_num"))
		row["cr_id"] = _clean_oracle_num(row.get("cr_id"))
		row["up_id"] = _clean_oracle_num(row.get("up_id"))
		row["ap_id"] = _clean_oracle_num(row.get("ap_id"))
		by_trans[trans_no] = row
	return list(by_trans.values()), sheet_row_counts


def _parse_detail_excel(file_url: str) -> tuple[list[dict], dict[str, int]]:
	rows, sheet_row_counts = _load_workbook_rows(file_url, DETAIL_EXCEL_MAP)
	seen: set[tuple[str, str, str]] = set()
	parsed: list[dict] = []
	for row in rows:
		trans_no = _cell_text(row.get("trans_num")).strip()
		inv_num = _cell_text(row.get("inv_num")).strip()
		if not trans_no:
			continue
		sr_num = _clean_oracle_num(row.get("sr_num"))
		dedupe_key = (trans_no, sr_num, inv_num)
		if dedupe_key in seen:
			continue
		seen.add(dedupe_key)
		row["trans_no"] = trans_no
		row["sr_num"] = sr_num
		row["visit_num"] = _clean_oracle_num(row.get("visit_num"))
		row["inv_num"] = inv_num
		row["dr_gl_code"] = _cell_text(row.get("dr_gl_code")).strip()
		row["cr_id"] = _clean_oracle_num(row.get("cr_id"))
		row["up_id"] = _clean_oracle_num(row.get("up_id"))
		parsed.append(row)
	return parsed, sheet_row_counts


def _group_detail_rows(detail_rows: list[dict]) -> dict[str, list[dict]]:
	grouped: dict[str, list[dict]] = {}
	for row in detail_rows:
		grouped.setdefault(row["trans_no"], []).append(row)
	return grouped


def _header_patient_name(header: dict) -> str | None:
	patient = _resolve_patient(header.get("patient_num"))
	if not patient:
		return None
	return frappe.db.get_value("Patient", patient, "patient_name")


def _resolve_unique_patient_visit(header: dict | None, detail_lines: list[dict]) -> str | None:
	header = header or {}
	patient = _resolve_patient(header.get("patient_num"))
	resolved: list[str] = []
	for line in detail_lines:
		visit_num = line.get("visit_num")
		if not visit_num:
			continue
		visit = _resolve_patient_visit(visit_num, patient)
		if visit:
			resolved.append(visit)
	unique = sorted(set(resolved))
	return unique[0] if len(unique) == 1 else None


def _resolve_parent_return_ip_service(detail_lines: list[dict]) -> tuple[str | None, int]:
	resolved = []
	for line in detail_lines:
		match = _resolve_return_ip_service(line.get("inv_num"))
		if match:
			resolved.append(match)
	unique = sorted(set(resolved))
	if len(unique) == 1:
		return unique[0], 1
	return None, len(unique)


def _build_service_rows(detail_lines: list[dict]) -> list[dict]:
	rows: list[dict] = []
	for line in _sort_detail_lines(detail_lines):
		inv_num = _cell_text(line.get("inv_num")).strip()
		total_amt = _amount(line.get("total_amt"))
		adjust_amt = _amount(line.get("adjust_amt"))
		discount_amt = max(total_amt - adjust_amt, 0)
		note = _cell_text(line.get("remarks_detail")).strip() or (
			_("Return against {0}", [inv_num]) if inv_num else _("Return line")
		)
		rows.append(
			{
				"date": _parse_date_field(line.get("inv_date")) or _parse_date_field(line.get("cr_date")),
				"service_name": inv_num or _("Returned Service"),
				"sr_no": line.get("sr_num") or "",
				"inv_type": _cell_text(line.get("inv_type")).strip(),
				"invoice_num": inv_num,
				"amount": adjust_amt or total_amt,
				"total_amount": total_amt or adjust_amt,
				"discount_amount": discount_amt,
				"net_amount": adjust_amt or total_amt,
				"gl_code": line.get("dr_gl_code") or "",
				"note": note,
			}
		)
	return rows


def _apply_header_fields(doc, header: dict, detail_lines: list[dict]) -> tuple[bool, int]:
	doc.trans_no = header["trans_no"]
	doc.file_number = _resolve_patient(header.get("patient_num")) or ""
	doc.patient_full_name = _header_patient_name(header) or ""
	doc.remarks = _cell_text(header.get("remarks_master")).strip()
	doc.trans_source = _cell_text(header.get("op_ip_source")).strip()
	doc.vch_status = _cell_text(header.get("vch_status")).strip()
	doc.cr_id = header.get("cr_id") or ""
	doc.cr_date = _format_legacy_datetime(header.get("cr_date")) or ""
	doc.up_id = header.get("up_id") or ""
	doc.up_date = _format_legacy_datetime(header.get("up_date")) or ""
	doc.ap_date = _format_legacy_datetime(header.get("ap_date")) or ""
	doc.category = doc.category or "Medical Service"
	doc.type = doc.type or "External Service"
	doc.set("return", 1)

	cost_center = _resolve_cost_center(header.get("branch_num"))
	if cost_center:
		doc.cost_center = cost_center

	header_total = _amount(header.get("total_amount"))
	doc.total_amount = header_total
	doc.additional_amount = 0
	doc.discount_amount = 0
	doc.net_amount = header_total

	patient_visit = _resolve_unique_patient_visit(header, detail_lines)
	if patient_visit:
		doc.patient_visit = patient_visit
	else:
		doc.patient_visit = ""

	return_ip_service, return_link_count = _resolve_parent_return_ip_service(detail_lines)
	doc.return_ip_service = return_ip_service or ""
	return bool(return_ip_service), return_link_count


def _preview_stats(headers: list[dict], details_by_trans: dict[str, list[dict]]) -> dict:
	trans_nos = sorted(set(row["trans_no"] for row in headers) | set(details_by_trans.keys()))
	existing_records = 0
	resolvable_patients = 0
	with_detail_lines = 0
	linked_return_ip_service = 0
	multi_return_invoices = 0
	detail_without_header = 0

	for trans_no in trans_nos:
		header = next((row for row in headers if row["trans_no"] == trans_no), None)
		lines = details_by_trans.get(trans_no) or []
		if frappe.db.exists(DOCTYPE, trans_no):
			existing_records += 1
		if header and _resolve_patient(header.get("patient_num")):
			resolvable_patients += 1
		if lines:
			with_detail_lines += 1
		if header is None and lines:
			detail_without_header += 1
		_, link_count = _resolve_parent_return_ip_service(lines)
		if link_count == 1:
			linked_return_ip_service += 1
		elif link_count > 1:
			multi_return_invoices += 1

	return {
		"transactions": len(trans_nos),
		"existing_records": existing_records,
		"new_records": len(trans_nos) - existing_records,
		"resolvable_patients": resolvable_patients,
		"transactions_with_detail_lines": with_detail_lines,
		"transactions_with_linked_return_ip_service": linked_return_ip_service,
		"transactions_with_multiple_return_invoices": multi_return_invoices,
		"detail_without_header": detail_without_header,
		"sample_trans_nos": trans_nos[:5],
	}


def _finalize_service_return_ip_service(doc) -> bool:
	submitted = False
	try:
		doc.flags.ignore_validate = True
		doc.flags.ignore_mandatory = True
		doc.flags.ignore_links = True
		doc.flags.skip_care_episode_guard = True
		doc.flags.from_legacy_import = True
		if doc.docstatus == 0:
			doc.submit()
			submitted = True
	except Exception:
		frappe.log_error(
			title=f"Service Return IP Service submit failed: {doc.name}",
			message=frappe.get_traceback(),
		)
	return submitted


def upsert_service_return_ip_service(header: dict, detail_lines: list[dict]) -> dict:
	trans_no = header.get("trans_no") or ""
	if not trans_no:
		return {"status": "skip_no_trans_no"}
	if not header:
		return {"status": "skip_no_header", "trans_no": trans_no}

	service_rows = _build_service_rows(detail_lines)
	existing = frappe.db.exists(DOCTYPE, trans_no)
	if existing:
		doc = frappe.get_doc(DOCTYPE, trans_no)
		action = "updated"
	else:
		doc = frappe.new_doc(DOCTYPE)
		doc.trans_no = trans_no
		action = "created"

	linked_return_ip_service, return_link_count = _apply_header_fields(doc, header, detail_lines)
	doc.set("services", [])
	for row in service_rows:
		doc.append("services", row)

	doc.flags.ignore_validate = True
	doc.flags.ignore_mandatory = True
	doc.flags.ignore_links = True
	doc.flags.ignore_validate_update_after_submit = True
	doc.flags.skip_care_episode_guard = True
	doc.flags.from_legacy_import = True

	if existing:
		doc.save(ignore_permissions=True)
	else:
		doc.insert(ignore_permissions=True)

	submitted = _finalize_service_return_ip_service(doc)

	return {
		"status": action,
		"trans_no": trans_no,
		"name": doc.name,
		"service_lines": len(service_rows),
		"linked_return_ip_service": linked_return_ip_service,
		"return_link_count": return_link_count,
		"submitted": submitted or doc.docstatus == 1,
	}


@frappe.whitelist()
def preview_service_return_import(header_file_url: str, detail_file_url: str) -> dict:
	_require_admin()
	if not (header_file_url or "").strip():
		frappe.throw(_("Please upload the SERVICE_RETURN_01 Excel file."))
	if not (detail_file_url or "").strip():
		frappe.throw(_("Please upload the SERVICE_RETURN_02 Excel file."))

	headers, header_sheet_counts = _parse_header_excel(header_file_url)
	details, detail_sheet_counts = _parse_detail_excel(detail_file_url)
	stats = _preview_stats(headers, _group_detail_rows(details))
	return {
		"header_rows": len(headers),
		"detail_rows": len(details),
		"header_sheet_row_counts": header_sheet_counts,
		"detail_sheet_row_counts": detail_sheet_counts,
		**stats,
	}


@frappe.whitelist()
def run_service_return_import(header_file_url: str, detail_file_url: str) -> dict:
	"""Import SERVICE_RETURN_01 + SERVICE_RETURN_02 into IP Service synchronously."""
	_require_admin()
	if not (header_file_url or "").strip():
		frappe.throw(_("Please upload the SERVICE_RETURN_01 Excel file."))
	if not (detail_file_url or "").strip():
		frappe.throw(_("Please upload the SERVICE_RETURN_02 Excel file."))

	headers, _ = _parse_header_excel(header_file_url)
	details, _ = _parse_detail_excel(detail_file_url)
	headers_by_trans = {row["trans_no"]: row for row in headers}
	details_by_trans = _group_detail_rows(details)
	trans_nos = sorted(set(headers_by_trans.keys()) | set(details_by_trans.keys()))

	created = updated = skipped = errors = 0
	linked_return_ip_service = 0
	multi_return_invoices = 0
	skip_no_header = 0

	for trans_no in trans_nos:
		header = headers_by_trans.get(trans_no)
		lines = details_by_trans.get(trans_no) or []
		if not header:
			skipped += 1
			skip_no_header += 1
			continue
		try:
			result = upsert_service_return_ip_service(header, lines)
			status = result.get("status")
			if status == "created":
				created += 1
			elif status == "updated":
				updated += 1
			else:
				skipped += 1
			if result.get("linked_return_ip_service"):
				linked_return_ip_service += 1
			elif (result.get("return_link_count") or 0) > 1:
				multi_return_invoices += 1
		except Exception:
			errors += 1
			frappe.log_error(title=f"SERVICE_RETURN import failed: {trans_no}", message=frappe.get_traceback())

	frappe.db.commit()
	return {
		"ok": True,
		"total": len(trans_nos),
		"created": created,
		"updated": updated,
		"skipped": skipped,
		"errors": errors,
		"skip_no_header": skip_no_header,
		"linked_return_ip_service": linked_return_ip_service,
		"multi_return_invoices": multi_return_invoices,
	}
