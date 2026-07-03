"""Import Oracle PATIENT_SICK_LEAVE_01 Excel into Sick Leave rows."""

from __future__ import annotations

from typing import Any

import frappe
from frappe import _
from frappe.utils import flt, getdate

from healthcare.api.ip_admission_discharge_import import _resolve_practitioner_by_doctors_id
from healthcare.api.patient_info_import import (
	_cell_text,
	_clean_oracle_num,
	_excel_file_path,
	_parse_date_value,
	_parse_datetime_value,
	_require_admin,
)
from healthcare.api.patient_visit_practitioner import practitioner_name_from_link
from healthcare.api.visit_diagnosis_sync import _resolve_inpatient_admission

DOCTYPE = "Sick Leave"

EXCEL_HEADER_MAP = {
	"TRANS_NUM": "trans_num",
	"SR_NUM": "sr_num",
	"PATIENT_NUM": "patient_num",
	"ADMISSION_NUM": "admission_num",
	"VISIT_NUM": "visit_num",
	"IP_OP_SOURCE": "ip_op_source",
	"FROM_DATE": "from_date",
	"TO_DATE": "to_date",
	"TOTAL_DAYS": "total_days",
	"DIAGNOSIS_DETAIL": "diagnosis_detail",
	"DOC_NUM": "doc_num",
	"BRANCH_NUM": "branch_num",
	"CR_ID": "cr_id",
	"CR_DATE": "cr_date",
	"UP_ID": "up_id",
	"UP_DATE": "up_date",
}


def _normalize_header(cell: Any) -> str:
	if cell is None:
		return ""
	text = str(cell).strip().upper().replace(" ", "_")
	return EXCEL_HEADER_MAP.get(text, text.lower())


def _format_legacy_datetime(value: Any) -> str | None:
	dt = _parse_datetime_value(value)
	if dt:
		return dt.strftime("%Y-%m-%d %H:%M:%S")
	parsed_date = _parse_date_value(value)
	if parsed_date:
		return parsed_date.strftime("%Y-%m-%d")
	text = _cell_text(value)
	return text or None


def _parse_date_field(value: Any):
	dt = _parse_datetime_value(value)
	if dt:
		return getdate(dt)
	parsed_date = _parse_date_value(value)
	if parsed_date:
		return parsed_date
	return None


def _format_date_text(value: Any) -> str | None:
	parsed = _parse_date_field(value)
	if parsed:
		return parsed.strftime("%Y-%m-%d")
	text = _cell_text(value)
	return text or None


def _resolve_patient(patient_num: Any) -> str | None:
	patient = _clean_oracle_num(patient_num)
	if not patient:
		return None
	if frappe.db.exists("Patient", patient):
		return patient
	return None


def _resolve_admission(admission_num: Any) -> str | None:
	admission = _clean_oracle_num(admission_num)
	if not admission:
		return None
	if frappe.db.exists("Inpatient Admission", admission):
		return admission
	return _resolve_inpatient_admission(admission)


def _parse_sheet_rows(ws) -> list[dict]:
	rows_iter = ws.iter_rows(values_only=True)
	try:
		header_row = next(rows_iter)
	except StopIteration:
		return []

	headers = [_normalize_header(h) for h in header_row]
	parsed: list[dict] = []
	for raw in rows_iter:
		if not raw or all(cell is None or str(cell).strip() == "" for cell in raw):
			continue
		row: dict[str, Any] = {}
		for idx, key in enumerate(headers):
			if not key or idx >= len(raw):
				continue
			row[key] = raw[idx]

		trans_no = _clean_oracle_num(row.get("trans_num"))
		if not trans_no:
			continue
		row["trans_no"] = trans_no
		row["sr_no"] = _clean_oracle_num(row.get("sr_num")) or None
		parsed.append(row)
	return parsed


def _parse_excel_rows(file_url: str) -> tuple[list[dict], dict[str, int]]:
	try:
		import openpyxl
	except ImportError:
		frappe.throw(
			_(
				"openpyxl is required to read Excel files. Install it in the bench environment: pip install openpyxl"
			)
		)

	path = _excel_file_path(file_url)
	wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
	by_key: dict[str, dict] = {}
	sheet_row_counts: dict[str, int] = {}
	try:
		for sheet_name in wb.sheetnames:
			sheet_rows = _parse_sheet_rows(wb[sheet_name])
			sheet_row_counts[sheet_name] = len(sheet_rows)
			for row in sheet_rows:
				by_key[row["trans_no"]] = row
	finally:
		wb.close()
	return list(by_key.values()), sheet_row_counts


def _build_fields(row: dict) -> dict[str, Any]:
	fields: dict[str, Any] = {
		"trans_no": row["trans_no"],
		"sr_no": row.get("sr_no"),
		"source": _cell_text(row.get("ip_op_source")) or None,
		"diagnosis": _cell_text(row.get("diagnosis_detail")) or None,
		"from_sick_01": 1,
		"legacy": 1,
		"cr_id": _clean_oracle_num(row.get("cr_id")) or None,
		"up_id": _clean_oracle_num(row.get("up_id")) or None,
	}

	from_date = _parse_date_field(row.get("from_date"))
	if from_date:
		fields["from_date"] = from_date

	to_date = _format_date_text(row.get("to_date"))
	if to_date:
		fields["to_date"] = to_date

	total_days = row.get("total_days")
	if total_days not in (None, ""):
		fields["days"] = _clean_oracle_num(total_days) or str(int(flt(total_days)))

	for date_field in ("cr_date", "up_date"):
		legacy_date = _format_legacy_datetime(row.get(date_field))
		if legacy_date:
			fields[date_field] = legacy_date

	patient = _resolve_patient(row.get("patient_num"))
	if patient:
		fields["patient"] = patient
		patient_name = frappe.db.get_value("Patient", patient, "patient_name")
		if patient_name:
			fields["patient_name"] = patient_name

	admission = _resolve_admission(row.get("admission_num"))
	if admission:
		fields["admission_no"] = admission

	practitioner = _resolve_practitioner_by_doctors_id(row.get("doc_num"))
	if practitioner:
		fields["doctor"] = practitioner
		doctor_name = practitioner_name_from_link(practitioner)
		if doctor_name:
			fields["doctor_name"] = doctor_name

	return {key: value for key, value in fields.items() if value not in (None, "")}


def _preview_stats(rows: list[dict]) -> dict:
	existing = 0
	resolvable_patients = 0
	resolvable_admissions = 0
	resolvable_doctors = 0
	for row in rows:
		if frappe.db.exists(DOCTYPE, row["trans_no"]):
			existing += 1
		if _resolve_patient(row.get("patient_num")):
			resolvable_patients += 1
		if _resolve_admission(row.get("admission_num")):
			resolvable_admissions += 1
		if _resolve_practitioner_by_doctors_id(row.get("doc_num")):
			resolvable_doctors += 1

	return {
		"existing_records": existing,
		"resolvable_patients": resolvable_patients,
		"resolvable_admissions": resolvable_admissions,
		"resolvable_doctors": resolvable_doctors,
		"sample_trans_nos": [row.get("trans_no") for row in rows[:5]],
	}


def upsert_sick_leave(row: dict) -> dict:
	fields = _build_fields(row)
	if not fields.get("trans_no"):
		return {"status": "skip", "trans_no": row.get("trans_no")}

	trans_no = fields["trans_no"]
	if frappe.db.exists(DOCTYPE, trans_no):
		doc = frappe.get_doc(DOCTYPE, trans_no)
		for key, value in fields.items():
			doc.set(key, value)
		doc.flags.ignore_mandatory = True
		doc.flags.ignore_links = True
		doc.flags.ignore_validate = True
		doc.save(ignore_permissions=True)
		action = "updated"
	else:
		doc = frappe.get_doc({"doctype": DOCTYPE, **fields})
		doc.flags.ignore_mandatory = True
		doc.flags.ignore_links = True
		doc.flags.ignore_validate = True
		doc.insert(ignore_permissions=True)
		action = "created"

	return {"status": action, "trans_no": trans_no, "name": doc.name}


@frappe.whitelist()
def preview_patient_sick_leave_import(file_url: str) -> dict:
	_require_admin()
	if not (file_url or "").strip():
		frappe.throw(_("Please upload the PATIENT_SICK_LEAVE_01 Excel file."))

	rows, sheet_row_counts = _parse_excel_rows(file_url)
	stats = _preview_stats(rows)
	raw_row_total = sum(sheet_row_counts.values())
	return {
		"excel_rows": len(rows),
		"raw_excel_rows": raw_row_total,
		"sheets": list(sheet_row_counts.keys()),
		"sheet_row_counts": sheet_row_counts,
		**stats,
	}


@frappe.whitelist()
def run_patient_sick_leave_import(file_url: str) -> dict:
	"""Import all rows synchronously (small file — no background job)."""
	_require_admin()
	if not (file_url or "").strip():
		frappe.throw(_("Please upload the PATIENT_SICK_LEAVE_01 Excel file."))

	rows, _ = _parse_excel_rows(file_url)
	created = updated = skipped = 0
	errors: list[str] = []

	for row in rows:
		try:
			result = upsert_sick_leave(row)
			status = result.get("status")
			if status == "created":
				created += 1
			elif status == "updated":
				updated += 1
			else:
				skipped += 1
		except Exception:
			errors.append(f"{row.get('trans_no')}: {frappe.get_traceback()}")
			frappe.log_error(title=f"PATIENT_SICK_LEAVE_01 import failed: {row.get('trans_no')}")

	frappe.db.commit()
	return {
		"ok": True,
		"total": len(rows),
		"created": created,
		"updated": updated,
		"skipped": skipped,
		"errors": len(errors),
	}
