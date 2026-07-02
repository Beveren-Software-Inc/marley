"""Import Oracle VISIT_00_03 Excel into legacy Lab Test rows (one lab per visit line)."""

from __future__ import annotations

import json
from typing import Any

import frappe
from frappe import _
from frappe.utils import cint, flt, getdate

from healthcare.api.lab_test_legacy_import import (
	_default_company,
	_format_legacy_date_str,
	_format_legacy_datetime,
	_resolve_cost_center,
	_to_currency,
)
from healthcare.api.patient_info_import import (
	_cell_text,
	_clean_oracle_num,
	_excel_file_path,
	_require_admin,
)
from healthcare.api.service_request_visit_import import ensure_patient_visit_for_legacy_service_import
from healthcare.api.visit_diagnosis_sync import _resolve_patient_visit

LAB_TEST_VISIT_IMPORT_BATCH_SIZE = 500
CACHE_TTL = 7200
CACHE_KEYS = {
	"file_url": "healthcare:data_migration:lab_test_visit_import:file_url",
	"row_keys": "healthcare:data_migration:lab_test_visit_import:row_keys",
	"rows": "healthcare:data_migration:lab_test_visit_import:rows",
}

EXCEL_HEADER_MAP = {
	"VISIT_NUM": "visit_num",
	"SR_NUM": "sr_num",
	"LAB_GROUP_NUM": "lab_group_num",
	"LAB_SUB_NUM": "lab_sub_num",
	"LAB_AMT_BOOK": "lab_amt_book",
	"LAB_AMT_ADD": "lab_amt_add",
	"LAB_AMT_DISC": "lab_amt_disc",
	"LAB_AMT_NET": "lab_amt_net",
	"FIELD1": "field1",
	"FIELD2": "field2",
	"FIELD3": "field3",
	"FIELD4": "field4",
	"FIELD5": "field5",
	"FIELD6": "field6",
	"FIELD7": "field7",
	"FIELD8": "field8",
	"FIELD9": "field9",
	"BRANCH_NUM": "branch_num",
	"CR_ID": "cr_id",
	"CR_DATE": "cr_date",
	"UP_ID": "up_id",
	"UP_DATE": "up_date",
}


def _normalize_header(cell: Any) -> str:
	if cell is None:
		return ""
	text = str(cell).strip().upper().replace(" ", "_")
	return EXCEL_HEADER_MAP.get(text, text.lower())


def _clean_lab_code(value: Any) -> str:
	return _cell_text(value).replace(",", "").strip()


def _clean_visit_num(value: Any) -> str:
	return _clean_oracle_num(value)


def _legacy_trans_num(visit_num: str, sr_num: str) -> str:
	return f"V03/{visit_num}/{sr_num}"


def _row_key(visit_num: str, sr_num: str) -> str:
	return f"{visit_num}::{sr_num}"


def _parse_sheet_rows(ws) -> list[dict]:
	rows_iter = ws.iter_rows(values_only=True)
	try:
		header_row = next(rows_iter)
	except StopIteration:
		return []

	headers = [_normalize_header(h) for h in header_row]
	parsed: list[dict] = []
	for raw in rows_iter:
		if not raw or all(cell is None or str(cell).strip() == "" for cell in raw):
			continue
		row: dict[str, Any] = {}
		for idx, key in enumerate(headers):
			if not key or idx >= len(raw):
				continue
			row[key] = raw[idx]
		visit_num = _clean_visit_num(row.get("visit_num"))
		sr_num = _clean_oracle_num(row.get("sr_num"))
		lab_group = _clean_lab_code(row.get("lab_group_num"))
		lab_sub = _clean_lab_code(row.get("lab_sub_num"))
		if not visit_num or not sr_num or (not lab_group and not lab_sub):
			continue
		row["visit_num"] = visit_num
		row["sr_num"] = sr_num
		row["lab_group_num"] = lab_group
		row["lab_sub_num"] = lab_sub
		row["cr_id"] = _clean_oracle_num(row.get("cr_id"))
		row["up_id"] = _clean_oracle_num(row.get("up_id"))
		parsed.append(row)
	return parsed


def _parse_excel_rows(file_url: str) -> tuple[list[dict], dict[str, int]]:
	try:
		import openpyxl
	except ImportError:
		frappe.throw(
			_("openpyxl is required to read Excel files. Install it in the bench environment: pip install openpyxl")
		)

	path = _excel_file_path(file_url)
	wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
	by_key: dict[str, dict] = {}
	sheet_row_counts: dict[str, int] = {}
	try:
		for sheet_name in wb.sheetnames:
			sheet_rows = _parse_sheet_rows(wb[sheet_name])
			sheet_row_counts[sheet_name] = len(sheet_rows)
			for row in sheet_rows:
				by_key[_row_key(row["visit_num"], row["sr_num"])] = row
	finally:
		wb.close()
	return list(by_key.values()), sheet_row_counts


def _build_lab_template_index() -> dict[str, str]:
	index: dict[str, str] = {}
	for row in frappe.get_all(
		"Lab Test Template",
		fields=["name", "item", "lab_test_code"],
		limit_page_length=0,
	):
		name = (row.name or "").strip()
		if not name:
			continue
		for raw in (row.name, row.item, row.lab_test_code):
			key = _clean_lab_code(raw).upper()
			if key:
				index[key] = name
	return index


def _resolve_lab_template(
	lab_group_num: Any,
	lab_sub_num: Any,
	*,
	template_index: dict[str, str] | None = None,
) -> str | None:
	"""Prefer LAB_GROUP_NUM; fall back to LAB_SUB_NUM for Lab Test Template."""
	candidates = [_clean_lab_code(lab_group_num), _clean_lab_code(lab_sub_num)]
	seen: set[str] = set()
	for code in candidates:
		if not code or code in seen:
			continue
		seen.add(code)
		keys = [code.upper(), code]
		if template_index is not None:
			for key in keys:
				if key in template_index:
					return template_index[key]
		else:
			for candidate in keys:
				if frappe.db.exists("Lab Test Template", candidate):
					return candidate
				for field in ("lab_test_code", "item"):
					name = frappe.db.get_value("Lab Test Template", {field: candidate}, "name")
					if name:
						return name
	return None


def _build_single_lab_test_line(row: dict, *, trans_num: str) -> dict:
	return {
		"trans_num": trans_num,
		"sr_num": row.get("sr_num") or "",
		"lab_group_num": row.get("lab_group_num") or "",
		"lab_sub_num": row.get("lab_sub_num") or "",
		"lab_amt_book": _to_currency(row.get("lab_amt_book")),
		"lab_amt_add": _to_currency(row.get("lab_amt_add")),
		"lab_amt_disc": _to_currency(row.get("lab_amt_disc")),
		"lab_amt_net": _to_currency(row.get("lab_amt_net")),
		"cr_id": row.get("cr_id") or "",
		"cr_date": _format_legacy_date_str(row.get("cr_date")),
		"up_id": row.get("up_id") or "",
		"up_date": _format_legacy_date_str(row.get("up_date")),
	}


def _apply_row_amounts(doc, row: dict) -> None:
	book = _to_currency(row.get("lab_amt_book"))
	add = _to_currency(row.get("lab_amt_add"))
	disc = _to_currency(row.get("lab_amt_disc"))
	net = _to_currency(row.get("lab_amt_net"))
	if net <= 0 and book > 0:
		net = max(book + add - disc, 0)

	doc.amount = book
	doc.discount_margin = "Amount"
	doc.discount_amount = disc
	doc.lab_amount_addition = add
	doc.grand_total = net if net else max(book - disc, 0)


def _apply_row_fields(
	doc,
	row: dict,
	*,
	patient_visit: str,
	patient: str,
	template_name: str | None,
	trans_num: str,
) -> None:
	doc.trans_num = trans_num
	doc.is_legacy_import = 1
	doc.template = template_name
	doc.patient = patient
	doc.patient_visit = patient_visit
	doc.legacy_visit_num = ""
	doc.sr_num = row.get("sr_num") or ""
	doc.cr_id = row.get("cr_id") or ""
	doc.cr_date = _format_legacy_date_str(row.get("cr_date"))
	doc.up_id = row.get("up_id") or ""
	doc.up_date = _format_legacy_date_str(row.get("up_date"))
	doc.branch_num = _clean_oracle_num(row.get("branch_num")) or _cell_text(row.get("branch_num"))

	trans_dt = _format_legacy_datetime(row.get("cr_date"))
	if trans_dt:
		doc.date = getdate(trans_dt)
		doc.time = trans_dt.time()
		doc.result_date = getdate(trans_dt)
		doc.transaction_date = trans_dt.strftime("%Y-%m-%d %H:%M:%S")

	cc = _resolve_cost_center(row.get("branch_num"))
	if cc:
		doc.cost_center = cc

	company = _default_company()
	if company:
		doc.company = company

	if patient and frappe.db.exists("Patient", patient):
		patient_doc = frappe.get_doc("Patient", patient)
		doc.patient_name = patient_doc.patient_name
		doc.patient_sex = patient_doc.sex or doc.patient_sex
		if patient_doc.dob:
			doc.patient_age = patient_doc.age

	if template_name and frappe.db.exists("Lab Test Template", template_name):
		template = frappe.get_doc("Lab Test Template", template_name)
		doc.lab_test_name = template.lab_test_name
		doc.legend_print_position = template.legend_print_position
		doc.result_legend = template.result_legend
		doc.worksheet_instructions = template.worksheet_instructions
	else:
		doc.lab_test_name = row.get("lab_sub_num") or row.get("lab_group_num") or trans_num

	_apply_row_amounts(doc, row)
	doc.status = "Completed"


def _find_existing_lab_test(trans_num: str, patient_visit: str, sr_num: str) -> str | None:
	if frappe.db.exists("Lab Test", trans_num):
		return trans_num
	rows = frappe.get_all(
		"Lab Test",
		filters={"is_legacy_import": 1, "patient_visit": patient_visit, "sr_num": sr_num},
		pluck="name",
		limit=1,
	)
	return rows[0] if rows else None


def _finalize_legacy_lab_test(doc) -> bool:
	submitted = False
	try:
		doc.flags.ignore_validate = True
		doc.flags.ignore_mandatory = True
		doc.flags.legacy_import = True
		if doc.docstatus == 0 and doc.status == "Completed":
			doc.reload()
			doc.flags.ignore_validate = True
			doc.submit()
			submitted = True
	except Exception:
		frappe.log_error(
			title=f"Legacy VISIT_00_03 lab submit failed: {doc.name}",
			message=frappe.get_traceback(),
		)
	return submitted


def upsert_lab_test_from_row(
	row: dict,
	*,
	template_index: dict[str, str] | None = None,
) -> dict:
	visit_num = row.get("visit_num")
	sr_num = row.get("sr_num")
	trans_num = _legacy_trans_num(visit_num, sr_num)

	visit_result = ensure_patient_visit_for_legacy_service_import(
		visit_num,
		order_date=row.get("cr_date"),
	)
	patient_visit = visit_result.get("visit")
	if not patient_visit:
		return {
			"status": visit_result.get("status") or "skip_no_visit",
			"visit_num": visit_num,
			"sr_num": sr_num,
		}

	patient = frappe.db.get_value("Patient Visit", patient_visit, "patient") or visit_num
	template_name = _resolve_lab_template(
		row.get("lab_group_num"),
		row.get("lab_sub_num"),
		template_index=template_index,
	)

	existing = _find_existing_lab_test(trans_num, patient_visit, sr_num)
	if existing:
		doc = frappe.get_doc("Lab Test", existing)
		if not doc.get("is_legacy_import"):
			return {"status": "skip_existing_non_legacy", "trans_num": trans_num}
		action = "updated"
	else:
		doc = frappe.new_doc("Lab Test")
		doc.trans_num = trans_num
		action = "created"

	_apply_row_fields(
		doc,
		row,
		patient_visit=patient_visit,
		patient=patient,
		template_name=template_name,
		trans_num=trans_num,
	)

	lab_line = _build_single_lab_test_line(row, trans_num=trans_num)
	doc.set("lab_test_lines", [])
	doc.append("lab_test_lines", lab_line)

	doc.flags.ignore_validate = True
	doc.flags.ignore_mandatory = True
	doc.flags.ignore_links = True
	doc.flags.legacy_import = True

	if existing:
		doc.save(ignore_permissions=True)
	else:
		doc.insert(ignore_permissions=True)

	submitted = _finalize_legacy_lab_test(doc)

	return {
		"status": action,
		"visit_num": visit_num,
		"sr_num": sr_num,
		"trans_num": trans_num,
		"name": doc.name,
		"template": template_name,
		"visit_created": visit_result.get("created"),
		"patient_created": visit_result.get("patient_created"),
		"submitted": submitted,
	}


def _preview_template_matches(
	rows: list[dict],
	*,
	template_index: dict[str, str] | None = None,
) -> dict:
	groups: set[str] = set()
	subs: set[str] = set()
	matched = 0
	unmatched: list[str] = []
	for row in rows:
		group = _clean_lab_code(row.get("lab_group_num"))
		sub = _clean_lab_code(row.get("lab_sub_num"))
		if group:
			groups.add(group)
		if sub:
			subs.add(sub)
		if _resolve_lab_template(group, sub, template_index=template_index):
			matched += 1
		else:
			code = sub or group
			if code and code not in unmatched:
				unmatched.append(code)

	return {
		"matching_templates": matched,
		"unmatched_row_count": len(rows) - matched,
		"unmatched_codes": unmatched[:25],
		"unique_lab_groups": len(groups),
		"unique_lab_subs": len(subs),
		"sample_lab_sub_nums": sorted(subs)[:10],
	}


def parse_and_cache_excel(file_url: str) -> dict:
	rows, sheet_row_counts = _parse_excel_rows(file_url)
	row_keys = [_row_key(row["visit_num"], row["sr_num"]) for row in rows]
	by_key = {_row_key(row["visit_num"], row["sr_num"]): row for row in rows}

	frappe.cache().set_value(CACHE_KEYS["file_url"], file_url, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(CACHE_KEYS["row_keys"], row_keys, expires_in_sec=CACHE_TTL)
	frappe.cache().set_value(
		CACHE_KEYS["rows"],
		json.dumps(by_key, default=str),
		expires_in_sec=CACHE_TTL,
	)

	visit_nums = {row["visit_num"] for row in rows}
	existing_visits = sum(1 for vn in visit_nums if _resolve_patient_visit(vn))
	template_index = _build_lab_template_index()
	template_stats = _preview_template_matches(rows, template_index=template_index)

	existing_lab = 0
	for row in rows:
		visit = _resolve_patient_visit(row["visit_num"])
		if visit and _find_existing_lab_test(
			_legacy_trans_num(row["visit_num"], row["sr_num"]),
			visit,
			row["sr_num"],
		):
			existing_lab += 1

	raw_row_total = sum(sheet_row_counts.values())
	return {
		"excel_rows": len(rows),
		"raw_excel_rows": raw_row_total,
		"sheets": list(sheet_row_counts.keys()),
		"sheet_row_counts": sheet_row_counts,
		"lab_tests": len(row_keys),
		"unique_visits": len(visit_nums),
		"existing_visits": existing_visits,
		"visits_to_create": len(visit_nums) - existing_visits,
		"existing_lab_tests": existing_lab,
		"matching_templates": template_stats["matching_templates"],
		"unmatched_row_count": template_stats["unmatched_row_count"],
		"unmatched_codes": template_stats["unmatched_codes"],
		"unique_lab_groups": template_stats["unique_lab_groups"],
		"unique_lab_subs": template_stats["unique_lab_subs"],
		"sample_lab_sub_nums": template_stats["sample_lab_sub_nums"],
		"lab_test_templates_in_db": len(template_index),
		"sample_visit_nums": sorted(visit_nums)[:5],
	}


def _load_cached_rows() -> dict[str, dict]:
	raw = frappe.cache().get_value(CACHE_KEYS["rows"])
	if not raw:
		return {}
	if isinstance(raw, dict):
		return raw
	return json.loads(raw)


@frappe.whitelist()
def preview_lab_test_visit_import(file_url: str) -> dict:
	_require_admin()
	if not (file_url or "").strip():
		frappe.throw(_("Please upload the VISIT_00_03 Excel file."))
	return parse_and_cache_excel(file_url)


def run_lab_test_visit_import_batch(offset: int = 0) -> dict:
	row_keys = frappe.cache().get_value(CACHE_KEYS["row_keys"]) or []
	rows_by_key = _load_cached_rows()
	if not row_keys or not rows_by_key:
		return {"processed": offset, "done": True, "batch_count": 0}

	batch_keys = row_keys[offset : offset + LAB_TEST_VISIT_IMPORT_BATCH_SIZE]
	if not batch_keys:
		return {"processed": offset, "done": True, "batch_count": 0}

	created = updated = skip_no_visit = skip_existing = visits_created = patients_created = submitted = 0
	errors: list[str] = []
	template_index = _build_lab_template_index()

	for key in batch_keys:
		row = rows_by_key.get(key) or {}
		try:
			result = upsert_lab_test_from_row(row, template_index=template_index)
			status = result.get("status")
			if status == "created":
				created += 1
			elif status == "updated":
				updated += 1
			elif status in ("skip_no_visit", "skip_no_patient"):
				skip_no_visit += 1
			elif status == "skip_existing_non_legacy":
				skip_existing += 1
			if result.get("visit_created"):
				visits_created += 1
			if result.get("patient_created"):
				patients_created += 1
			if result.get("submitted"):
				submitted += 1
		except Exception:
			errors.append(f"{key}: {frappe.get_traceback()}")
			frappe.log_error(title=f"Lab test VISIT_00_03 import failed: {key}")

	frappe.db.commit()
	processed = offset + len(batch_keys)
	return {
		"processed": processed,
		"done": len(batch_keys) < LAB_TEST_VISIT_IMPORT_BATCH_SIZE,
		"batch_count": len(batch_keys),
		"created": created,
		"updated": updated,
		"skip_no_visit": skip_no_visit,
		"skip_existing_non_legacy": skip_existing,
		"visits_created": visits_created,
		"patients_created": patients_created,
		"submitted": submitted,
		"errors": len(errors),
	}
